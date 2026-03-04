"""
==============================================================================
ANÁLISE EXECUTIVA COMPLETA - RECUPERAÇÃO ALTA 26.1
==============================================================================
Script: 9_Analise_Executiva_Completa.py
Objetivo: Responder perguntas de negócio sobre performance por canal
Autor: Carlos Porto de Melo / Claude AI
Data: 2025-12-11
==============================================================================
"""

import pandas as pd
import numpy as np
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.formatting.rule import DataBarRule

# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CAMINHO_ENTRADA = os.path.join(CAMINHO_BASE, 'Data', 'hubspot_leads.csv')
PASTA_OUTPUT = os.path.join(CAMINHO_BASE, 'Outputs')
os.makedirs(PASTA_OUTPUT, exist_ok=True)

DATA_HOJE = datetime.now().strftime('%Y-%m-%d')
CAMINHO_SAIDA = os.path.join(PASTA_OUTPUT, f'Analise_Executiva_Completa_{DATA_HOJE}.xlsx')

# Estilos
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
ALERT_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
SUCCESS_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
DANGER_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def aplicar_estilo_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

def ajustar_largura_colunas(ws):
    from openpyxl.utils import get_column_letter
    for col_idx, column in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column:
            try:
                if cell.value and not isinstance(cell, type(None)):
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 40)

print(f"[INIT] Gerando Análise Executiva Completa...")
print(f"[INFO] Arquivo de entrada: {CAMINHO_ENTRADA}")

# ==============================================================================
# CARREGAMENTO E PREPARAÇÃO
# ==============================================================================
try:
    df = pd.read_csv(CAMINHO_ENTRADA, encoding='utf-8')
except UnicodeDecodeError:
    df = pd.read_csv(CAMINHO_ENTRADA, encoding='latin1')

df.columns = df.columns.str.strip()
df['Data de criação'] = pd.to_datetime(df['Data de criação'], errors='coerce')
df['Ano'] = df['Data de criação'].dt.year
df['Mes'] = df['Data de criação'].dt.month
df['Mes_Ref'] = df['Data de criação'].dt.to_period('M')

# Flag de matrícula
df['Matricula'] = df['Etapa do negócio'].str.contains('MATRÍCULA', case=False, na=False).astype(int)

# Classificação de canal (simplificada e correta)
def classificar_canal(row):
    fonte = str(row['Fonte original do tráfego']).lower()
    if 'off-line' in fonte:
        return 'Offline (Unidade/Balcão)'
    elif 'social pago' in fonte:
        return 'Meta Ads (Social Pago)'
    elif 'pesquisa paga' in fonte:
        return 'Google Ads (Pesquisa Paga)'
    elif 'pesquisa orgânica' in fonte:
        return 'Google Orgânico (SEO)'
    elif 'social orgânico' in fonte:
        return 'Social Orgânico'
    elif 'tráfego direto' in fonte:
        return 'Tráfego Direto'
    else:
        return 'Outros'

df['Canal'] = df.apply(classificar_canal, axis=1)

print(f"[OK] Dados carregados: {len(df):,} leads")

# ==============================================================================
# ANÁLISES
# ==============================================================================

# 1. PERFORMANCE POR CANAL 2025
df_2025 = df[df['Ano'] == 2025].copy()

perf_canal = df_2025.groupby('Canal').agg(
    Leads=('Canal', 'count'),
    Matriculas=('Matricula', 'sum')
).reset_index()
perf_canal['Taxa_Conversao'] = perf_canal['Matriculas'] / perf_canal['Leads']
perf_canal['Share_Leads'] = perf_canal['Leads'] / perf_canal['Leads'].sum()
perf_canal['Share_Matriculas'] = perf_canal['Matriculas'] / perf_canal['Matriculas'].sum()
perf_canal = perf_canal.sort_values('Matriculas', ascending=False)

# 2. ANÁLISE YOY (Jan-Nov para comparação justa)
df_2024_ytd = df[(df['Ano'] == 2024) & (df['Mes'] <= 11)]
df_2025_ytd = df[(df['Ano'] == 2025) & (df['Mes'] <= 11)]

def calc_metricas(d):
    return d.groupby('Canal').agg(
        Leads=('Canal', 'count'),
        Matriculas=('Matricula', 'sum')
    )

m24 = calc_metricas(df_2024_ytd)
m25 = calc_metricas(df_2025_ytd)

yoy = pd.DataFrame({
    'Leads_2024': m24['Leads'],
    'Mat_2024': m24['Matriculas'],
    'Leads_2025': m25['Leads'],
    'Mat_2025': m25['Matriculas']
}).fillna(0).astype(int)

yoy['Conv_2024'] = yoy['Mat_2024'] / yoy['Leads_2024']
yoy['Conv_2025'] = yoy['Mat_2025'] / yoy['Leads_2025']
yoy['Var_Leads'] = (yoy['Leads_2025'] / yoy['Leads_2024']) - 1
yoy['Var_Mat'] = (yoy['Mat_2025'] / yoy['Mat_2024']) - 1
yoy['Var_Conv_pp'] = yoy['Conv_2025'] - yoy['Conv_2024']
yoy = yoy.sort_values('Mat_2025', ascending=False).reset_index()

# 3. TREND MENSAL OFFLINE
df_offline = df[df['Canal'] == 'Offline (Unidade/Balcão)']
trend_24 = df_offline[df_offline['Ano'] == 2024].groupby('Mes').agg(
    Leads_24=('Mes', 'count'),
    Mat_24=('Matricula', 'sum')
)
trend_25 = df_offline[df_offline['Ano'] == 2025].groupby('Mes').agg(
    Leads_25=('Mes', 'count'),
    Mat_25=('Matricula', 'sum')
)
trend_offline = trend_24.join(trend_25, how='outer').fillna(0).astype(int).reset_index()
trend_offline['Var_Leads'] = (trend_offline['Leads_25'] / trend_offline['Leads_24']) - 1
trend_offline['Var_Mat'] = (trend_offline['Mat_25'] / trend_offline['Mat_24']) - 1
meses = {1:'Jan', 2:'Fev', 3:'Mar', 4:'Abr', 5:'Mai', 6:'Jun', 
         7:'Jul', 8:'Ago', 9:'Set', 10:'Out', 11:'Nov', 12:'Dez'}
trend_offline['Mes_Nome'] = trend_offline['Mes'].map(meses)

# 4. PERFORMANCE POR UNIDADE (OFFLINE)
df_off_25 = df[(df['Canal'] == 'Offline (Unidade/Balcão)') & (df['Ano'] == 2025)]
perf_unidade = df_off_25.groupby('Unidade Desejada').agg(
    Leads=('Unidade Desejada', 'count'),
    Matriculas=('Matricula', 'sum')
).reset_index()
perf_unidade['Taxa_Conversao'] = perf_unidade['Matriculas'] / perf_unidade['Leads']
perf_unidade = perf_unidade.sort_values('Leads', ascending=False)

# Identificar unidades críticas (alta volume, baixa conversão)
unidades_criticas = perf_unidade[
    (perf_unidade['Leads'] >= 100) & 
    (perf_unidade['Taxa_Conversao'] < 0.25)
].sort_values('Leads', ascending=False)

# Identificar unidades top (alta conversão)
unidades_top = perf_unidade[
    (perf_unidade['Leads'] >= 100) & 
    (perf_unidade['Taxa_Conversao'] >= 0.40)
].sort_values('Taxa_Conversao', ascending=False)

# 5. DETALHAMENTO FONTE OFFLINE
detalhe_offline = df_off_25.groupby('Detalhamento da fonte original do tráfego 1').agg(
    Leads=('Detalhamento da fonte original do tráfego 1', 'count'),
    Matriculas=('Matricula', 'sum')
).reset_index()
detalhe_offline['Taxa_Conversao'] = detalhe_offline['Matriculas'] / detalhe_offline['Leads']
detalhe_offline = detalhe_offline.sort_values('Leads', ascending=False)

# 6. ANÁLISE DE GAP - ONDE ATACAR
# Calcular gap de matrículas por canal
totais = yoy[['Canal', 'Mat_2024', 'Mat_2025', 'Var_Mat']].copy()
totais['Gap_Absoluto'] = totais['Mat_2025'] - totais['Mat_2024']
totais['Meta_Recuperacao'] = (totais['Gap_Absoluto'].abs() * (totais['Gap_Absoluto'] < 0)).astype(int)
totais = totais.sort_values('Gap_Absoluto')

# ==============================================================================
# EXPORTAÇÃO EXCEL
# ==============================================================================
print(f"[EXPORTANDO] Gerando Excel com múltiplas abas...")

wb = Workbook()

# --- ABA 1: RESUMO EXECUTIVO ---
ws1 = wb.active
ws1.title = '1_Resumo_Executivo'

# Título
ws1['A1'] = 'ANÁLISE EXECUTIVA - RECUPERAÇÃO ALTA 26.1'
ws1['A1'].font = Font(bold=True, size=16, color='1F4E79')
ws1.merge_cells('A1:G1')

ws1['A3'] = f'Data de Geração: {DATA_HOJE}'
ws1['A4'] = f'Período de Análise: 2024-2025 (YTD até mês {df_2025["Mes"].max()})'

# KPIs principais
ws1['A6'] = 'INDICADORES CHAVE 2025'
ws1['A6'].font = Font(bold=True, size=12)

kpis = [
    ['Métrica', '2024 (YTD)', '2025 (YTD)', 'Variação'],
    ['Total de Leads', yoy['Leads_2024'].sum(), yoy['Leads_2025'].sum(), 
     f"{((yoy['Leads_2025'].sum()/yoy['Leads_2024'].sum())-1)*100:.1f}%"],
    ['Total de Matrículas', yoy['Mat_2024'].sum(), yoy['Mat_2025'].sum(),
     f"{((yoy['Mat_2025'].sum()/yoy['Mat_2024'].sum())-1)*100:.1f}%"],
    ['Taxa de Conversão Geral', f"{yoy['Mat_2024'].sum()/yoy['Leads_2024'].sum()*100:.1f}%",
     f"{yoy['Mat_2025'].sum()/yoy['Leads_2025'].sum()*100:.1f}%",
     f"{((yoy['Mat_2025'].sum()/yoy['Leads_2025'].sum())-(yoy['Mat_2024'].sum()/yoy['Leads_2024'].sum()))*100:.1f}pp"]
]

for i, row in enumerate(kpis, start=7):
    for j, val in enumerate(row, start=1):
        ws1.cell(row=i, column=j, value=val)
        if i == 7:
            ws1.cell(row=i, column=j).fill = HEADER_FILL
            ws1.cell(row=i, column=j).font = HEADER_FONT

# Insights principais
ws1['A13'] = 'PRINCIPAIS INSIGHTS'
ws1['A13'].font = Font(bold=True, size=12)

insights = [
    '⚠️ Meta Ads: Queda de 53% em leads e 49% em matrículas YOY - URGENTE',
    '✅ Offline: Crescimento de 13% em leads, mas queda de 11% em matrículas (conversão caiu 10.8pp)',
    '⚠️ Google Orgânico (SEO): Queda de 32% em leads e 34% em matrículas',
    '✅ Google Ads: Estável (+5.5% leads), boa oportunidade de escala',
    f'🎯 Gap total de matrículas YOY: {totais["Gap_Absoluto"].sum():,} (precisamos recuperar)'
]

for i, insight in enumerate(insights, start=14):
    ws1.cell(row=i, column=1, value=insight)

ajustar_largura_colunas(ws1)

# --- ABA 2: PERFORMANCE POR CANAL ---
ws2 = wb.create_sheet('2_Performance_Canal')

headers = ['Canal', 'Leads', 'Matrículas', 'Taxa Conversão', 'Share Leads', 'Share Matrículas']
for j, h in enumerate(headers, start=1):
    ws2.cell(row=1, column=j, value=h)
aplicar_estilo_header(ws2)

for i, row in perf_canal.iterrows():
    ws2.cell(row=i+2, column=1, value=row['Canal'])
    ws2.cell(row=i+2, column=2, value=row['Leads'])
    ws2.cell(row=i+2, column=3, value=row['Matriculas'])
    ws2.cell(row=i+2, column=4, value=row['Taxa_Conversao'])
    ws2.cell(row=i+2, column=4).number_format = '0.0%'
    ws2.cell(row=i+2, column=5, value=row['Share_Leads'])
    ws2.cell(row=i+2, column=5).number_format = '0.0%'
    ws2.cell(row=i+2, column=6, value=row['Share_Matriculas'])
    ws2.cell(row=i+2, column=6).number_format = '0.0%'

ajustar_largura_colunas(ws2)

# --- ABA 3: ANÁLISE YOY ---
ws3 = wb.create_sheet('3_Analise_YOY')

headers = ['Canal', 'Leads 2024', 'Mat 2024', 'Conv 2024', 'Leads 2025', 'Mat 2025', 'Conv 2025', 
           'Var Leads', 'Var Mat', 'Var Conv (pp)']
for j, h in enumerate(headers, start=1):
    ws3.cell(row=1, column=j, value=h)
aplicar_estilo_header(ws3)

for idx, (i, row) in enumerate(yoy.iterrows(), start=2):
    ws3.cell(row=idx, column=1, value=row['Canal'])
    ws3.cell(row=idx, column=2, value=row['Leads_2024'])
    ws3.cell(row=idx, column=3, value=row['Mat_2024'])
    ws3.cell(row=idx, column=4, value=row['Conv_2024'])
    ws3.cell(row=idx, column=4).number_format = '0.0%'
    ws3.cell(row=idx, column=5, value=row['Leads_2025'])
    ws3.cell(row=idx, column=6, value=row['Mat_2025'])
    ws3.cell(row=idx, column=7, value=row['Conv_2025'])
    ws3.cell(row=idx, column=7).number_format = '0.0%'
    ws3.cell(row=idx, column=8, value=row['Var_Leads'])
    ws3.cell(row=idx, column=8).number_format = '0.0%'
    ws3.cell(row=idx, column=9, value=row['Var_Mat'])
    ws3.cell(row=idx, column=9).number_format = '0.0%'
    ws3.cell(row=idx, column=10, value=row['Var_Conv_pp'])
    ws3.cell(row=idx, column=10).number_format = '0.0%'
    
    # Destacar variações negativas
    if row['Var_Mat'] < -0.1:
        ws3.cell(row=idx, column=9).fill = DANGER_FILL

ajustar_largura_colunas(ws3)

# --- ABA 4: TREND OFFLINE ---
ws4 = wb.create_sheet('4_Trend_Offline_Mensal')

headers = ['Mês', 'Leads 2024', 'Mat 2024', 'Leads 2025', 'Mat 2025', 'Var Leads', 'Var Mat']
for j, h in enumerate(headers, start=1):
    ws4.cell(row=1, column=j, value=h)
aplicar_estilo_header(ws4)

for idx, (i, row) in enumerate(trend_offline.iterrows(), start=2):
    ws4.cell(row=idx, column=1, value=row['Mes_Nome'])
    ws4.cell(row=idx, column=2, value=row['Leads_24'])
    ws4.cell(row=idx, column=3, value=row['Mat_24'])
    ws4.cell(row=idx, column=4, value=row['Leads_25'])
    ws4.cell(row=idx, column=5, value=row['Mat_25'])
    ws4.cell(row=idx, column=6, value=row['Var_Leads'] if pd.notna(row['Var_Leads']) and row['Var_Leads'] != np.inf else 0)
    ws4.cell(row=idx, column=6).number_format = '0.0%'
    ws4.cell(row=idx, column=7, value=row['Var_Mat'] if pd.notna(row['Var_Mat']) and row['Var_Mat'] != np.inf else 0)
    ws4.cell(row=idx, column=7).number_format = '0.0%'

ajustar_largura_colunas(ws4)

# --- ABA 5: UNIDADES CRÍTICAS ---
ws5 = wb.create_sheet('5_Unidades_Criticas')

ws5['A1'] = 'UNIDADES COM BAIXA CONVERSÃO (>100 leads, <25% conversão)'
ws5['A1'].font = Font(bold=True, size=12, color='C00000')
ws5.merge_cells('A1:D1')

headers = ['Unidade', 'Leads', 'Matrículas', 'Taxa Conversão']
for j, h in enumerate(headers, start=1):
    ws5.cell(row=3, column=j, value=h)
aplicar_estilo_header(ws5, row=3)

for idx, (i, row) in enumerate(unidades_criticas.iterrows(), start=4):
    ws5.cell(row=idx, column=1, value=row['Unidade Desejada'])
    ws5.cell(row=idx, column=2, value=row['Leads'])
    ws5.cell(row=idx, column=3, value=row['Matriculas'])
    ws5.cell(row=idx, column=4, value=row['Taxa_Conversao'])
    ws5.cell(row=idx, column=4).number_format = '0.0%'
    ws5.cell(row=idx, column=4).fill = DANGER_FILL

# Adicionar unidades top
start_row = len(unidades_criticas) + 7
ws5.cell(row=start_row, column=1, value='UNIDADES TOP PERFORMANCE (>100 leads, >=40% conversão)')
ws5.cell(row=start_row, column=1).font = Font(bold=True, size=12, color='006400')
ws5.merge_cells(f'A{start_row}:D{start_row}')

for j, h in enumerate(headers, start=1):
    ws5.cell(row=start_row+2, column=j, value=h)
aplicar_estilo_header(ws5, row=start_row+2)

for idx, (i, row) in enumerate(unidades_top.iterrows(), start=start_row+3):
    ws5.cell(row=idx, column=1, value=row['Unidade Desejada'])
    ws5.cell(row=idx, column=2, value=row['Leads'])
    ws5.cell(row=idx, column=3, value=row['Matriculas'])
    ws5.cell(row=idx, column=4, value=row['Taxa_Conversao'])
    ws5.cell(row=idx, column=4).number_format = '0.0%'
    ws5.cell(row=idx, column=4).fill = SUCCESS_FILL

ajustar_largura_colunas(ws5)

# --- ABA 6: RECOMENDAÇÕES ESTRATÉGICAS ---
ws6 = wb.create_sheet('6_Recomendacoes')

ws6['A1'] = 'RECOMENDAÇÕES ESTRATÉGICAS'
ws6['A1'].font = Font(bold=True, size=16, color='1F4E79')

recomendacoes = [
    ['', ''],
    ['PERGUNTA 1: O cadastro da unidade diminuiu, precisamos atacar mais em ações offline?', ''],
    ['', ''],
    ['RESPOSTA:', 'NÃO NECESSARIAMENTE. O volume de leads OFFLINE CRESCEU 13% YOY.'],
    ['', 'O problema é a CONVERSÃO que caiu 10.8 pontos percentuais (de 49.3% para 38.5%).'],
    ['', ''],
    ['DIAGNÓSTICO:', '• Leads estão sendo gerados mas não convertidos'],
    ['', '• 13 unidades têm >100 leads e conversão <25% (ver aba 5)'],
    ['', '• Problema provável: qualidade do atendimento/follow-up, não volume de leads'],
    ['', ''],
    ['AÇÃO RECOMENDADA:', '1. NÃO aumentar captação offline por enquanto'],
    ['', '2. FOCAR em treinamento de secretárias nas unidades críticas'],
    ['', '3. Implementar processo de follow-up estruturado'],
    ['', '4. Benchmark: replicar práticas das unidades TOP (Cuiabá 90.6%, Asa Sul 81.8%)'],
    ['', ''],
    ['', ''],
    ['PERGUNTA 2: Soltamos incentivo de cadastro de leads para secretárias?', ''],
    ['', ''],
    ['RESPOSTA:', 'SIM, MAS COM FOCO DIFERENTE.'],
    ['', ''],
    ['ATUAL:', '• Secretárias já estão cadastrando leads (97.9% via CRM_UI = manual)'],
    ['', '• O volume está crescendo, então o incentivo não deve ser por QUANTIDADE'],
    ['', ''],
    ['PROPOSTA:', '1. INCENTIVO POR CONVERSÃO: bonificar por matrícula fechada, não por lead cadastrado'],
    ['', '2. GAMIFICAÇÃO: ranking mensal de conversão entre unidades'],
    ['', '3. META HÍBRIDA: base fixa + variável por taxa de conversão'],
    ['', ''],
    ['', ''],
    ['ONDE ATACAR PARA BATER A META:', ''],
    ['', ''],
    ['PRIORIDADE 1 (URGENTE):', 'META ADS - Queda de 49% em matrículas'],
    ['', '• Revisar segmentação e criativos'],
    ['', '• Possível exaustão de audiência ou aumento de custos'],
    ['', '• Gap: -133 matrículas vs 2024'],
    ['', ''],
    ['PRIORIDADE 2:', 'OFFLINE - Conversão caindo'],
    ['', '• Focar nas 13 unidades críticas (potencial de +200 matrículas)'],
    ['', '• Unidades prioritárias: SBC, SJC, Morumbi, Santo André'],
    ['', ''],
    ['PRIORIDADE 3:', 'GOOGLE ORGÂNICO - Queda de 32% em leads'],
    ['', '• Revisar estratégia de SEO'],
    ['', '• Verificar posicionamento para termos-chave'],
    ['', ''],
    ['OPORTUNIDADE:', 'GOOGLE ADS - Estável e com boa conversão (5.4%)'],
    ['', '• Canal com potencial de escala'],
    ['', '• Considerar aumento de investimento'],
]

for i, (col1, col2) in enumerate(recomendacoes, start=2):
    ws6.cell(row=i, column=1, value=col1)
    ws6.cell(row=i, column=2, value=col2)
    if 'PERGUNTA' in col1 or 'RESPOSTA:' in col1 or 'PRIORIDADE' in col1 or 'OPORTUNIDADE:' in col1:
        ws6.cell(row=i, column=1).font = Font(bold=True, color='1F4E79')
    if 'DIAGNÓSTICO:' in col1 or 'AÇÃO RECOMENDADA:' in col1 or 'ATUAL:' in col1 or 'PROPOSTA:' in col1:
        ws6.cell(row=i, column=1).font = Font(bold=True)

ws6.column_dimensions['A'].width = 35
ws6.column_dimensions['B'].width = 70

# --- ABA 7: DADOS BRUTOS ---
ws7 = wb.create_sheet('7_Dados_Suporte')

# Performance por unidade (todas)
headers = ['Unidade', 'Leads', 'Matrículas', 'Taxa Conversão']
for j, h in enumerate(headers, start=1):
    ws7.cell(row=1, column=j, value=h)
aplicar_estilo_header(ws7)

for idx, (i, row) in enumerate(perf_unidade.head(50).iterrows(), start=2):
    ws7.cell(row=idx, column=1, value=row['Unidade Desejada'])
    ws7.cell(row=idx, column=2, value=row['Leads'])
    ws7.cell(row=idx, column=3, value=row['Matriculas'])
    ws7.cell(row=idx, column=4, value=row['Taxa_Conversao'])
    ws7.cell(row=idx, column=4).number_format = '0.0%'

ajustar_largura_colunas(ws7)

# Salvar
wb.save(CAMINHO_SAIDA)

print(f"\n{'='*70}")
print(f"[SUCESSO] Análise Executiva gerada: {CAMINHO_SAIDA}")
print(f"{'='*70}")
print("\nABAS DISPONÍVEIS:")
print("1. Resumo Executivo - KPIs e insights principais")
print("2. Performance por Canal - Métricas 2025")
print("3. Análise YOY - Comparativo 2024 vs 2025")
print("4. Trend Offline Mensal - Evolução mês a mês")
print("5. Unidades Críticas - Onde focar ações")
print("6. Recomendações - Respostas às perguntas de negócio")
print("7. Dados Suporte - Performance detalhada por unidade")
