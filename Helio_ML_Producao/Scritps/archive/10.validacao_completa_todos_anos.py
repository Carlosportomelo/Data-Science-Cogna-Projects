"""
==============================================================================
VALIDAÇÃO COMPLETA - TODAS AS FONTES DISCRIMINADAS (SEM "OUTROS")
==============================================================================
Script: 11_Validacao_Fontes_Discriminadas.py
Objetivo: Mostrar TODAS as fontes originais do HubSpot sem agrupamento
Data: 2025-12-11
==============================================================================
"""

import pandas as pd
import numpy as np
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CAMINHO_ENTRADA = os.path.join(CAMINHO_BASE, 'Data', 'hubspot_leads.csv')
PASTA_OUTPUT = os.path.join(CAMINHO_BASE, 'Outputs')
os.makedirs(PASTA_OUTPUT, exist_ok=True)

DATA_HOJE = datetime.now().strftime('%Y-%m-%d')
CAMINHO_SAIDA = os.path.join(PASTA_OUTPUT, f'Validacao_Fontes_Discriminadas_{DATA_HOJE}.xlsx')

# Estilos
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
TOTAL_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def df_to_sheet(ws, df, start_row=1, header=True, format_percent_cols=None, format_number_cols=None):
    """Escreve DataFrame na planilha com formatação"""
    if format_percent_cols is None:
        format_percent_cols = []
    if format_number_cols is None:
        format_number_cols = []
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=header), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            
            if r_idx == start_row and header:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
            else:
                # Formatar colunas de percentual
                if c_idx in format_percent_cols:
                    cell.number_format = '0.00%'
                # Formatar colunas numéricas
                if c_idx in format_number_cols:
                    cell.number_format = '#,##0'

def ajustar_largura(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 40)

def add_total_row(ws, df, start_row):
    """Adiciona linha de total"""
    # Calcular a linha de total considerando se há cabeçalho escrito a partir de start_row
    # Quando header=True e start_row=1, os dados ocupam start_row+1 .. start_row+len(df)
    total_row = start_row + len(df) + 1
    ws.cell(row=total_row, column=1, value='TOTAL')
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    
    for col in range(2, ws.max_column + 1):
        try:
            # Soma apenas colunas numéricas
            col_values = [ws.cell(row=r, column=col).value for r in range(start_row + 1, total_row)]
            col_values = [v for v in col_values if isinstance(v, (int, float)) and not pd.isna(v)]
            if col_values:
                ws.cell(row=total_row, column=col, value=sum(col_values))
        except:
            pass
        ws.cell(row=total_row, column=col).font = Font(bold=True)
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = THIN_BORDER

print(f"[INIT] Gerando Validação com Fontes Discriminadas")
print(f"[INFO] Entrada: {CAMINHO_ENTRADA}")

# ==============================================================================
# CARREGAMENTO
# ==============================================================================
try:
    df = pd.read_csv(CAMINHO_ENTRADA, encoding='utf-8')
except:
    df = pd.read_csv(CAMINHO_ENTRADA, encoding='latin1')

df.columns = df.columns.str.strip()
# Tentar parse robusto das datas: primeiro tentativa padrão, depois com dayfirst=True
raw_dates = df['Data de criação'].astype(str)
parsed_default = pd.to_datetime(raw_dates, errors='coerce')
parsed_dayfirst = pd.to_datetime(raw_dates, dayfirst=True, errors='coerce')
# Preencher com a primeira parse válida encontrada
df['Data de criação'] = parsed_default.fillna(parsed_dayfirst)
df['Ano'] = df['Data de criação'].dt.year
df['Mes'] = df['Data de criação'].dt.month
df['Matricula'] = df['Etapa do negócio'].str.contains('MATRÍCULA', case=False, na=False).astype(int)

# IMPORTANTE: Preencher nulos com identificação clara
df['Fonte original do tráfego'] = df['Fonte original do tráfego'].fillna('(Sem fonte - Nulo)')

# Também tratar o detalhamento
df['Detalhamento da fonte original do tráfego 1'] = df['Detalhamento da fonte original do tráfego 1'].fillna('(Sem detalhamento)')

print(f"[OK] Dados carregados: {len(df):,} leads")

# Verificar fontes
print("\n[CHECK] Fontes únicas encontradas:")
for fonte in df['Fonte original do tráfego'].unique():
    count = len(df[df['Fonte original do tráfego'] == fonte])
    print(f"  - {fonte}: {count:,}")

# ==============================================================================
# CRIAR WORKBOOK
# ==============================================================================
wb = Workbook()

# --- ABA 1: RESUMO GERAL ---
ws1 = wb.active
ws1.title = '1_Resumo_Geral'

resumo = df.groupby('Ano').agg(
    Total_Leads=('Ano', 'count'),
    Matriculas=('Matricula', 'sum')
).reset_index()
resumo['Taxa_Conversao'] = resumo['Matriculas'] / resumo['Total_Leads']

df_to_sheet(ws1, resumo, format_percent_cols=[4], format_number_cols=[2, 3])
add_total_row(ws1, resumo, 1)
ajustar_largura(ws1)

# --- ABA 2: TODAS AS FONTES (TOTAL) ---
ws2 = wb.create_sheet('2_Fontes_Total')

fonte_total = df.groupby('Fonte original do tráfego').agg(
    Total_Leads=('Fonte original do tráfego', 'count'),
    Matriculas=('Matricula', 'sum')
).reset_index()
fonte_total['Taxa_Conversao'] = fonte_total['Matriculas'] / fonte_total['Total_Leads']
fonte_total['Share_Leads'] = fonte_total['Total_Leads'] / fonte_total['Total_Leads'].sum()
fonte_total = fonte_total.sort_values('Total_Leads', ascending=False)
fonte_total.columns = ['Fonte Original (HubSpot)', 'Total Leads', 'Matrículas', 'Taxa Conv', 'Share']

df_to_sheet(ws2, fonte_total, format_percent_cols=[4, 5], format_number_cols=[2, 3])
add_total_row(ws2, fonte_total, 1)
ajustar_largura(ws2)

# --- ABA 3: FONTES POR ANO (LEADS) ---
ws3 = wb.create_sheet('3_Fontes_Ano_Leads')

fonte_ano_leads = df.pivot_table(
    index='Fonte original do tráfego',
    columns='Ano',
    values='Matricula',
    aggfunc='count',
    fill_value=0
).reset_index()
fonte_ano_leads.columns = ['Fonte Original'] + [str(c) for c in fonte_ano_leads.columns[1:]]
fonte_ano_leads['TOTAL'] = fonte_ano_leads.iloc[:, 1:].sum(axis=1)
fonte_ano_leads = fonte_ano_leads.sort_values('TOTAL', ascending=False)

df_to_sheet(ws3, fonte_ano_leads, format_number_cols=list(range(2, len(fonte_ano_leads.columns) + 1)))
add_total_row(ws3, fonte_ano_leads, 1)
ajustar_largura(ws3)

# --- ABA 4: FONTES POR ANO (MATRÍCULAS) ---
ws4 = wb.create_sheet('4_Fontes_Ano_Matriculas')

fonte_ano_mat = df.pivot_table(
    index='Fonte original do tráfego',
    columns='Ano',
    values='Matricula',
    aggfunc='sum',
    fill_value=0
).reset_index()
fonte_ano_mat.columns = ['Fonte Original'] + [str(c) for c in fonte_ano_mat.columns[1:]]
fonte_ano_mat['TOTAL'] = fonte_ano_mat.iloc[:, 1:].sum(axis=1)
fonte_ano_mat = fonte_ano_mat.sort_values('TOTAL', ascending=False)

df_to_sheet(ws4, fonte_ano_mat, format_number_cols=list(range(2, len(fonte_ano_mat.columns) + 1)))
add_total_row(ws4, fonte_ano_mat, 1)
ajustar_largura(ws4)

# --- ABA 5: TAXA DE CONVERSÃO POR FONTE/ANO ---
ws5 = wb.create_sheet('5_Fontes_Ano_Conversao')

conv_data = []
for fonte in df['Fonte original do tráfego'].unique():
    row_data = {'Fonte Original': fonte}
    for ano in sorted(df['Ano'].dropna().unique()):
        subset = df[(df['Fonte original do tráfego'] == fonte) & (df['Ano'] == ano)]
        if len(subset) > 0:
            conv = subset['Matricula'].sum() / len(subset)
        else:
            conv = 0
        row_data[int(ano)] = conv
    conv_data.append(row_data)

conv_df = pd.DataFrame(conv_data)
# Ordenar por volume total
totais = df.groupby('Fonte original do tráfego').size().to_dict()
conv_df['_sort'] = conv_df['Fonte Original'].map(totais)
conv_df = conv_df.sort_values('_sort', ascending=False).drop('_sort', axis=1)

df_to_sheet(ws5, conv_df, format_percent_cols=list(range(2, len(conv_df.columns) + 1)))
ajustar_largura(ws5)

# --- ABA 6: OFFLINE DETALHADO ---
ws6 = wb.create_sheet('6_Offline_Detalhado')

df_offline = df[df['Fonte original do tráfego'] == 'Fontes off-line']

offline_det = df_offline.groupby('Detalhamento da fonte original do tráfego 1').agg(
    Total_Leads=('Detalhamento da fonte original do tráfego 1', 'count'),
    Matriculas=('Matricula', 'sum')
).reset_index()
offline_det['Taxa_Conversao'] = offline_det['Matriculas'] / offline_det['Total_Leads']
offline_det['Share'] = offline_det['Total_Leads'] / offline_det['Total_Leads'].sum()
offline_det = offline_det.sort_values('Total_Leads', ascending=False)
offline_det.columns = ['Detalhamento Fonte', 'Total Leads', 'Matrículas', 'Taxa Conv', 'Share']

df_to_sheet(ws6, offline_det, format_percent_cols=[4, 5], format_number_cols=[2, 3])
add_total_row(ws6, offline_det, 1)
ajustar_largura(ws6)

# --- ABA 7: OFFLINE DETALHAMENTO x ANO ---
ws7 = wb.create_sheet('7_Offline_Det_Ano')

offline_det_ano = df_offline.pivot_table(
    index='Detalhamento da fonte original do tráfego 1',
    columns='Ano',
    values='Matricula',
    aggfunc='count',
    fill_value=0
).reset_index()
offline_det_ano.columns = ['Detalhamento'] + [str(c) for c in offline_det_ano.columns[1:]]
offline_det_ano['TOTAL'] = offline_det_ano.iloc[:, 1:].sum(axis=1)
offline_det_ano = offline_det_ano.sort_values('TOTAL', ascending=False)

df_to_sheet(ws7, offline_det_ano, format_number_cols=list(range(2, len(offline_det_ano.columns) + 1)))
add_total_row(ws7, offline_det_ano, 1)
ajustar_largura(ws7)

# --- ABA 8: OFFLINE POR ANO ---
ws8 = wb.create_sheet('8_Offline_Por_Ano')

offline_ano = df_offline.groupby('Ano').agg(
    Total_Leads=('Ano', 'count'),
    Matriculas=('Matricula', 'sum')
).reset_index()
offline_ano['Taxa_Conversao'] = offline_ano['Matriculas'] / offline_ano['Total_Leads']

df_to_sheet(ws8, offline_ano, format_percent_cols=[4], format_number_cols=[2, 3])
add_total_row(ws8, offline_ano, 1)
ajustar_largura(ws8)

# --- ABA 9: OFFLINE MENSAL POR ANO ---
ws9 = wb.create_sheet('9_Offline_Mensal')

meses_nome = {1:'Jan', 2:'Fev', 3:'Mar', 4:'Abr', 5:'Mai', 6:'Jun', 
              7:'Jul', 8:'Ago', 9:'Set', 10:'Out', 11:'Nov', 12:'Dez'}

offline_mensal = df_offline.pivot_table(
    index='Mes',
    columns='Ano',
    values='Matricula',
    aggfunc='count',
    fill_value=0
).reset_index()
offline_mensal['Mes'] = offline_mensal['Mes'].map(meses_nome)
offline_mensal['TOTAL'] = offline_mensal.iloc[:, 1:].sum(axis=1)

df_to_sheet(ws9, offline_mensal, format_number_cols=list(range(2, len(offline_mensal.columns) + 1)))
add_total_row(ws9, offline_mensal, 1)
ajustar_largura(ws9)

# --- ABA 10: TODAS UNIDADES OFFLINE ---
ws10 = wb.create_sheet('10_Offline_Unidades')

unid_total = df_offline.groupby('Unidade Desejada').agg(
    Total_Leads=('Unidade Desejada', 'count'),
    Matriculas=('Matricula', 'sum')
).reset_index()
unid_total['Taxa_Conversao'] = unid_total['Matriculas'] / unid_total['Total_Leads']
unid_total = unid_total.sort_values('Total_Leads', ascending=False)

df_to_sheet(ws10, unid_total, format_percent_cols=[4], format_number_cols=[2, 3])
ajustar_largura(ws10)

# --- ABA 11: UNIDADES x ANO (TOP 40) ---
ws11 = wb.create_sheet('11_Unidades_Ano')

top_unidades = unid_total.head(40)['Unidade Desejada'].tolist()

unid_ano = df_offline[df_offline['Unidade Desejada'].isin(top_unidades)].pivot_table(
    index='Unidade Desejada',
    columns='Ano',
    values='Matricula',
    aggfunc='count',
    fill_value=0
).reset_index()
unid_ano['TOTAL'] = unid_ano.iloc[:, 1:].sum(axis=1)
unid_ano = unid_ano.sort_values('TOTAL', ascending=False)

df_to_sheet(ws11, unid_ano, format_number_cols=list(range(2, len(unid_ano.columns) + 1)))
ajustar_largura(ws11)

# --- ABA 12: UNIDADES 2025 ---
ws12 = wb.create_sheet('12_Unidades_2025')

df_off_25 = df_offline[df_offline['Ano'] == 2025]
unid_25 = df_off_25.groupby('Unidade Desejada').agg(
    Leads=('Unidade Desejada', 'count'),
    Matriculas=('Matricula', 'sum')
).reset_index()
unid_25['Taxa_Conversao'] = unid_25['Matriculas'] / unid_25['Leads']
unid_25 = unid_25.sort_values('Leads', ascending=False)

df_to_sheet(ws12, unid_25, format_percent_cols=[4], format_number_cols=[2, 3])

# Destacar baixa conversão
for row in range(2, ws12.max_row + 1):
    conv = ws12.cell(row=row, column=4).value
    leads = ws12.cell(row=row, column=2).value
    if conv and leads and conv < 0.25 and leads >= 100:
        for col in range(1, 5):
            ws12.cell(row=row, column=col).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

ajustar_largura(ws12)

# --- ABA 13: SOCIAL PAGO DETALHADO ---
ws13 = wb.create_sheet('13_SocialPago_Detalhado')

df_social = df[df['Fonte original do tráfego'] == 'Social pago']

social_det = df_social.groupby('Detalhamento da fonte original do tráfego 1').agg(
    Total_Leads=('Detalhamento da fonte original do tráfego 1', 'count'),
    Matriculas=('Matricula', 'sum')
).reset_index()
social_det['Taxa_Conversao'] = social_det['Matriculas'] / social_det['Total_Leads']
social_det['Share'] = social_det['Total_Leads'] / social_det['Total_Leads'].sum()
social_det = social_det.sort_values('Total_Leads', ascending=False)
social_det.columns = ['Detalhamento', 'Total Leads', 'Matrículas', 'Taxa Conv', 'Share']

df_to_sheet(ws13, social_det.head(30), format_percent_cols=[4, 5], format_number_cols=[2, 3])
ajustar_largura(ws13)

# --- ABA 14: PESQUISA PAGA DETALHADO ---
ws14 = wb.create_sheet('14_PesquisaPaga_Detalhado')

df_pesquisa = df[df['Fonte original do tráfego'] == 'Pesquisa paga']

pesq_det = df_pesquisa.groupby('Detalhamento da fonte original do tráfego 1').agg(
    Total_Leads=('Detalhamento da fonte original do tráfego 1', 'count'),
    Matriculas=('Matricula', 'sum')
).reset_index()
pesq_det['Taxa_Conversao'] = pesq_det['Matriculas'] / pesq_det['Total_Leads']
pesq_det['Share'] = pesq_det['Total_Leads'] / pesq_det['Total_Leads'].sum()
pesq_det = pesq_det.sort_values('Total_Leads', ascending=False)
pesq_det.columns = ['Detalhamento', 'Total Leads', 'Matrículas', 'Taxa Conv', 'Share']

df_to_sheet(ws14, pesq_det.head(30), format_percent_cols=[4, 5], format_number_cols=[2, 3])
ajustar_largura(ws14)

# --- ABA 15: ETAPAS DO NEGÓCIO ---
ws15 = wb.create_sheet('15_Etapas_Negocio')

etapas = df.groupby('Etapa do negócio').agg(
    Total=('Etapa do negócio', 'count')
).reset_index()
etapas['Share'] = etapas['Total'] / etapas['Total'].sum()
etapas = etapas.sort_values('Total', ascending=False)

df_to_sheet(ws15, etapas, format_percent_cols=[3], format_number_cols=[2])
ajustar_largura(ws15)

# ==============================================================================
# SALVAR (tenta salvar; se arquivo estiver em uso, grava com sufixo timestamp)
# ==============================================================================
try:
    wb.save(CAMINHO_SAIDA)
    saved_path = CAMINHO_SAIDA
except PermissionError:
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    alt_name = f'Validacao_Fontes_Discriminadas_{DATA_HOJE}_{ts}.xlsx'
    alt_path = os.path.join(PASTA_OUTPUT, alt_name)
    try:
        wb.save(alt_path)
        saved_path = alt_path
        print(f"[WARN] Arquivo original estava em uso. Salvo em: {alt_path}")
    except Exception as e:
        print(f"[ERROR] Não foi possível salvar o arquivo: {e}")
        raise

print(f"\n{'='*70}")
print(f"[SUCESSO] Arquivo gerado: {saved_path}")
print(f"{'='*70}")
print("\nABAS DISPONÍVEIS:")
print("1.  Resumo Geral por Ano")
print("2.  Todas as Fontes (Total) - SEM AGRUPAMENTO")
print("3.  Fontes por Ano (Leads)")
print("4.  Fontes por Ano (Matrículas)")
print("5.  Taxa de Conversão por Fonte/Ano")
print("6.  Offline - Detalhamento da Fonte")
print("7.  Offline - Detalhamento x Ano")
print("8.  Offline - Por Ano")
print("9.  Offline - Mensal por Ano")
print("10. Offline - Todas as Unidades")
print("11. Offline - Unidades x Ano (Top 40)")
print("12. Offline - Unidades 2025")
print("13. Social Pago - Detalhamento")
print("14. Pesquisa Paga - Detalhamento")
print("15. Etapas do Negócio")