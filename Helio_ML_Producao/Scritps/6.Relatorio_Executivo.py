"""
==============================================================================
RELATÓRIO EXECUTIVO - DASHBOARD GERENCIAL (C-LEVEL)
==============================================================================
Script: 6.Relatorio_Executivo.py
Objetivo: Gerar visão macro consolidada para diretoria.
          Foco: Potencial Financeiro, Qualidade de Safra e Eficiência de Canal.
Data: 2025-12-19
==============================================================================
"""

import pandas as pd
import os
import glob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASTA_OUTPUT = os.path.join(CAMINHO_BASE, 'Outputs')
PASTA_DADOS_INPUT = os.path.join(PASTA_OUTPUT, 'Dados_Scored')
PASTA_RELATORIOS_EXEC = os.path.join(PASTA_OUTPUT, 'Relatorios_Executivos')

os.makedirs(PASTA_RELATORIOS_EXEC, exist_ok=True)

# --- PREMISSAS FINANCEIRAS ---
TICKET_PROPRIA = 1222.25
TICKET_FRANQUIA = 907.49

# Lista de Unidades Próprias (para garantir classificação)
UNIDADES_PROPRIAS = [
    'VILA LEOPOLDINA', 'SANTANA', 'PINHEIROS', 'PERDIZES', 
    'PACAEMBU', 'MORUMBI', 'JARDINS', 'ITAIM'
]

print("="*80)
print("GERADOR DE RELATÓRIO EXECUTIVO (C-LEVEL)")
print("="*80)

# ==============================================================================
# 1. CARREGAR DADOS
# ==============================================================================
def carregar_dados():
    padrao = os.path.join(PASTA_DADOS_INPUT, 'leads_scored_*.csv')
    arquivos = glob.glob(padrao)
    
    if not arquivos:
        print(f"[ERRO] Nenhum arquivo de dados encontrado em: {PASTA_DADOS_INPUT}")
        return None
    
    arquivo_recente = max(arquivos, key=os.path.getmtime)
    print(f"[LENDO] Base: {os.path.basename(arquivo_recente)}")
    
    df = pd.read_csv(arquivo_recente, sep=';', encoding='utf-8-sig')
    
    # Garantir classificação de unidade
    if 'Tipo_Unidade' not in df.columns:
        df['Tipo_Unidade'] = df['Unidade Desejada'].apply(
            lambda u: 'Própria' if any(p in str(u).upper() for p in UNIDADES_PROPRIAS) else 'Franqueada'
        )
    
    return df

df = carregar_dados()
if df is None:
    exit()

# ==============================================================================
# 2. CÁLCULOS FINANCEIROS (PIPELINE)
# ==============================================================================
print("[PROCESSANDO] Calculando projeções financeiras...")

# Definir Ticket
df['Ticket_Medio'] = df['Tipo_Unidade'].apply(lambda x: TICKET_PROPRIA if x == 'Própria' else TICKET_FRANQUIA)

# Cálculo LTV (Mesma lógica do script de unidades)
taxa_anual = 1.07
fator_ano1 = 12
fator_ano2 = 12 * taxa_anual
fator_ano3 = 12 * (taxa_anual**2)
fator_ano4 = 12 * (taxa_anual**3)
fator_ano5 = 12 * (taxa_anual**4)
fator_ano6_semestral = 6 * (taxa_anual**5)

mult_66m = fator_ano1 + fator_ano2 + fator_ano3 + fator_ano4 + fator_ano5 + fator_ano6_semestral

# Receita Ponderada (Probabilidade * Valor)
df['Pipeline_1M'] = df['Ticket_Medio'] * df['Probabilidade_Conversao']
df['Pipeline_12M'] = df['Pipeline_1M'] * 12
df['Pipeline_LTV'] = df['Pipeline_1M'] * mult_66m

# ==============================================================================
# 3. GERAÇÃO DO EXCEL EXECUTIVO
# ==============================================================================
wb = Workbook()

# Estilos
HEADER_FILL = PatternFill(start_color='800000', end_color='800000', fill_type='solid') # Vermelho Escuro Executivo
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
SUBTOTAL_FILL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
DESC_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
DESC_FONT = Font(italic=True, size=9, color='555555')

# Dicionário de Descrições (Glossário nas colunas)
DICIO_COLUNAS = {
    # KPIs
    'Total Leads': 'Volume total de leads processados na base.',
    'Leads Qualificados (Nota 4-5)': 'Quantidade de leads com score 4 ou 5 (Alta qualidade).',
    '% Qualificados': 'Percentual de leads qualificados sobre o total.',
    'Pipeline (1 Mês)': 'Faturamento projetado para o 1º mês (Ticket x Probabilidade).',
    'Pipeline (LTV Total)': 'Faturamento projetado para todo o ciclo de vida (66 meses).',
    
    # Por Tipo
    'Tipo_Unidade': 'Classificação da unidade (Própria ou Franqueada).',
    'Leads': 'Quantidade de leads recebidos.',
    'Pipeline_1M': 'Soma da Receita Ponderada (1 Mês).',
    'Pipeline_LTV': 'Soma da Receita Ponderada (LTV Total).',
    
    # Por Nota
    'Score (Nota)': 'Nota de qualidade atribuída pelo modelo (1=Baixo, 5=Alto).',
    'Share': 'Participação percentual no volume total.',
    'Potencial_LTV': 'Soma do potencial financeiro (LTV) deste grupo.',
    
    # Por Canal
    'Fonte_Cat': 'Canal de origem do lead (Marketing).',
    'Share_Leads': 'Participação do canal no volume total de leads.',
    'Prob_Media': 'Probabilidade média de conversão dos leads deste canal.',
    
    # Top Unidades
    'Unidade Desejada': 'Unidade de interesse do lead.',
    'Leads_Totais': 'Total de leads recebidos pela unidade.',
    'Leads_Qualificados': 'Quantidade de leads com nota 4 ou 5.',
    'Pipeline_12M': 'Receita projetada para os primeiros 12 meses.',
    
    # Detalhe
    'Qtd': 'Quantidade de leads neste agrupamento.',
    'Ticket_Medio': 'Valor do ticket médio considerado.',
    'Rec_1M': 'Receita projetada (1 Mês).',
    'Rec_12M': 'Receita projetada (12 Meses).',
    'Rec_LTV': 'Receita projetada (LTV Total).'
}

def formatar_planilha(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

def escrever_tabela(ws, df_tabela, titulo, row_start=1):
    ws.cell(row=row_start, column=1, value=titulo).font = Font(bold=True, size=14, color='800000')
    
    # Header
    for c_idx, col_name in enumerate(df_tabela.columns, 1):
        # Título da Coluna
        cell = ws.cell(row=row_start+1, column=c_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Descrição da Coluna (Linha abaixo)
        desc = DICIO_COLUNAS.get(col_name, "")
        cell_desc = ws.cell(row=row_start+2, column=c_idx, value=desc)
        cell_desc.fill = DESC_FILL
        cell_desc.font = DESC_FONT
        cell_desc.border = THIN_BORDER
        cell_desc.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

    # Dados
    for r_idx, row in enumerate(dataframe_to_rows(df_tabela, index=False, header=False), start=row_start+3):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            
            # Formatação inteligente
            if isinstance(val, float):
                if val > 1000: # Provavelmente dinheiro
                    cell.number_format = '#,##0.00'
                elif val < 1: # Provavelmente porcentagem
                    cell.number_format = '0.0%'
                else:
                    cell.number_format = '#,##0.00'
    
    return row_start + len(df_tabela) + 5

# --- ABA 1: DASHBOARD GERAL ---
ws1 = wb.active

# 1.1 Big Numbers
kpis = pd.DataFrame([{
    'Total Leads': len(df),
    'Leads Qualificados (Nota 4-5)': len(df[df['Nota_1a5'] >= 4]),
    '% Qualificados': len(df[df['Nota_1a5'] >= 4]) / len(df),
    'Pipeline (1 Mês)': df['Pipeline_1M'].sum(),
    'Pipeline (LTV Total)': df['Pipeline_LTV'].sum()
}])

row_cursor = escrever_tabela(ws1, kpis, "KPIs GLOBAIS (Pipeline Ponderado)", 1)

# 1.2 Visão por Tipo de Unidade
por_tipo = df.groupby('Tipo_Unidade').agg(
    Leads=('Record ID', 'count'),
    Pipeline_1M=('Pipeline_1M', 'sum'),
    Pipeline_LTV=('Pipeline_LTV', 'sum')
).reset_index()
row_cursor = escrever_tabela(ws1, por_tipo, "VISÃO POR TIPO DE UNIDADE", row_cursor)

# 1.3 Qualidade da Safra (Distribuição d Notas)
por_nota = df.groupby('Nota_1a5').agg(
    Share=('Record ID', lambda x: len(x)/len(df)),
    Potencial_LTV=('Pipeline_LTV', 'sum')
).reset_index().rename(columns={'Nota_1a5': 'Score (Nota)'})
row_cursor = escrever_tabela(ws1, por_nota, "QUALIDADE DA SAFRA (DISTRIBUIÇÃO DE SCORE)", row_cursor)

formatar_planilha(ws1)

# --- ABA 2: MARKETING & CANAIS ---
ws2 = wb.create_sheet("Marketing_Canais")

# 2.1 Performance por Canal
por_canal = df.groupby('Fonte_Cat').agg(
    Leads=('Record ID', 'count'),
    Prob_Media=('Probabilidade_Conversao', 'mean'),
    Pipeline_LTV=('Pipeline_LTV', 'sum')
).reset_index().sort_values('Pipeline_LTV', ascending=False)

# Adicionar Share
por_canal['Share_Leads'] = por_canal['Leads'] / por_canal['Leads'].sum()
# Ordenar colunas
por_canal = por_canal[['Fonte_Cat', 'Leads', 'Share_Leads', 'Prob_Media', 'Pipeline_LTV']]

escrever_tabela(ws2, por_canal, "PERFORMANCE FINANCEIRA POR CANAL DE AQUISIÇÃO", 1)
formatar_planilha(ws2)

# --- ABA 3: RANKING DE OPORTUNIDADES ---
ws3 = wb.create_sheet("Ranking_Oportunidades")
# 3.1 Todas as Unidades por Potencial Financeiro
top_unidades = df.groupby(['Unidade Desejada', 'Tipo_Unidade']).agg(
    Leads_Totais=('Record ID', 'count'),
    Leads_Qualificados=('Nota_1a5', lambda x: (x>=4).sum()),
    Pipeline_12M=('Pipeline_12M', 'sum'),
    Pipeline_LTV=('Pipeline_LTV', 'sum')
).reset_index().sort_values('Pipeline_LTV', ascending=False)

escrever_tabela(ws3, top_unidades, "TODAS AS UNIDADES POR POTENCIAL DE RECEITA (LTV)", 1)
formatar_planilha(ws3)

# --- ABA 4: DETALHE FINANCEIRO (PROJEÇÃO) ---
ws4 = wb.create_sheet("Detalhe_Financeiro")

# Agrupamento completo para controller
detalhe_fin = df.groupby(['Tipo_Unidade', 'Nota_1a5']).agg(
    Qtd=('Record ID', 'count'),
    Ticket_Medio=('Ticket_Medio', 'mean'),
    Rec_1M=('Pipeline_1M', 'sum'),
    Rec_12M=('Pipeline_12M', 'sum'),
    Rec_LTV=('Pipeline_LTV', 'sum')
).reset_index()

escrever_tabela(ws4, detalhe_fin, "DETALHAMENTO FINANCEIRO POR SCORE E TIPO", 1)
formatar_planilha(ws4)

# ==============================================================================
# 4. SALVAR
# ==============================================================================
nome_arquivo = f"Relatorio_Executivo_CLevel_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
caminho_final = os.path.join(PASTA_RELATORIOS_EXEC, nome_arquivo)

try:
    wb.save(caminho_final)
    print(f"\n[SUCESSO] Relatório Executivo gerado: {caminho_final}")
except PermissionError:
    print(f"\n[ERRO] O arquivo está aberto. Feche o Excel e tente novamente.")
    # Tentar salvar com timestamp
    caminho_alt = caminho_final.replace('.xlsx', f'_{datetime.now().strftime("%H%M%S")}.xlsx')
    wb.save(caminho_alt)
    print(f"[SALVO] Salvo como cópia: {caminho_alt}")

print("="*80)
print("RESUMO DOS INDICADORES:")
print(f"Total Pipeline (LTV Ponderado): R$ {df['Pipeline_LTV'].sum():,.2f}")
print(f"Total Leads Qualificados: {len(df[df['Nota_1a5'] >= 4])}")
print("="*80)
