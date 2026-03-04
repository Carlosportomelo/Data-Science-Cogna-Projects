"""
Relatório Gerencial Executivo
Gera um Excel e um resumo Markdown com: visão geral, métricas de performance do modelo,
resumo por segmento, projeção financeira sumarizada e caveats/data quality.
Destinatários: Gerente de Marketing, Gerente Comercial, Diretoria.
"""
import os
import glob
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASTA_DADOS = os.path.join(BASE, 'Outputs', 'Dados_Scored')
# Pasta específica para relatórios gerenciais
PASTA_REPORT = os.path.join(BASE, 'Outputs', 'Relatorio_Gerencial')
os.makedirs(PASTA_REPORT, exist_ok=True)

TICKET_PROPRIA = 1222.25
TICKET_FRANQUIA = 907.49

DATA = datetime.now().strftime('%Y-%m-%d')

# Helper: latest file by pattern
def latest_file(pattern):
    files = glob.glob(pattern)
    return max(files, key=os.path.getctime) if files else None

# Localizar base leads scored
leads_pattern = os.path.join(PASTA_DADOS, 'leads_scored_*.csv')
leads_file = latest_file(leads_pattern)
if not leads_file:
    print('Arquivo leads_scored não encontrado em Outputs/Dados_Scored. Rode o scoring primeiro.')
    raise SystemExit(1)

# Carregar
try:
    df = pd.read_csv(leads_file, sep=';', encoding='utf-8-sig')
except:
    df = pd.read_csv(leads_file, sep=';', encoding='latin1')

# Load history AUC
hist_path = os.path.join(PASTA_DADOS, 'Historico_Evolucao_IA.csv')
auc_latest = None
if os.path.exists(hist_path):
    try:
        df_log = pd.read_csv(hist_path, sep=';', encoding='utf-8-sig')
        if not df_log.empty and 'AUC_Performance' in df_log.columns:
            auc_latest = df_log.iloc[-1]['AUC_Performance']
    except:
        auc_latest = None

# Load blind test (aggregate) if exists
blind_pattern = os.path.join(BASE, 'Outputs', 'archive', 'blind_test_results_*.csv')
blind_file = latest_file(blind_pattern)
blind_metrics = None
if blind_file:
    try:
        blind_metrics = pd.read_csv(blind_file, encoding='utf-8-sig')
    except:
        blind_metrics = None

# Basic stats
total_leads = len(df)
scored_cols = ['Probabilidade_Conversao', 'Nota_1a5']
scored_available = all(c in df.columns for c in scored_cols)

nota_counts = df['Nota_1a5'].value_counts().sort_index() if 'Nota_1a5' in df.columns else pd.Series()
nota_share = (nota_counts / nota_counts.sum()).round(3) if not nota_counts.empty else nota_counts
prob_mean_by_nota = df.groupby('Nota_1a5')['Probabilidade_Conversao'].mean().round(4) if scored_available else pd.Series()

# Segment summary
seg_summary = df.groupby('Segmento_ML').agg(
    Leads=('Record ID', 'count'),
    Prob_Media=('Probabilidade_Conversao', 'mean')
).reset_index()

# Financial projection quick
if 'Probabilidade_Conversao' in df.columns and 'Tipo_Unidade' in df.columns:
    def ticket(tipo):
        return TICKET_PROPRIA if 'Própria' in str(tipo) else TICKET_FRANQUIA
    df['Ticket'] = df['Tipo_Unidade'].apply(ticket)
    df['Esp_1M'] = df['Ticket'] * df['Probabilidade_Conversao']
    proj = df.groupby('Nota_1a5').agg(
        Qtd=('Record ID', 'count'),
        Esp_1M=('Esp_1M', 'sum')
    ).reset_index()
    proj_total = proj['Esp_1M'].sum()
else:
    proj = pd.DataFrame()
    proj_total = 0.0

# Data quality checks
required_cols = ['Record ID', 'Etapa do negócio', 'Data de criação', 'Probabilidade_Conversao', 'Nota_1a5']
missing_cols = [c for c in required_cols if c not in df.columns]
null_rates = (df.isna().mean() * 100).round(2).sort_values(ascending=False)

# Caveats (automatic)
caveats = []
if 'Número de atividades de vendas' in df.columns:
    pct_balcao = (df['Número de atividades de vendas'] < 3).mean() * 100
    caveats.append(f"{pct_balcao:.1f}% dos leads têm <3 atividades (vendas de balcão; não rastreáveis)")
if 'Probabilidade_Conversao' not in df.columns:
    caveats.append('Probabilidade_Conversao ausente — scoring não foi aplicado')
if missing_cols:
    caveats.append('Colunas obrigatórias faltando: ' + ', '.join(missing_cols))
if auc_latest is None:
    caveats.append('AUC não encontrado no histórico — ver Historico_Evolucao_IA.csv')

# Gerar Excel executivo
wb = Workbook()
ws = wb.active
ws.title = 'Resumo_Executivo'

ws['A1'] = 'Relatório Gerencial - Lead Scoring'
ws['A1'].font = Font(bold=True, size=14)
ws['A2'] = f'Data do Relatório: {DATA}'
ws['A3'] = f'Fonte: {os.path.basename(leads_file)}'

# KPIs
ws['A5'] = 'KPIs'
ws['A5'].font = Font(bold=True)
ws['A6'] = f'Total de leads no arquivo: {total_leads:,}'
if auc_latest is not None:
    ws['A7'] = f'AUC (última execução): {auc_latest:.4f}'
else:
    ws['A7'] = 'AUC (última execução): N/A'

if blind_metrics is not None:
    # aggregate global accuracy if present
    try:
        acc_global = blind_metrics.loc[blind_metrics['Segmento']=='GLOBAL','Acuracia'].values
        if acc_global.size > 0:
            ws['A8'] = f'Blind Test (global) - Acurácia: {acc_global[0]:.2%}'
        else:
            # fallback: compute mean
            ws['A8'] = f'Blind Test (por segmento) disponível ({len(blind_metrics)} linhas)'
    except Exception:
        ws['A8'] = f'Blind Test disponível: {os.path.basename(blind_file)}'

# Distribuição de notas (tabela)
start = 10
ws.cell(row=start-1, column=1, value='Distribuição de Notas (1-5)').font = Font(bold=True)
if not nota_counts.empty:
    for i, (nota, cnt) in enumerate(nota_counts.items(), start=start):
        ws.cell(row=i, column=1, value=f'Nota {nota}')
        ws.cell(row=i, column=2, value=int(cnt))
        ws.cell(row=i, column=3, value=float(nota_share.get(nota, 0)))

# Segment sheet
ws_seg = wb.create_sheet('Resumo_Segmentos')
ws_seg.append(['Segmento', 'Leads', 'Prob. Média'])
for r in dataframe_to_rows(seg_summary, index=False, header=False):
    ws_seg.append(r)

# Financial sheet
ws_fin = wb.create_sheet('Financeiro')
ws_fin.append(['Nota', 'Qtd', 'Esp_1M'])
for r in dataframe_to_rows(proj, index=False, header=False):
    ws_fin.append(r)
ws_fin.append([])
ws_fin.append(['Total Estimado 1M', proj_total])

# Caveats sheet
ws_c = wb.create_sheet('Caveats')
ws_c['A1'] = 'Observações e Caveats'
ws_c['A1'].font = Font(bold=True)
for i, txt in enumerate(caveats, start=2):
    ws_c.cell(row=i, column=1, value=txt)

# Save Excel
out_xlsx = os.path.join(PASTA_REPORT, f'Relatorio_Gerencial_{DATA}.xlsx')
wb.save(out_xlsx)

# Save simple markdown summary (legível)
md_lines = []
md_lines.append(f'# Relatório Gerencial - Lead Scoring ({DATA})\n')
md_lines.append(f'- Fonte: {os.path.basename(leads_file)}')
md_lines.append(f'- Total de leads: {total_leads:,}')
if auc_latest is not None:
    md_lines.append(f'- AUC (última execução): {auc_latest:.4f}')
if blind_metrics is not None:
    md_lines.append(f'- Blind test: arquivo {os.path.basename(blind_file)} disponível com {len(blind_metrics):,} linhas')
md_lines.append('\n## Distribuição de Notas')
for nota, cnt in nota_counts.items():
    md_lines.append(f'- Nota {nota}: {cnt:,} ({nota_share.get(nota,0):.1%})')

md_lines.append('\n## Resumo por Segmento')
# usar coluna correta (Segmento_ML) conforme gerado acima
seg_col = 'Segmento_ML' if 'Segmento_ML' in seg_summary.columns else ( 'Segmento' if 'Segmento' in seg_summary.columns else seg_summary.columns[0])
for _, row in seg_summary.iterrows():
    md_lines.append(f"- {row[seg_col]}: {int(row['Leads']):,} leads | Prob. média: {row['Prob_Media']:.2%}")

md_lines.append('\n## Projeção Financeira (1m)')
md_lines.append(f'- Estimativa total (1M): R$ {proj_total:,.2f}')
md_lines.append('\n## Caveats')
for c in caveats:
    md_lines.append(f'- {c}')

out_md = os.path.join(PASTA_REPORT, f'Relatorio_Gerencial_{DATA}.md')
with open(out_md, 'w', encoding='utf-8') as f:
    f.write('\n'.join(md_lines))

print(f'Relatórios gerados:\n - {out_xlsx}\n - {out_md}')
