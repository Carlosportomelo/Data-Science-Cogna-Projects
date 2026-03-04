# -*- coding: utf-8 -*-
"""
SCRIPT DE ANÁLISE DE PERFORMANCE - RED BALLOON
Versão: FINAL - Com classificação EXATA por lista
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / 'outputs'

# ===== CLASSIFICAÇÃO EXATA - SOMENTE ESTAS LISTAS =====
UNIDADES_PROPRIAS = {
    'VILA LEOPOLDINA', 'MORUMBI', 'ITAIM BIBI', 'PACAEMBU', 'PINHEIROS', 
    'JARDINS', 'PERDIZES', 'SANTANA'
}

UNIDADES_FRANQUEADAS = {
    'TATUAPE', 'SALVADOR', 'GUARULHOS', 'JUNDIAI', 'ASA SUL', 'SAO BERNARDO DO CAMPO',
    'SANTO ANDRE', 'CURITIBA', 'SAO CAETANO DO SUL', 'POMPEIA', 'MOOCA', 'BOA VIAGEM',
    'GOIANIA', 'UBERLANDIA', 'TIJUCA', 'ALPHAVILLE', 'LOURDES', 'TAQUARAL',
    'RIBEIRAO PRETO', 'GRACAS', 'PIRACICABA', 'SAO JOSE DOS CAMPOS', 'JACAREPAGUA',
    'IPIRANGA', 'PAMPULHA', 'CASA FORTE', 'CAMPO BELO', 'VILA SAO FRANCISCO',
    'ALDEOTA', 'MOGI DAS CRUZES', 'AGUA FRIA', 'PARAISO', 'MOEMA', 'CUIABA',
    'NOVA CAMPINAS', 'VALINHOS', 'SAUDE', 'ATIBAIA', 'BUTANTA', 'SION',
    'SANTO AMARO', 'GRANJA VIANA', 'SOROCABA', 'LIMEIRA', 'LONDRINA', 'TUCURUVI',
    'AMERICANA', 'INDAIATUBA', 'BARRA JARDIM OCEANICO', 'BAURU', 'PORTO ALEGRE',
    'CHACARA KLABIN', 'ICARAI', 'NOVA LIMA', 'CAMPO GRANDE', 'BOSQUE DA BARRA',
    'VILA MARIANA', 'MANAUS', 'ANAPOLIS', 'BARRA BLUE SQUARE', 'IPANEMA',
    'ARACATUBA', 'REGIAO OCEANICA', 'SAO JOSE DO RIO PRETO', 'ITU'
}

def find_csv():
    """Encontra o arquivo CSV de base em múltiplos locais, procurando por um padrão."""
    paths_to_check = [
        BASE_DIR / 'data',  # Pasta data do projeto
        Path.cwd(),  # Diretório atual
        Path.cwd().parent,  # Diretório pai
        Path.home(),  # Pasta do usuário
        Path.home() / 'Downloads', # Pasta de Downloads
    ]
    
    unique_paths = sorted(list(set(paths_to_check)))
    
    logger.info("Procurando por 'base_leads*.csv' nos diretórios:")
    for p in unique_paths:
        logger.info(f" -> {p}")

    for p in unique_paths:
        if p.exists():
            found_files = list(p.glob('base_leads*.csv'))
            if found_files:
                latest_file = sorted(found_files, key=lambda x: x.stat().st_mtime, reverse=True)[0]
                logger.info(f"✅ Arquivo encontrado: {latest_file}")
                return latest_file
    
    path_list_str = "\n".join(f" -> {p}" for p in unique_paths)
    raise FileNotFoundError(f"Nenhum arquivo 'base_leads*.csv' encontrado. Verifique se o arquivo existe em um dos locais:\n{path_list_str}")

def format_header(ws, row, cols, bg="1F4E78"):
    for col_num, title in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = title
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def format_cells(ws, start_row, end_row, cols_count):
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    for row in range(start_row, end_row + 1):
        for col in range(1, cols_count + 1):
            ws.cell(row=row, column=col).border = border

print("=" * 100)
print("ANÁLISE DE PERFORMANCE - RED BALLOON")
print("=" * 100)

try:
    # Usar arquivo especificado pelo usuário
    input_file = Path(r"C:\Users\a483650\Projetos\_DADOS_CENTRALIZADOS\hubspot\hubspot_leads_ATUAL.csv")
    df = None
    for encoding in ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']:
        try:
            df = pd.read_csv(input_file, sep=',', encoding=encoding)
            logger.info(f"✓ Arquivo lido com {encoding}")
            break
        except Exception as e:
            logger.warning(f"Falha ao ler com {encoding}: {e}")
            continue
    if df is None:
        raise ValueError("Não foi possível ler o arquivo CSV com os encodings testados.")
    df.columns = df.columns.str.strip()
    # Checar colunas obrigatórias
    required_cols = ['Etapa do negócio', 'Unidade Desejada']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Colunas obrigatórias ausentes no arquivo: {missing_cols}\nColunas encontradas: {list(df.columns)}")

    # === Remover arquivos antigos de relatório e dados ===
    import os
    import glob
    # Apagar relatórios antigos
    for f in glob.glob(str(OUTPUT_DIR / 'RELATORIO_PERFORMANCE_*.xlsx')):
        try:
            os.remove(f)
            logger.info(f"Arquivo antigo removido: {f}")
        except Exception as e:
            logger.warning(f"Não foi possível remover {f}: {e}")
    # Apagar arquivos de dados antigos
    for f in glob.glob(str(OUTPUT_DIR / 'DADOS_LOOKER_*.xlsx')):
        try:
            os.remove(f)
            logger.info(f"Arquivo antigo removido: {f}")
        except Exception as e:
            logger.warning(f"Não foi possível remover {f}: {e}")

    # --- CÁLCULO DOS TOTAIS BRUTOS (CENÁRIO TOTAL) ---
    # Normalização temporária para contagem global antes dos filtros
    _etapa_raw = df['Etapa do negócio'].astype(str).str.strip().str.upper()
    raw_conv = _etapa_raw.str.contains('MATRÍCULA CONCLUÍDA', na=False).sum()
    raw_perd = _etapa_raw.str.contains('PERDIDO', na=False).sum()
    raw_total = len(df)

    # ===== REMOVER REGISTROS DE 'BILINGUAL' NA ETAPA DO NEGÓCIO =====
    # Esses registros serão ignorados na geração do relatório final
    if 'Etapa do negócio' in df.columns:
        mask_bilingual = df['Etapa do negócio'].fillna('').astype(str).str.lower().str.contains('biling', na=False)
        count_bilingual = int(mask_bilingual.sum())
        if count_bilingual > 0:
            logger.info(f"Ignorando {count_bilingual} registros com 'bilingual' na coluna 'Etapa do negócio'.")
            df = df[~mask_bilingual].copy()

    # --- CÁLCULOS PARA A ABA DE CONTEXTO (ANTES DE QUALQUER FILTRO) ---
    total_leads_bruto = len(df)
    
    # 1. Contar leads sem unidade designada
    leads_sem_unidade = df['Unidade Desejada'].isna().sum()
    # Normalização para contagem precisa
    # Preenche NaN com vazio, converte para string, remove espaços e põe em maiúsculo
    series_unidade = df['Unidade Desejada'].fillna('').astype(str).str.strip().str.upper()
    
    # 2. Contar leads de unidades não analisadas
    # 1. Registros Incompletos: Campo de unidade vazio ou nulo
    mask_incompletos = (series_unidade == '') | (series_unidade == 'NAN')
    leads_sem_unidade = mask_incompletos.sum()
    
    # 2. Unidades Não Monitoradas: Campo preenchido, mas unidade não está na lista
    UNIDADES_ANALISADAS = UNIDADES_PROPRIAS.union(UNIDADES_FRANQUEADAS)
    unidades_para_analise = df['Unidade Desejada'].dropna().astype(str).str.strip().str.upper()
    
    unidades_encontradas = set(unidades_para_analise.unique())
    lista_unidades_nao_analisadas = sorted(list(unidades_encontradas - UNIDADES_ANALISADAS))
    # Filtra apenas os que têm unidade preenchida para verificar se é válida
    series_completas = series_unidade[~mask_incompletos]
    mask_nao_listadas = ~series_completas.isin(UNIDADES_ANALISADAS)
    count_unidades_nao_analisadas = mask_nao_listadas.sum()
    
    leads_de_unidades_nao_listadas_mask = unidades_para_analise.isin(lista_unidades_nao_analisadas)
    count_unidades_nao_analisadas = int(leads_de_unidades_nao_listadas_mask.sum())
    lista_unidades_nao_analisadas = sorted(series_completas[mask_nao_listadas].unique())
    # --- FIM DOS CÁLCULOS DE CONTEXTO ---

    print("\n" + "=" * 60)
    print("DIAGNÓSTICO DE DADOS (RECONCILIAÇÃO):")
    print(f"1. Total de linhas no arquivo (Bruto): {total_leads_bruto:,.0f}")
    print(f"2. Excluídos - Sem 'Unidade Desejada': -{leads_sem_unidade:,.0f}")
    print(f"3. Excluídos - Unidades fora do escopo: -{count_unidades_nao_analisadas:,.0f}")
    if count_unidades_nao_analisadas > 0:
        print(f"   (Ex: {', '.join(lista_unidades_nao_analisadas[:3])}...)")

    # ===== AGORA, APLICAR FILTROS NO DATAFRAME PRINCIPAL =====
    df = df.dropna(subset=['Unidade Desejada']).copy()
    df['Unidade Desejada'] = df['Unidade Desejada'].astype(str).str.strip().str.upper()
    df = df[df['Unidade Desejada'].isin(UNIDADES_ANALISADAS)].copy()
    
    # ===== VERIFICAÇÃO DE DATAS E REMOÇÃO DE INCONSISTÊNCIAS =====
    col_create = next((c for c in df.columns if c.lower() in ['create date', 'data de criação', 'data de criacao', 'createdate']), None)
    col_close = next((c for c in df.columns if c.lower() in ['close date', 'data de fechamento', 'closedate', 'closed date', 'close_date']), None)
    
    count_date_errors = 0
    date_errors_summary = pd.DataFrame()
    
    if col_create:
        # Converter para datetime
        df['dt_create'] = pd.to_datetime(df[col_create], dayfirst=True, errors='coerce')
        
        if col_close:
            df['dt_close'] = pd.to_datetime(df[col_close], dayfirst=True, errors='coerce')
            
            # Identificar inconsistências: Fechamento < Criação
            mask_valid_dates = df['dt_create'].notna() & df['dt_close'].notna()
            mask_error = mask_valid_dates & (df['dt_close'] < df['dt_create'])
            count_date_errors = mask_error.sum()
            
            if count_date_errors > 0:
                errors_df = df[mask_error].copy()
                date_errors_summary = errors_df.groupby('Unidade Desejada').size().reset_index(name='Qtd_Erros')
                date_errors_summary = date_errors_summary.sort_values('Qtd_Erros', ascending=False)
                
                # REMOVER ERROS DA BASE PRINCIPAL
                df = df[~mask_error].copy()
                logger.info(f"ℹ️ Removidos {count_date_errors} leads com data de fechamento anterior à criação.")

    print(f"4. Excluídos - Inconsistência de Datas: -{count_date_errors:,.0f}")
    print(f"5. Leads Válidos (Base da Análise): {len(df):,.0f}")
    print("=" * 60 + "\n")
    
    logger.info(f"✓ Análise focada em {len(df):,} leads válidos das {len(UNIDADES_ANALISADAS)} unidades especificadas.")

    df['Número de atividades de vendas'] = pd.to_numeric(df['Número de atividades de vendas'], errors='coerce').fillna(0).astype(int)
    
    # ===== APLICAR CLASSIFICAÇÃO EXATA =====
    def classify(unidade):
        if unidade in UNIDADES_PROPRIAS:
            return 'Própria'
        return 'Franqueada'  # Não precisa de 'elif' ou 'else', pois já filtramos
    
    df['Tipo'] = df['Unidade Desejada'].apply(classify)
    
    print(f"\nOK: Unidades Próprias: {len(UNIDADES_PROPRIAS)}")
    print(f"OK: Unidades Franqueadas: {len(UNIDADES_FRANQUEADAS)}")
    print(f"OK: Total: {len(UNIDADES_PROPRIAS) + len(UNIDADES_FRANQUEADAS)}")
    
    # Identificar conversões
    df['Etapa_Norm'] = df['Etapa do negócio'].astype(str).str.strip().str.upper()
    df['Convertido'] = df['Etapa_Norm'].str.contains('MATRÍCULA CONCLUÍDA', na=False).astype(int)
    df['Perdido'] = df['Etapa_Norm'].str.contains('PERDIDO', na=False).astype(int)
    df = df.drop('Etapa_Norm', axis=1)
    
    # Análise por unidade
    analise = df.groupby(['Tipo', 'Unidade Desejada']).agg({
        'Record ID': 'count',
        'Convertido': 'sum',
        'Perdido': 'sum',
        'Número de atividades de vendas': 'mean',
        'Proprietário do negócio': lambda x: x.mode()[0] if len(x.mode()) > 0 else 'N/A'
    }).rename(columns={'Record ID': 'Total', 'Convertido': 'Matriculas', 'Perdido': 'Perdidos',
                       'Número de atividades de vendas': 'Ativ_Media_Total', 'Proprietário do negócio': 'Responsável'})
    
    ativ_media_mat = df[df['Convertido'] == 1].groupby(['Tipo', 'Unidade Desejada'])['Número de atividades de vendas'].mean()
    ativ_media_perd = df[df['Perdido'] == 1].groupby(['Tipo', 'Unidade Desejada'])['Número de atividades de vendas'].mean()
    analise = analise.join(ativ_media_mat.rename('Ativ_Media_Mat')).fillna({'Ativ_Media_Mat': 0})
    analise = analise.join(ativ_media_perd.rename('Ativ_Media_Perd')).fillna({'Ativ_Media_Perd': 0})
    
    analise['Taxa_Conv'] = (analise['Matriculas'] / analise['Total'] * 100).round(1)
    analise['Taxa_Perda'] = (analise['Perdidos'] / analise['Total'] * 100).round(1)
    analise['Leads_p_Matricula'] = (analise['Total'] / analise['Matriculas']).round(1)
    
    # ===== ANÁLISE POR CANAL (FONTE DE TRÁFEGO) =====
    col_fonte = None
    
    # 1. Busca Exata (Case Insensitive)
    cols_map = {c.lower(): c for c in df.columns}
    candidates = [
        'fonte original', 'original source', 'original source type',
        'latest source', 'ultima fonte', 'última fonte',
        'fonte', 'origem', 'canal', 'source', 'midia', 'mídia'
    ]
    
    for c in candidates:
        if c in cols_map:
            col_fonte = cols_map[c]
            break
            
    # 2. Busca Parcial (se não encontrou exata) - Procura palavras chave no nome da coluna
    if not col_fonte:
        keywords = ['source', 'fonte', 'origem']
        for col in df.columns:
            col_lower = col.lower()
            if any(k in col_lower for k in keywords):
                col_fonte = col
                break
    
    # ===== PREPARAR DADOS PARA O LOOKER =====
    logger.info("Preparando dataframe para exportação (Looker)...")
    
    df_looker = df.copy()
    
    # Adicionar períodos de tempo, se a data de criação existir
    if 'dt_create' in df_looker.columns:
        df_looker['Ano_Mes_Criacao'] = df_looker['dt_create'].dt.to_period('M').astype(str)
        df_looker['Ano_Semana_Criacao'] = df_looker['dt_create'].dt.to_period('W').astype(str)
        # Adicionar coluna Semana_nome: S1, S2, ...
        def semana_nome(dt):
            if pd.isnull(dt):
                return ''
            dia = dt.day
            return f"S{((dia - 1) // 7) + 1}"
        df_looker['Semana_nome'] = df_looker['dt_create'].apply(semana_nome)
    
    # Selecionar e renomear colunas
    colunas_looker = {
        'Record ID': 'ID_Lead',
        'dt_create': 'Data_Criacao',
        'dt_close': 'Data_Fechamento',
        'Unidade Desejada': 'Unidade',
        'Tipo': 'Tipo_Unidade',
        'Etapa do negócio': 'Etapa_Negocio',
        'Proprietário do negócio': 'Proprietario_Negocio',
        'Número de atividades de vendas': 'Num_Atividades_Venda',
        'Convertido': 'Flag_Convertido',
        'Perdido': 'Flag_Perdido',
        'Ano_Mes_Criacao': 'Ano_Mes_Criacao',
        'Ano_Semana_Criacao': 'Ano_Semana_Criacao',
        'Semana_nome': 'Semana_nome'
    }
    # Adicionar a coluna de fonte dinamicamente se ela foi encontrada
    if col_fonte:
        colunas_looker[col_fonte] = 'Fonte_Lead'

    # Manter apenas colunas que de fato existem no dataframe para evitar erros
    colunas_presentes = {k: v for k, v in colunas_looker.items() if k in df_looker.columns}
    df_looker = df_looker[list(colunas_presentes.keys())].rename(columns=colunas_presentes)
    
    # Assegurar que colunas de data não tenham timezone para a escrita no Excel
    if 'Data_Criacao' in df_looker.columns:
        df_looker['Data_Criacao'] = pd.to_datetime(df_looker['Data_Criacao']).dt.tz_localize(None)
    if 'Data_Fechamento' in df_looker.columns:
        df_looker['Data_Fechamento'] = pd.to_datetime(df_looker['Data_Fechamento']).dt.tz_localize(None)


    # ===== RESUMO POR TIPO - AQUI É O CRÍTICO =====
    resumo = df.groupby('Tipo').agg({
        'Unidade Desejada': 'nunique',
        'Record ID': 'count',
        'Convertido': 'sum',
        'Perdido': 'sum',
        'Número de atividades de vendas': 'mean'
    }).rename(columns={
        'Unidade Desejada': 'Qtd_Unidades',
        'Record ID': 'Total_Leads',
        'Convertido': 'Matriculas',
        'Perdido': 'Perdidos',
        'Número de atividades de vendas': 'Ativ_Media_Total'
    })
    
    ativ_media_mat_tipo = df[df['Convertido'] == 1].groupby('Tipo')['Número de atividades de vendas'].mean()
    ativ_media_perd_tipo = df[df['Perdido'] == 1].groupby('Tipo')['Número de atividades de vendas'].mean()
    resumo = resumo.join(ativ_media_mat_tipo.rename('Ativ_Media_Mat')).fillna({'Ativ_Media_Mat': 0})
    resumo = resumo.join(ativ_media_perd_tipo.rename('Ativ_Media_Perd')).fillna({'Ativ_Media_Perd': 0})
    
    resumo['Taxa_Conv_%'] = (resumo['Matriculas'] / resumo['Total_Leads'] * 100).round(1)
    resumo['Taxa_Perda_%'] = (resumo['Perdidos'] / resumo['Total_Leads'] * 100).round(1)
    resumo['Leads_p_Matricula'] = (resumo['Total_Leads'] / resumo['Matriculas']).round(1)
    
    # ===== CONTINUAÇÃO ANÁLISE POR CANAL =====
    if col_fonte:
        logger.info(f"✓ Coluna de fonte de tráfego identificada: {col_fonte}")
        df[col_fonte] = df[col_fonte].fillna('Não Identificado').astype(str).str.strip()
        
        analise_canais = df.groupby(col_fonte).agg({
            'Record ID': 'count',
            'Convertido': 'sum',
            'Perdido': 'sum',
            'Número de atividades de vendas': 'mean'
        }).rename(columns={'Record ID': 'Total', 'Convertido': 'Matriculas', 'Perdido': 'Perdidos', 'Número de atividades de vendas': 'Ativ_Media_Total'})
        
        ativ_media_mat_canal = df[df['Convertido'] == 1].groupby(col_fonte)['Número de atividades de vendas'].mean()
        ativ_media_perd_canal = df[df['Perdido'] == 1].groupby(col_fonte)['Número de atividades de vendas'].mean()
        analise_canais = analise_canais.join(ativ_media_mat_canal.rename('Ativ_Media_Mat')).fillna({'Ativ_Media_Mat': 0})
        analise_canais = analise_canais.join(ativ_media_perd_canal.rename('Ativ_Media_Perd')).fillna({'Ativ_Media_Perd': 0})
        
        analise_canais['Taxa_Conv'] = (analise_canais['Matriculas'] / analise_canais['Total'] * 100).round(1)
        analise_canais['Taxa_Perda'] = (analise_canais['Perdidos'] / analise_canais['Total'] * 100).round(1)
        analise_canais['Leads_p_Matricula'] = (analise_canais['Total'] / analise_canais['Matriculas']).round(1)
        analise_canais = analise_canais.sort_values('Total', ascending=False)
    else:
        print("\nAVISO: Coluna de fonte de tráfego NÃO encontrada.")
        print("Colunas disponíveis no arquivo:")
        for c in sorted(df.columns):
            print(f" - {c}")
        logger.warning("Nenhuma coluna de fonte de tráfego encontrada.")
        analise_canais = pd.DataFrame()

    # ===== ANÁLISE TEMPORAL (JÁ COM DADOS LIMPOS) =====
    timeline_data_mensal = pd.DataFrame()
    timeline_data_semanal = pd.DataFrame()
    timeline_unidade_mensal = pd.DataFrame()

    if 'dt_create' in df.columns:
        logger.info("Gerando análises temporais...")
        df_timeline_base = df[df['dt_create'].notna()].copy()
        
        # --- ANÁLISE MENSAL GERAL ---
        df_timeline_base['Periodo_M'] = df_timeline_base['dt_create'].dt.to_period('M')
        timeline_data_mensal = df_timeline_base.groupby('Periodo_M').agg({
            'Record ID': 'count', 'Convertido': 'sum', 'Perdido': 'sum'
        }).rename(columns={'Record ID': 'Total', 'Convertido': 'Matriculas', 'Perdido': 'Perdidos'})
        timeline_data_mensal['Taxa_Conv'] = (timeline_data_mensal['Matriculas'] / timeline_data_mensal['Total'] * 100).round(1)
        timeline_data_mensal['Taxa_Perda'] = (timeline_data_mensal['Perdidos'] / timeline_data_mensal['Total'] * 100).round(1)
        timeline_data_mensal.index = timeline_data_mensal.index.astype(str)

        # --- ANÁLISE SEMANAL GERAL ---
        df_timeline_base['Periodo_S'] = df_timeline_base['dt_create'].dt.to_period('W')
        timeline_data_semanal = df_timeline_base.groupby('Periodo_S').agg({
            'Record ID': 'count', 'Convertido': 'sum', 'Perdido': 'sum'
        }).rename(columns={'Record ID': 'Total', 'Convertido': 'Matriculas', 'Perdido': 'Perdidos'})
        timeline_data_semanal['Taxa_Conv'] = (timeline_data_semanal['Matriculas'] / timeline_data_semanal['Total'] * 100).round(1)
        timeline_data_semanal['Taxa_Perda'] = (timeline_data_semanal['Perdidos'] / timeline_data_semanal['Total'] * 100).round(1)
        timeline_data_semanal.index = timeline_data_semanal.index.astype(str)

        # --- ANÁLISE MENSAL POR UNIDADE ---
        timeline_unidade_mensal = df_timeline_base.groupby(['Unidade Desejada', 'Periodo_M']).agg({
            'Record ID': 'count', 'Convertido': 'sum', 'Perdido': 'sum'
        }).rename(columns={'Record ID': 'Total', 'Convertido': 'Matriculas', 'Perdido': 'Perdidos'})
        timeline_unidade_mensal['Taxa_Conv'] = (timeline_unidade_mensal['Matriculas'] / timeline_unidade_mensal['Total'] * 100).round(1)
        timeline_unidade_mensal = timeline_unidade_mensal.reset_index()
        timeline_unidade_mensal['Periodo_M'] = timeline_unidade_mensal['Periodo_M'].astype(str)
        timeline_unidade_mensal = timeline_unidade_mensal.sort_values(['Unidade Desejada', 'Periodo_M'], ascending=[True, True])

    # ===== ANÁLISE DE ETAPAS DO NEGÓCIO =====
    logger.info("Gerando aba de etapas do negócio...")
    col_etapa = 'Etapa do negócio'
    df[col_etapa] = df[col_etapa].fillna('Não Preenchido').astype(str).str.strip()
    
    # Análise Geral
    analise_etapas_geral = df.groupby(col_etapa).agg(Total_Leads=('Record ID', 'count')).sort_values('Total_Leads', ascending=False)
    
    # Análise por Unidade
    analise_etapas_unidade = df.groupby(['Unidade Desejada', col_etapa]).agg(Total_Leads=('Record ID', 'count')).reset_index()
    analise_etapas_unidade = analise_etapas_unidade.sort_values(['Unidade Desejada', 'Total_Leads'], ascending=[True, False])


    # Métricas gerais
    total = len(df)
    conv = int(df['Convertido'].sum())
    perd = int(df['Perdido'].sum())
    
    print(f"\n- Total de Leads: {total:,}")
    print(f"- Matrículas: {conv:,} ({conv/total*100:.1f}%)")
    print(f"- Perdidos: {perd:,} ({perd/total*100:.1f}%)")
    print(f"- Em Pipeline: {total - conv - perd:,}")
    
    # Cores para Formatação Condicional
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde (Bom/Eficiente)
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Amarelo (Atenção)

    # ===== GERAR EXCEL =====
    date_str = datetime.now().strftime("%Y%m%d")
    
    # Garante que a pasta de saída exista
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_file = OUTPUT_DIR / f"base_funil_unids_proprias_{date_str}.xlsx"
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # ORDEM DAS ABAS:
    # 1. Resumo Executivo
    # 2. Evolução Mensal (Geral)
    # 3. Evolução Semanal (Geral)
    # 4. Evolução por Unidade
    # 5. Etapas do Negócio
    # 6. Próprias vs Franqueadas
    # 7. Análise de Canais
    # 8. Detalhamento Completo
    # 9. Canais por Unidade (Completo)
    # 10. Contexto da Análise
    # 11. Inconsistências de Datas

    # 1. Resumo Executivo
    ws = wb.create_sheet("Resumo Executivo")
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    
    # Título Principal
    ws['A1'] = "RELATÓRIO DE PERFORMANCE - RED BALLOON"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A1:C2')

    # Data de Geração
    ws['A3'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A3'].font = Font(italic=True, size=10, color="595959")
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A3:C3')

    # Subtítulo
    ws['A4'] = "Métricas Gerais"
    ws['A4'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A4'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A4'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A4:C4')

    # Cabeçalhos das Colunas
    ws['B5'] = "Cenário Válido (Analítico)"
    ws['C5'] = "Cenário Total (Bruto)"
    for col in ['B', 'C']:
        ws[f'{col}5'].font = Font(bold=True, color="595959")
        ws[f'{col}5'].alignment = Alignment(horizontal="center")
        ws[f'{col}5'].border = Border(bottom=Side(style='thin'))

    # Dados
    summary_data = [
        ("Total de Leads", total, raw_total, '#,##0'),
        ("Matrículas", conv, raw_conv, '#,##0'),
        ("Taxa de Conversão", conv / total if total > 0 else 0, raw_conv / raw_total if raw_total > 0 else 0, '0.0%'),
        ("Perdidos", perd, raw_perd, '#,##0'),
        ("Taxa de Perda", perd / total if total > 0 else 0, raw_perd / raw_total if raw_total > 0 else 0, '0.0%'),
        ("Em Pipeline", total - conv - perd, raw_total - raw_conv - raw_perd, '#,##0')
    ]

    row = 6
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for label, val_valid, val_raw, num_format in summary_data:
        cell_label = ws[f'A{row}']
        cell_valid = ws[f'B{row}']
        cell_raw = ws[f'C{row}']
        
        cell_label.value = label
        cell_label.font = Font(bold=True)
        cell_label.border = thin_border
        
        cell_valid.value = val_valid
        cell_valid.number_format = num_format
        cell_valid.alignment = Alignment(horizontal="right")
        cell_valid.border = thin_border

        cell_raw.value = val_raw
        cell_raw.number_format = num_format
        cell_raw.alignment = Alignment(horizontal="right")
        cell_raw.border = thin_border

        if row % 2 != 0:
            cell_label.fill = light_gray_fill
            cell_valid.fill = light_gray_fill
            cell_raw.fill = light_gray_fill
            
        row += 1

    # 2. Evolução Mensal (Geral)
    if not timeline_data_mensal.empty:
        ws_time = wb.create_sheet("Evolução Mensal (Geral)")
        for col in ['A', 'B', 'C', 'D', 'E', 'F']: ws_time.column_dimensions[col].width = 15
        
        ws_time['A1'] = "EVOLUÇÃO MENSAL DE LEADS (GERAL)"
        ws_time['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_time['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        ws_time.merge_cells('A1:F1')
        ws_time['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        row = 3
        format_header(ws_time, row, ["Mês/Ano", "Total Leads", "Matrículas", "% Conv.", "Perdidos", "% Perda"], "1F4E78")
        
        row = 4
        for periodo, row_data in timeline_data_mensal.iterrows():
            ws_time[f'A{row}'] = str(periodo)
            ws_time[f'B{row}'] = int(row_data['Total'])
            ws_time[f'C{row}'] = int(row_data['Matriculas'])
            ws_time[f'D{row}'] = row_data['Taxa_Conv']
            ws_time[f'E{row}'] = int(row_data['Perdidos'])
            ws_time[f'F{row}'] = row_data['Taxa_Perda']
            
            ws_time[f'D{row}'].number_format = '0.0"%"'
            ws_time[f'F{row}'].number_format = '0.0"%"'
            row += 1
        format_cells(ws_time, 3, row - 1, 6)

    # 3. Evolução Semanal (Geral)
    if not timeline_data_semanal.empty:
        ws_week = wb.create_sheet("Evolução Semanal (Geral)")
        for col in ['A', 'B', 'C', 'D', 'E', 'F']: ws_week.column_dimensions[col].width = 20
        
        ws_week['A1'] = "EVOLUÇÃO SEMANAL DE LEADS (GERAL)"
        ws_week['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_week['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        ws_week.merge_cells('A1:F1')
        ws_week['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        row = 3
        format_header(ws_week, row, ["Semana", "Total Leads", "Matrículas", "% Conv.", "Perdidos", "% Perda"], "1F4E78")
        
        row = 4
        for periodo, row_data in timeline_data_semanal.iterrows():
            ws_week[f'A{row}'] = str(periodo)
            ws_week[f'B{row}'] = int(row_data['Total'])
            ws_week[f'C{row}'] = int(row_data['Matriculas'])
            ws_week[f'D{row}'] = row_data['Taxa_Conv']
            ws_week[f'E{row}'] = int(row_data['Perdidos'])
            ws_week[f'F{row}'] = row_data['Taxa_Perda']
            
            ws_week[f'D{row}'].number_format = '0.0"%"'
            ws_week[f'F{row}'].number_format = '0.0"%"'
            row += 1
        format_cells(ws_week, 3, row - 1, 6)

    # 4. Evolução por Unidade
    if not timeline_unidade_mensal.empty:
        ws_unit_time = wb.create_sheet("Evolução por Unidade")
        ws_unit_time.column_dimensions['A'].width = 30
        ws_unit_time.column_dimensions['B'].width = 15
        for col in ['C', 'D', 'E', 'F']: ws_unit_time.column_dimensions[col].width = 12
        
        ws_unit_time['A1'] = "EVOLUÇÃO MENSAL POR UNIDADE"
        ws_unit_time['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_unit_time['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        ws_unit_time.merge_cells('A1:F1')
        ws_unit_time['A1'].alignment = Alignment(horizontal="center", vertical="center")

        row = 3
        format_header(ws_unit_time, row, ["Unidade", "Mês/Ano", "Total", "Matrículas", "% Conv.", "Perdidos"], "1F4E78")
        
        row = 4
        for _, row_data in timeline_unidade_mensal.iterrows():
            ws_unit_time[f'A{row}'] = str(row_data['Unidade Desejada'])
            ws_unit_time[f'B{row}'] = str(row_data['Periodo_M'])
            ws_unit_time[f'C{row}'] = int(row_data['Total'])
            ws_unit_time[f'D{row}'] = int(row_data['Matriculas'])
            ws_unit_time[f'E{row}'] = row_data['Taxa_Conv']
            ws_unit_time[f'F{row}'] = int(row_data['Perdidos'])
            ws_unit_time[f'E{row}'].number_format = '0.0"%"'
            row += 1
        format_cells(ws_unit_time, 3, row - 1, 6)

    # 5. Etapas do Negócio
    ws_etapas = wb.create_sheet("Etapas do Negócio")
    ws_etapas.column_dimensions['A'].width = 40
    ws_etapas.column_dimensions['B'].width = 20
    ws_etapas.column_dimensions['D'].width = 30
    ws_etapas.column_dimensions['E'].width = 40
    ws_etapas.column_dimensions['F'].width = 20

    ws_etapas['A1'] = "LEADS POR ETAPA DO NEGÓCIO (GERAL)"
    ws_etapas['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_etapas['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_etapas.merge_cells('A1:B1')
    
    format_header(ws_etapas, 3, ["Etapa do Negócio", "Total Leads"], "595959")
    row = 4
    for etapa, row_data in analise_etapas_geral.iterrows():
        ws_etapas[f'A{row}'] = str(etapa)
        ws_etapas[f'B{row}'] = int(row_data['Total_Leads'])
        row += 1
    format_cells(ws_etapas, 3, row-1, 2)

    ws_etapas['D1'] = "LEADS POR ETAPA DO NEGÓCIO (POR UNIDADE)"
    ws_etapas['D1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_etapas['D1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_etapas.merge_cells('D1:F1')
    
    format_header(ws_etapas, 3, ["Unidade", "Etapa do Negócio", "Total Leads"], "595959")
    row = 4
    for _, row_data in analise_etapas_unidade.iterrows():
        ws_etapas[f'D{row}'] = str(row_data['Unidade Desejada'])
        ws_etapas[f'E{row}'] = str(row_data['Etapa do negócio'])
        ws_etapas[f'F{row}'] = int(row_data['Total_Leads'])
        row += 1
    format_cells(ws_etapas, 3, row-1, 6)
    
    # 6. Próprias vs Franqueadas
    ws_comp = wb.create_sheet("Próprias vs Franqueadas")
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws_comp.column_dimensions[col].width = 16
    
    ws_comp['A1'] = "COMPARATIVO: PRÓPRIAS vs FRANQUEADAS"
    ws_comp['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_comp['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws_comp.merge_cells('A1:K1')
    ws_comp['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    row = 3
    format_header(ws_comp, row, ["Tipo", "Qtd Unid.", "Total Leads", "Matrículas", "% Conv.", "Leads p/ Matrícula", "Perdidos", "% Perda", "Ativ. Méd. (Geral)", "Ativ. Méd. (Matr.)", "Ativ. Méd. (Perd.)"], "70AD47")
    
    row = 4
    for tipo, row_data in resumo.iterrows():
        ws_comp[f'A{row}'] = str(tipo)
        ws_comp[f'B{row}'] = int(row_data['Qtd_Unidades'])
        ws_comp[f'C{row}'] = int(row_data['Total_Leads'])
        ws_comp[f'D{row}'] = int(row_data['Matriculas'])
        ws_comp[f'E{row}'] = row_data['Taxa_Conv_%']
        
        leads_p_mat = row_data['Leads_p_Matricula']
        if pd.isna(leads_p_mat) or leads_p_mat == float('inf') or leads_p_mat == 0:
            ws_comp[f'F{row}'] = "N/A"
        else:
            ws_comp[f'F{row}'] = f"1 a cada {leads_p_mat}"

        ws_comp[f'G{row}'] = int(row_data['Perdidos'])
        ws_comp[f'H{row}'] = row_data['Taxa_Perda_%']
        ws_comp[f'I{row}'] = row_data['Ativ_Media_Total']
        ws_comp[f'J{row}'] = row_data['Ativ_Media_Mat']
        ws_comp[f'K{row}'] = row_data['Ativ_Media_Perd']
        
        # Formatação Condicional: Verde se Matrícula < Perdido (Eficiência)
        if row_data['Ativ_Media_Mat'] > 0 and row_data['Ativ_Media_Perd'] > 0:
            if row_data['Ativ_Media_Mat'] < row_data['Ativ_Media_Perd']:
                ws_comp[f'J{row}'].fill = green_fill
            else:
                ws_comp[f'J{row}'].fill = yellow_fill

        ws_comp[f'C{row}'].number_format = '#,##0'
        ws_comp[f'D{row}'].number_format = '#,##0'
        ws_comp[f'E{row}'].number_format = '0.0"%"'
        ws_comp[f'G{row}'].number_format = '#,##0'
        ws_comp[f'H{row}'].number_format = '0.0"%"'
        ws_comp[f'I{row}'].number_format = '0.0'
        ws_comp[f'J{row}'].number_format = '0.0'
        ws_comp[f'K{row}'].number_format = '0.0'
        row += 1
    
    format_cells(ws_comp, 4, row - 1, 11)
    
    # 4. Análise de Canais (Geral)
    if not analise_canais.empty:
        ws_canais = wb.create_sheet("Análise de Canais")
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws_canais.column_dimensions[col].width = 16
        ws_canais.column_dimensions['A'].width = 35 # Canal name usually longer
        
        ws_canais['A1'] = "PERFORMANCE POR CANAL DE AQUISIÇÃO"
        ws_canais['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_canais['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        ws_canais.merge_cells('A1:J1')
        ws_canais['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        row = 3
        format_header(ws_canais, row, ["Canal / Fonte", "Total Leads", "Matrículas", "% Conv.", "Leads p/ Matrícula", "Perdidos", "% Perda", "Ativ. Méd. (Geral)", "Ativ. Méd. (Matr.)", "Ativ. Méd. (Perd.)"], "70AD47")
        
        row = 4
        for canal, row_data in analise_canais.iterrows():
            ws_canais[f'A{row}'] = str(canal)
            ws_canais[f'B{row}'] = int(row_data['Total'])
            ws_canais[f'C{row}'] = int(row_data['Matriculas'])
            ws_canais[f'D{row}'] = row_data['Taxa_Conv']
            
            leads_p_mat = row_data['Leads_p_Matricula']
            if pd.isna(leads_p_mat) or leads_p_mat == float('inf') or leads_p_mat == 0:
                ws_canais[f'E{row}'] = "N/A"
            else:
                ws_canais[f'E{row}'] = f"1 a cada {leads_p_mat}"
                
            ws_canais[f'F{row}'] = int(row_data['Perdidos'])
            ws_canais[f'G{row}'] = row_data['Taxa_Perda']
            ws_canais[f'H{row}'] = row_data['Ativ_Media_Total']
            ws_canais[f'I{row}'] = row_data['Ativ_Media_Mat']
            ws_canais[f'J{row}'] = row_data['Ativ_Media_Perd']
            
            # Formatação Condicional
            if row_data['Ativ_Media_Mat'] > 0 and row_data['Ativ_Media_Perd'] > 0:
                if row_data['Ativ_Media_Mat'] < row_data['Ativ_Media_Perd']:
                    ws_canais[f'I{row}'].fill = green_fill
                else:
                    ws_canais[f'I{row}'].fill = yellow_fill

            ws_canais[f'B{row}'].number_format = '#,##0'
            ws_canais[f'C{row}'].number_format = '#,##0'
            ws_canais[f'D{row}'].number_format = '0.0"%"'
            ws_canais[f'F{row}'].number_format = '#,##0'
            ws_canais[f'G{row}'].number_format = '0.0"%"'
            ws_canais[f'H{row}'].number_format = '0.0'
            ws_canais[f'I{row}'].number_format = '0.0'
            ws_canais[f'J{row}'].number_format = '0.0'
            
            row += 1
            
        format_cells(ws_canais, 4, row - 1, 10)
        
        # --- ADICIONAR GRÁFICO DE CANAIS ---
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Volume de Leads por Canal"
        chart.y_axis.title = "Total Leads"
        chart.x_axis.title = "Canal"
        
        data = Reference(ws_canais, min_col=2, min_row=3, max_row=row-1)
        cats = Reference(ws_canais, min_col=1, min_row=4, max_row=row-1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10 # Altura do gráfico
        chart.width = 25  # Largura do gráfico
        
        ws_canais.add_chart(chart, "L4")

    # 5. Detalhamento Completo
    ws_det = wb.create_sheet("Detalhamento Completo")
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws_det.column_dimensions[col].width = 16
    
    ws_det['A1'] = "DETALHAMENTO POR UNIDADE"
    ws_det['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_det['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws_det.merge_cells('A1:L1')
    ws_det['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    row = 3
    format_header(ws_det, row, ["Unidade", "Tipo", "Total", "Matrículas", "% Conv.", "Leads p/ Matrícula", "Perdidos", "% Perda", "Ativ. Méd. (Geral)", "Ativ. Méd. (Matr.)", "Ativ. Méd. (Perd.)", "Responsável"], "70AD47")
    
    row = 4
    for (tipo, unidade), row_data in analise.sort_values('Total', ascending=False).iterrows():
        ws_det[f'A{row}'] = str(unidade)
        ws_det[f'B{row}'] = tipo
        ws_det[f'C{row}'] = int(row_data['Total'])
        ws_det[f'D{row}'] = int(row_data['Matriculas'])
        ws_det[f'E{row}'] = row_data['Taxa_Conv']
        
        leads_p_mat = row_data['Leads_p_Matricula']
        if pd.isna(leads_p_mat) or leads_p_mat == float('inf') or leads_p_mat == 0:
            ws_det[f'F{row}'] = "N/A"
        else:
            ws_det[f'F{row}'] = f"1 a cada {leads_p_mat}"

        ws_det[f'G{row}'] = int(row_data['Perdidos'])
        ws_det[f'H{row}'] = row_data['Taxa_Perda']
        ws_det[f'I{row}'] = row_data['Ativ_Media_Total']
        ws_det[f'J{row}'] = row_data['Ativ_Media_Mat']
        ws_det[f'K{row}'] = row_data['Ativ_Media_Perd']
        ws_det[f'L{row}'] = str(row_data['Responsável'])
        
        # Formatação Condicional
        if row_data['Ativ_Media_Mat'] > 0 and row_data['Ativ_Media_Perd'] > 0:
            if row_data['Ativ_Media_Mat'] < row_data['Ativ_Media_Perd']:
                ws_det[f'J{row}'].fill = green_fill
            else:
                ws_det[f'J{row}'].fill = yellow_fill
        
        ws_det[f'C{row}'].number_format = '#,##0'
        ws_det[f'D{row}'].number_format = '#,##0'
        ws_det[f'E{row}'].number_format = '0.0"%"'
        ws_det[f'G{row}'].number_format = '#,##0'
        ws_det[f'H{row}'].number_format = '0.0"%"'
        ws_det[f'I{row}'].number_format = '0.0'
        ws_det[f'J{row}'].number_format = '0.0'
        ws_det[f'K{row}'].number_format = '0.0'
        row += 1
    
    format_cells(ws_det, 4, row - 1, 12)
    
    # 6. Canais por Unidade (Completo)
    if col_fonte:
        logger.info("Gerando aba unificada de canais por unidade...")
        
        ws_all = wb.create_sheet("Canais por Unidade (Completo)")
        ws_all.column_dimensions['A'].width = 25 # Unidade
        ws_all.column_dimensions['B'].width = 30 # Canal
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws_all.column_dimensions[col].width = 15
            
        ws_all['A1'] = "DETALHAMENTO DE CANAIS POR UNIDADE"
        ws_all['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_all['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws_all.merge_cells('A1:K1')
        ws_all['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        row = 3
        format_header(ws_all, row, ["Unidade", "Canal / Fonte", "Total Leads", "Matrículas", "% Conv.", "Leads p/ Matrícula", "Perdidos", "% Perda", "Ativ. Méd. (Geral)", "Ativ. Méd. (Matr.)", "Ativ. Méd. (Perd.)"], "70AD47")
        
        # Agrupar dados
        analise_unificada_base = df.groupby(['Unidade Desejada', col_fonte]).agg({
            'Record ID': 'count',
            'Convertido': 'sum',
            'Perdido': 'sum',
            'Número de atividades de vendas': 'mean'
        }).rename(columns={'Record ID': 'Total', 'Convertido': 'Matriculas', 'Perdido': 'Perdidos', 'Número de atividades de vendas': 'Ativ_Media_Total'})
        
        ativ_media_mat_unif = df[df['Convertido'] == 1].groupby(['Unidade Desejada', col_fonte])['Número de atividades de vendas'].mean()
        ativ_media_perd_unif = df[df['Perdido'] == 1].groupby(['Unidade Desejada', col_fonte])['Número de atividades de vendas'].mean()
        
        analise_unificada = analise_unificada_base.join(ativ_media_mat_unif.rename('Ativ_Media_Mat')).fillna({'Ativ_Media_Mat': 0})
        analise_unificada = analise_unificada.join(ativ_media_perd_unif.rename('Ativ_Media_Perd')).fillna({'Ativ_Media_Perd': 0})
        analise_unificada = analise_unificada.reset_index()
        
        analise_unificada['Taxa_Conv'] = (analise_unificada['Matriculas'] / analise_unificada['Total'] * 100).round(1)
        analise_unificada['Taxa_Perda'] = (analise_unificada['Perdidos'] / analise_unificada['Total'] * 100).round(1)
        analise_unificada['Leads_p_Matricula'] = (analise_unificada['Total'] / analise_unificada['Matriculas']).round(1)
        
        # Ordenar por Unidade e depois por Total Leads
        analise_unificada = analise_unificada.sort_values(['Unidade Desejada', 'Total'], ascending=[True, False])
        
        row = 4
        for _, row_data in analise_unificada.iterrows():
            ws_all[f'A{row}'] = str(row_data['Unidade Desejada'])
            ws_all[f'B{row}'] = str(row_data[col_fonte])
            ws_all[f'C{row}'] = int(row_data['Total'])
            ws_all[f'D{row}'] = int(row_data['Matriculas'])
            ws_all[f'E{row}'] = row_data['Taxa_Conv']
            
            leads_p_mat = row_data['Leads_p_Matricula']
            if pd.isna(leads_p_mat) or leads_p_mat == float('inf') or leads_p_mat == 0:
                ws_all[f'F{row}'] = "N/A"
            else:
                ws_all[f'F{row}'] = f"1 a cada {leads_p_mat}"
                
            ws_all[f'G{row}'] = int(row_data['Perdidos'])
            ws_all[f'H{row}'] = row_data['Taxa_Perda']
            ws_all[f'I{row}'] = row_data['Ativ_Media_Total']
            ws_all[f'J{row}'] = row_data['Ativ_Media_Mat']
            ws_all[f'K{row}'] = row_data['Ativ_Media_Perd']
            
            # Formatação Condicional
            if row_data['Ativ_Media_Mat'] > 0 and row_data['Ativ_Media_Perd'] > 0:
                if row_data['Ativ_Media_Mat'] < row_data['Ativ_Media_Perd']:
                    ws_all[f'J{row}'].fill = green_fill
                else:
                    ws_all[f'J{row}'].fill = yellow_fill

            ws_all[f'C{row}'].number_format = '#,##0'
            ws_all[f'D{row}'].number_format = '#,##0'
            ws_all[f'E{row}'].number_format = '0.0"%"'
            ws_all[f'G{row}'].number_format = '#,##0'
            ws_all[f'H{row}'].number_format = '0.0"%"'
            ws_all[f'I{row}'].number_format = '0.0'
            ws_all[f'J{row}'].number_format = '0.0'
            ws_all[f'K{row}'].number_format = '0.0'
            
            row += 1
        format_cells(ws_all, 4, row - 1, 11)

    # 7. Contexto da Análise
    ws_contexto = wb.create_sheet("Contexto da Análise")
    ws_contexto.column_dimensions['A'].width = 50
    ws_contexto.column_dimensions['B'].width = 20

    ws_contexto['A1'] = "CONTEXTO E METODOLOGIA DA ANÁLISE"
    ws_contexto['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_contexto['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws_contexto['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_contexto.merge_cells('A1:B2')

    # Explanation text
    ws_contexto['A4'] = "Para garantir a consistência e relevância dos dados, o relatório foca na performance das unidades ativas e em leads que possuem as informações essenciais preenchidas. Abaixo, o detalhamento da filtragem aplicada sobre a base de dados original."
    ws_contexto['A4'].alignment = Alignment(wrap_text=True, vertical="top")
    ws_contexto.merge_cells('A4:B5')
    ws_contexto.row_dimensions[4].height = 45

    # Reconciliation Table Title
    ws_contexto['A7'] = "Reconciliação do Total de Leads"
    ws_contexto['A7'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_contexto['A7'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_contexto['A7'].alignment = Alignment(horizontal="center")
    ws_contexto.merge_cells('A7:B7')

    # Reconciliation Table Data
    reconciliation_data = {
        "Total de leads no arquivo original (Bruto)": (total_leads_bruto, '#,##0'),
        "(-) Registros Incompletos (Sem Unidade)": (leads_sem_unidade, '#,##0'),
        "(-) Unidades não monitoradas (Fora do escopo)": (count_unidades_nao_analisadas, '#,##0'),
        "(-) Inconsistência de Datas (Fechamento < Criação)": (count_date_errors, '#,##0'),
        "(=) Leads Válidos (Base da Análise)": (total, '#,##0'),
    }

    row_ctx = 8
    for label, (value, num_format) in reconciliation_data.items():
        cell_label = ws_contexto[f'A{row_ctx}']
        cell_value = ws_contexto[f'B{row_ctx}']
        
        cell_label.value = label
        cell_value.value = value
        cell_value.number_format = num_format
        cell_value.alignment = Alignment(horizontal="right")
        cell_label.border = thin_border
        cell_value.border = thin_border
        if "(=)" in label:
            cell_label.font = Font(bold=True)
            cell_value.font = Font(bold=True)
        if row_ctx % 2 != 0:
            cell_label.fill = light_gray_fill
            cell_value.fill = light_gray_fill
        row_ctx += 1

    # List of excluded units
    if lista_unidades_nao_analisadas:
        row_ctx += 1
        ws_contexto[f'A{row_ctx}'] = f"Unidades Encontradas mas Não Analisadas ({len(lista_unidades_nao_analisadas)})"
        ws_contexto[f'A{row_ctx}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws_contexto[f'A{row_ctx}'].fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
        ws_contexto[f'A{row_ctx}'].alignment = Alignment(horizontal="center")
        ws_contexto.merge_cells(f'A{row_ctx}:B{row_ctx}')
        row_ctx += 1
        ws_contexto[f'A{row_ctx}'] = ", ".join(lista_unidades_nao_analisadas)
        ws_contexto[f'A{row_ctx}'].alignment = Alignment(wrap_text=True, vertical="top")
        ws_contexto[f'A{row_ctx}'].font = Font(italic=True, size=10)
        ws_contexto.merge_cells(f'A{row_ctx}:B{row_ctx+3}')

    # 8. Inconsistências de Datas
    if not date_errors_summary.empty:
        ws_err = wb.create_sheet("Inconsistências de Datas")
        ws_err.column_dimensions['A'].width = 35
        ws_err.column_dimensions['B'].width = 20
        
        ws_err['A1'] = "REGISTROS COM DATAS INCONSISTENTES"
        ws_err['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_err['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        ws_err.merge_cells('A1:B1')
        ws_err['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        ws_err['A2'] = "Critério: Data de Fechamento anterior à Data de Criação"
        ws_err.merge_cells('A2:B2')
        ws_err['A2'].font = Font(italic=True)
        ws_err['A2'].alignment = Alignment(horizontal="center")
        
        row = 4
        format_header(ws_err, row, ["Unidade", "Qtd. Erros"], "595959")
        
        row = 5
        for _, row_data in date_errors_summary.iterrows():
            ws_err[f'A{row}'] = str(row_data['Unidade Desejada'])
            ws_err[f'B{row}'] = int(row_data['Qtd_Erros'])
            ws_err[f'B{row}'].number_format = '#,##0'
            row += 1
        format_cells(ws_err, 5, row - 1, 2)

    wb.save(output_file)
    
    print("\n" + "=" * 100)
    print("RESUMO POR TIPO (CORRETO):")
    print("=" * 100)
    print(resumo[['Qtd_Unidades', 'Total_Leads', 'Matriculas', 'Taxa_Conv_%', 'Taxa_Perda_%']].to_string())
    print("\n" + "=" * 100)
    print("PROCESSO CONCLUÍDO COM SUCESSO!")
    print(f"Arquivo de Relatório: {output_file}")
    print("=" * 100)

    # ===== GERAR EXCEL PARA LOOKER =====
    try:
        output_looker_file = OUTPUT_DIR / f"base_funil_unids_proprias_{date_str}.xlsx"
        logger.info(f"Gerando arquivo de dados para o Looker: {output_looker_file}")
        
        # Garantir que todas as colunas do df_looker sejam compatíveis com Excel
        for col in df_looker.columns:
            if df_looker[col].dtype == 'object':
                df_looker[col] = df_looker[col].apply(lambda x: str(x) if not pd.isnull(x) else "")
        with pd.ExcelWriter(output_looker_file, engine='openpyxl') as writer:
            df_looker.to_excel(writer, sheet_name='Base_Leads_Enriquecida', index=False)
            analise.reset_index().to_excel(writer, sheet_name='Resumo_Unidades', index=False)
        
        print(f"Arquivo para Looker gerado: {output_looker_file}")
        print("=" * 100)

    except Exception as e:
        logger.error(f"Erro ao gerar o arquivo para o Looker: {e}")

except Exception as e:
    print(f"\nERRO: {str(e)}")
    logger.error(str(e), exc_info=True)
    exit(1)