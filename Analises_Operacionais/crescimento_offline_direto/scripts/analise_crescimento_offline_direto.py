"""
Script de Análise de Crescimento de Leads Offline e Tráfego Direto
===================================================================

Objetivo:
Analisar de forma SIMPLES o crescimento de leads offline e tráfego direto
comparando os relatórios consolidados de diferentes datas.

Este script:
1. Lê os relatórios consolidados existentes
2. Extrai APENAS o ciclo 26.1 de "Fontes off-line" e "Tráfego direto"
3. Mostra quanto cresceu ou caiu entre cada data

Exemplo de saída:
- 30/01/2026: 15,709 leads
- 06/02/2026: 16,958 leads (+1,249 | +7.9%)
- 09/02/2026: 17,063 leads (+105 | +0.6%)

Autor: Análise Operacional
Data: Fevereiro 2026
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===========================
# CONFIGURAÇÕES GLOBAIS
# ===========================

# Define o diretório base do projeto
BASE_DIR = Path(__file__).parent.parent
SCRIPT_DIR = BASE_DIR / 'scripts'
DATA_DIR = BASE_DIR / 'data'
OUTPUT_DIR = BASE_DIR / 'outputs'

# Diretório dos relatórios consolidados
RELATORIOS_DIR = BASE_DIR.parent / 'eficiencia_canal' / 'outputs' / 'analise_consolidada_leads_matriculas'

# Arquivos de relatórios a serem analisados (na ordem cronológica correta)
RELATORIOS = {
    '30/01/2026': 'analise_consolidada_leads_matriculas_2026-01-30.xlsx',
    '06/02/2026': 'analise_consolidada_leads_matriculas_2026-02-06.xlsx',
    '09/02/2026': 'analise_consolidada_leads_matriculas_2026-02-09.xlsx'
}

# Canais de interesse (nome exato como aparece nos relatórios)
CANAL_OFFLINE = 'Fontes off-line'
CANAL_DIRETO = 'Tráfego direto'

# ===========================
# FUNÇÕES AUXILIARES
# ===========================

def aplicar_estilo_cabecalho(ws, row_num, fill_color='4472C4'):
    """Aplica estilo ao cabeçalho de uma tabela"""
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    font = Font(bold=True, color='FFFFFF', size=11)
    alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[row_num]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
            cell.border = border

def aplicar_bordas(ws, start_row, end_row, start_col, end_col):
    """Aplica bordas em um intervalo de células"""
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = border

def formatar_numero(value):
    """Formata números para exibição"""
    if pd.isna(value):
        return 0
    return int(value) if isinstance(value, (int, float)) else value

def formatar_percentual(value):
    """Formata percentuais para exibição"""
    if pd.isna(value) or value == float('inf') or value == float('-inf'):
        return 'N/A'
    return f"{value:.1f}%"

# ===========================
# FUNÇÕES DE ANÁLISE
# ===========================

def ler_relatorio(caminho_arquivo, data_label):
    """Lê um relatório consolidado e extrai dados de Offline e Tráfego Direto"""
    print(f"\n   [INFO] Lendo relatório de {data_label}")
    print(f"   [INFO] Arquivo: {caminho_arquivo.name}")
    
    if not caminho_arquivo.exists():
        print(f"   [ERRO] Arquivo não encontrado: {caminho_arquivo}")
        return None
    
    try:
        # Lê a aba "Leads (Canais)"
        df_canais = pd.read_excel(caminho_arquivo, sheet_name='Leads (Canais)')
        
        # Procura as linhas de Offline e Tráfego Direto
        mask_offline = df_canais['Fonte original do tráfego'] == CANAL_OFFLINE
        mask_direto = df_canais['Fonte original do tráfego'] == CANAL_DIRETO
        
        if not mask_offline.any():
            print(f"   [AVISO] Canal '{CANAL_OFFLINE}' não encontrado")
            dados_offline = None
        else:
            dados_offline = df_canais[mask_offline].iloc[0].to_dict()
            print(f"   [OK] Dados de Offline encontrados")
        
        if not mask_direto.any():
            print(f"   [AVISO] Canal '{CANAL_DIRETO}' não encontrado")
            dados_direto = None
        else:
            dados_direto = df_canais[mask_direto].iloc[0].to_dict()
            print(f"   [OK] Dados de Tráfego Direto encontrados")
        
        # Lê a aba "Leads (Unidades)"
        df_unidades = pd.read_excel(caminho_arquivo, sheet_name='Leads (Unidades)')
        print(f"   [OK] Dados de Unidades carregados")
        
        return {
            'data': data_label,
            'arquivo': caminho_arquivo.name,
            'offline': dados_offline,
            'direto': dados_direto,
            'dataframe_canais': df_canais,
            'dataframe_unidades': df_unidades
        }
    
    except Exception as e:
        print(f"   [ERRO] Erro ao ler arquivo: {str(e)}")
        return None

def carregar_todos_relatorios():
    """Carrega todos os relatórios especificados"""
    print("\n" + "="*80)
    print("CARREGANDO RELATÓRIOS CONSOLIDADOS")
    print("="*80)
    
    relatorios_dados = {}
    
    for data_label, nome_arquivo in RELATORIOS.items():
        caminho = RELATORIOS_DIR / nome_arquivo
        dados = ler_relatorio(caminho, data_label)
        
        if dados:
            relatorios_dados[data_label] = dados
        else:
            print(f"   [AVISO] Relatório de {data_label} não pôde ser carregado")
    
    print(f"\n   [RESUMO] Total de relatórios carregados: {len(relatorios_dados)}/{len(RELATORIOS)}")
    
    return relatorios_dados

def extrair_valores_ciclos(dados_canal):
    """Extrai os valores dos ciclos de um canal"""
    if dados_canal is None:
        return {}
    
    # Ciclos padrão que podem aparecer nos relatórios
    ciclos_possiveis = ['23.1', '24.1', '25.1', '26.1']
    valores = {}
    
    for ciclo in ciclos_possiveis:
        if ciclo in dados_canal:
            valores[ciclo] = dados_canal[ciclo]
        else:
            valores[ciclo] = 0
    
    # Adiciona total geral se existir
    if 'TOTAL GERAL' in dados_canal:
        valores['TOTAL GERAL'] = dados_canal['TOTAL GERAL']
    
    return valores

def criar_tabela_comparativa():
    """Cria tabela comparativa simples entre os relatórios"""
    print("\n" + "="*80)
    print("GERANDO ANÁLISE COMPARATIVA")
    print("="*80)
    
    relatorios = carregar_todos_relatorios()
    
    if len(relatorios) == 0:
        print("   [ERRO] Nenhum relatório foi carregado com sucesso")
        return None, None
    
    # Prepara listas para armazenar dados
    dados_comparacao = []
    
    # Ordena as datas cronologicamente
    datas_ordenadas = sorted(relatorios.keys(), key=lambda x: datetime.strptime(x, '%d/%m/%Y'))
    
    print(f"\n   [INFO] Analisando relatórios na ordem: {' → '.join(datas_ordenadas)}")
    
    # Para cada relatório, extrai APENAS o ciclo 26.1
    for data in datas_ordenadas:
        info = relatorios[data]
        
        # Extrai valores de Offline - ciclo 26.1
        valores_offline = extrair_valores_ciclos(info['offline'])
        offline_total = valores_offline.get('26.1', 0)
        
        # Extrai valores de Tráfego Direto - ciclo 26.1
        valores_direto = extrair_valores_ciclos(info['direto'])
        direto_total = valores_direto.get('26.1', 0)
        
        # Soma tudo
        total = offline_total + direto_total
        
        dados_comparacao.append({
            'Data_Relatorio': data,
            'Offline': offline_total,
            'Trafego_Direto': direto_total,
            'Total': total
        })
    
    # Cria DataFrame
    df_comparacao = pd.DataFrame(dados_comparacao)
    
    print("\n   [RESUMO] Total de Leads por Data:")
    print("="*60)
    for _, row in df_comparacao.iterrows():
        print(f"   {row['Data_Relatorio']}: {int(row['Total']):>6,} leads "
              f"(Offline: {int(row['Offline']):>6,} | Direto: {int(row['Trafego_Direto']):>6,})")
    print("="*60)
    
    return df_comparacao, relatorios

def calcular_crescimento_entre_datas(df_comparacao):
    """Calcula o crescimento simples entre as datas"""
    print("\n" + "="*80)
    print("CRESCIMENTO ENTRE DATAS")
    print("="*80)
    
    crescimento_data = []
    
    # Calcula crescimento entre datas consecutivas
    for i in range(1, len(df_comparacao)):
        row_anterior = df_comparacao.iloc[i-1]
        row_atual = df_comparacao.iloc[i]
        
        # Total
        total_anterior = row_anterior['Total']
        total_atual = row_atual['Total']
        cresc_total_abs = total_atual - total_anterior
        cresc_total_pct = ((total_atual - total_anterior) / total_anterior * 100) if total_anterior > 0 else 0
        
        # Offline
        offline_anterior = row_anterior['Offline']
        offline_atual = row_atual['Offline']
        cresc_offline_abs = offline_atual - offline_anterior
        cresc_offline_pct = ((offline_atual - offline_anterior) / offline_anterior * 100) if offline_anterior > 0 else 0
        
        # Direto
        direto_anterior = row_anterior['Trafego_Direto']
        direto_atual = row_atual['Trafego_Direto']
        cresc_direto_abs = direto_atual - direto_anterior
        cresc_direto_pct = ((direto_atual - direto_anterior) / direto_anterior * 100) if direto_anterior > 0 else 0
        
        crescimento_data.append({
            'Periodo': f"{row_anterior['Data_Relatorio']} → {row_atual['Data_Relatorio']}",
            'Total_Anterior': int(total_anterior),
            'Total_Atual': int(total_atual),
            'Crescimento_Total': int(cresc_total_abs),
            'Crescimento_Total_Pct': cresc_total_pct,
            'Offline_Anterior': int(offline_anterior),
            'Offline_Atual': int(offline_atual),
            'Crescimento_Offline': int(cresc_offline_abs),
            'Crescimento_Offline_Pct': cresc_offline_pct,
            'Direto_Anterior': int(direto_anterior),
            'Direto_Atual': int(direto_atual),
            'Crescimento_Direto': int(cresc_direto_abs),
            'Crescimento_Direto_Pct': cresc_direto_pct
        })
        
        print(f"\n   {row_anterior['Data_Relatorio']} → {row_atual['Data_Relatorio']}:")
        print(f"   {'─'*60}")
        print(f"   Total:           {int(total_anterior):>6,} → {int(total_atual):>6,}  |  {int(cresc_total_abs):>+7,}  ({cresc_total_pct:>+6.1f}%)")
        print(f"   Offline:         {int(offline_anterior):>6,} → {int(offline_atual):>6,}  |  {int(cresc_offline_abs):>+7,}  ({cresc_offline_pct:>+6.1f}%)")
        print(f"   Tráfego Direto:  {int(direto_anterior):>6,} → {int(direto_atual):>6,}  |  {int(cresc_direto_abs):>+7,}  ({cresc_direto_pct:>+6.1f}%)")
    
    df_crescimento = pd.DataFrame(crescimento_data)
    
    return df_crescimento

def analisar_crescimento_por_unidade(relatorios):
    """Analisa crescimento por unidade entre as datas"""
    print("\n" + "="*80)
    print("ANÁLISE POR UNIDADE")
    print("="*80)
    
    # Ordena as datas cronologicamente
    datas_ordenadas = sorted(relatorios.keys(), key=lambda x: datetime.strptime(x, '%d/%m/%Y'))
    
    # Dicionário para armazenar dados por unidade
    dados_unidades = {}
    
    # Para cada relatório
    for data in datas_ordenadas:
        info = relatorios[data]
        df_unidades = info['dataframe_unidades']
        
        # Filtra apenas Offline e Direto
        mask = df_unidades['Fonte original do tráfego'].isin([CANAL_OFFLINE, CANAL_DIRETO])
        df_filtrado = df_unidades[mask]
        
        # Agrupa por unidade e soma apenas o ciclo 26.1
        unidades_totais = df_filtrado.groupby('Unidade')['26.1'].sum().to_dict()
        
        # Armazena
        for unidade, total in unidades_totais.items():
            if unidade not in dados_unidades:
                dados_unidades[unidade] = {}
            dados_unidades[unidade][data] = total
    
    # Cria DataFrame comparativo
    df_comparativo = pd.DataFrame(dados_unidades).T
    df_comparativo = df_comparativo[datas_ordenadas]  # Ordena colunas
    df_comparativo = df_comparativo.fillna(0).astype(int)
    
    # Calcula crescimentos
    for i in range(1, len(datas_ordenadas)):
        data_ant = datas_ordenadas[i-1]
        data_atual = datas_ordenadas[i]
        
        col_cresc = f"Crescimento {data_ant} → {data_atual}"
        col_pct = f"Crescimento % {data_ant} → {data_atual}"
        
        df_comparativo[col_cresc] = df_comparativo[data_atual] - df_comparativo[data_ant]
        df_comparativo[col_pct] = ((df_comparativo[data_atual] - df_comparativo[data_ant]) / 
                                    df_comparativo[data_ant] * 100).fillna(0)
        df_comparativo[col_pct] = df_comparativo[col_pct].round(1)
    
    # Reseta index para ter Unidade como coluna
    df_comparativo = df_comparativo.reset_index()
    df_comparativo = df_comparativo.rename(columns={'index': 'Unidade'})
    
    # Ordena por total da última data (decrescente)
    df_comparativo = df_comparativo.sort_values(datas_ordenadas[-1], ascending=False)
    
    print(f"\n   [INFO] Total de unidades analisadas: {len(df_comparativo)}")
    print(f"\n   [TOP 10] Unidades com mais leads na última data ({datas_ordenadas[-1]}):")
    print("="*60)
    for idx, row in df_comparativo.head(10).iterrows():
        print(f"   {row['Unidade']:<25} {int(row[datas_ordenadas[-1]]):>6,} leads")
    print("="*60)
    
    return df_comparativo, datas_ordenadas

# ===========================
# FUNÇÕES DE EXPORTAÇÃO
# ===========================

def exportar_para_excel(df_comparacao, df_crescimento, df_unidades, datas_ordenadas):
    """Exporta dados para um arquivo Excel simples e formatado"""
    print("\n" + "="*80)
    print("EXPORTANDO PARA EXCEL")
    print("="*80)
    
    # Cria nome do arquivo com timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = OUTPUT_DIR / f'crescimento_offline_direto_{timestamp}.xlsx'
    
    # Garante que o diretório existe
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Cria o Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # 1. ABA RESUMO - Total por Data
        print("   [INFO] Criando aba: Resumo por Data")
        df_resumo = df_comparacao[['Data_Relatorio', 'Offline', 'Trafego_Direto', 'Total']].copy()
        df_resumo.columns = ['Data Relatório', 'Offline', 'Tráfego Direto', 'Total']
        df_resumo.to_excel(writer, sheet_name='Resumo por Data', index=False, startrow=2)
        
        ws = writer.sheets['Resumo por Data']
        ws['A1'] = 'TOTAL DE LEADS OFFLINE E TRÁFEGO DIRETO POR DATA'
        ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
        ws['A2'] = 'Valores consolidados (Total Geral de todos os ciclos)'
        ws['A2'].font = Font(italic=True, size=10, color='666666')
        aplicar_estilo_cabecalho(ws, 3, '4472C4')
        
        # Ajusta largura das colunas
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        
        # 2. ABA CRESCIMENTO - Variação entre datas
        print("   [INFO] Criando aba: Crescimento")
        df_cresc_export = df_crescimento[['Periodo', 'Total_Anterior', 'Total_Atual', 
                                          'Crescimento_Total', 'Crescimento_Total_Pct']].copy()
        df_cresc_export.columns = ['Período', 'Total Anterior', 'Total Atual', 
                                    'Crescimento', 'Crescimento %']
        df_cresc_export['Crescimento %'] = df_cresc_export['Crescimento %'].apply(lambda x: f"{x:.1f}%")
        df_cresc_export.to_excel(writer, sheet_name='Crescimento', index=False, startrow=2)
        
        ws = writer.sheets['Crescimento']
        ws['A1'] = 'CRESCIMENTO ENTRE DATAS'
        ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
        ws['A2'] = 'Análise de variação do total de leads'
        ws['A2'].font = Font(italic=True, size=10, color='666666')
        aplicar_estilo_cabecalho(ws, 3, '70AD47')
        
        # Ajusta largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        
        # 3. ABA DETALHAMENTO - Crescimento por canal
        print("   [INFO] Criando aba: Detalhamento por Canal")
        df_detalhe = df_crescimento[['Periodo', 
                                      'Offline_Anterior', 'Offline_Atual', 'Crescimento_Offline', 'Crescimento_Offline_Pct',
                                      'Direto_Anterior', 'Direto_Atual', 'Crescimento_Direto', 'Crescimento_Direto_Pct']].copy()
        df_detalhe.columns = ['Período', 
                              'Offline Anterior', 'Offline Atual', 'Crescimento Offline', 'Crescimento Offline %',
                              'Direto Anterior', 'Direto Atual', 'Crescimento Direto', 'Crescimento Direto %']
        df_detalhe['Crescimento Offline %'] = df_detalhe['Crescimento Offline %'].apply(lambda x: f"{x:.1f}%")
        df_detalhe['Crescimento Direto %'] = df_detalhe['Crescimento Direto %'].apply(lambda x: f"{x:.1f}%")
        df_detalhe.to_excel(writer, sheet_name='Detalhamento por Canal', index=False, startrow=2)
        
        ws = writer.sheets['Detalhamento por Canal']
        ws['A1'] = 'DETALHAMENTO DO CRESCIMENTO POR CANAL'
        ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
        ws['A2'] = 'Análise separada: Offline e Tráfego Direto'
        ws['A2'].font = Font(italic=True, size=10, color='666666')
        aplicar_estilo_cabecalho(ws, 3, 'FF6B35')
        
        # Ajusta largura das colunas
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws.column_dimensions[col].width = 18
        
        # 4. ABA ANÁLISE POR UNIDADE
        print("   [INFO] Criando aba: Análise por Unidade")
        df_unidades.to_excel(writer, sheet_name='Análise por Unidade', index=False, startrow=2)
        
        ws = writer.sheets['Análise por Unidade']
        ws['A1'] = 'ANÁLISE DE CRESCIMENTO POR UNIDADE'
        ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
        ws['A2'] = 'Total de leads Offline + Tráfego Direto por unidade'
        ws['A2'].font = Font(italic=True, size=10, color='666666')
        aplicar_estilo_cabecalho(ws, 3, '9B59B6')
        
        # Ajusta largura das colunas
        ws.column_dimensions['A'].width = 25
        for i in range(2, len(df_unidades.columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18
    
    print(f"\n   [SUCESSO] Relatório gerado: {output_file}")
    print(f"   [INFO] Tamanho do arquivo: {output_file.stat().st_size / 1024:.1f} KB")
    
    return output_file

def gerar_relatorio_texto(df_comparacao, df_crescimento):
    """Gera um relatório em texto simples para visualização rápida"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = OUTPUT_DIR / f'relatorio_crescimento_{timestamp}.txt'
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("="*80 + "\n")
        f.write("RELATÓRIO DE CRESCIMENTO - LEADS OFFLINE E TRÁFEGO DIRETO\n")
        f.write("="*80 + "\n\n")
        
        f.write(f"Data de Geração: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
        f.write(f"Relatórios Analisados: {', '.join(RELATORIOS.keys())}\n")
        
        f.write("\n" + "-"*80 + "\n")
        f.write("TOTAL DE LEADS POR DATA\n")
        f.write("-"*80 + "\n\n")
        
        for _, row in df_comparacao.iterrows():
            f.write(f"{row['Data_Relatorio']:>12}:  {int(row['Total']):>6,} leads  ")
            f.write(f"(Offline: {int(row['Offline']):>6,}  |  Direto: {int(row['Trafego_Direto']):>6,})\n")
        
        f.write("\n" + "-"*80 + "\n")
        f.write("CRESCIMENTO ENTRE DATAS\n")
        f.write("-"*80 + "\n\n")
        
        for _, row in df_crescimento.iterrows():
            f.write(f"{row['Periodo']}\n")
            f.write(f"  Total:           {row['Total_Anterior']:>6,} → {row['Total_Atual']:>6,}  |  ")
            f.write(f"{row['Crescimento_Total']:>+7,}  ({row['Crescimento_Total_Pct']:>+6.1f}%)\n")
            f.write(f"  Offline:         {row['Offline_Anterior']:>6,} → {row['Offline_Atual']:>6,}  |  ")
            f.write(f"{row['Crescimento_Offline']:>+7,}  ({row['Crescimento_Offline_Pct']:>+6.1f}%)\n")
            f.write(f"  Tráfego Direto:  {row['Direto_Anterior']:>6,} → {row['Direto_Atual']:>6,}  |  ")
            f.write(f"{row['Crescimento_Direto']:>+7,}  ({row['Crescimento_Direto_Pct']:>+6.1f}%)\n\n")
        
        f.write("="*80 + "\n")
        f.write("FIM DO RELATÓRIO\n")
        f.write("="*80 + "\n")
    
    print(f"   [SUCESSO] Relatório texto gerado: {output_file}")
    
    return output_file

# ===========================
# FUNÇÃO PRINCIPAL
# ===========================

def main():
    """Função principal que executa todo o fluxo de análise"""
    print("\n")
    print("="*80)
    print("ANÁLISE DE CRESCIMENTO - LEADS OFFLINE E TRÁFEGO DIRETO")
    print("="*80)
    print(f"Início da execução: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("="*80)
    
    try:
        # 1. Carrega e compara relatórios
        df_comparacao, relatorios = criar_tabela_comparativa()
        
        if df_comparacao is None or len(df_comparacao) == 0:
            print("\n[ERRO] Não foi possível gerar a análise comparativa!")
            return
        
        # 2. Calcula crescimento
        df_crescimento = calcular_crescimento_entre_datas(df_comparacao)
        
        # 3. Análise por unidade
        df_unidades, datas_ordenadas = analisar_crescimento_por_unidade(relatorios)
        
        # 4. Exporta para Excel
        arquivo_excel = exportar_para_excel(df_comparacao, df_crescimento, df_unidades, datas_ordenadas)
        
        # 5. Gera relatório texto
        arquivo_texto = gerar_relatorio_texto(df_comparacao, df_crescimento)
        
        print("\n" + "="*80)
        print("PROCESSO CONCLUÍDO COM SUCESSO!")
        print("="*80)
        print(f"\n   Arquivos gerados:")
        print(f"   - Excel: {arquivo_excel.name}")
        print(f"   - Texto: {arquivo_texto.name}")
        print(f"\n   Relatórios analisados: {len(relatorios)}")
        print(f"   Total de comparações: {len(df_crescimento)}")
        print(f"   Unidades analisadas: {len(df_unidades)}")
        print("\n" + "="*80)
        
    except Exception as e:
        print("\n" + "="*80)
        print("ERRO DURANTE A EXECUÇÃO")
        print("="*80)
        print(f"[ERRO] {str(e)}")
        import traceback
        traceback.print_exc()
        print("="*80)

if __name__ == "__main__":
    main()
