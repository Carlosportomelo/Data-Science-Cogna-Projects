import pandas as pd
import numpy as np
import warnings
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

# Carregar os dados
file_path = r'c:\Users\a483650\Projetos\03_Analises_Operacionais\ICP\DATA\Escolas_Novas_2025.xlsx'
df = pd.read_excel(file_path, sheet_name='Escolas')

# Limpar dados - remover linhas vazias
df_clean = df.dropna(subset=['Nome da Escola']).copy()

# Processar dados
# 1. Classificar segmentos
df_clean['Seg_limpo'] = df_clean['Seg'].fillna('')
segmentos_total = []

for idx, row in df_clean.iterrows():
    seg = str(row['Seg']).lower()
    tem_infantil = 'infantil' in seg
    tem_fundamental = 'fundamental' in seg or 'fund' in seg
    tem_medio = 'médio' in seg or 'medio' in seg or 'ensino médio' in seg
    
    if tem_infantil and tem_fundamental and tem_medio:
        segmentos_total.append('Completo')
    else:
        segmentos_total.append('Parcial')

df_clean['Cobertura_Segmentos'] = segmentos_total

# 2. Classificar Brownfield vs Greenfield - CLASSIFICAÇÃO CORRETA
# Brownfield = cidades COM franqueado Red Balloon (origem menciona franqueado/franquia)
# Greenfield = cidades SEM franqueado Red Balloon
df_clean['Tipo_Campo'] = df_clean['Origem'].apply(
    lambda x: 'Brownfield' if pd.notna(x) and ('franqueado' in str(x).lower() or 'franquia' in str(x).lower()) else 'Greenfield'
)

# 3. Classificar se tinha inglês anterior (nova coluna separada)
df_clean['Tinha_Ingles_Anterior'] = df_clean['Inglês'].apply(
    lambda x: 'Sim' if pd.notna(x) and 'sim' in str(x).lower() else 'Não/Insatisfeito' if pd.notna(x) else 'N/D'
)

# 4. Extrair horas
def extrair_horas(texto):
    if pd.isna(texto):
        return None
    texto_str = str(texto).lower()
    if '5' in texto_str:
        return 5
    elif '4' in texto_str:
        return 4
    elif '3' in texto_str:
        return 3
    elif '2' in texto_str:
        return 2
    return None

df_clean['Horas_Numerico'] = df_clean['Horas '].apply(extrair_horas)
df_clean['Mais_3_horas'] = df_clean['Horas_Numerico'].apply(lambda x: 'Sim' if pd.notna(x) and x > 3 else 'Não' if pd.notna(x) else 'N/D')

# 5. Converter faturamento para numérico
df_clean['Faturamento_Numerico'] = pd.to_numeric(df_clean['Faturamento'], errors='coerce')

# 6. Classificar porte
def classificar_porte(alunos):
    if pd.isna(alunos):
        return 'N/D'
    if alunos < 300:
        return 'Pequeno'
    elif alunos < 600:
        return 'Médio'
    else:
        return 'Grande'

df_clean['Porte'] = df_clean['Alunado'].apply(classificar_porte)

# 7. Score de Upsell (apenas para pedagógicas)
df_clean['Score_Upsell'] = 0

# Pontuação por segmentos parciais
df_clean.loc[df_clean['Cobertura_Segmentos'] == 'Parcial', 'Score_Upsell'] += 3

# Pontuação por horas baixas
df_clean.loc[df_clean['Horas_Numerico'] <= 3, 'Score_Upsell'] += 2

# Pontuação por alunado alto
df_clean.loc[df_clean['Alunado'] > 500, 'Score_Upsell'] += 3
df_clean.loc[(df_clean['Alunado'] > 300) & (df_clean['Alunado'] <= 500), 'Score_Upsell'] += 2

# Preparar Excel de saída
output_file = r'c:\Users\a483650\Projetos\03_Analises_Operacionais\ICP\OUTPUTS\Analise_ICP_Processada.xlsx'

# Função para formatar planilha
def formatar_planilha(sheet, titulo_cor='4472C4', tem_valores_monetarios=False):
    """
    Formata uma planilha com estilos profissionais
    """
    # Cores
    cor_header = PatternFill(start_color=titulo_cor, end_color=titulo_cor, fill_type='solid')
    cor_linha_par = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Bordas
    borda_fina = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Fonte
    fonte_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    fonte_normal = Font(name='Calibri', size=10)
    
    # Alinhamento
    alinhamento_centro = Alignment(horizontal='center', vertical='center')
    alinhamento_esquerda = Alignment(horizontal='left', vertical='center')
    
    # Formatar cabeçalho (linha 1)
    for cell in sheet[1]:
        cell.fill = cor_header
        cell.font = fonte_header
        cell.border = borda_fina
        cell.alignment = alinhamento_centro
    
    # Formatar linhas de dados
    for idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
        # Cor alternada
        if idx % 2 == 0:
            fill = cor_linha_par
        else:
            fill = PatternFill(fill_type=None)
        
        for cell in row:
            cell.border = borda_fina
            cell.font = fonte_normal
            cell.fill = fill
            
            # Alinhamento baseado no tipo
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
                # Formatação de números
                if tem_valores_monetarios and cell.column > 1:
                    # Verificar se é uma coluna monetária
                    header_value = str(sheet.cell(1, cell.column).value).lower()
                    if 'fat' in header_value or 'receita' in header_value or 'valor' in header_value:
                        cell.number_format = 'R$ #,##0.00'
                    elif cell.value and cell.value > 100:
                        cell.number_format = '#,##0.00'
            else:
                cell.alignment = alinhamento_esquerda
    
    # Ajustar largura das colunas
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Congelar primeira linha
    sheet.freeze_panes = 'A2'

# Criar sheets
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Sheet 1: Dados completos processados
    df_export = df_clean[[
        'Nome da Escola', 'Comercial', 'Perfil', 'Alunado', 'Porte',
        'Horas ', 'Horas_Numerico', 'Mais_3_horas',
        'Seg', 'Cobertura_Segmentos',
        'Faturamento', 'Faturamento_Numerico',
        'Origem', 'Tipo_Campo',
        'Inglês', 'Tinha_Ingles_Anterior',
        'Score_Upsell',
        'Motivo de ganho de contrato', 'Observação'
    ]]
    df_export.to_excel(writer, sheet_name='Dados Processados', index=False)
    
    # Sheet 2: Análise por Perfil
    perfil_analise = df_clean.groupby('Perfil').agg({
        'Nome da Escola': 'count',
        'Faturamento_Numerico': ['sum', 'mean', 'median'],
        'Alunado': ['sum', 'mean']
    }).round(2)
    perfil_analise.columns = ['Qtd Escolas', 'Fat. Total', 'Fat. Médio', 'Fat. Mediano', 'Alunos Total', 'Alunos Médio']
    perfil_analise.to_excel(writer, sheet_name='Análise por Perfil')
    
    # Sheet 3: Análise por Origem
    origem_analise = df_clean.groupby('Origem').agg({
        'Nome da Escola': 'count',
        'Faturamento_Numerico': 'sum'
    }).rename(columns={'Nome da Escola': 'Número de Escolas', 'Faturamento_Numerico': 'Faturamento Total'})
    origem_analise = origem_analise.sort_values('Faturamento Total', ascending=False)
    origem_analise.to_excel(writer, sheet_name='Análise por Origem')
    
    # Sheet 4: Brownfield vs Greenfield (CORRIGIDO)
    campo_analise = df_clean.groupby('Tipo_Campo').agg({
        'Nome da Escola': 'count',
        'Faturamento_Numerico': 'sum',
        'Alunado': 'sum'
    }).rename(columns={'Nome da Escola': 'Número de Escolas', 'Faturamento_Numerico': 'Faturamento Total'})
    campo_analise.to_excel(writer, sheet_name='Brownfield vs Greenfield')
    
    # Sheet 4b: Análise de Inglês Anterior (nova sheet)
    ingles_analise = df_clean.groupby('Tinha_Ingles_Anterior').agg({
        'Nome da Escola': 'count',
        'Faturamento_Numerico': 'sum',
        'Alunado': 'sum'
    }).rename(columns={'Nome da Escola': 'Número de Escolas', 'Faturamento_Numerico': 'Faturamento Total'})
    ingles_analise.to_excel(writer, sheet_name='Inglês Anterior')
    
    # Sheet 5: Prioridades de Upsell (Pedagógicas)
    pedagogicas = df_clean[df_clean['Perfil'] == 'Pedagógica'].copy()
    pedagogicas_sorted = pedagogicas.sort_values('Score_Upsell', ascending=False)
    upsell_export = pedagogicas_sorted[[
        'Nome da Escola', 'Alunado', 'Faturamento_Numerico', 
        'Horas_Numerico', 'Seg', 'Cobertura_Segmentos', 'Score_Upsell', 'Tipo_Campo', 'Comercial'
    ]]
    upsell_export.to_excel(writer, sheet_name='Prioridades Upsell', index=False)
    
    # Sheet 6: Resumo Executivo
    resumo_data = {
        'Métrica': [
            'Total de Escolas',
            'Faturamento Total',
            'Ticket Médio',
            'Alunos Totais',
            '--- BROWNFIELD/GREENFIELD ---',
            'Escolas Brownfield (c/ franqueado)',
            'Escolas Greenfield (s/ franqueado)',
            '--- INGLÊS ANTERIOR ---',
            'Escolas c/ Inglês Anterior',
            'Escolas s/ Inglês Anterior',
            '--- OUTROS ---',
            'Escolas com >3h',
            'Perfil Mais Rentável (Médio)',
            'Melhor Canal (Origem)',
        ],
        'Valor': [
            len(df_clean),
            f"R$ {df_clean['Faturamento_Numerico'].sum():,.2f}",
            f"R$ {df_clean['Faturamento_Numerico'].mean():,.2f}",
            int(df_clean['Alunado'].sum()),
            '---',
            f"{len(df_clean[df_clean['Tipo_Campo'] == 'Brownfield'])} escolas",
            f"{len(df_clean[df_clean['Tipo_Campo'] == 'Greenfield'])} escolas",
            '---',
            f"{len(df_clean[df_clean['Tinha_Ingles_Anterior'] == 'Sim'])} escolas",
            f"{len(df_clean[df_clean['Tinha_Ingles_Anterior'] != 'Sim'])} escolas",
            '---',
            len(df_clean[df_clean['Mais_3_horas'] == 'Sim']),
            'Comercial (R$ 246.670)',
            'Indicação Franqueado',
        ]
    }
    resumo_df = pd.DataFrame(resumo_data)
    resumo_df.to_excel(writer, sheet_name='Resumo Executivo', index=False)
    
    # APLICAR FORMATAÇÃO EM TODAS AS SHEETS
    workbook = writer.book
    
    # Sheet 1: Dados Processados (azul padrão)
    formatar_planilha(workbook['Dados Processados'], titulo_cor='4472C4', tem_valores_monetarios=True)
    
    # Sheet 2: Análise por Perfil (verde)
    formatar_planilha(workbook['Análise por Perfil'], titulo_cor='70AD47', tem_valores_monetarios=True)
    
    # Sheet 3: Análise por Origem (laranja)
    formatar_planilha(workbook['Análise por Origem'], titulo_cor='ED7D31', tem_valores_monetarios=True)
    
    # Sheet 4: Brownfield vs Greenfield (roxo)
    formatar_planilha(workbook['Brownfield vs Greenfield'], titulo_cor='9966CC', tem_valores_monetarios=True)
    
    # Sheet 4b: Inglês Anterior (azul claro)
    formatar_planilha(workbook['Inglês Anterior'], titulo_cor='5B9BD5', tem_valores_monetarios=True)
    
    # Sheet 5: Prioridades Upsell (vermelho)
    sheet_upsell = workbook['Prioridades Upsell']
    formatar_planilha(sheet_upsell, titulo_cor='C00000', tem_valores_monetarios=True)
    
    # Adicionar formatação condicional no Score
    from openpyxl.formatting.rule import ColorScaleRule
    score_column = None
    for idx, cell in enumerate(sheet_upsell[1], start=1):
        if cell.value == 'Score_Upsell':
            score_column = get_column_letter(idx)
            break
    
    if score_column:
        # Escala de cor: vermelho (baixo) -> amarelo -> verde (alto)
        color_scale = ColorScaleRule(
            start_type='min', start_color='F8696B',
            mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='63BE7B'
        )
        sheet_upsell.conditional_formatting.add(
            f'{score_column}2:{score_column}{sheet_upsell.max_row}',
            color_scale
        )
    
    # Sheet 6: Resumo Executivo (azul escuro)
    sheet_resumo = workbook['Resumo Executivo']
    formatar_planilha(sheet_resumo, titulo_cor='1F4E78', tem_valores_monetarios=False)
    
    # Ajustar largura específica para o resumo
    sheet_resumo.column_dimensions['A'].width = 40
    sheet_resumo.column_dimensions['B'].width = 35

print("=" * 80)
print("✅ EXCEL FORMATADO CRIADO COM SUCESSO! (VERSÃO CORRIGIDA)")
print("=" * 80)
print(f"\nArquivo: {output_file}")
print("\n🔄 CORREÇÃO APLICADA:")
print("  • Brownfield = Cidades COM franqueado Red Balloon")
print("  • Greenfield = Cidades SEM franqueado Red Balloon")
print("  • Nova sheet 'Inglês Anterior' analisa se escola tinha inglês antes")
print("\n🎨 FORMATAÇÃO APLICADA:")
print("  • Cabeçalhos coloridos por sheet (cores diferentes)")
print("  • Linhas alternadas em cinza claro")
print("  • Bordas em todas as células")
print("  • Largura automática das colunas")
print("  • Valores monetários formatados (R$)")
print("  • Formatação condicional no Score de Upsell")
print("  • Primeira linha congelada para scroll")
print("  • Fonte Calibri profissional")
print("\n📊 Sheets criadas:")
print("  1. 🔵 Dados Processados - Todos os dados com classificações CORRETAS")
print("  2. 🟢 Análise por Perfil - Métricas agregadas por perfil")
print("  3. 🟠 Análise por Origem - Métricas por canal de aquisição")
print("  4. 🟣 Brownfield vs Greenfield - CORRIGIDO (com/sem franqueado)")
print("  5. 🔷 Inglês Anterior - Nova análise sobre inglês prévio")
print("  6. 🔴 Prioridades Upsell - Lista ordenada (com Brownfield/Greenfield)")
print("  7. 🔶 Resumo Executivo - KPIs principais atualizados")
print("\n" + "=" * 80)
