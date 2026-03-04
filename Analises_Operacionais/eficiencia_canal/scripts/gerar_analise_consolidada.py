"""
gerar_analise_consolidada.py
─────────────────────────────────────────────────────────────
Gera o relatório consolidado de Leads × Matrículas × Erros
por ciclo comercial Red Balloon, a partir do CSV do HubSpot.

Uso: python gerar_analise_consolidada.py
─────────────────────────────────────────────────────────────
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys
from datetime import datetime

# ════════════════════════════════════════════════════════════
# CONFIGURAÇÃO — ajuste aqui se necessário
# ════════════════════════════════════════════════════════════

INPUT_PATH = Path(r"C:\Users\a483650\Projetos\_DADOS_CENTRALIZADOS\hubspot\hubspot_leads_atual.csv")

OUTPUT_DIR = Path(r"C:\Users\a483650\Projetos\03_Analises_Operacionais\eficiencia_canal\outputs")
OUTPUT_FILE = OUTPUT_DIR / "analise_consolidada_leads_matriculas.xlsx"

# Ciclos comerciais (nome: (data_início, data_fim))
# Use None em data_fim para cortar na data máxima do CSV (ciclo em andamento).
CICLOS = {
    "23.1": ("2022-10-01", "2023-02-20"),
    "24.1": ("2023-10-01", "2024-02-20"),
    "25.1": ("2024-10-01", "2025-02-20"),
    "26.1": ("2025-10-01", "2026-02-20"),
}

# Comparação padrão
CICLO_ATUAL = "26.1"
CICLO_ANTERIOR = "25.1"

# Etapa que define matrícula
ETAPA_MATRICULA = "MATRÍCULA CONCLUÍDA"

# Fontes que aparecem individualmente (demais → "Sem Fonte original do tráfego")
FONTES_MAPEADAS = [
    "Fontes off-line",
    "Pesquisa orgânica",
    "Pesquisa paga",
    "Referências",
    "Referências de IA",
    "Social orgânico",
    "Social pago",
    "Tráfego direto",
    "E-mail marketing",
    "Outras campanhas",
]
LABEL_SEM_FONTE = "Sem Fonte original do tráfego"


# ════════════════════════════════════════════════════════════
# CARGA E PREPARAÇÃO
# ════════════════════════════════════════════════════════════

def carregar_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        print(f"❌ Arquivo não encontrado: {path}")
        sys.exit(1)

    df = pd.read_csv(path, low_memory=False)
    df["Data de criação"] = pd.to_datetime(df["Data de criação"], errors="coerce")
    df["Data de fechamento"] = pd.to_datetime(df["Data de fechamento"], errors="coerce")

    nulos = df["Data de criação"].isna().sum()
    if nulos:
        print(f"  ⚠ {nulos} registros sem 'Data de criação' removidos")
        df.dropna(subset=["Data de criação"], inplace=True)

    return df


def preparar(df: pd.DataFrame) -> pd.DataFrame:
    # Atribuir ciclo
    df["Ciclo"] = pd.Series(dtype="object", index=df.index)
    for ciclo, (inicio, fim) in CICLOS.items():
        mask = (
            (df["Data de criação"] >= inicio)
            & (df["Data de criação"] <= fim + " 23:59:59")
        )
        df.loc[mask, "Ciclo"] = ciclo

    fora = df["Ciclo"].isna().sum()
    if fora:
        print(f"  ℹ {fora:,} registros fora dos ciclos (ignorados)")
    df.dropna(subset=["Ciclo"], inplace=True)

    # Classificar fonte
    df["Fonte original do tráfego"] = np.where(
        df["Fonte original do tráfego"].isin(FONTES_MAPEADAS),
        df["Fonte original do tráfego"],
        LABEL_SEM_FONTE,
    )

    # Flags
    df["is_matricula"] = df["Etapa do negócio"].str.contains(
        ETAPA_MATRICULA, case=False, na=False
    )
    df["is_erro"] = df["Data de fechamento"].notna() & (
        df["Data de fechamento"] < df["Data de criação"]
    )

    return df


# ════════════════════════════════════════════════════════════
# GERAÇÃO DAS TABELAS
# ════════════════════════════════════════════════════════════

def pivot_por(df: pd.DataFrame, filtro: str | None, grupo: str | list, ciclos: list) -> pd.DataFrame:
    sub = df[df[filtro]] if filtro else df

    if isinstance(grupo, str):
        grupo = [grupo]

    pv = sub.groupby(grupo + ["Ciclo"]).size().unstack(fill_value=0)

    for c in ciclos:
        if c not in pv.columns:
            pv[c] = 0
    pv = pv[ciclos]

    pv["TOTAL GERAL"] = pv.sum(axis=1)

    if CICLO_ANTERIOR in pv.columns and CICLO_ATUAL in pv.columns:
        pv[f"Dif ({CICLO_ATUAL} vs {CICLO_ANTERIOR})"] = pv[CICLO_ATUAL] - pv[CICLO_ANTERIOR]
        pv[f"Var % ({CICLO_ATUAL} vs {CICLO_ANTERIOR})"] = np.where(
            pv[CICLO_ANTERIOR] == 0,
            np.where(pv[CICLO_ATUAL] == 0, 0, 1),
            (pv[CICLO_ATUAL] - pv[CICLO_ANTERIOR]) / pv[CICLO_ANTERIOR],
        )

    result = pv.sort_index().reset_index()
    if "Unidade Desejada" in result.columns:
        result.rename(columns={"Unidade Desejada": "Unidade"}, inplace=True)
    return result


def gerar_resumo(df: pd.DataFrame, ciclos: list) -> pd.DataFrame:
    rows = []
    for i, ciclo in enumerate(ciclos):
        sub = df[df["Ciclo"] == ciclo]
        leads = len(sub)
        mat = int(sub["is_matricula"].sum())
        erros = int(sub["is_erro"].sum())

        pct_erros = erros / mat if mat else 0
        conversao = mat / leads if leads else 0

        if i == 0:
            cresc_l = cresc_m = cresc_e = 0
        else:
            prev = df[df["Ciclo"] == ciclos[i - 1]]
            pl, pm, pe = len(prev), int(prev["is_matricula"].sum()), int(prev["is_erro"].sum())
            cresc_l = (leads - pl) / pl if pl else 0
            cresc_m = (mat - pm) / pm if pm else 0
            cresc_e = (erros - pe) / pe if pe else 0

        rows.append({
            "Ciclo": ciclo,
            "Leads": leads,
            "Matriculas": mat,
            "Erros (Fech < Criação)": erros,
            "% Erros": pct_erros,
            "Conversao": conversao,
            "Cresc. Leads %": cresc_l,
            "Cresc. Matrículas %": cresc_m,
            "Cresc. Erros %": cresc_e,
        })
    return pd.DataFrame(rows)


# ════════════════════════════════════════════════════════════
# FORMATAÇÃO EXCEL
# ════════════════════════════════════════════════════════════

H_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
H_FILL = PatternFill("solid", fgColor="2F5496")
D_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))


def escrever_aba(ws, df: pd.DataFrame, pct_cols: list[str] | None = None):
    pct_cols = pct_cols or []

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font, cell.fill = H_FONT, H_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font, cell.border = D_FONT, THIN_BORDER
            col_name = df.columns[ci - 1]

            if col_name in pct_cols or col_name.startswith("Var %"):
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="center")
            elif isinstance(val, (int, float, np.integer, np.floating)) and col_name != "Ciclo":
                cell.number_format = "#,##0"

    for ci in range(1, len(df.columns) + 1):
        header_len = len(str(df.columns[ci - 1]))
        data_len = df.iloc[:, ci - 1].astype(str).str.len().max() if len(df) else 0
        ws.column_dimensions[get_column_letter(ci)].width = min(max(header_len, data_len) + 4, 30)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


# ════════════════════════════════════════════════════════════
# EXECUÇÃO
# ════════════════════════════════════════════════════════════

def main():
    t0 = datetime.now()
    print("=" * 60)
    print("  Análise Consolidada — Leads × Matrículas × Erros")
    print("=" * 60)

    print(f"\n📂 Input: {INPUT_PATH}")
    df = carregar_csv(INPUT_PATH)
    print(f"  {len(df):,} registros")

    print("\n⚙  Preparando dados...")
    df = preparar(df)
    ciclos = sorted(df["Ciclo"].unique())
    print(f"  Ciclos: {ciclos}")
    print(f"  Matrículas: {df['is_matricula'].sum():,}")
    print(f"  Erros: {df['is_erro'].sum():,}")

    print("\n📊 Gerando tabelas...")
    fonte_col = "Fonte original do tráfego"
    unidade_fonte = ["Unidade Desejada", "Fonte original do tráfego"]

    resumo = gerar_resumo(df, ciclos)
    mat_canais = pivot_por(df, "is_matricula", fonte_col, ciclos)
    mat_unidades = pivot_por(df, "is_matricula", unidade_fonte, ciclos)
    leads_canais = pivot_por(df, None, fonte_col, ciclos)
    leads_unidades = pivot_por(df, None, unidade_fonte, ciclos)
    erros_canais = pivot_por(df, "is_erro", fonte_col, ciclos)
    erros_unidades = pivot_por(df, "is_erro", unidade_fonte, ciclos)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    pct_resumo = ["% Erros", "Conversao", "Cresc. Leads %", "Cresc. Matrículas %", "Cresc. Erros %"]

    abas = [
        ("Resumo e Insights", resumo, pct_resumo),
        ("Matrículas (Canais)", mat_canais, []),
        ("Matrículas (Unidades)", mat_unidades, []),
        ("Leads (Canais)", leads_canais, []),
        ("Leads (Unidades)", leads_unidades, []),
        ("Erros (Canais)", erros_canais, []),
        ("Erros (Unidades)", erros_unidades, []),
    ]

    for i, (nome, dados, pct) in enumerate(abas):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = nome
        escrever_aba(ws, dados, pct)
        print(f"  ✓ {nome} ({len(dados)} linhas)")

    wb.save(str(OUTPUT_FILE))

    elapsed = (datetime.now() - t0).total_seconds()
    print(f"\n💾 Output: {OUTPUT_FILE}")
    print(f"⏱  {elapsed:.1f}s")
    print("=" * 60)


if __name__ == "__main__":
    main()