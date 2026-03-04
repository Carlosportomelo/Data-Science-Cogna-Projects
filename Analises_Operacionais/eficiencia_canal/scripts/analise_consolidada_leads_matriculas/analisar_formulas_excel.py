import openpyxl
from openpyxl.utils import get_column_letter

excel_06 = r'c:\Users\a483650\Projetos\analise_eficiencia_canal\outputs\analise_consolidada_leads_matriculas\analise_consolidada_leads_matriculas_2026-02-06.xlsx'

wb = openpyxl.load_workbook(excel_06)
ws = wb['Resumo e Insights']

print("╔" + "═" * 150 + "╗")
print("║" + "ANÁLISE DAS FÓRMULAS DO EXCEL 06/02".center(150) + "║")
print("╚" + "═" * 150 + "╝\n")

print("CÉLULAS DO SHEET 'Resumo e Insights':\n")

for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=15):
    for cell in row:
        if cell.value:
            col_letter = cell.column_letter
            row_num = cell.row
            value = cell.value
            
            print(f"[{col_letter}{row_num}] = {value}")
            
            # Se a célula tem fórmula, mostrar a fórmula
            if isinstance(value, str) and (value.startswith('=') or cell.data_type == 'f'):
                print(f"     (Fórmula bruta: {cell.value})")

print("\n" + "═" * 150)
print("DETALHAMENTO DE CADA CÉLULA COM FÓRMULA:\n")

for row in range(1, 20):
    for col in range(1, 20):
        cell = ws.cell(row=row, column=col)
        if cell.data_type == 'f':  # Se é fórmula
            print(f"[{cell.column_letter}{cell.row}] = {cell.value}")
            print(f"         (Valor: {cell.value})")
            print()
