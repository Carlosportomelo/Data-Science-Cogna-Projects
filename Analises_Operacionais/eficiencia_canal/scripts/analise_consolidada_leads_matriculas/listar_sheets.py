import pandas as pd
import openpyxl

excel_30 = r'c:\Users\a483650\Projetos\analise_eficiencia_canal\outputs\analise_consolidada_leads_matriculas\analise_consolidada_leads_matriculas_2026-01-30.xlsx'
excel_06 = r'c:\Users\a483650\Projetos\analise_eficiencia_canal\outputs\analise_consolidada_leads_matriculas\analise_consolidada_leads_matriculas_2026-02-06.xlsx'

print("EXCEL 30/01 - Sheets disponíveis:")
wb_30 = openpyxl.load_workbook(excel_30)
print(wb_30.sheetnames)

print("\nEXCEL 06/02 - Sheets disponíveis:")
wb_06 = openpyxl.load_workbook(excel_06)
print(wb_06.sheetnames)

# Ler primeira sheet de cada
print("\n" + "="*100)
print("EXCEL 30/01 - Primeira sheet:")
df_30 = pd.read_excel(excel_30, sheet_name=0)
print(df_30.head(20))

print("\n" + "="*100)
print("EXCEL 06/02 - Primeira sheet:")
df_06 = pd.read_excel(excel_06, sheet_name=0)
print(df_06.head(20))
