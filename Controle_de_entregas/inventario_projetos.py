"""
Script para gerar inventário completo de projetos de dados
Analisa a estrutura de C:/Users/a483650/Projetos e cria uma planilha Excel organizada
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Definição do inventário de projetos
projetos_data = {
    'ID': [],
    'Nome do Projeto': [],
    'Categoria': [],
    'Objetivo': [],
    'Status': [],
    'Fontes de Dados': [],
    'Scripts Principais': [],
    'Outputs Gerados': [],
    'Stakeholders': [],
    'Frequência Atualização': [],
    'Dashboards/Looker': [],
    'Tecnologias': [],
    'Observações': []
}

# Projeto 1: Projeto Helio
projetos_data['ID'].append('PRJ-001')
projetos_data['Nome do Projeto'].append('Projeto Helio (Lead Scoring ML)')
projetos_data['Categoria'].append('Machine Learning / CRM Analytics')
projetos_data['Objetivo'].append('Sistema de scoring de leads usando ML (Random Forest) para prever probabilidade de conversão em matrículas. Pipeline completo com geração de listas operacionais e relatórios gerenciais.')
projetos_data['Status'].append('Produção')
projetos_data['Fontes de Dados'].append('HubSpot (Leads, Negócios Perdidos), Base Interna (Matrículas Finais)')
projetos_data['Scripts Principais'].append('0.Master_Pipeline.py (Orquestrador), 1.ML_Lead_Scoring.py, 4.Analise_Unidades.py, 6.Relatorio_Executivo.py, 7.Gerar_Listas_Unidades.py, 10.Validacao_Conversao_Helio.py, 5.Consolidador_Historico.py')
projetos_data['Outputs Gerados'].append('Scoring por Lead, Listas Operacionais por Unidade, Relatórios Executivos, Validação de Conversão, Histórico de Evolução')
projetos_data['Stakeholders'].append('Unidades de Negócio, Gerência Comercial, TI')
projetos_data['Frequência Atualização'].append('Diário/Sob Demanda')
projetos_data['Dashboards/Looker'].append('N/A')
projetos_data['Tecnologias'].append('Python, Scikit-learn (Random Forest), Pandas, Openpyxl')
projetos_data['Observações'].append('Pipeline automatizado com 5 etapas: Core ML, Operacional, Gerencial, Histórico, Limpeza. Backup em Caixa Forte.')

# Projeto 2: Análise Performance Mídia Paga
projetos_data['ID'].append('PRJ-002')
projetos_data['Nome do Projeto'].append('Análise Performance Mídia Paga')
projetos_data['Categoria'].append('Marketing Analytics / Performance Marketing')
projetos_data['Objetivo'].append('Pipeline integrado de análise de performance de mídia paga: Meta Ads + Google Ads + HubSpot CRM. Normalização de funil, atribuição de investimento por lead (prorrateio), relatórios de RVO e Matrículas por Ciclo.')
projetos_data['Status'].append('Produção')
projetos_data['Fontes de Dados'].append('Meta Ads, Google Ads, HubSpot CRM')
projetos_data['Scripts Principais'].append('1.analise_performance_meta.py, 2.analise_performance_google.py, 3.analise_performance_hubspot_FINAL_ID.py, 4.gerar_cenarios_preditivos.py')
projetos_data['Outputs Gerados'].append('Hubspot_Meta_Base.xlsx (Base Consolidada), Relatórios de CPL, RVO, Custo por Negócio')
projetos_data['Stakeholders'].append('Marketing, Comercial, BI')
projetos_data['Frequência Atualização'].append('Semanal/Quinzenal')
projetos_data['Dashboards/Looker'].append('Looker Studio (Meta Dashboard, HubSpot Dashboard)')
projetos_data['Tecnologias'].append('Python, Pandas, Excel')
projetos_data['Observações'].append('Investimentos oficiais do Meta Ads. Prorrateio proporcional ao número de negócios. Fluxo: Meta > HubSpot > Validação.')

# Projeto 3: Análise Eficiência Canal
projetos_data['ID'].append('PRJ-003')
projetos_data['Nome do Projeto'].append('Análise Eficiência Canal')
projetos_data['Categoria'].append('Sales Analytics / Channel Performance')
projetos_data['Objetivo'].append('Análise forense e auditoria de eficiência dos canais de captação (Out-Dez 2025). Compara dados de HubSpot com planilhas financeiras, validação de ciclos e geração de matriz de desempenho.')
projetos_data['Status'].append('Concluído/Auditoria')
projetos_data['Fontes de Dados'].append('HubSpot, Planilhas Financeiras (Fechamentos Mensais), Backup de Bases')
projetos_data['Scripts Principais'].append('analise_performance.py, analise_forense_corrigida.py, gerar_matriz_desempenho.py, gerar_relatorio_final_auditoria.py, comparar_excels_com_hubspot.py, auditoria_integridade_dados.py')
projetos_data['Outputs Gerados'].append('Relatórios de Auditoria, Matriz de Desempenho, Comparativos de Ciclos')
projetos_data['Stakeholders'].append('Financeiro, Comercial, Auditoria')
projetos_data['Frequência Atualização'].append('Por Ciclo (Trimestral/Semestral)')
projetos_data['Dashboards/Looker'].append('N/A')
projetos_data['Tecnologias'].append('Python, Pandas, Openpyxl')
projetos_data['Observações'].append('Foco em Ciclo 26.1 (Out/25-Jan/26). Múltiplos scripts de validação e investigação. Complexo: 32 scripts.')

# Projeto 4: Auditoria de Bases
projetos_data['ID'].append('PRJ-004')
projetos_data['Nome do Projeto'].append('Auditoria de Bases (Leads Sumidos)')
projetos_data['Categoria'].append('Data Quality / Auditoria')
projetos_data['Objetivo'].append('Investigação de leads sumidos/duplicados entre bases HubSpot e internas. Análise de impactos por secretarias e unidades.')
projetos_data['Status'].append('Concluído')
projetos_data['Fontes de Dados'].append('HubSpot (leads, negócios), Contagem de Negócios, Matrículas')
projetos_data['Scripts Principais'].append('analise_duplicados.py, analise_duplicatas_rigorosa.py, analise_leads_sumidos.py, gerar_excel_leads_sumidos.py, ranking_impactos.py')
projetos_data['Outputs Gerados'].append('ANALISE_DUPLICATAS_LEADS_SUMIDOS.xlsx, RANKING_SECRETARIAS_UNIDADES_IMPACTADAS.xlsx, LEADS_SUMIDOS_ANALISE_COMPLETA.xlsx')
projetos_data['Stakeholders'].append('TI, Qualidade de Dados, Operações')
projetos_data['Frequência Atualização'].append('Sob Demanda (Investigação)')
projetos_data['Dashboards/Looker'].append('N/A')
projetos_data['Tecnologias'].append('Python, Pandas')
projetos_data['Observações'].append('Análise profunda de IDs e duplicatas rigorosas. Impacto em matrículas e negócios perdidos.')

# Projeto 5: Correção Preenchimento Meta Call Center
projetos_data['ID'].append('PRJ-005')
projetos_data['Nome do Projeto'].append('Correção Preenchimento Meta Call Center')
projetos_data['Categoria'].append('Data Quality / Saneamento')
projetos_data['Objetivo'].append('Saneamento e correção de dados de meta/descobertas em negócios do Call Center. Validação de preenchimento correto.')
projetos_data['Status'].append('Concluído')
projetos_data['Fontes de Dados'].append('Novos Negócios (Exportações do HubSpot em datas específicas)')
projetos_data['Scripts Principais'].append('Script/saneamento_leads.py, Script/analise_meta_nascimento.py, Script/diagnostico_meta.py, Script/validar_amostra_diogo.py')
projetos_data['Outputs Gerados'].append('Novos Negócios_com_descobertas_CORRIGIDO.xlsx, Relatórios de Cobertura/Mudanças/NãoEncontrados')
projetos_data['Stakeholders'].append('Call Center, Operações, CRM')
projetos_data['Frequência Atualização'].append('Única (Projeto de Correção)')
projetos_data['Dashboards/Looker'].append('N/A')
projetos_data['Tecnologias'].append('Python, Pandas, Openpyxl')
projetos_data['Observações'].append('Foco em datas 04/02 e 05/02. Análise de Meta de Nascimento e preenchimento de campos.')

# Projeto 6: Análise Curva de Alunos
projetos_data['ID'].append('PRJ-006')
projetos_data['Nome do Projeto'].append('Análise Curva de Alunos')
projetos_data['Categoria'].append('Business Analytics / Student Lifecycle')
projetos_data['Objetivo'].append('Análise de comportamento da curva de alunos. Geração de base para Looker Studio.')
projetos_data['Status'].append('Ativo')
projetos_data['Fontes de Dados'].append('DATA/ (Bases internas de alunos)')
projetos_data['Scripts Principais'].append('SCRIPT/gerador_base_looker.py')
projetos_data['Outputs Gerados'].append('Base para Looker Studio (Análise de Retenção/Evolução de Alunos)')
projetos_data['Stakeholders'].append('Educacional, BI, Estratégia')
projetos_data['Frequência Atualização'].append('Mensal/Trimestral')
projetos_data['Dashboards/Looker'].append('Looker Studio')
projetos_data['Tecnologias'].append('Python, Pandas')
projetos_data['Observações'].append('Projeto enxuto: 1 script principal. Foco em visualização no Looker.')

# Projeto 7: Análise Cultura Inglesa CEFR
projetos_data['ID'].append('PRJ-007')
projetos_data['Nome do Projeto'].append('Análise Cultura Inglesa CEFR')
projetos_data['Categoria'].append('Educacional / Web Scraping')
projetos_data['Objetivo'].append('Raspagem de dados de certificados da Cultura Inglesa. Análise de níveis CEFR.')
projetos_data['Status'].append('Concluído')
projetos_data['Fontes de Dados'].append('Web Scraping (Certificados Online)')
projetos_data['Scripts Principais'].append('N/A (Dados já coletados)')
projetos_data['Outputs Gerados'].append('Dados_Brutos_Certificados_Raspados.csv')
projetos_data['Stakeholders'].append('Educacional, Pesquisa')
projetos_data['Frequência Atualização'].append('Única (Coleta Pontual)')
projetos_data['Dashboards/Looker'].append('N/A')
projetos_data['Tecnologias'].append('Python, Web Scraping')
projetos_data['Observações'].append('Dados brutos coletados. Possível base para análise comparativa CEFR.')

# Projeto 8: Pesquisa Correlação CEFR Cultura Inglesa
projetos_data['ID'].append('PRJ-008')
projetos_data['Nome do Projeto'].append('Pesquisa Correlação CEFR')
projetos_data['Categoria'].append('Pesquisa / Análise Estatística')
projetos_data['Objetivo'].append('Pesquisa de correlação entre desempenho interno e padrões CEFR da Cultura Inglesa.')
projetos_data['Status'].append('Ativo/Pesquisa')
projetos_data['Fontes de Dados'].append('Subfolder: Pesquisa_Correlacao_CEFR_Cultura_inglesa/')
projetos_data['Scripts Principais'].append('N/A (Em desenvolvimento)')
projetos_data['Outputs Gerados'].append('N/A')
projetos_data['Stakeholders'].append('Educacional, Pesquisa, Qualidade')
projetos_data['Frequência Atualização'].append('Projeto de Pesquisa')
projetos_data['Dashboards/Looker'].append('N/A')
projetos_data['Tecnologias'].append('Python, Estatística')
projetos_data['Observações'].append('Subprojeto dentro do mesmo diretório. Pesquisa acadêmica/estratégica.')

# Projeto 9: Análise Performance Valor The News
projetos_data['ID'].append('PRJ-009')
projetos_data['Nome do Projeto'].append('Análise Performance Valor The News')
projetos_data['Categoria'].append('Marketing Analytics / Content Performance')
projetos_data['Objetivo'].append('Análise de performance e valor gerado por leads oriundos da plataforma The News.')
projetos_data['Status'].append('Ativo')
projetos_data['Fontes de Dados'].append('DATA/ (Leads The News, Conversões)')
projetos_data['Scripts Principais'].append('SCRIPTS/analisar_leads.py, SCRIPTS/gerar_relatorio.py')
projetos_data['Outputs Gerados'].append('OUTPUTS/ (Relatórios de Performance e ROI)')
projetos_data['Stakeholders'].append('Marketing, Parcerias, BI')
projetos_data['Frequência Atualização'].append('Mensal')
projetos_data['Dashboards/Looker'].append('Potencial (Looker Studio)')
projetos_data['Tecnologias'].append('Python, Pandas')
projetos_data['Observações'].append('Foco em avaliar valor do canal The News. 2 scripts principais.')

# Projeto 10: Comparativo Unidades
projetos_data['ID'].append('PRJ-010')
projetos_data['Nome do Projeto'].append('Comparativo Unidades')
projetos_data['Categoria'].append('Business Intelligence / Benchmarking')
projetos_data['Objetivo'].append('Análise comparativa de performance entre unidades (franquias vs próprias).')
projetos_data['Status'].append('Ativo')
projetos_data['Fontes de Dados'].append('data/ (Consolidação de múltiplas unidades)')
projetos_data['Scripts Principais'].append('scripts/analise_divergencia.py')
projetos_data['Outputs Gerados'].append('outputs/ (Relatórios Comparativos)')
projetos_data['Stakeholders'].append('Diretoria, Franquias, Operações')
projetos_data['Frequência Atualização'].append('Mensal/Trimestral')
projetos_data['Dashboards/Looker'].append('Potencial')
projetos_data['Tecnologias'].append('Python, Pandas')
projetos_data['Observações'].append('Benchmarking de unidades. Insights para gestão estratégica.')

# Projeto 11: Apresentação Radar
projetos_data['ID'].append('PRJ-011')
projetos_data['Nome do Projeto'].append('Apresentação Radar')
projetos_data['Categoria'].append('Visualização / Apresentação')
projetos_data['Objetivo'].append('Apresentações visuais em formato radar/balloon para posts Instagram.')
projetos_data['Status'].append('Concluído')
projetos_data['Fontes de Dados'].append('N/A (Templates HTML)')
projetos_data['Scripts Principais'].append('balloon-instagram.html')
projetos_data['Outputs Gerados'].append('Posts/ (Visualizações para redes sociais)')
projetos_data['Stakeholders'].append('Marketing, Comunicação')
projetos_data['Frequência Atualização'].append('Sob Demanda')
projetos_data['Dashboards/Looker'].append('N/A')
projetos_data['Tecnologias'].append('HTML, CSS, JavaScript (Visualização)')
projetos_data['Observações'].append('Projeto de comunicação visual. Não é analytics tradicional.')

# Projeto 12: Ambiente Teste Isolado
projetos_data['ID'].append('PRJ-012')
projetos_data['Nome do Projeto'].append('Ambiente Teste Isolado (2025-12-15)')
projetos_data['Categoria'].append('Desenvolvimento / Testing')
projetos_data['Objetivo'].append('Ambiente isolado para testes de scripts e validações antes de produção.')
projetos_data['Status'].append('Ambiente de Teste')
projetos_data['Fontes de Dados'].append('Data/ (Cópias de segurança para teste)')
projetos_data['Scripts Principais'].append('Scripts/ (Versões de teste), EXECUTAR_TESTE.bat, GERAR_PACOTE_FINAL.bat')
projetos_data['Outputs Gerados'].append('Outputs/, Entrega_Final_Producao/')
projetos_data['Stakeholders'].append('Desenvolvimento, QA')
projetos_data['Frequência Atualização'].append('Sob Demanda (Teste)')
projetos_data['Dashboards/Looker'].append('N/A')
projetos_data['Tecnologias'].append('Python, Batch Scripts')
projetos_data['Observações'].append('Sandbox para desenvolvimento seguro. Evita impactos em produção.')

# Projeto 13: Análise Performance Mídia Orgânica
projetos_data['ID'].append('PRJ-013')
projetos_data['Nome do Projeto'].append('Análise Performance Mídia Orgânica')
projetos_data['Categoria'].append('Marketing Analytics / Organic Channels')
projetos_data['Objetivo'].append('Análise de performance dos canais orgânicos (SEO, Social, Direct).')
projetos_data['Status'].append('Planejado/Vazio')
projetos_data['Fontes de Dados'].append('N/A')
projetos_data['Scripts Principais'].append('N/A')
projetos_data['Outputs Gerados'].append('N/A')
projetos_data['Stakeholders'].append('Marketing, BI')
projetos_data['Frequência Atualização'].append('A definir')
projetos_data['Dashboards/Looker'].append('A definir')
projetos_data['Tecnologias'].append('A definir')
projetos_data['Observações'].append('Pasta vazia. Projeto futuro ou descontinuado.')

# Projeto 14: Análise Performance RVO
projetos_data['ID'].append('PRJ-014')
projetos_data['Nome do Projeto'].append('Análise Performance RVO')
projetos_data['Categoria'].append('Sales Analytics / RVO Tracking')
projetos_data['Objetivo'].append('Análise de RVO (Realizado vs Objetivo) por período.')
projetos_data['Status'].append('Planejado/Vazio')
projetos_data['Fontes de Dados'].append('N/A')
projetos_data['Scripts Principais'].append('N/A')
projetos_data['Outputs Gerados'].append('N/A')
projetos_data['Stakeholders'].append('Comercial, BI')
projetos_data['Frequência Atualização'].append('A definir')
projetos_data['Dashboards/Looker'].append('A definir')
projetos_data['Tecnologias'].append('A definir')
projetos_data['Observações'].append('Pasta vazia. Possível consolidação com outros projetos de RVO.')

# Projeto 15: Projeto Helio Teste
projetos_data['ID'].append('PRJ-015')
projetos_data['Nome do Projeto'].append('Projeto Helio Teste')
projetos_data['Categoria'].append('Desenvolvimento / Testing')
projetos_data['Objetivo'].append('Versão de teste/desenvolvimento do Projeto Helio.')
projetos_data['Status'].append('Desenvolvimento')
projetos_data['Fontes de Dados'].append('N/A (Cópia do PRJ-001)')
projetos_data['Scripts Principais'].append('N/A')
projetos_data['Outputs Gerados'].append('N/A')
projetos_data['Stakeholders'].append('Desenvolvimento')
projetos_data['Frequência Atualização'].append('Sob Demanda')
projetos_data['Dashboards/Looker'].append('N/A')
projetos_data['Tecnologias'].append('Python, ML')
projetos_data['Observações'].append('Ambiente de testes para Helio. Não usar em produção.')

# Criar DataFrame
df = pd.DataFrame(projetos_data)

# Criar planilha Excel com formatação profissional
wb = Workbook()
ws = wb.active
ws.title = "Inventário Projetos"

# Estilos
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
alignment_center = Alignment(horizontal="center", vertical="top", wrap_text=True)
alignment_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Adicionar cabeçalho do documento
ws.merge_cells('A1:M1')
ws['A1'] = 'INVENTÁRIO DE PROJETOS DE DADOS - DIRETÓRIO C:\\Users\\a483650\\Projetos'
ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells('A2:M2')
ws['A2'] = f'Data de Geração: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
ws['A2'].font = Font(size=10, italic=True)
ws['A2'].alignment = Alignment(horizontal="center")

# Adicionar dados
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 4):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        
        if r_idx == 4:  # Cabeçalho
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment_center
        else:
            if c_idx == 1:  # ID
                cell.alignment = alignment_center
            else:
                cell.alignment = alignment_left

# Ajustar largura das colunas
column_widths = {
    'A': 10,   # ID
    'B': 35,   # Nome
    'C': 25,   # Categoria
    'D': 50,   # Objetivo
    'E': 15,   # Status
    'F': 40,   # Fontes
    'G': 50,   # Scripts
    'H': 40,   # Outputs
    'I': 25,   # Stakeholders
    'J': 20,   # Frequência
    'K': 25,   # Dashboards
    'L': 30,   # Tecnologias
    'M': 50    # Observações
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Ajustar altura das linhas de dados
for row in range(5, ws.max_row + 1):
    ws.row_dimensions[row].height = 60

# Congelar painéis
ws.freeze_panes = 'A5'

# Adicionar aba de resumo
ws_resumo = wb.create_sheet("Resumo Executivo")

# Criar resumo por categoria
resumo_categoria = df['Categoria'].value_counts().reset_index()
resumo_categoria.columns = ['Categoria', 'Quantidade']

resumo_status = df['Status'].value_counts().reset_index()
resumo_status.columns = ['Status', 'Quantidade']

# Cabeçalho do resumo
ws_resumo.merge_cells('A1:D1')
ws_resumo['A1'] = 'RESUMO EXECUTIVO - PROJETOS DE DADOS'
ws_resumo['A1'].font = Font(bold=True, size=14, color="1F4E78")
ws_resumo['A1'].alignment = Alignment(horizontal="center")

# Resumo por categoria
ws_resumo['A3'] = 'DISTRIBUIÇÃO POR CATEGORIA'
ws_resumo['A3'].font = Font(bold=True, size=12)

row_start = 4
for r_idx, row in enumerate(dataframe_to_rows(resumo_categoria, index=False, header=True), row_start):
    for c_idx, value in enumerate(row, 1):
        cell = ws_resumo.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        if r_idx == row_start:
            cell.fill = header_fill
            cell.font = header_font

# Resumo por status
row_start = row_start + len(resumo_categoria) + 3
ws_resumo[f'A{row_start}'] = 'DISTRIBUIÇÃO POR STATUS'
ws_resumo[f'A{row_start}'].font = Font(bold=True, size=12)

row_start += 1
for r_idx, row in enumerate(dataframe_to_rows(resumo_status, index=False, header=True), row_start):
    for c_idx, value in enumerate(row, 1):
        cell = ws_resumo.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        if r_idx == row_start:
            cell.fill = header_fill
            cell.font = header_font

# Estatísticas gerais
row_start = row_start + len(resumo_status) + 3
ws_resumo[f'A{row_start}'] = 'ESTATÍSTICAS GERAIS'
ws_resumo[f'A{row_start}'].font = Font(bold=True, size=12)

stats = [
    ['Total de Projetos', len(df)],
    ['Projetos em Produção', len(df[df['Status'] == 'Produção'])],
    ['Projetos Ativos', len(df[df['Status'] == 'Ativo'])],
    ['Projetos com Dashboard Looker', len(df[df['Dashboards/Looker'].str.contains('Looker', na=False)])],
    ['Projetos Python', len(df[df['Tecnologias'].str.contains('Python', na=False)])],
]

row_start += 1
for i, (label, value) in enumerate(stats, row_start):
    ws_resumo[f'A{i}'] = label
    ws_resumo[f'B{i}'] = value
    ws_resumo[f'A{i}'].font = Font(bold=True)
    ws_resumo[f'A{i}'].border = border
    ws_resumo[f'B{i}'].border = border

ws_resumo.column_dimensions['A'].width = 35
ws_resumo.column_dimensions['B'].width = 20

# Salvar
import os
output_path = os.path.join(os.path.dirname(__file__), 'INVENTARIO_PROJETOS_DATA_ANALYTICS.xlsx')
wb.save(output_path)

print(f"✅ Inventário criado com sucesso!")
print(f"📊 Total de projetos documentados: {len(df)}")
print(f"📁 Arquivo salvo em: {output_path}")
print(f"\n📋 Resumo:")
print(f"   - Projetos em Produção: {len(df[df['Status'] == 'Produção'])}")
print(f"   - Projetos Ativos: {len(df[df['Status'] == 'Ativo'])}")
print(f"   - Projetos Concluídos: {len(df[df['Status'] == 'Concluído'])}")
print(f"   - Projetos em Desenvolvimento: {len(df[df['Status'].str.contains('Desenvolvimento|Teste', na=False)])}")
