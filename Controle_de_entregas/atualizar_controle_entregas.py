"""
Script para atualizar a planilha de controle de entregas
com o ecossistema completo de data science criado
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Dados do ecossistema criado
entregas = [
    {
        "Categoria": "🏗️ INFRAESTRUTURA",
        "Entrega": "_DADOS_CENTRALIZADOS/",
        "Descrição": "Repositório único centralizado para todas as bases de dados oficiais. Única fonte de verdade (Single Source of Truth) para HubSpot, Matrículas e Meta Ads.",
        "Uso": "• Eliminar duplicação de dados\n• Garantir que todos os projetos usem a mesma versão das bases\n• Facilitar atualização (atualiza em 1 lugar, todos os projetos se beneficiam)\n• Manter histórico de versões antigas",
        "Atualização": "Sob demanda (quando receber novos exports do HubSpot/Google Sheets)",
        "Localização": "C:\\Users\\a483650\\Projetos\\_DADOS_CENTRALIZADOS\\"
    },
    {
        "Categoria": "🏗️ INFRAESTRUTURA",
        "Entrega": "_SCRIPTS_COMPARTILHADOS/",
        "Descrição": "Pasta com scripts utilitários que servem para todo o ecossistema de projetos.",
        "Uso": "• sincronizar_bases.py: Sincroniza bases do central para todos os projetos\n• validar_reorganizacao.py: Testa se projetos conseguem acessar dados\n• inventario_projetos.py: Gera Excel com inventário completo\n• analisar_duplicacoes.py: Identifica arquivos duplicados com MD5 hash",
        "Atualização": "Conforme necessidade de manutenção",
        "Localização": "C:\\Users\\a483650\\Projetos\\_SCRIPTS_COMPARTILHADOS\\"
    },
    {
        "Categoria": "🏗️ INFRAESTRUTURA",
        "Entrega": "_ARQUIVO/",
        "Descrição": "Pasta para projetos inativos/descontinuados que não devem ser deletados mas não estão em uso ativo.",
        "Uso": "• Manter histórico de projetos antigos\n• Liberar espaço visual na estrutura principal\n• Permitir consulta futura se necessário",
        "Atualização": "Estático (arquivamento)",
        "Localização": "C:\\Users\\a483650\\Projetos\\_ARQUIVO\\"
    },
    {
        "Categoria": "🚀 PRODUÇÃO",
        "Entrega": "01_Helio_ML_Producao",
        "Descrição": "Sistema de Machine Learning para Lead Scoring. Prevê probabilidade de conversão de leads usando Random Forest. Gera relatórios automáticos por unidade e dashboards de performance.",
        "Uso": "• Scoring diário/semanal de leads do HubSpot\n• Priorização de contatos para o time comercial\n• Análise de features importantes para conversão\n• Relatórios para gestores de cada unidade\n• Dashboard de precisão do modelo",
        "Atualização": "Semanal (ou sob demanda)",
        "Localização": "C:\\Users\\a483650\\Projetos\\01_Helio_ML_Producao\\"
    },
    {
        "Categoria": "🚀 PRODUÇÃO",
        "Entrega": "02_Pipeline_Midia_Paga",
        "Descrição": "Pipeline de análise de performance de campanhas de Meta Ads (Facebook/Instagram). Cruza dados de anúncios com conversões do HubSpot para ROI real.",
        "Uso": "• Análise de performance de campanhas Meta\n• Cálculo de ROI real (investimento vs receita gerada)\n• Identificação de criativos e segmentações mais eficientes\n• Recomendações para otimização de budget",
        "Atualização": "Semanal",
        "Localização": "C:\\Users\\a483650\\Projetos\\02_Pipeline_Midia_Paga\\"
    },
    {
        "Categoria": "📊 ANÁLISES OPERACIONAIS",
        "Entrega": "03_Analises_Operacionais/",
        "Descrição": "Categoria agrupando 4 projetos de análises pontuais/ad-hoc que suportam decisões operacionais.",
        "Uso": "• analise_geral_ep/: Análise exploratória geral de EP\n• eficiencia_canal/: Analisa eficiência de canais de aquisição\n• Analises_Gustavo/: Análises diversas solicitadas pelo Gustavo\n• analise_performance_midiapaga/ (backup): Versão antiga do pipeline",
        "Atualização": "Sob demanda (análises pontuais)",
        "Localização": "C:\\Users\\a483650\\Projetos\\03_Analises_Operacionais\\"
    },
    {
        "Categoria": "🔍 AUDITORIAS DE QUALIDADE",
        "Entrega": "04_Auditorias_Qualidade/",
        "Descrição": "Categoria agrupando 2 projetos de auditoria de dados para garantir qualidade e confiabilidade das bases.",
        "Uso": "• auditoria_leads_sumidos/: Investiga leads que somem do pipeline\n• auditoria_taxas/: Valida cálculos de taxas de conversão",
        "Atualização": "Mensal (ou quando suspeitar de problemas nos dados)",
        "Localização": "C:\\Users\\a483650\\Projetos\\04_Auditorias_Qualidade\\"
    },
    {
        "Categoria": "📚 PESQUISAS EDUCACIONAIS",
        "Entrega": "05_Pesquisas_Educacionais/",
        "Descrição": "Categoria agrupando 2 projetos de pesquisa e análise do perfil educacional dos alunos.",
        "Uso": "• Analise_Perfil_Alunos/: Análise demográfica e comportamental\n• Perfil_Educacional_Alunos/: Foco em background educacional",
        "Atualização": "Trimestral",
        "Localização": "C:\\Users\\a483650\\Projetos\\05_Pesquisas_Educacionais\\"
    },
    {
        "Categoria": "📋 DOCUMENTAÇÃO",
        "Entrega": "INVENTARIO_PROJETOS_DATA_ANALYTICS.xlsx",
        "Descrição": "Planilha Excel com inventário completo de todos os arquivos do ecossistema: tipo, tamanho, última modificação, duplicações.",
        "Uso": "• Visão geral de todos os arquivos CSV/XLSX\n• Identificar arquivos não utilizados\n• Rastrear crescimento dos dados ao longo do tempo\n• Planejar limpezas futuras",
        "Atualização": "Mensal (reexecutar inventario_projetos.py)",
        "Localização": "C:\\Users\\a483650\\Projetos\\Controle_de_entregas\\INVENTARIO_PROJETOS_DATA_ANALYTICS.xlsx"
    },
    {
        "Categoria": "📋 DOCUMENTAÇÃO",
        "Entrega": "RELATORIO_DUPLICACOES.xlsx",
        "Descrição": "Relatório Excel listando todos os grupos de arquivos duplicados identificados por hash MD5.",
        "Uso": "• Identificar oportunidades de economia de espaço\n• Planejar estratégia de limpeza\n• Evitar deletar arquivos ainda em uso",
        "Atualização": "Mensal (reexecutar analisar_duplicacoes.py)",
        "Localização": "C:\\Users\\a483650\\Projetos\\Controle_de_entregas\\RELATORIO_DUPLICACOES.xlsx"
    },
    {
        "Categoria": "📋 DOCUMENTAÇÃO",
        "Entrega": "RELATORIO_LIMPEZA_FINAL.md",
        "Descrição": "Documento Markdown detalhando toda a limpeza executada: arquivos deletados, espaço liberado, validações realizadas.",
        "Uso": "• Registrar histórico da reorganização\n• Localizar backup de segurança\n• Auditar decisões de deleção\n• Restaurar arquivos se necessário",
        "Atualização": "Estático (registro histórico)",
        "Localização": "C:\\Users\\a483650\\Projetos\\Controle_de_entregas\\RELATORIO_LIMPEZA_FINAL.md"
    },
    {
        "Categoria": "🔧 SCRIPTS DE AUTOMAÇÃO",
        "Entrega": "REORGANIZAR_MASTER.ps1",
        "Descrição": "Script PowerShell master que orquestra toda a reorganização em 3 fases: preparação, consolidação e limpeza.",
        "Uso": "• Replicar estrutura em outros ambientes\n• Documentar processo de reorganização\n• Facilitar futuras reorganizações",
        "Atualização": "Conforme evolução da estrutura",
        "Localização": "C:\\Users\\a483650\\Projetos\\Controle_de_entregas\\REORGANIZAR_MASTER.ps1"
    }
]

# Criar DataFrame
df = pd.DataFrame(entregas)

# Salvar em Excel com formatação
output_path = r"C:\Users\a483650\Projetos\Controle_de_entregas\CONTROLE_ENTREGAS_ECOSSISTEMA_DATA_SCIENCE.xlsx"
df.to_excel(output_path, index=False, sheet_name="Ecossistema")

# Aplicar formatação avançada
wb = load_workbook(output_path)
ws = wb["Ecossistema"]

# Definir estilos
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Cores por categoria
category_colors = {
    "🏗️ INFRAESTRUTURA": "DCE6F1",
    "🚀 PRODUÇÃO": "C6E0B4",
    "📊 ANÁLISES OPERACIONAIS": "FFE699",
    "🔍 AUDITORIAS DE QUALIDADE": "F4B084",
    "📚 PESQUISAS EDUCACIONAIS": "D5A6BD",
    "📋 DOCUMENTAÇÃO": "B4C7E7",
    "🔧 SCRIPTS DE AUTOMAÇÃO": "A9D08E"
}

# Formatar cabeçalho
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Formatar dados
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    categoria = ws.cell(row_idx, 1).value
    color = category_colors.get(categoria, "FFFFFF")
    
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# Ajustar larguras das colunas
ws.column_dimensions['A'].width = 25  # Categoria
ws.column_dimensions['B'].width = 35  # Entrega
ws.column_dimensions['C'].width = 60  # Descrição
ws.column_dimensions['D'].width = 60  # Uso
ws.column_dimensions['E'].width = 20  # Atualização
ws.column_dimensions['F'].width = 50  # Localização

# Ajustar altura das linhas
ws.row_dimensions[1].height = 30
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 100

# Adicionar aba de resumo executivo
ws_resumo = wb.create_sheet("Resumo Executivo", 0)

resumo_data = [
    ["ECOSSISTEMA DE DATA SCIENCE - VISÃO GERAL"],
    [""],
    ["Data de Atualização:", datetime.now().strftime("%d/%m/%Y %H:%M")],
    [""],
    ["📊 ESTATÍSTICAS DO ECOSSISTEMA"],
    ["Métrica", "Valor"],
    ["Total de Projetos Ativos", "8 projetos"],
    ["Projetos em Produção", "2 (Helio ML + Pipeline Mídia Paga)"],
    ["Projetos Arquivados", "5 projetos"],
    ["Espaço Total", "~1.09 GB (após limpeza de 330 MB)"],
    ["Bases Centralizadas", "5 bases oficiais"],
    ["Scripts Compartilhados", "4 utilitários principais"],
    [""],
    ["🎯 PRINCÍPIOS DO ECOSSISTEMA"],
    ["Princípio", "Implementação"],
    ["Single Source of Truth", "_DADOS_CENTRALIZADOS/ contém TODAS as bases oficiais"],
    ["DRY (Don't Repeat Yourself)", "Scripts compartilhados reutilizáveis"],
    ["Separação por Função", "5 categorias: Produção, Análises, Auditorias, Pesquisas, Arquivo"],
    ["Versionamento", "Histórico de versões antigas em subpastas historico/"],
    ["Nomenclatura Clara", "Prefixos numerados (01_, 02_) para prioridade"],
    ["Automação", "Scripts de sincronização e validação automáticos"],
    [""],
    ["🔄 FLUXO DE TRABALHO RECOMENDADO"],
    ["Passo", "Ação"],
    ["1. Atualizar dados", "Copiar novos exports para _DADOS_CENTRALIZADOS/"],
    ["2. Sincronizar", "Executar: python _SCRIPTS_COMPARTILHADOS/sincronizar_bases.py"],
    ["3. Validar", "Executar: python _SCRIPTS_COMPARTILHADOS/validar_reorganizacao.py"],
    ["4. Executar projetos", "Rodar scripts de produção (01_Helio, 02_Pipeline)"],
    ["5. Análises ad-hoc", "Usar projetos em 03_Analises_Operacionais/"],
    [""],
    ["⚠️ REGRAS DE OURO"],
    ["Regra"],
    ["❌ NUNCA editar dados diretamente nos projetos - sempre no _DADOS_CENTRALIZADOS/"],
    ["❌ NUNCA deletar pastas sem backup - usar _BACKUP_SEGURANCA_*/"],
    ["✅ SEMPRE sincronizar após atualizar dados centrais"],
    ["✅ SEMPRE validar após mudanças estruturais"],
    ["✅ SEMPRE manter apenas 1 versão ATUAL de cada base"],
    [""],
    ["🚀 COMANDOS RÁPIDOS"],
    ["Ação", "Comando"],
    ["Sincronizar bases", "python C:\\Users\\a483650\\Projetos\\_SCRIPTS_COMPARTILHADOS\\sincronizar_bases.py"],
    ["Validar estrutura", "python C:\\Users\\a483650\\Projetos\\_SCRIPTS_COMPARTILHADOS\\validar_reorganizacao.py"],
    ["Inventário completo", "python C:\\Users\\a483650\\Projetos\\Controle_de_entregas\\inventario_projetos.py"],
    ["Analisar duplicatas", "python C:\\Users\\a483650\\Projetos\\Controle_de_entregas\\analisar_duplicacoes.py"],
]

for row_data in resumo_data:
    ws_resumo.append(row_data)

# Formatar resumo executivo
title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
title_font = Font(bold=True, color="FFFFFF", size=14)
ws_resumo['A1'].fill = title_fill
ws_resumo['A1'].font = title_font
ws_resumo.merge_cells('A1:B1')

section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_font = Font(bold=True, color="FFFFFF", size=12)

for row in [5, 14, 23, 33, 39]:
    ws_resumo[f'A{row}'].fill = section_fill
    ws_resumo[f'A{row}'].font = section_font
    ws_resumo.merge_cells(f'A{row}:B{row}')

ws_resumo.column_dimensions['A'].width = 40
ws_resumo.column_dimensions['B'].width = 70

wb.save(output_path)
print(f"✅ Planilha criada com sucesso: {output_path}")
print(f"📊 Total de entregas documentadas: {len(entregas)}")
print(f"📂 Abas criadas: Resumo Executivo + Ecossistema")
