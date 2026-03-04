"""
==============================================================================
VALIDAÇÃO DE CONVERSÕES OCULTAS - HELIO
==============================================================================
Script: 14.Validacao_Conversao_Oculta.py
Objetivo: Identificar leads que tinham nota no HubSpot, NÃO estavam com
          confirmação de matrícula, mas que de fato matricularam ao serem
          encontrados na base matriculas_finais.

          INCLUI ANÁLISE DE:
          1. Leads em qualificação/negociação que viraram matrícula
          2. Leads em Visita Agendada que viraram matrícula
          3. Leads em Visita Realizada que viraram matrícula

          Isso revela conversões que aconteceram "fora do radar" do HubSpot,
          validando a eficácia do modelo de scoring.

Data: 2026-01-30
==============================================================================
"""

import pandas as pd
import os
import glob
import unicodedata
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASTA_DATA = os.path.join(BASE_DIR, 'Data')
PASTA_OUTPUT = os.path.join(BASE_DIR, 'Outputs')
PASTA_RELATORIOS_ML = os.path.join(PASTA_OUTPUT, 'Relatorios_ML')
PASTA_DB_HELIO = os.path.join(PASTA_OUTPUT, 'DB_HELIO')

# Arquivos de entrada
CAMINHO_MATRICULAS = os.path.join(PASTA_DATA, 'matriculas_finais_limpo.csv')
CAMINHO_HUBSPOT = os.path.join(PASTA_DATA, 'hubspot_leads.csv')

# Arquivo de saída
DATA_HOJE = datetime.now().strftime('%Y-%m-%d')
CAMINHO_SAIDA = os.path.join(PASTA_DB_HELIO, f'Conversoes_Ocultas_{DATA_HOJE}.xlsx')

os.makedirs(PASTA_DB_HELIO, exist_ok=True)

# Marcos históricos para análise
MARCOS = {
    'Marco1_12Dez_2025': {
        'pasta': os.path.join(PASTA_RELATORIOS_ML, 'Relatório 12-12-25'),
        'descricao': '12/12/2025 - Baseline Pré-Helio',
        'data': datetime(2025, 12, 12)
    },
    'Marco2_05Jan_2026': {
        'pasta': os.path.join(PASTA_RELATORIOS_ML, 'Relatório 01-05-26'),
        'descricao': '05/01/2026 - Baseline Pré-Helio',
        'data': datetime(2026, 1, 5)
    },
    'Marco3_20Jan_2026': {
        'pasta': os.path.join(PASTA_RELATORIOS_ML, 'start'),
        'descricao': '20/01/2026 - Lançamento Helio',
        'data': datetime(2026, 1, 20)
    }
}

# Estilos Excel
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
FILL_SUCESSO = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FILL_ALERTA = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

# ==============================================================================
# FUNÇÕES AUXILIARES
# ==============================================================================

def normalizar_nome(nome):
    """
    Normaliza nome para matching entre bases.
    Remove acentos, sufixos de idade/unidade, e padroniza formato.
    """
    if pd.isna(nome):
        return ""
    nome = str(nome).upper().strip()
    # Remove sufixos como " - 10 - UNIDADE" ou " - UNIDADE"
    nome = re.sub(r'\s*-\s*\d+\s*-\s*.*$', '', nome)
    nome = re.sub(r'\s*-\s*[A-Z\s]+$', '', nome)
    # Remove acentos
    nome = unicodedata.normalize('NFKD', nome)
    nome = ''.join(c for c in nome if not unicodedata.combining(c))
    # Normaliza espaços
    nome = ' '.join(nome.split())
    return nome


def carregar_matriculas_finais():
    """
    Carrega base de matrículas finais (source of truth).
    Retorna um set com nomes normalizados e um DataFrame para análise.
    """
    print("\n[1] Carregando base de matrículas finais...")

    if not os.path.exists(CAMINHO_MATRICULAS):
        print(f"    [ERRO] Arquivo não encontrado: {CAMINHO_MATRICULAS}")
        return set(), pd.DataFrame()

    # Tentar diferentes encodings
    for encoding in ['utf-8-sig', 'utf-8', 'latin1']:
        try:
            df = pd.read_csv(CAMINHO_MATRICULAS, sep=';', encoding=encoding)
            break
        except:
            continue
    else:
        try:
            df = pd.read_csv(CAMINHO_MATRICULAS, sep=None, engine='python', encoding='utf-8')
        except Exception as e:
            print(f"    [ERRO] Falha ao ler arquivo: {e}")
            return set(), pd.DataFrame()

    df.columns = df.columns.str.strip()

    # Identificar colunas
    col_nome = next((c for c in df.columns if 'nome' in c.lower() and 'aluno' in c.lower()), None)
    col_tipo = next((c for c in df.columns if 'tipo' in c.lower()), None)
    col_unidade = next((c for c in df.columns if 'unidade' in c.lower()), None)
    col_status = next((c for c in df.columns if 'status' in c.lower() and 'p1' in c.lower()), None)

    if not col_nome:
        print("    [ERRO] Coluna 'Nome do Aluno' não encontrada")
        return set(), pd.DataFrame()

    # Filtrar apenas alunos NOVOS (excluir REMA)
    total_original = len(df)
    if col_tipo:
        df = df[~df[col_tipo].astype(str).str.contains('REMA', case=False, na=False)]

    # Normalizar nomes
    df['Nome_Normalizado'] = df[col_nome].apply(normalizar_nome)

    set_matriculas = set(df['Nome_Normalizado'].unique())

    print(f"    Total de matrículas: {len(set_matriculas)} (de {total_original} registros)")
    if col_tipo:
        print(f"    (Excluídos {total_original - len(df)} rematrículas)")

    return set_matriculas, df


def carregar_leads_marco(pasta_marco, nome_marco):
    """
    Carrega todos os leads pontuados de um marco específico.
    Consolida dados de todas as planilhas Lead_Scoring_*.xlsx.
    """
    if not os.path.exists(pasta_marco):
        print(f"    [!] Pasta não encontrada: {pasta_marco}")
        return pd.DataFrame()

    arquivos = glob.glob(os.path.join(pasta_marco, 'Lead_Scoring_*.xlsx'))

    # Excluir arquivos temporários
    arquivos = [a for a in arquivos if not os.path.basename(a).startswith('~$')]

    if not arquivos:
        print(f"    [!] Nenhum arquivo Lead_Scoring encontrado")
        return pd.DataFrame()

    dfs = []

    for arquivo in arquivos:
        try:
            xl = pd.ExcelFile(arquivo)

            # Abas que contêm leads pontuados
            abas_relevantes = ['2_Top500_Nota5', '3_Nota4', '4_Top10_Por_Unidade', '5_Em_Qualificacao']

            for aba in abas_relevantes:
                if aba in xl.sheet_names:
                    # Buscar linha do header dinamicamente
                    df_preview = pd.read_excel(xl, sheet_name=aba, header=None, nrows=20)
                    header_idx = -1

                    for idx, row in df_preview.iterrows():
                        row_str = row.astype(str).str.lower()
                        if row_str.str.contains('record id|nome do negócio', regex=True, na=False).any():
                            header_idx = idx
                            break

                    if header_idx != -1:
                        df_temp = pd.read_excel(xl, sheet_name=aba, header=header_idx)
                        df_temp.columns = df_temp.columns.astype(str).str.strip()
                        df_temp['Origem_Aba'] = aba
                        df_temp['Arquivo_Origem'] = os.path.basename(arquivo)
                        dfs.append(df_temp)

        except Exception as e:
            print(f"    [!] Erro ao ler {os.path.basename(arquivo)}: {e}")

    if dfs:
        df_consolidado = pd.concat(dfs, ignore_index=True)
        return df_consolidado

    return pd.DataFrame()


def identificar_colunas(df):
    """Identifica colunas relevantes no DataFrame."""
    colunas = {
        'nome': None,
        'etapa': None,
        'nota': None,
        'unidade': None,
        'record_id': None,
        'probabilidade': None
    }

    for col in df.columns:
        col_lower = col.lower()
        if 'nome' in col_lower and 'negócio' in col_lower:
            colunas['nome'] = col
        elif 'nome' in col_lower and not colunas['nome']:
            colunas['nome'] = col
        elif 'etapa' in col_lower:
            colunas['etapa'] = col
        elif 'nota' in col_lower and '1a5' in col_lower:
            colunas['nota'] = col
        elif 'nota' in col_lower and not colunas['nota']:
            colunas['nota'] = col
        elif 'unidade' in col_lower and 'desejada' in col_lower:
            colunas['unidade'] = col
        elif 'unidade' in col_lower and not colunas['unidade']:
            colunas['unidade'] = col
        elif 'record' in col_lower and 'id' in col_lower:
            colunas['record_id'] = col
        elif 'probabilidade' in col_lower:
            colunas['probabilidade'] = col

    return colunas


def df_to_sheet(ws, df, start_row=1, highlight_col=None):
    """Adiciona DataFrame ao sheet com formatação."""
    if df.empty:
        return

    # Limpar caracteres ilegais
    df_safe = df.copy()
    for col in df_safe.select_dtypes(include=['object']):
        df_safe[col] = df_safe[col].astype(str).apply(
            lambda x: re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', x)
        )

    for r_idx, row in enumerate(dataframe_to_rows(df_safe, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER

            if r_idx == start_row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center', wrap_text=True)


def ajustar_largura(ws):
    """Auto-ajusta largura das colunas."""
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


# ==============================================================================
# EXECUÇÃO PRINCIPAL
# ==============================================================================

def main():
    print("="*80)
    print("VALIDAÇÃO DE CONVERSÕES OCULTAS - HELIO")
    print("="*80)
    print(f"Data de Execução: {DATA_HOJE}")
    print(f"Objetivo: Identificar leads com nota que viraram matrícula")
    print(f"          sem confirmação prévia no HubSpot")
    print("="*80)

    # 1. Carregar matrículas finais
    set_matriculas, df_matriculas = carregar_matriculas_finais()

    if not set_matriculas:
        print("\n[ERRO] Não foi possível carregar matrículas. Abortando.")
        return

    # 2. Processar cada marco
    print("\n[2] Processando marcos históricos...")

    resultados = {}
    todos_convertidos = []

    # Resultados separados por tipo de etapa
    resultados_qualificacao = {}
    resultados_visita_agendada = {}
    resultados_visita_realizada = {}

    for marco_id, info in MARCOS.items():
        print(f"\n  >> {info['descricao']}:")

        df_leads = carregar_leads_marco(info['pasta'], marco_id)

        if df_leads.empty:
            print(f"     Nenhum lead encontrado")
            continue

        # Identificar colunas
        cols = identificar_colunas(df_leads)

        if not cols['nome']:
            print(f"     [!] Coluna de nome não encontrada")
            continue

        print(f"     Total leads carregados: {len(df_leads)}")

        # Normalizar nome dos leads
        df_leads['Nome_Normalizado'] = df_leads[cols['nome']].apply(normalizar_nome)

        # Classificar etapa do lead
        if cols['etapa']:
            df_leads['Etapa_Lower'] = df_leads[cols['etapa']].astype(str).str.lower()

            # Máscaras por tipo de etapa
            mask_matricula = df_leads['Etapa_Lower'].str.contains('matrícula|matricula|ganho|won', na=False)
            mask_visita_realizada = df_leads['Etapa_Lower'].str.contains('visita realizada', na=False)
            mask_visita_agendada = df_leads['Etapa_Lower'].str.contains('visita agendada|visita marcada', na=False) & ~mask_visita_realizada
            mask_qualificacao = ~mask_matricula & ~mask_visita_realizada & ~mask_visita_agendada
        else:
            df_leads['Etapa_Lower'] = ''
            mask_matricula = pd.Series([False] * len(df_leads))
            mask_visita_realizada = pd.Series([False] * len(df_leads))
            mask_visita_agendada = pd.Series([False] * len(df_leads))
            mask_qualificacao = pd.Series([True] * len(df_leads))

        # Cruzar TODOS os leads com matrículas finais
        df_leads['Virou_Matricula'] = df_leads['Nome_Normalizado'].isin(set_matriculas)

        # =====================================================================
        # ANÁLISE 1: Leads em QUALIFICAÇÃO/NEGOCIAÇÃO que viraram matrícula
        # =====================================================================
        df_qualificacao = df_leads[mask_qualificacao].copy()
        df_qualif_convertidos = df_qualificacao[df_qualificacao['Virou_Matricula']].copy()

        # Remover duplicatas
        if cols['nota'] and not df_qualif_convertidos.empty:
            df_qualif_convertidos[cols['nota']] = pd.to_numeric(df_qualif_convertidos[cols['nota']], errors='coerce')
            df_qualif_convertidos = df_qualif_convertidos.sort_values(cols['nota'], ascending=False)
            df_qualif_convertidos = df_qualif_convertidos.drop_duplicates(subset=['Nome_Normalizado'], keep='first')

        n_qualif = len(df_qualif_convertidos)
        print(f"     Em Qualificação/Negociação -> Matrícula: {n_qualif}")

        if n_qualif > 0:
            df_qualif_convertidos['Marco'] = info['descricao']
            df_qualif_convertidos['Tipo_Conversao'] = 'Qualificação -> Matrícula'
            resultados_qualificacao[marco_id] = {
                'descricao': info['descricao'],
                'total': len(df_qualificacao),
                'convertidos': n_qualif,
                'df': df_qualif_convertidos
            }
            todos_convertidos.append(df_qualif_convertidos)

        # =====================================================================
        # ANÁLISE 2: Leads em VISITA AGENDADA que viraram matrícula
        # =====================================================================
        df_vis_agendada = df_leads[mask_visita_agendada].copy()
        df_vis_ag_convertidos = df_vis_agendada[df_vis_agendada['Virou_Matricula']].copy()

        # Remover duplicatas
        if cols['nota'] and not df_vis_ag_convertidos.empty:
            df_vis_ag_convertidos[cols['nota']] = pd.to_numeric(df_vis_ag_convertidos[cols['nota']], errors='coerce')
            df_vis_ag_convertidos = df_vis_ag_convertidos.sort_values(cols['nota'], ascending=False)
            df_vis_ag_convertidos = df_vis_ag_convertidos.drop_duplicates(subset=['Nome_Normalizado'], keep='first')

        n_vis_ag = len(df_vis_ag_convertidos)
        print(f"     Em Visita Agendada -> Matrícula: {n_vis_ag}")

        if n_vis_ag > 0:
            df_vis_ag_convertidos['Marco'] = info['descricao']
            df_vis_ag_convertidos['Tipo_Conversao'] = 'Visita Agendada -> Matrícula'
            resultados_visita_agendada[marco_id] = {
                'descricao': info['descricao'],
                'total': len(df_vis_agendada),
                'convertidos': n_vis_ag,
                'df': df_vis_ag_convertidos
            }
            todos_convertidos.append(df_vis_ag_convertidos)

        # =====================================================================
        # ANÁLISE 3: Leads em VISITA REALIZADA que viraram matrícula
        # =====================================================================
        df_vis_realizada = df_leads[mask_visita_realizada].copy()
        df_vis_re_convertidos = df_vis_realizada[df_vis_realizada['Virou_Matricula']].copy()

        # Remover duplicatas
        if cols['nota'] and not df_vis_re_convertidos.empty:
            df_vis_re_convertidos[cols['nota']] = pd.to_numeric(df_vis_re_convertidos[cols['nota']], errors='coerce')
            df_vis_re_convertidos = df_vis_re_convertidos.sort_values(cols['nota'], ascending=False)
            df_vis_re_convertidos = df_vis_re_convertidos.drop_duplicates(subset=['Nome_Normalizado'], keep='first')

        n_vis_re = len(df_vis_re_convertidos)
        print(f"     Em Visita Realizada -> Matrícula: {n_vis_re}")

        if n_vis_re > 0:
            df_vis_re_convertidos['Marco'] = info['descricao']
            df_vis_re_convertidos['Tipo_Conversao'] = 'Visita Realizada -> Matrícula'
            resultados_visita_realizada[marco_id] = {
                'descricao': info['descricao'],
                'total': len(df_vis_realizada),
                'convertidos': n_vis_re,
                'df': df_vis_re_convertidos
            }
            todos_convertidos.append(df_vis_re_convertidos)

        # =====================================================================
        # TOTAIS DO MARCO (para compatibilidade com estrutura anterior)
        # =====================================================================
        df_sem_confirmacao = df_leads[~mask_matricula].copy()
        df_convertidos = df_sem_confirmacao[df_sem_confirmacao['Virou_Matricula']].copy()

        if cols['nota'] and not df_convertidos.empty:
            df_convertidos[cols['nota']] = pd.to_numeric(df_convertidos[cols['nota']], errors='coerce')
            df_convertidos = df_convertidos.sort_values(cols['nota'], ascending=False)
            df_convertidos = df_convertidos.drop_duplicates(subset=['Nome_Normalizado'], keep='first')

        n_convertidos = len(df_convertidos)

        if n_convertidos > 0:
            taxa = (n_convertidos / len(df_sem_confirmacao)) * 100 if len(df_sem_confirmacao) > 0 else 0
            df_convertidos['Marco'] = info['descricao']
            df_convertidos['Data_Marco'] = info['data']

            resultados[marco_id] = {
                'descricao': info['descricao'],
                'total_leads': len(df_leads),
                'sem_confirmacao': len(df_sem_confirmacao),
                'convertidos': n_convertidos,
                'taxa': taxa,
                'df': df_convertidos,
                'colunas': cols,
                'qualificacao': n_qualif,
                'visita_agendada': n_vis_ag,
                'visita_realizada': n_vis_re
            }

        print(f"     >>> TOTAL CONVERSÕES OCULTAS: {n_convertidos}")

    # 3. Consolidar resultados
    print("\n" + "="*80)
    print("RESUMO DAS CONVERSÕES OCULTAS")
    print("="*80)

    total_geral = 0
    total_qualificacao = 0
    total_visita_agendada = 0
    total_visita_realizada = 0

    for marco_id, dados in resultados.items():
        print(f"\n{dados['descricao']}:")
        print(f"  - Total leads com nota: {dados['total_leads']}")
        print(f"  - Sem confirmação no HubSpot: {dados['sem_confirmacao']}")
        print(f"  - VIRARAM MATRÍCULA: {dados['convertidos']}")
        print(f"    > De Qualificação/Negociação: {dados.get('qualificacao', 0)}")
        print(f"    > De Visita Agendada: {dados.get('visita_agendada', 0)}")
        print(f"    > De Visita Realizada: {dados.get('visita_realizada', 0)}")
        print(f"  - Taxa de conversão: {dados['taxa']:.2f}%")

        total_geral += dados['convertidos']
        total_qualificacao += dados.get('qualificacao', 0)
        total_visita_agendada += dados.get('visita_agendada', 0)
        total_visita_realizada += dados.get('visita_realizada', 0)

    print(f"\n{'='*80}")
    print(f"TOTAL DE CONVERSÕES OCULTAS: {total_geral}")
    print(f"  > De Qualificação/Negociação: {total_qualificacao}")
    print(f"  > De Visita Agendada: {total_visita_agendada}")
    print(f"  > De Visita Realizada: {total_visita_realizada}")
    print(f"{'='*80}")

    # 4. Análise por Nota
    if todos_convertidos:
        df_todos = pd.concat(todos_convertidos, ignore_index=True)

        # Identificar coluna de nota
        col_nota = None
        for col in df_todos.columns:
            if 'nota' in col.lower() and '1a5' in col.lower():
                col_nota = col
                break
            elif 'nota' in col.lower():
                col_nota = col

        if col_nota:
            print("\n[ANÁLISE POR NOTA]")
            df_todos[col_nota] = pd.to_numeric(df_todos[col_nota], errors='coerce')
            dist_nota = df_todos.groupby(col_nota).size()

            for nota, count in dist_nota.items():
                if pd.notna(nota):
                    print(f"  Nota {int(nota)}: {count} conversões")

    # 5. Gerar Excel
    print("\n[3] Gerando relatório Excel...")

    wb = Workbook()

    # --- ABA 1: RESUMO ---
    ws_resumo = wb.active
    ws_resumo.title = 'Resumo'

    ws_resumo['A1'] = 'RELATÓRIO DE CONVERSÕES OCULTAS - HELIO'
    ws_resumo['A1'].font = Font(bold=True, size=14)
    ws_resumo['A2'] = f'Data de Execução: {DATA_HOJE}'
    ws_resumo['A3'] = ''
    ws_resumo['A4'] = 'O que são Conversões Ocultas?'
    ws_resumo['A4'].font = Font(bold=True)
    ws_resumo['A5'] = 'Leads que tinham nota no HubSpot, NÃO estavam com confirmação de matrícula,'
    ws_resumo['A6'] = 'mas que de fato matricularam (encontrados na base matriculas_finais).'
    ws_resumo['A7'] = ''
    ws_resumo['A8'] = 'RESUMO POR TIPO DE ETAPA ORIGINAL:'
    ws_resumo['A8'].font = Font(bold=True)

    # Tabela resumo detalhada
    resumo_data = []
    for marco_id, dados in resultados.items():
        resumo_data.append({
            'Marco': dados['descricao'],
            'Total Leads': dados['total_leads'],
            'Sem Confirmação': dados['sem_confirmacao'],
            'De Qualificação': dados.get('qualificacao', 0),
            'De Visita Agendada': dados.get('visita_agendada', 0),
            'De Visita Realizada': dados.get('visita_realizada', 0),
            'Total Conversões': dados['convertidos'],
            'Taxa (%)': round(dados['taxa'], 2)
        })

    if resumo_data:
        df_resumo = pd.DataFrame(resumo_data)
        df_to_sheet(ws_resumo, df_resumo, start_row=10)

        # Total
        row_total = 10 + len(resumo_data) + 1
        ws_resumo.cell(row=row_total, column=1, value='TOTAL').font = Font(bold=True)
        ws_resumo.cell(row=row_total, column=4, value=total_qualificacao).font = Font(bold=True)
        ws_resumo.cell(row=row_total, column=5, value=total_visita_agendada).font = Font(bold=True)
        ws_resumo.cell(row=row_total, column=6, value=total_visita_realizada).font = Font(bold=True)
        ws_resumo.cell(row=row_total, column=7, value=total_geral).font = Font(bold=True)

    ajustar_largura(ws_resumo)

    # --- ABA 2: TODOS OS CONVERTIDOS ---
    if todos_convertidos:
        ws_todos = wb.create_sheet('Todas_Conversoes')

        df_todos = pd.concat(todos_convertidos, ignore_index=True)

        # Selecionar colunas relevantes
        cols_export = []
        for col in df_todos.columns:
            col_lower = col.lower()
            if any(x in col_lower for x in ['nome', 'etapa', 'nota', 'unidade', 'record', 'probabilidade', 'marco']):
                cols_export.append(col)

        if cols_export:
            df_export = df_todos[cols_export].copy()
            df_to_sheet(ws_todos, df_export)
            ajustar_largura(ws_todos)

    # --- ABA: VISITA AGENDADA -> MATRÍCULA ---
    ws_vis_ag = wb.create_sheet('Visita_Agendada_Matricula')
    ws_vis_ag['A1'] = 'LEADS EM VISITA AGENDADA QUE VIRARAM MATRÍCULA'
    ws_vis_ag['A1'].font = Font(bold=True, size=12)
    ws_vis_ag['A2'] = '(Não atualizados no HubSpot)'

    if resultados_visita_agendada:
        dfs_vis_ag = [d['df'] for d in resultados_visita_agendada.values()]
        if dfs_vis_ag:
            df_vis_ag_all = pd.concat(dfs_vis_ag, ignore_index=True)
            # Remover duplicatas globais
            df_vis_ag_all = df_vis_ag_all.drop_duplicates(subset=['Nome_Normalizado'], keep='first')

            cols_export = []
            for col in df_vis_ag_all.columns:
                col_lower = col.lower()
                if any(x in col_lower for x in ['nome', 'etapa', 'nota', 'unidade', 'record', 'marco']):
                    cols_export.append(col)

            if cols_export:
                df_to_sheet(ws_vis_ag, df_vis_ag_all[cols_export], start_row=4)

    ajustar_largura(ws_vis_ag)

    # --- ABA: VISITA REALIZADA -> MATRÍCULA ---
    ws_vis_re = wb.create_sheet('Visita_Realizada_Matricula')
    ws_vis_re['A1'] = 'LEADS EM VISITA REALIZADA QUE VIRARAM MATRÍCULA'
    ws_vis_re['A1'].font = Font(bold=True, size=12)
    ws_vis_re['A2'] = '(Não atualizados no HubSpot)'

    if resultados_visita_realizada:
        dfs_vis_re = [d['df'] for d in resultados_visita_realizada.values()]
        if dfs_vis_re:
            df_vis_re_all = pd.concat(dfs_vis_re, ignore_index=True)
            # Remover duplicatas globais
            df_vis_re_all = df_vis_re_all.drop_duplicates(subset=['Nome_Normalizado'], keep='first')

            cols_export = []
            for col in df_vis_re_all.columns:
                col_lower = col.lower()
                if any(x in col_lower for x in ['nome', 'etapa', 'nota', 'unidade', 'record', 'marco']):
                    cols_export.append(col)

            if cols_export:
                df_to_sheet(ws_vis_re, df_vis_re_all[cols_export], start_row=4)

    ajustar_largura(ws_vis_re)

    # --- ABA: QUALIFICAÇÃO -> MATRÍCULA ---
    ws_qualif = wb.create_sheet('Qualificacao_Matricula')
    ws_qualif['A1'] = 'LEADS EM QUALIFICAÇÃO/NEGOCIAÇÃO QUE VIRARAM MATRÍCULA'
    ws_qualif['A1'].font = Font(bold=True, size=12)
    ws_qualif['A2'] = '(Não atualizados no HubSpot)'

    if resultados_qualificacao:
        dfs_qualif = [d['df'] for d in resultados_qualificacao.values()]
        if dfs_qualif:
            df_qualif_all = pd.concat(dfs_qualif, ignore_index=True)
            # Remover duplicatas globais
            df_qualif_all = df_qualif_all.drop_duplicates(subset=['Nome_Normalizado'], keep='first')

            cols_export = []
            for col in df_qualif_all.columns:
                col_lower = col.lower()
                if any(x in col_lower for x in ['nome', 'etapa', 'nota', 'unidade', 'record', 'marco']):
                    cols_export.append(col)

            if cols_export:
                df_to_sheet(ws_qualif, df_qualif_all[cols_export], start_row=4)

    ajustar_largura(ws_qualif)

    # --- ABAS POR MARCO (COMPLETO) ---
    for marco_id, dados in resultados.items():
        ws_marco = wb.create_sheet(marco_id[:31])  # Limite de 31 caracteres para nome de aba

        df_marco = dados['df']
        cols = dados['colunas']

        # Selecionar colunas relevantes
        cols_export = []
        for key, col in cols.items():
            if col and col in df_marco.columns:
                cols_export.append(col)

        if cols_export:
            df_export = df_marco[cols_export].copy()
            df_to_sheet(ws_marco, df_export)
            ajustar_largura(ws_marco)

    # --- ABA: ANÁLISE POR NOTA ---
    if todos_convertidos:
        ws_nota = wb.create_sheet('Analise_Por_Nota')
        ws_nota['A1'] = 'CONVERSÕES OCULTAS POR NOTA'
        ws_nota['A1'].font = Font(bold=True, size=12)

        df_todos = pd.concat(todos_convertidos, ignore_index=True)

        # Encontrar coluna de nota
        col_nota = None
        for col in df_todos.columns:
            if 'nota' in col.lower() and '1a5' in col.lower():
                col_nota = col
                break
            elif 'nota' in col.lower():
                col_nota = col

        if col_nota:
            df_todos[col_nota] = pd.to_numeric(df_todos[col_nota], errors='coerce')

            analise_nota = df_todos.groupby(col_nota).agg(
                Quantidade=('Nome_Normalizado', 'count')
            ).reset_index()
            analise_nota.columns = ['Nota', 'Quantidade']
            analise_nota['Percentual (%)'] = (analise_nota['Quantidade'] / analise_nota['Quantidade'].sum() * 100).round(1)

            df_to_sheet(ws_nota, analise_nota, start_row=3)
            ajustar_largura(ws_nota)

    # Salvar
    try:
        wb.save(CAMINHO_SAIDA)
        print(f"\n[OK] Relatório salvo: {CAMINHO_SAIDA}")
    except Exception as e:
        print(f"\n[ERRO] Falha ao salvar: {e}")
        # Tentar salvar com nome alternativo
        alt_path = os.path.join(PASTA_DB_HELIO, f'Conversoes_Ocultas_{DATA_HOJE}_backup.xlsx')
        try:
            wb.save(alt_path)
            print(f"[OK] Salvo em caminho alternativo: {alt_path}")
        except:
            pass

    print("\n[FIM DA VALIDAÇÃO]")
    print("="*80)


if __name__ == "__main__":
    main()
