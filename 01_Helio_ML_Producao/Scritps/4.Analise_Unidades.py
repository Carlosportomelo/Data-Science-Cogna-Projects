"""
==============================================================================
ANÁLISE DE UNIDADES - RELATÓRIOS GERENCIAIS
==============================================================================
Script: 2.Analise_Unidades.py
Objetivo: Gerar relatórios detalhados por unidade (Próprias vs Franqueadas)
          com ranking de qualidade e abas individuais.
Data: 2025-12-19
==============================================================================
"""

import pandas as pd
import os
import glob
import shutil
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASTA_OUTPUT = os.path.join(CAMINHO_BASE, 'Outputs')
PASTA_DADOS_INPUT = os.path.join(PASTA_OUTPUT, 'Dados_Scored')
PASTA_RELATORIOS_UNIDADES = os.path.join(PASTA_OUTPUT, 'Relatorios_Unidades')

os.makedirs(PASTA_RELATORIOS_UNIDADES, exist_ok=True)

# --- CONSTANTES FINANCEIRAS ---
TICKET_PROPRIA = 1222.25
TICKET_FRANQUIA = 907.49

# Função de Backup
def gerenciar_backup(pasta):
    pasta_backup = os.path.join(pasta, 'Backup')
    os.makedirs(pasta_backup, exist_ok=True)
    
    for item in os.listdir(pasta):
        caminho_item = os.path.join(pasta, item)
        if item == 'Backup': continue
        
        if os.path.isfile(caminho_item):
            destino = os.path.join(pasta_backup, item)
            if os.path.exists(destino):
                try: os.remove(destino)
                except: pass
            try: shutil.move(caminho_item, destino)
            except: pass

print("[BACKUP] Organizando relatórios antigos em Backup...")
gerenciar_backup(PASTA_RELATORIOS_UNIDADES)

DATA_HOJE = datetime.now().strftime('%Y-%m-%d')

# Lista de Unidades Próprias (Mesma do Script 1)
UNIDADES_PROPRIAS = [
    'VILA LEOPOLDINA', 'SANTANA', 'PINHEIROS', 'PERDIZES', 
    'PACAEMBU', 'MORUMBI', 'JARDINS', 'ITAIM'
]

print("="*80)
print("GERADOR DE RELATÓRIOS POR UNIDADE")
print("="*80)

# ==============================================================================
# 1. CARREGAR DADOS
# ==============================================================================
def carregar_ultimo_arquivo():
    # Procura pelo arquivo CSV mais recente gerado pelo script 1
    padrao = os.path.join(PASTA_DADOS_INPUT, 'leads_scored_*.csv')
    arquivos = glob.glob(padrao)
    
    if not arquivos:
        print(f"[ERRO] Nenhum arquivo 'leads_scored_*.csv' encontrado em: {PASTA_DADOS_INPUT}")
        print("   Execute o script '1.ML_Lead_Scoring.py' primeiro.")
        return None
    
    arquivo_recente = max(arquivos, key=os.path.getmtime)
    print(f"[LENDO] Lendo arquivo base: {os.path.basename(arquivo_recente)}")
    
    # Lê o CSV (o script 1 exporta com separador ';')
    return pd.read_csv(arquivo_recente, sep=';', encoding='utf-8-sig')

df = carregar_ultimo_arquivo()
if df is None:
    exit()

# Adicionar Data do Relatório para histórico (Snapshot)
df['Data_Relatorio'] = DATA_HOJE

# --- Garantir e parsear coluna de Data de criação do lead (para export)
if 'Data de criação' in df.columns:
    try:
        df['Data de criação'] = pd.to_datetime(df['Data de criação'], dayfirst=True, errors='coerce')
    except Exception:
        pass

# Garantir que a coluna Tipo_Unidade existe (caso venha de versão antiga)
if 'Tipo_Unidade' not in df.columns:
    print("[AVISO] Coluna 'Tipo_Unidade' não encontrada. Recalculando...")
    def classificar_tipo_unidade(unidade):
        unidade_upper = str(unidade).upper()
        for propria in UNIDADES_PROPRIAS:
            if propria in unidade_upper:
                return 'Própria'
        return 'Franqueada'
    df['Tipo_Unidade'] = df['Unidade Desejada'].apply(classificar_tipo_unidade)

# ==============================================================================
# 2. FUNÇÃO DE GERAÇÃO DE EXCEL
# ==============================================================================
def gerar_relatorio_excel(df_grupo, nome_arquivo, titulo_relatorio):
    print(f"   → Gerando: {nome_arquivo} ({len(df_grupo)} leads)...")
    wb = Workbook()
    
    # Estilos
    HEADER_FILL = PatternFill(start_color='CE0E2D', end_color='CE0E2D', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF')
    THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # --- ABA 1: RANKING GERAL ---
    ws1 = wb.active
    ws1.title = "Ranking_Qualidade"
    
    # Pivot Table: Unidade vs Nota
    pivot = pd.crosstab(df_grupo['Unidade Desejada'], df_grupo['Nota_1a5'])
    
    # Garantir colunas 1 a 5
    for i in range(1, 6):
        if i not in pivot.columns:
            pivot[i] = 0
    pivot = pivot[[1, 2, 3, 4, 5]]
    
    # Métricas
    pivot['Total_Leads'] = pivot.sum(axis=1)
    pivot['Qualificados (4+5)'] = pivot[4] + pivot[5]
    pivot['% Qualificados'] = (pivot['Qualificados (4+5)'] / pivot['Total_Leads']).fillna(0)
    
    # Ordenar por Volume Total
    ranking = pivot.sort_values('Total_Leads', ascending=False).reset_index()
    
    # Escrever no Excel
    ws1['A1'] = f"RANKING DE QUALIDADE - {titulo_relatorio.upper()}"
    ws1['A1'].font = Font(bold=True, size=14)
    
    cols = ['Unidade', 'Nota 1', 'Nota 2', 'Nota 3', 'Nota 4', 'Nota 5', 'Total Leads', 'Qtd Qualificados (4+5)', '% Qualificados']
    ws1.append([])
    ws1.append(cols)
    
    for row in dataframe_to_rows(ranking, index=False, header=False):
        ws1.append(row)
        
    # Formatação Aba 1
    for cell in ws1[3]: # Header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        
    for row in ws1.iter_rows(min_row=4, max_row=ws1.max_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = THIN_BORDER
            if cell.column == 9: # %
                cell.number_format = '0.0%'
                
    ws1.column_dimensions['A'].width = 35
    for c in range(2, 10):
        ws1.column_dimensions[get_column_letter(c)].width = 15

    # --- ABAS POR UNIDADE ---
    unidades = sorted(df_grupo['Unidade Desejada'].dropna().unique())
    
    for unidade in unidades:
        # Sanitizar nome da aba
        nome_aba = str(unidade).replace('/', '-').replace('\\', '').replace('?', '').replace('*', '').replace('[', '').replace(']', '')[:30]
        ws = wb.create_sheet(title=nome_aba)
        
        # Filtrar e Selecionar Colunas
        df_unidade = df_grupo[df_grupo['Unidade Desejada'] == unidade].copy()
        cols_view = ['Data_Relatorio', 'Data de criação', 'Record ID', 'Nome do negócio', 'Etapa do negócio', 'Nota_1a5', 'Prioridade_Atendimento', 'Probabilidade_Conversao', 'Fonte_Cat', 'Número de atividades de vendas', 'Dias_Desde_Criacao', 'Proprietário do negócio']
        cols_existentes = [c for c in cols_view if c in df_unidade.columns]
        
        df_view = df_unidade[cols_existentes].sort_values('Probabilidade_Conversao', ascending=False)
        
        # Detectar índice da coluna de data para formatação no Excel
        date_col_idx = None
        if 'Data de criação' in df_view.columns:
            date_col_idx = list(df_view.columns).index('Data de criação') + 1

        ws['A1'] = f"LEADS: {unidade}"
        ws['A1'].font = Font(bold=True, size=12)

        for r_idx, row in enumerate(dataframe_to_rows(df_view, index=False, header=True), start=3):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 3:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                else:
                    # Aplicar formatação de data quando for a coluna de criação
                    if date_col_idx is not None and c_idx == date_col_idx:
                        try:
                            cell.number_format = 'DD/MM/YYYY'
                        except Exception:
                            pass

    # --- ABA PROJEÇÃO FINANCEIRA (NOVO) ---
    ws_fin = wb.create_sheet("Projecao_Financeira")
    ws_fin['A1'] = f"PROJEÇÃO FINANCEIRA - {titulo_relatorio.upper()}"
    ws_fin['A1'].font = Font(bold=True, size=14)
    
    # Definir ticket base
    ticket_base = TICKET_PROPRIA if 'Própria' in titulo_relatorio else TICKET_FRANQUIA
    ws_fin['A2'] = f"Premissas: Ticket Médio R$ {ticket_base:,.2f} | LTV: 66 Meses (Com discriminação de Ajuste 7% a.a.)"
    
    # Calcular
    df_fin = df_grupo.copy()
    
    # Lógica de LTV (Cópia do Script 1)
    taxa_anual = 1.07
    fator_ano1 = 12
    fator_ano2 = 12 * taxa_anual
    fator_ano3 = 12 * (taxa_anual**2)
    fator_ano4 = 12 * (taxa_anual**3)
    fator_ano5 = 12 * (taxa_anual**4)
    fator_ano6_semestral = 6 * (taxa_anual**5)
    
    mult_24m = fator_ano1 + fator_ano2
    mult_66m_com_ajuste = fator_ano1 + fator_ano2 + fator_ano3 + fator_ano4 + fator_ano5 + fator_ano6_semestral
    mult_66m_sem_ajuste = 66

    df_fin['Esp_1M'] = ticket_base * df_fin['Probabilidade_Conversao']
    df_fin['Esp_12M'] = df_fin['Esp_1M'] * 12
    df_fin['Esp_24M'] = df_fin['Esp_1M'] * mult_24m
    df_fin['LTV_Base'] = df_fin['Esp_1M'] * mult_66m_sem_ajuste
    df_fin['LTV_Total'] = df_fin['Esp_1M'] * mult_66m_com_ajuste
    df_fin['LTV_Ajuste'] = df_fin['LTV_Total'] - df_fin['LTV_Base']
    
    # Pivot por Unidade
    pivot_fin = df_fin.groupby('Unidade Desejada').agg(
        Leads=('Record ID', 'count'),
        Rec_1M=('Esp_1M', 'sum'),
        Rec_12M=('Esp_12M', 'sum'),
        Rec_24M=('Esp_24M', 'sum'),
        LTV_Base=('LTV_Base', 'sum'),
        LTV_Ajuste=('LTV_Ajuste', 'sum'),
        LTV_Total=('LTV_Total', 'sum')
    ).reset_index().sort_values('LTV_Total', ascending=False)
    
    # Escrever
    cols = ['Unidade', 'Qtd Leads', 'Rec. 1 Mês', 'Rec. 1 Ano', 'Rec. 2 Anos', 'LTV Base (66m)', 'Ajuste (7% a.a.)', 'LTV Total']
    ws_fin.append([])
    ws_fin.append(cols)
    
    # Header style
    for cell in ws_fin[4]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # Rows
    current_row = 5
    for r_idx, row in enumerate(dataframe_to_rows(pivot_fin, index=False, header=False), start=current_row):
        ws_fin.cell(row=r_idx, column=1, value=row[0]).border = THIN_BORDER
        ws_fin.cell(row=r_idx, column=2, value=row[1]).border = THIN_BORDER
        
        # Values (cols 3 to 8)
        for c_offset in range(6):
            val = row[2 + c_offset]
            cell = ws_fin.cell(row=r_idx, column=3 + c_offset, value=val)
            cell.number_format = '#,##0.00'
            cell.border = THIN_BORDER
        current_row = r_idx
        
    # Total Row
    total_row = current_row + 1
    ws_fin.cell(row=total_row, column=1, value="TOTAL GERAL").font = Font(bold=True)
    ws_fin.cell(row=total_row, column=2, value=pivot_fin['Leads'].sum()).font = Font(bold=True)
    
    totais_vals = [
        pivot_fin['Rec_1M'].sum(), pivot_fin['Rec_12M'].sum(), pivot_fin['Rec_24M'].sum(),
        pivot_fin['LTV_Base'].sum(), pivot_fin['LTV_Ajuste'].sum(), pivot_fin['LTV_Total'].sum()
    ]
    
    for i, val in enumerate(totais_vals):
        cell = ws_fin.cell(row=total_row, column=3 + i, value=val)
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'
        
    # Borders for total
    for c in range(1, 9):
        ws_fin.cell(row=total_row, column=c).border = THIN_BORDER
        ws_fin.cell(row=total_row, column=c).fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        
    # --- TABELA POR NOTA (INVESTIMENTO POTENCIAL) ---
    row_nota_titulo = total_row + 3
    ws_fin.cell(row=row_nota_titulo, column=1, value="POTENCIAL POR SCORE (NOTA)").font = Font(bold=True, size=12)

    pivot_nota = df_fin.groupby('Nota_1a5').agg(
        Leads=('Record ID', 'count'),
        Rec_1M=('Esp_1M', 'sum'),
        Rec_12M=('Esp_12M', 'sum'),
        Rec_24M=('Esp_24M', 'sum'),
        LTV_Base=('LTV_Base', 'sum'),
        LTV_Ajuste=('LTV_Ajuste', 'sum'),
        LTV_Total=('LTV_Total', 'sum')
    ).reset_index().sort_values('Nota_1a5')

    cols_nota = ['Nota', 'Qtd Leads', 'Rec. 1 Mês', 'Rec. 1 Ano', 'Rec. 2 Anos', 'LTV Base (66m)', 'Ajuste (7% a.a.)', 'LTV Total']
    
    # Header Nota
    row_nota_header = row_nota_titulo + 1
    for c_idx, col_name in enumerate(cols_nota, start=1):
        cell = ws_fin.cell(row=row_nota_header, column=c_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # Rows Nota
    for r_idx, row in enumerate(dataframe_to_rows(pivot_nota, index=False, header=False), start=row_nota_header + 1):
        ws_fin.cell(row=r_idx, column=1, value=row[0]).border = THIN_BORDER
        ws_fin.cell(row=r_idx, column=2, value=row[1]).border = THIN_BORDER
        for c_offset in range(6):
            val = row[2 + c_offset]
            cell = ws_fin.cell(row=r_idx, column=3 + c_offset, value=val)
            cell.number_format = '#,##0.00'
            cell.border = THIN_BORDER

    ws_fin.column_dimensions['A'].width = 35
    for c in range(3, 9):
        ws_fin.column_dimensions[get_column_letter(c)].width = 18
    
    # Save with protection
    try:
        wb.save(os.path.join(PASTA_RELATORIOS_UNIDADES, nome_arquivo))
    except PermissionError:
        print(f"\n[ERRO] O arquivo parece estar ABERTO no Excel: {nome_arquivo}")
        novo_nome = nome_arquivo.replace('.xlsx', f'_Copia_{datetime.now().strftime("%H%M%S")}.xlsx')
        wb.save(os.path.join(PASTA_RELATORIOS_UNIDADES, novo_nome))
        print(f"  → Salvo como: {novo_nome}")

# ==============================================================================
# 3. EXECUÇÃO
# ==============================================================================
gerar_relatorio_excel(df[df['Tipo_Unidade'] == 'Própria'], f'Relatorio_Unidades_Proprias_{DATA_HOJE}.xlsx', 'Unidades Próprias')
gerar_relatorio_excel(df[df['Tipo_Unidade'] == 'Franqueada'], f'Relatorio_Unidades_Franqueadas_{DATA_HOJE}.xlsx', 'Unidades Franqueadas')

print("\n[OK] Processo concluído com sucesso!")