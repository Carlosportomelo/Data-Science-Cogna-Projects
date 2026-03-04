"""
==============================================================================
AUDITORIA DE MATCH ENTRE BASES - HELIO
==============================================================================
Script: 13.Auditoria_Match_Bases.py
Objetivo: Auditar e visualizar matches entre HubSpot, Helio e Matriculas Finais

BASES ANALISADAS:
1. hubspot_leads.csv - Leads do CRM
2. leads_scored_*.csv - Leads pontuados pelo Helio (marcos 12/12, 05/01, 20/01)
3. matriculas_finais_limpo.csv - Matriculas efetivas (Source of Truth)

Data: 2026-01-30
==============================================================================
"""

import pandas as pd
import numpy as np
import os
import glob
import unicodedata
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# ==============================================================================
# CONFIGURACAO
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASTA_DATA = os.path.join(CAMINHO_BASE, 'Data')
PASTA_OUTPUT = os.path.join(CAMINHO_BASE, 'Outputs')
PASTA_AUDITORIA = os.path.join(PASTA_OUTPUT, 'Auditoria')

os.makedirs(PASTA_AUDITORIA, exist_ok=True)

DATA_HOJE = datetime.now().strftime('%Y-%m-%d')

# Arquivos
CAMINHO_HUBSPOT = os.path.join(PASTA_DATA, 'hubspot_leads.csv')
CAMINHO_MATRICULAS = os.path.join(PASTA_DATA, 'matriculas_finais_limpo.csv')

# Marcos do Helio
MARCOS_HELIO = {
    '12/12/2025': os.path.join(PASTA_OUTPUT, 'archive', 'leads_scored_2025-12-12.csv'),
    '05/01/2026': os.path.join(PASTA_OUTPUT, 'Dados_Scored', 'Backup', 'leads_scored_2026-01-05.csv'),
    '20/01/2026': os.path.join(PASTA_OUTPUT, 'Dados_Scored', 'Backup', 'leads_scored_2026-01-20.csv'),
}

# Estilos Excel
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
FILL_SUCESSO = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FILL_ERRO = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
FILL_ALERTA = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

# ==============================================================================
# FUNCOES DE NORMALIZACAO
# ==============================================================================
def normalizar_nome_matricula(nome):
    """Normaliza nome da base de matriculas."""
    if pd.isna(nome): return ""
    nome = str(nome).upper().strip()
    nome = unicodedata.normalize('NFKD', nome)
    nome = ''.join(c for c in nome if not unicodedata.combining(c))
    nome = ' '.join(nome.split())
    return nome

def normalizar_nome_hubspot(nome):
    """Normaliza nome do HubSpot (remove sufixos como ' - 10 - UNIDADE')."""
    if pd.isna(nome): return ""
    nome = str(nome).upper().strip()
    # Remove sufixos como " - 10 - UNIDADE"
    nome = re.sub(r'\s*-\s*\d+\s*-\s*.*$', '', nome)
    nome = unicodedata.normalize('NFKD', nome)
    nome = ''.join(c for c in nome if not unicodedata.combining(c))
    nome = ' '.join(nome.split())
    return nome

def df_to_sheet(ws, df, start_row=1, freeze_panes=False):
    """Adiciona DataFrame ao sheet com formatacao."""
    if df.empty:
        ws.cell(row=start_row, column=1, value="Sem dados")
        return

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if r_idx == start_row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center', wrap_text=True)

    if freeze_panes:
        ws.freeze_panes = ws.cell(row=start_row+1, column=1)

def ajustar_largura(ws):
    """Auto-ajusta largura das colunas."""
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

# ==============================================================================
# CARREGAMENTO DAS BASES
# ==============================================================================
print("="*80)
print("AUDITORIA DE MATCH ENTRE BASES - HELIO")
print("="*80)

# 1. Carregar Matriculas Finais (Source of Truth)
print("\n[1/4] Carregando base de Matriculas Finais...")
df_mat = pd.read_csv(CAMINHO_MATRICULAS, sep=';', encoding='utf-8-sig')
df_mat.columns = df_mat.columns.str.strip()

# Filtrar apenas NOVOS
df_mat = df_mat[df_mat['Tipo'] == 'NOVO'].copy()
df_mat['Nome_Norm'] = df_mat['Nome do Aluno'].apply(normalizar_nome_matricula)
df_mat['P1_Paga'] = df_mat['Status P1'].str.upper().isin(['PAGA', 'BOLSA'])

set_matriculados = set(df_mat['Nome_Norm'].unique())
set_p1_paga = set(df_mat[df_mat['P1_Paga']]['Nome_Norm'].unique())

print(f"   Total matriculados (NOVO): {len(set_matriculados)}")
print(f"   Com P1 paga/bolsa: {len(set_p1_paga)}")

# 2. Carregar HubSpot Leads
print("\n[2/4] Carregando base HubSpot...")
try:
    df_hub = pd.read_csv(CAMINHO_HUBSPOT, encoding='utf-8')
except:
    df_hub = pd.read_csv(CAMINHO_HUBSPOT, encoding='latin1')

df_hub.columns = df_hub.columns.str.strip()
df_hub['Nome_Norm'] = df_hub['Nome do negocio'].apply(normalizar_nome_hubspot) if 'Nome do negocio' in df_hub.columns else df_hub['Nome do negócio'].apply(normalizar_nome_hubspot)

set_hubspot = set(df_hub['Nome_Norm'].unique())
print(f"   Total leads HubSpot: {len(df_hub)}")
print(f"   Nomes unicos: {len(set_hubspot)}")

# 3. Carregar Marcos do Helio
print("\n[3/4] Carregando marcos do Helio...")
marcos_dados = {}

for marco_nome, arquivo in MARCOS_HELIO.items():
    if os.path.exists(arquivo):
        df_marco = pd.read_csv(arquivo, sep=';', encoding='utf-8-sig')
        nome_col = 'Nome do negocio' if 'Nome do negocio' in df_marco.columns else 'Nome do negócio'
        df_marco['Nome_Norm'] = df_marco[nome_col].apply(normalizar_nome_hubspot)
        marcos_dados[marco_nome] = df_marco
        print(f"   {marco_nome}: {len(df_marco)} leads carregados")
    else:
        print(f"   {marco_nome}: Arquivo nao encontrado")

# ==============================================================================
# ANALISE DE MATCHES
# ==============================================================================
print("\n[4/4] Analisando matches...")

# Match HubSpot x Matriculas
match_hub_mat = set_hubspot.intersection(set_matriculados)
print(f"\n   HUBSPOT x MATRICULAS:")
print(f"   - Leads no HubSpot: {len(set_hubspot)}")
print(f"   - Matriculados: {len(set_matriculados)}")
print(f"   - Match encontrado: {len(match_hub_mat)} ({len(match_hub_mat)/len(set_matriculados)*100:.1f}% dos matriculados)")

# Matriculados que NAO estao no HubSpot
mat_sem_hubspot = set_matriculados - set_hubspot
print(f"   - Matriculados SEM registro no HubSpot: {len(mat_sem_hubspot)}")

# ==============================================================================
# GERAR RELATORIO EXCEL
# ==============================================================================
print("\n" + "="*80)
print("Gerando relatorio Excel...")

wb = Workbook()

# --- ABA 1: RESUMO GERAL ---
ws_resumo = wb.active
ws_resumo.title = '1_Resumo_Geral'
ws_resumo['A1'] = 'AUDITORIA DE MATCH ENTRE BASES - HELIO'
ws_resumo['A1'].font = Font(bold=True, size=14)
ws_resumo['A2'] = f'Data: {DATA_HOJE}'
ws_resumo['A3'] = f'Objetivo: Validar matches entre HubSpot, Helio e Matriculas Finais'

# Tabela de resumo
resumo_data = [
    ['Base', 'Total Registros', 'Nomes Unicos', 'Descricao'],
    ['Matriculas Finais', len(df_mat), len(set_matriculados), 'Source of Truth - Alunos NOVOS'],
    ['HubSpot Leads', len(df_hub), len(set_hubspot), 'CRM - Todos os leads'],
]

for marco_nome, df_marco in marcos_dados.items():
    resumo_data.append([f'Helio {marco_nome}', len(df_marco), len(df_marco['Nome_Norm'].unique()), f'Leads pontuados em {marco_nome}'])

for r_idx, row in enumerate(resumo_data, start=5):
    for c_idx, value in enumerate(row, start=1):
        cell = ws_resumo.cell(row=r_idx, column=c_idx, value=value)
        cell.border = THIN_BORDER
        if r_idx == 5:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

# Tabela de matches
ws_resumo['A12'] = 'ANALISE DE MATCHES'
ws_resumo['A12'].font = Font(bold=True, size=12)

match_data = [
    ['Comparacao', 'Base A', 'Base B', 'Match', '% de B em A'],
    ['HubSpot x Matriculas', len(set_hubspot), len(set_matriculados), len(match_hub_mat), f'{len(match_hub_mat)/len(set_matriculados)*100:.1f}%'],
]

for marco_nome, df_marco in marcos_dados.items():
    set_marco = set(df_marco['Nome_Norm'].unique())
    match_marco = set_marco.intersection(set_matriculados)
    match_data.append([f'Helio {marco_nome} x Matriculas', len(set_marco), len(set_matriculados), len(match_marco), f'{len(match_marco)/len(set_matriculados)*100:.1f}%'])

for r_idx, row in enumerate(match_data, start=13):
    for c_idx, value in enumerate(row, start=1):
        cell = ws_resumo.cell(row=r_idx, column=c_idx, value=value)
        cell.border = THIN_BORDER
        if r_idx == 13:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

ajustar_largura(ws_resumo)

# --- ABA 2: MATRICULADOS NO HUBSPOT ---
ws_mat_hub = wb.create_sheet('2_Matriculados_no_HubSpot')
ws_mat_hub['A1'] = 'MATRICULADOS QUE ESTAO NO HUBSPOT (MATCH)'
ws_mat_hub['A1'].font = Font(bold=True, size=12)
ws_mat_hub['A2'] = f'Total: {len(match_hub_mat)} alunos'

# Criar DataFrame com detalhes
df_match = df_mat[df_mat['Nome_Norm'].isin(match_hub_mat)].copy()
df_match = df_match[['Nome do Aluno', 'Status P1', 'Unidade', 'Tipo']].copy()
df_match.columns = ['Nome_Aluno', 'Status_P1', 'Unidade', 'Tipo']

# Adicionar info do HubSpot
hub_info = df_hub[df_hub['Nome_Norm'].isin(match_hub_mat)].copy()
hub_info = hub_info.drop_duplicates(subset=['Nome_Norm'], keep='first')

# Identificar colunas de etapa
etapa_col = 'Etapa do negócio' if 'Etapa do negócio' in hub_info.columns else 'Etapa do negocio'
if etapa_col in hub_info.columns:
    hub_lookup = dict(zip(hub_info['Nome_Norm'], hub_info[etapa_col]))
    df_match['Etapa_HubSpot'] = df_match['Nome_Aluno'].apply(normalizar_nome_matricula).map(hub_lookup)

df_to_sheet(ws_mat_hub, df_match, start_row=4, freeze_panes=True)
ajustar_largura(ws_mat_hub)

# --- ABA 3: MATRICULADOS SEM HUBSPOT ---
ws_mat_sem = wb.create_sheet('3_Matriculados_SEM_HubSpot')
ws_mat_sem['A1'] = 'MATRICULADOS QUE NAO ESTAO NO HUBSPOT'
ws_mat_sem['A1'].font = Font(bold=True, size=12)
ws_mat_sem['A2'] = f'Total: {len(mat_sem_hubspot)} alunos'
ws_mat_sem['A3'] = 'Observacao: Estes alunos podem ter entrado por canais offline ou nomes diferentes'

df_sem_hub = df_mat[df_mat['Nome_Norm'].isin(mat_sem_hubspot)].copy()
df_sem_hub = df_sem_hub[['Nome do Aluno', 'Status P1', 'Unidade', 'Tipo']].copy()

df_to_sheet(ws_mat_sem, df_sem_hub, start_row=5, freeze_panes=True)
ajustar_largura(ws_mat_sem)

# --- ABA 4-6: ANALISE POR MARCO DO HELIO ---
for marco_nome, df_marco in marcos_dados.items():
    marco_clean = marco_nome.replace('/', '-')
    ws_marco = wb.create_sheet(f'4_Helio_{marco_clean}')

    ws_marco['A1'] = f'ANALISE HELIO - MARCO {marco_nome}'
    ws_marco['A1'].font = Font(bold=True, size=12)

    # Match com matriculas
    df_marco['Matriculou'] = df_marco['Nome_Norm'].isin(set_matriculados)
    df_marco['P1_Paga'] = df_marco['Nome_Norm'].isin(set_p1_paga)
    df_marco['Converteu'] = df_marco['Matriculou'] | df_marco['P1_Paga']

    total = len(df_marco)
    total_conv = df_marco['Converteu'].sum()

    ws_marco['A3'] = f'Total leads pontuados: {total}'
    ws_marco['A4'] = f'Convertidos (matricula/P1): {total_conv} ({total_conv/total*100:.1f}%)'

    # Tabela por nota
    ws_marco['A6'] = 'PERFORMANCE POR NOTA'
    ws_marco['A6'].font = Font(bold=True)

    perf_data = [['Nota', 'Total', 'Convertidos', 'Taxa %', 'Acerto']]
    for nota in sorted(df_marco['Nota_1a5'].unique()):
        subset = df_marco[df_marco['Nota_1a5'] == nota]
        n = len(subset)
        conv = subset['Converteu'].sum()
        taxa = conv/n*100 if n > 0 else 0
        acerto = 'BOM' if (nota >= 4 and taxa > 1) or (nota < 4 and taxa < 1) else 'VERIFICAR'
        perf_data.append([int(nota), n, conv, f'{taxa:.1f}%', acerto])

    for r_idx, row in enumerate(perf_data, start=7):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_marco.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if r_idx == 7:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

    # Lista de convertidos com nota alta (ACERTOS)
    ws_marco['A15'] = 'ACERTOS: Nota 4-5 que Converteram'
    ws_marco['A15'].font = Font(bold=True, color='006400')

    nome_col = 'Nome do negocio' if 'Nome do negocio' in df_marco.columns else 'Nome do negócio'
    acertos = df_marco[(df_marco['Nota_1a5'] >= 4) & df_marco['Converteu']][[nome_col, 'Nota_1a5', 'Probabilidade_Conversao', 'Unidade Desejada']].copy()
    acertos.columns = ['Nome', 'Nota', 'Probabilidade', 'Unidade']
    acertos = acertos.sort_values('Probabilidade', ascending=False).head(50)

    df_to_sheet(ws_marco, acertos, start_row=16, freeze_panes=True)

    # Lista de erros (Nota alta que perdeu ou Nota baixa que converteu)
    start_erros = 16 + len(acertos) + 3
    ws_marco.cell(row=start_erros, column=1, value='SURPRESAS: Nota 1-2 que Converteram').font = Font(bold=True, color='8B0000')

    surpresas = df_marco[(df_marco['Nota_1a5'] <= 2) & df_marco['Converteu']][[nome_col, 'Nota_1a5', 'Probabilidade_Conversao', 'Unidade Desejada']].copy()
    surpresas.columns = ['Nome', 'Nota', 'Probabilidade', 'Unidade']

    df_to_sheet(ws_marco, surpresas, start_row=start_erros+1)

    ajustar_largura(ws_marco)

# --- ABA 7: LISTA COMPLETA DE MATCHES ---
ws_matches = wb.create_sheet('7_Lista_Matches_Completa')
ws_matches['A1'] = 'LISTA COMPLETA DE MATCHES (MATRICULADOS ENCONTRADOS)'
ws_matches['A1'].font = Font(bold=True, size=12)

# Consolidar todos os matches
all_matches = []
for marco_nome, df_marco in marcos_dados.items():
    nome_col = 'Nome do negocio' if 'Nome do negocio' in df_marco.columns else 'Nome do negócio'
    matches = df_marco[df_marco['Nome_Norm'].isin(set_matriculados)].copy()
    matches['Marco'] = marco_nome
    if len(matches) > 0:
        cols_keep = [nome_col, 'Nota_1a5', 'Probabilidade_Conversao', 'Unidade Desejada', 'Marco']
        cols_exist = [c for c in cols_keep if c in matches.columns]
        all_matches.append(matches[cols_exist])

if all_matches:
    df_all_matches = pd.concat(all_matches, ignore_index=True)
    df_all_matches = df_all_matches.drop_duplicates(subset=[df_all_matches.columns[0]], keep='first')
    df_all_matches = df_all_matches.sort_values('Nota_1a5', ascending=False)

    # Adicionar info da matricula
    mat_lookup = dict(zip(df_mat['Nome_Norm'], df_mat['Status P1']))
    df_all_matches['Status_P1'] = df_all_matches.iloc[:, 0].apply(normalizar_nome_hubspot).map(mat_lookup)

    df_to_sheet(ws_matches, df_all_matches, start_row=3, freeze_panes=True)

ajustar_largura(ws_matches)

# --- ABA 8: ESTATISTICAS CONSOLIDADAS ---
ws_stats = wb.create_sheet('8_Estatisticas_Consolidadas')
ws_stats['A1'] = 'ESTATISTICAS CONSOLIDADAS DE PERFORMANCE'
ws_stats['A1'].font = Font(bold=True, size=12)

stats_data = [
    ['Metrica', 'Valor', 'Observacao'],
    ['Total Matriculados (NOVO)', len(set_matriculados), 'Base de verdade'],
    ['Com P1 Paga', len(set_p1_paga), 'Conversao confirmada'],
    ['Presentes no HubSpot', len(match_hub_mat), f'{len(match_hub_mat)/len(set_matriculados)*100:.1f}% dos matriculados'],
    ['', '', ''],
    ['PERFORMANCE DO MODELO', '', ''],
]

# Calcular metricas por marco
for marco_nome, df_marco in marcos_dados.items():
    df_marco['Converteu'] = df_marco['Nome_Norm'].isin(set_matriculados)

    # Taxa por nota 5
    n5 = df_marco[df_marco['Nota_1a5'] == 5]
    taxa_n5 = n5['Converteu'].sum() / len(n5) * 100 if len(n5) > 0 else 0

    # Taxa por nota 1
    n1 = df_marco[df_marco['Nota_1a5'] == 1]
    taxa_n1 = n1['Converteu'].sum() / len(n1) * 100 if len(n1) > 0 else 0

    # Lift
    lift = taxa_n5 / taxa_n1 if taxa_n1 > 0 else 0

    stats_data.append([f'Marco {marco_nome}', '', ''])
    stats_data.append([f'  Taxa Nota 5', f'{taxa_n5:.1f}%', f'{n5["Converteu"].sum()} de {len(n5)}'])
    stats_data.append([f'  Taxa Nota 1', f'{taxa_n1:.1f}%', f'{n1["Converteu"].sum()} de {len(n1)}'])
    stats_data.append([f'  Lift (N5/N1)', f'{lift:.1f}x', 'Quanto melhor Nota 5 performa vs Nota 1'])

for r_idx, row in enumerate(stats_data, start=3):
    for c_idx, value in enumerate(row, start=1):
        cell = ws_stats.cell(row=r_idx, column=c_idx, value=value)
        cell.border = THIN_BORDER
        if r_idx == 3:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

ajustar_largura(ws_stats)

# Salvar arquivo
output_file = os.path.join(PASTA_AUDITORIA, f'Auditoria_Match_Bases_{DATA_HOJE}.xlsx')
wb.save(output_file)

print(f"\n{'='*80}")
print(f"RELATORIO GERADO COM SUCESSO!")
print(f"{'='*80}")
print(f"Arquivo: {output_file}")
print(f"\nABAS DISPONIVEIS:")
print("  1. Resumo Geral - Visao consolidada das bases")
print("  2. Matriculados no HubSpot - Lista de matches")
print("  3. Matriculados SEM HubSpot - Alunos sem registro no CRM")
print("  4-6. Analise por Marco - Performance do Helio em cada marco")
print("  7. Lista Completa de Matches - Todos os convertidos encontrados")
print("  8. Estatisticas Consolidadas - Metricas de performance")
print("="*80)
