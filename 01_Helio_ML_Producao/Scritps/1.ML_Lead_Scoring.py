"""
==============================================================================
MODELO DE ML - SCORING DE PROBABILIDADE DE CONVERSÃO (COM DATA ÚLTIMA ATIVIDADE)
==============================================================================
Script: 1.ML_Lead_Scoring.py
Objetivo: Treinar modelo (blindado de viés temporal) + Relatório com Data de Atividade.
Data: 2026-01-21
==============================================================================
"""

import pandas as pd
import numpy as np
import os
import shutil
import glob
import pickle
from datetime import datetime
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import roc_auc_score
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASTA_DATA = os.path.join(CAMINHO_BASE, 'Data')

def encontrar_arquivo_recente(pasta, padrao):
    arquivos = glob.glob(os.path.join(pasta, padrao))
    if not arquivos: return None
    return max(arquivos, key=os.path.getmtime)

CAMINHO_LEADS = encontrar_arquivo_recente(PASTA_DATA, 'hubspot_leads*.csv') or os.path.join(PASTA_DATA, 'hubspot_leads.csv')
CAMINHO_PERDIDOS = encontrar_arquivo_recente(PASTA_DATA, 'hubspot_negocios_perdidos*.csv') or os.path.join(PASTA_DATA, 'hubspot_negocios_perdidos.csv')
# ATUALIZAÇÃO: Usar base limpa de matrículas finais (Nome do Aluno = Nome do negócio)
CAMINHO_MATRICULAS = os.path.join(PASTA_DATA, 'matriculas_finais_limpo.csv')

PASTA_OUTPUT = os.path.join(CAMINHO_BASE, 'Outputs')
PASTA_MODELS = os.path.join(CAMINHO_BASE, 'models')
PASTA_DADOS = os.path.join(PASTA_OUTPUT, 'Dados_Scored')
PASTA_RELATORIOS = os.path.join(PASTA_OUTPUT, 'Relatorios_ML')

os.makedirs(PASTA_DADOS, exist_ok=True)
os.makedirs(PASTA_RELATORIOS, exist_ok=True)
os.makedirs(PASTA_MODELS, exist_ok=True)

# --- CONSTANTES FINANCEIRAS ---
TICKET_PROPRIA = 1222.25
TICKET_FRANQUIA = 907.49
UNIDADES_PROPRIAS = [
    'VILA LEOPOLDINA', 'SANTANA', 'PINHEIROS', 'PERDIZES', 
    'PACAEMBU', 'MORUMBI', 'JARDINS', 'ITAIM'
]

# Backup
def gerenciar_backup(pasta, ignorar=None):
    if ignorar is None: ignorar = []
    # Caixa Forte: Preserva o histórico para análise de evolução (Quem subiu/desceu)
    pasta_backup = os.path.join(pasta, 'Backup', 'Historico_Caixa_Forte')
    os.makedirs(pasta_backup, exist_ok=True)
    for item in os.listdir(pasta):
        caminho_item = os.path.join(pasta, item)
        if item == 'Backup' or item in ignorar: continue
        if os.path.isfile(caminho_item):
            destino = os.path.join(pasta_backup, item)
            if os.path.exists(destino):
                try: os.remove(destino)
                except: pass
            try: shutil.move(caminho_item, destino)
            except: pass

print("[BACKUP] Preservando histórico na Caixa Forte...")
gerenciar_backup(PASTA_DADOS, ignorar=['Historico_Evolucao_IA.csv'])
gerenciar_backup(PASTA_RELATORIOS)

DATA_HOJE = datetime.now().strftime('%Y-%m-%d')

print("="*80)
print("MODELO DE ML - LEAD SCORING (COM RASTREAMENTO DE ATIVIDADE)")
print("="*80)

# ==============================================================================
# 1. CARREGAMENTO DOS DADOS
# ==============================================================================
print("\n[1/7] Carregando dados...")
if not os.path.exists(CAMINHO_LEADS):
    print(f"Erro: Arquivo {CAMINHO_LEADS} não encontrado.")
    exit()

try: df = pd.read_csv(CAMINHO_LEADS, encoding='utf-8')
except: df = pd.read_csv(CAMINHO_LEADS, encoding='latin1')
df.columns = df.columns.str.strip()

# --- CARREGAMENTO EXTRA: PAULÍNIA ---
arquivos_paulinia = glob.glob(os.path.join(PASTA_DATA, '*paulinia*.csv')) + glob.glob(os.path.join(PASTA_DATA, '*Paulinia*.csv'))
if arquivos_paulinia:
    print(f"\n  [EXTRA] Encontrado(s) {len(arquivos_paulinia)} arquivo(s) de Paulínia.")
    for arq_p in set(arquivos_paulinia): # set para evitar duplicatas se o glob pegar o mesmo arquivo
        print(f"    -> Integrando: {os.path.basename(arq_p)}")
        try: df_p = pd.read_csv(arq_p, encoding='utf-8')
        except: df_p = pd.read_csv(arq_p, encoding='latin1')
        df_p.columns = df_p.columns.str.strip()
        df = pd.concat([df, df_p], ignore_index=True)
        print(f"    -> +{len(df_p)} leads adicionados à base principal.")

if os.path.exists(CAMINHO_PERDIDOS):
    try: df_perdidos = pd.read_csv(CAMINHO_PERDIDOS, encoding='utf-8')
    except: df_perdidos = pd.read_csv(CAMINHO_PERDIDOS, encoding='latin1')
    df_perdidos.columns = df_perdidos.columns.str.strip()
    
    if 'Record ID' in df.columns:
        df['Record ID'] = df['Record ID'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
    if 'Record ID' in df_perdidos.columns:
        df_perdidos['Record ID'] = df_perdidos['Record ID'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
        df_perdidos = df_perdidos.drop_duplicates('Record ID')
    
    cols_enrich = ['Record ID']
    motivo_col = [c for c in df_perdidos.columns if 'motivo' in c.lower()]
    if motivo_col: cols_enrich.append(motivo_col[0])
    
    df_enrich = df_perdidos[cols_enrich].copy()
    if len(cols_enrich) > 1: df_enrich.columns = ['Record ID', 'Motivo_Perda']
    df_enrich['In_Lost_Base'] = True
    
    df = df.merge(df_enrich, on='Record ID', how='left')
    df['In_Lost_Base'] = df['In_Lost_Base'].fillna(False)
else:
    print("  -> Base de perdidos não encontrada.")

# --- CARREGAMENTO DA NOVA BASE DE MATRÍCULAS (SOURCE OF TRUTH) ---
# ATUALIZAÇÃO: matriculas_finais_limpo.csv com Nome do Aluno = Nome do negócio
print("\n  [INFO] Carregando base de Matrículas Finais (Verdade Absoluta)...")
set_matriculados_finais = set()
set_p1_paga = set()  # Novo: set separado para P1 paga
if os.path.exists(CAMINHO_MATRICULAS):
    try:
        df_mat = pd.read_csv(CAMINHO_MATRICULAS, sep=';', encoding='utf-8-sig')
    except:
        try:
            df_mat = pd.read_csv(CAMINHO_MATRICULAS, sep=';', encoding='latin1')
        except:
            df_mat = pd.read_csv(CAMINHO_MATRICULAS, sep=None, engine='python', encoding='utf-8')

    df_mat.columns = df_mat.columns.str.strip()
    print(f"    -> Colunas encontradas: {list(df_mat.columns)}")

    # Estrutura esperada: Nome do Aluno, Status P1, Unidade, Tipo
    col_nome = next((c for c in df_mat.columns if 'nome' in c.lower() and 'aluno' in c.lower()), None)
    col_status_p1 = next((c for c in df_mat.columns if 'status' in c.lower() and 'p1' in c.lower()), None)
    col_unidade = next((c for c in df_mat.columns if 'unidade' in c.lower()), None)
    col_tipo = next((c for c in df_mat.columns if 'tipo' in c.lower()), None)

    if col_nome:
        print(f"    -> Coluna de nome detectada: '{col_nome}'")

        # Filtrar apenas alunos NOVOS (Excluir REMA)
        if col_tipo:
            antes = len(df_mat)
            df_mat = df_mat[~df_mat[col_tipo].astype(str).str.contains('REMA', case=False, na=False)]
            print(f"    -> Filtrado REMA: {antes - len(df_mat)} removidos, {len(df_mat)} restantes.")

        # Normalizar nome para matching (remove acentos, maiúsculas, espaços extras)
        import unicodedata
        def normalizar_nome(nome):
            if pd.isna(nome): return ""
            nome = str(nome).upper().strip()
            # Remove acentos
            nome = unicodedata.normalize('NFKD', nome)
            nome = ''.join(c for c in nome if not unicodedata.combining(c))
            # Remove caracteres especiais e espaços extras
            nome = ' '.join(nome.split())
            return nome

        df_mat['Nome_Normalizado'] = df_mat[col_nome].apply(normalizar_nome)

        # Todos na base são matriculados (independente do status P1)
        set_matriculados_finais = set(df_mat['Nome_Normalizado'].unique())
        print(f"    -> {len(set_matriculados_finais)} matrículas efetivas carregadas.")

        # Filtrar P1 = PAGA ou BOLSA para set separado
        if col_status_p1:
            df_p1_paga = df_mat[df_mat[col_status_p1].astype(str).str.upper().isin(['PAGA', 'BOLSA'])]
            set_p1_paga = set(df_p1_paga['Nome_Normalizado'].unique())
            print(f"    -> {len(set_p1_paga)} com P1 paga/bolsa (conversão confirmada).")
    else:
        print("    [AVISO] Coluna 'Nome do Aluno' não identificada em matriculas_finais_limpo.csv")
else:
    print("    [AVISO] Arquivo matriculas_finais_limpo.csv não encontrado. Usando apenas HubSpot.")

# ==============================================================================
# 2. PREPARAÇÃO DOS DADOS
# ==============================================================================
print("\n[2/7] Preparando features e CORRIGINDO DATAS...")

# Datas Principais
col_data = 'Data de criação'
df[col_data] = pd.to_datetime(df[col_data], dayfirst=True, errors='coerce')
df['Data de fechamento'] = pd.to_datetime(df['Data de fechamento'], dayfirst=True, errors='coerce')

# --- NOVA COLUNA: DATA ÚLTIMA ATIVIDADE (Apenas para Relatório, NÃO para Modelo) ---
# Tenta encontrar colunas com nomes prováveis
cols_atividade = [c for c in df.columns if 'atividade' in c.lower() and 'data' in c.lower()]
col_last_activity = 'Data da última atividade' # Default
if cols_atividade:
    col_last_activity = cols_atividade[0] # Pega a primeira que encontrar parecida
    print(f"  -> Coluna de atividade detectada: {col_last_activity}")

if col_last_activity in df.columns:
    df[col_last_activity] = pd.to_datetime(df[col_last_activity], dayfirst=True, errors='coerce')
    
    # CORREÇÃO: Preencher datas vazias com a Data de Criação (Fallback)
    nulos = df[col_last_activity].isna().sum()
    if nulos > 0:
        print(f"  [!]  Preenchendo {nulos} datas de atividade vazias com a Data de Criação.")
        df[col_last_activity] = df[col_last_activity].fillna(df['Data de criação'])
else:
    df[col_last_activity] = df['Data de criação'] # Se coluna não existir, usa criação como padrão

df['Ano'] = df['Data de criação'].dt.year
df['Mes'] = df['Data de criação'].dt.month
df['DiaSemana'] = df['Data de criação'].dt.dayofweek
df['Trimestre'] = df['Data de criação'].dt.quarter

# --- CORREÇÃO DE VAZAMENTO ---
df['Dias_Desde_Criacao'] = (pd.Timestamp.now() - df['Data de criação']).dt.days
df['Dias_Desde_Criacao'] = df['Dias_Desde_Criacao'].fillna(0).clip(lower=0)

# Tratar Nulos
df['Fonte original do tráfego'] = df['Fonte original do tráfego'].fillna('(Sem fonte)')
df['Detalhamento da fonte original do tráfego 1'] = df['Detalhamento da fonte original do tráfego 1'].fillna('(Sem detalhe)')
df['Número de atividades de vendas'] = pd.to_numeric(df['Número de atividades de vendas'], errors='coerce').fillna(0)
df['Unidade Desejada'] = df['Unidade Desejada'].fillna('(Sem unidade)')
for col in ['Ano', 'Mes', 'DiaSemana', 'Trimestre']: df[col] = df[col].fillna(0)

# Densidade
df['Densidade_Atividade'] = df['Número de atividades de vendas'] / (df['Dias_Desde_Criacao'] + 1)

# Categorização
def classificar_atividade_comercial(n):
    if n < 3: return '0_Outlier_Sem_Registro'
    elif n <= 6: return '1_Inicial_3_a_6'
    elif n <= 12: return '2_Engajamento_7_a_12'
    elif n <= 20: return '3_Persistencia_13_a_20'
    elif n <= 35: return '4_Alta_Persistencia_21_a_35'
    elif n <= 60: return '5_Recuperacao_Longa_36_a_60'
    else: return '6_Saturacao_Extrema_60+'

df['Faixa_Atividade'] = df['Número de atividades de vendas'].apply(classificar_atividade_comercial)

def categorizar_fonte(fonte):
    fonte = str(fonte).lower()
    if 'off-line' in fonte or 'offline' in fonte: return 'Offline'
    elif 'social pago' in fonte: return 'Meta_Ads'
    elif 'pesquisa paga' in fonte: return 'Google_Ads'
    elif 'pesquisa orgânica' in fonte: return 'Google_Organico'
    elif 'social orgânico' in fonte: return 'Social_Organico'
    elif 'tráfego direto' in fonte: return 'Direto'
    else: return 'Outros'

df['Fonte_Cat'] = df['Fonte original do tráfego'].apply(categorizar_fonte)

def definir_segmento_ml(fonte_cat):
    if fonte_cat == 'Offline': return 'Offline'
    elif fonte_cat in ['Google_Ads', 'Google_Organico']: return 'Digital_Busca'
    elif fonte_cat in ['Meta_Ads', 'Social_Organico']: return 'Digital_Social'
    else: return 'Digital_Direto'

df['Segmento_ML'] = df['Fonte_Cat'].apply(definir_segmento_ml)

def categorizar_detalhe_offline(row):
    if row['Fonte_Cat'] != 'Offline': return 'N/A'
    detalhe = str(row['Detalhamento da fonte original do tráfego 1']).upper()
    if 'CRM_UI' in detalhe: return 'CRM_UI'
    elif 'CONTACTS' in detalhe: return 'CONTACTS'
    elif 'IMPORT' in detalhe: return 'IMPORT'
    else: return 'Outros_Offline'

df['Detalhe_Offline'] = df.apply(categorizar_detalhe_offline, axis=1)

# ==============================================================================
# 3. DEFINIÇÃO DO TARGET (COM FEEDBACK LOOP)
# ==============================================================================
print("\n[3/7] Definindo target com feedback loop...")

# ATUALIZAÇÃO: Criar chave de cruzamento por Nome Normalizado (Nome do Aluno = Nome do negócio)
import unicodedata
import re

def normalizar_nome_hubspot(nome):
    if pd.isna(nome): return ""
    nome = str(nome).upper().strip()
    # Remove sufixos como " - 10 - UNIDADE"
    nome = re.sub(r'\s*-\s*\d+\s*-\s*.*$', '', nome)
    # Remove acentos
    nome = unicodedata.normalize('NFKD', nome)
    nome = ''.join(c for c in nome if not unicodedata.combining(c))
    # Remove caracteres especiais e espaços extras
    nome = ' '.join(nome.split())
    return nome

df['Nome_Normalizado'] = df['Nome do negócio'].apply(normalizar_nome_hubspot)
df['In_Matriculas_Finais'] = df['Nome_Normalizado'].isin(set_matriculados_finais)
df['In_P1_Paga'] = df['Nome_Normalizado'].isin(set_p1_paga)

# --- FEEDBACK LOOP: Carregar validacao anterior para aprender com erros ---
CAMINHO_VALIDACAO = os.path.join(PASTA_OUTPUT, 'DB_HELIO')
arquivos_validacao = glob.glob(os.path.join(CAMINHO_VALIDACAO, 'Evolucao_Temporal_Helio_*.xlsx'))
set_falsos_negativos = set()  # Leads nota baixa que converteram (erro do modelo)
set_verdadeiros_positivos = set()  # Leads nota alta que converteram (acerto)

if arquivos_validacao:
    ultimo_validacao = max(arquivos_validacao, key=os.path.getmtime)
    print(f"  [FEEDBACK] Carregando validacao: {os.path.basename(ultimo_validacao)}")
    try:
        df_val = pd.read_excel(ultimo_validacao, sheet_name='Base_Completa')

        # Identificar Falsos Negativos: Nota 1-3 que matricularam (ERRO - modelo subestimou)
        mask_fn = (df_val['Nota_Original_1a5'] < 4) & (df_val['Resultado_Conversao'] == 'Matriculou')
        if 'Nome_Negocio' in df_val.columns:
            set_falsos_negativos = set(df_val[mask_fn]['Nome_Negocio'].apply(normalizar_nome_hubspot).unique())
            print(f"    -> {len(set_falsos_negativos)} Falsos Negativos identificados (nota baixa que converteu)")

        # Identificar Verdadeiros Positivos: Nota 4-5 que matricularam (ACERTO)
        mask_tp = (df_val['Nota_Original_1a5'] >= 4) & (df_val['Resultado_Conversao'] == 'Matriculou')
        if 'Nome_Negocio' in df_val.columns:
            set_verdadeiros_positivos = set(df_val[mask_tp]['Nome_Negocio'].apply(normalizar_nome_hubspot).unique())
            print(f"    -> {len(set_verdadeiros_positivos)} Verdadeiros Positivos confirmados")
    except Exception as e:
        print(f"    [AVISO] Nao foi possivel carregar validacao: {e}")
else:
    print("  [INFO] Nenhuma validacao anterior encontrada. Treinando sem feedback loop.")

# Marcar leads com feedback
df['Is_Falso_Negativo'] = df['Nome_Normalizado'].isin(set_falsos_negativos)
df['Is_Verdadeiro_Positivo'] = df['Nome_Normalizado'].isin(set_verdadeiros_positivos)

def classificar_etapa(row):
    # 1. A Verdade Absoluta: Está na base de matrículas finais?
    if row['In_Matriculas_Finais']: return 'SUCESSO'

    # 2. P1 Paga também é considerado SUCESSO (próximo de matricular)
    if row['In_P1_Paga']: return 'SUCESSO'

    # 3. Feedback Loop: Falso Negativo confirmado vira SUCESSO
    if row['Is_Falso_Negativo']: return 'SUCESSO'

    etapa = str(row['Etapa do negócio']).lower()
    in_lost = row.get('In_Lost_Base', False)

    # 4. Fallback HubSpot
    if 'matrícula' in etapa: return 'SUCESSO'
    elif 'perdido' in etapa or in_lost: return 'PERDIDO'
    elif 'qualificação' in etapa: return 'QUALIFICACAO'
    elif 'pausa' in etapa: return 'PAUSA'
    else: return 'ANDAMENTO'

df['Etapa_Class'] = df.apply(classificar_etapa, axis=1)

# Log de matches encontrados
n_match_mat = df['In_Matriculas_Finais'].sum()
n_match_p1 = df['In_P1_Paga'].sum()
n_fn = df['Is_Falso_Negativo'].sum()
print(f"  [MATCH] {n_match_mat} leads encontrados na base de matriculas finais")
print(f"  [MATCH] {n_match_p1} leads com P1 paga")
print(f"  [FEEDBACK] {n_fn} Falsos Negativos agora marcados como SUCESSO (correcao do modelo)")

df_treino = df[df['Etapa_Class'].isin(['SUCESSO', 'PERDIDO'])].copy()
df_treino['Target'] = (df_treino['Etapa_Class'] == 'SUCESSO').astype(int)

# --- PESO PARA CASOS CONFIRMADOS (SAMPLE WEIGHT) - OTIMIZADO PARA PREVISÃO ---
# Estratégia: Balancear entre aprender erros SEM causar overfitting
# Falsos Negativos agora têm peso MODERADO (6.0) ao invés de 10.0 para evitar overfitting
df_treino['Sample_Weight'] = 1.0
df_treino.loc[df_treino['In_Matriculas_Finais'] == True, 'Sample_Weight'] = 3.0  # Peso 3x para matriculas reais (verdade absoluta)
df_treino.loc[df_treino['Is_Falso_Negativo'] == True, 'Sample_Weight'] = 6.0  # Peso 6x para FN - BALANCEADO (era 5, passou por 10, agora 6)
df_treino.loc[df_treino['Is_Verdadeiro_Positivo'] == True, 'Sample_Weight'] = 2.0  # Peso 2x para acertos confirmados

print(f"  [PESO] Distribuicao de pesos no treino (OTIMIZADO PARA PREVISÃO):")
print(f"    -> Peso 1.0 (padrao): {(df_treino['Sample_Weight'] == 1.0).sum()} leads")
print(f"    -> Peso 2.0 (VP confirmado): {(df_treino['Sample_Weight'] == 2.0).sum()} leads")
print(f"    -> Peso 3.0 (matricula real): {(df_treino['Sample_Weight'] == 3.0).sum()} leads")
print(f"    -> Peso 6.0 (FN - aprender erro moderadamente): {(df_treino['Sample_Weight'] == 6.0).sum()} leads")

df_score = df[df['Etapa_Class'].isin(['QUALIFICACAO', 'PAUSA'])].copy()

# ==============================================================================
# 4. ENCODING
# ==============================================================================
print("\n[4/7] Encoding...")
FEATURES_CAT = ['Fonte_Cat', 'Detalhe_Offline', 'Unidade Desejada', 'Faixa_Atividade']
# IMPORTANTE: Data da última atividade NÃO entra aqui para não enviesar o treino
FEATURES_NUM_BASE = ['Mes', 'DiaSemana', 'Trimestre', 'Dias_Desde_Criacao', 'Densidade_Atividade']

encoders = {}
for col in FEATURES_CAT:
    le = LabelEncoder()
    le.fit(df[col].astype(str))
    encoders[col] = le
    df_treino[f'{col}_enc'] = le.transform(df_treino[col].astype(str))
    df_score[f'{col}_enc'] = le.transform(df_score[col].astype(str))

# ==============================================================================
# 5. TREINAMENTO (COM SAMPLE WEIGHTS)
# ==============================================================================
print("\n[5/7] Treinando modelos com pesos ajustados...")

def treinar_modelo_segmentado(nome, df_segmento, features_segmento):
    print(f"  >>> Treinando: {nome}")
    if len(df_segmento) < 50: return None, 0, None
    X = df_segmento[features_segmento].values
    y = df_segmento['Target'].values

    # Extrair sample weights se existirem
    if 'Sample_Weight' in df_segmento.columns:
        sample_weights = df_segmento['Sample_Weight'].values
    else:
        sample_weights = np.ones(len(y))

    X_tr, X_te, y_tr, y_te, w_tr, w_te = train_test_split(
        X, y, sample_weights, test_size=0.2, random_state=42, stratify=y
    )

    clf = RandomForestClassifier(n_estimators=100, max_depth=12, min_samples_split=20,
                                 class_weight='balanced', random_state=42, n_jobs=-1)

    # Treinar COM sample_weight para dar mais importancia aos casos confirmados
    clf.fit(X_tr, y_tr, sample_weight=w_tr)

    auc = roc_auc_score(y_te, clf.predict_proba(X_te)[:, 1])
    imp = pd.DataFrame({'Feature': features_segmento, 'Importance': clf.feature_importances_}).sort_values('Importance', ascending=False)

    # Contar quantos casos pesados foram usados no treino
    n_pesados = (w_tr > 1.0).sum()
    print(f"      AUC: {auc:.4f} | Casos com peso extra no treino: {n_pesados}")
    return clf, auc, imp

modelos_dict = {}
for segmento in df['Segmento_ML'].unique():
    if segmento == 'Offline':
        cols_cat_enc = [f'{c}_enc' for c in ['Detalhe_Offline', 'Unidade Desejada', 'Faixa_Atividade']]
        df_tr_seg = df_treino[(df_treino['Segmento_ML'] == segmento) & (df_treino['Número de atividades de vendas'] >= 3)].copy()
        feats = cols_cat_enc + FEATURES_NUM_BASE + ['Número de atividades de vendas']
    else:
        cols_cat_enc = [f'{c}_enc' for c in ['Fonte_Cat', 'Unidade Desejada', 'Faixa_Atividade']]
        df_tr_seg = df_treino[df_treino['Segmento_ML'] == segmento].copy()
        feats = cols_cat_enc + FEATURES_NUM_BASE
        
    modelo, auc, imp = treinar_modelo_segmentado(segmento, df_tr_seg, feats)
    if modelo: modelos_dict[segmento] = {'model': modelo, 'auc': auc, 'features': feats, 'importance': imp}

# ==============================================================================
# 6. SCORING
# ==============================================================================
print("\n[6/7] Aplicando Scoring...")
df_score['Probabilidade_Conversao'] = 0.0

for segmento, dados in modelos_dict.items():
    mask = df_score['Segmento_ML'] == segmento
    if mask.any():
        X_score = df_score.loc[mask, dados['features']].values
        df_score.loc[mask, 'Probabilidade_Conversao'] = dados['model'].predict_proba(X_score)[:, 1]

# ESTRATÉGIA PREDITIVA: Manter thresholds ORIGINAIS mas com ajuste inteligente
# Razão: Thresholds fixos estão calibrados pelo modelo. Não devemos alterá-los artificialmente.
# Ao invés disso, vamos fazer ajustes APÓS o modelo rodar, baseados em contexto (atividade).
df_score['Nota_1a5'] = pd.cut(df_score['Probabilidade_Conversao'], bins=[0, 0.2, 0.4, 0.6, 0.8, 1.0], labels=[1, 2, 3, 4, 5], include_lowest=True).astype(int)

# AJUSTE INTELIGENTE: Reboost para leads com "voz fraca" (baixa atividade recente)
# Estes leads podem estar dormindo, não necessariamente perdidos
# Aplicamos reboost APENAS em leads que têm indica‌dores positivos (FORA de perdido)
if 'Data da última atividade' in df_score.columns:
    try:
        df_score['Dias_Sem_Atividade'] = (pd.Timestamp.now() - pd.to_datetime(df_score['Data da última atividade'], errors='coerce')).dt.days
        
        # Leads que estão em QUALIFICAÇÃO ou PAUSA (não em PERDIDO) + atividade baixa + nota baixa
        # = Candidatos a "ressuscitar"
        mask_reboost = (
            (df_score['Dias_Sem_Atividade'] > 30) & 
            (df_score['Nota_1a5'] == 2) &  # Apenas Nota 2, não Nota 1 (muito conservador)
            (df_score['Probabilidade_Conversao'] > 0.15)  # Modelo deu pelo menos 15%
        )
        
        # Reboost: aumenta 1 nota (2 → 3), mas MANTÉM a probabilidade como dado
        df_score.loc[mask_reboost, 'Nota_1a5'] = 3
        
        if mask_reboost.sum() > 0:
            print(f"  [REBOOST ATIVIDADE] {mask_reboost.sum()} leads Nota 2 → Nota 3 (atividade baixa + prob > 15%)")
    except Exception as e:
        print(f"  [AVISO] Não foi possível aplicar reboost de atividade: {e}")

corte_super = 0.90
def definir_prioridade(row):
    if row['Probabilidade_Conversao'] >= corte_super and row['Nota_1a5'] == 5: return '1_Super_Qualificado_TOP'
    elif row['Nota_1a5'] == 5: return '2_Alta_Prioridade_ALTA'
    elif row['Nota_1a5'] == 4: return '3_Media_Prioridade_MEDIA'
    else: return '4_Nutricao_Baixa'

df_score['Prioridade_Atendimento'] = df_score.apply(definir_prioridade, axis=1)
df_score['Tipo_Unidade'] = df_score['Unidade Desejada'].apply(lambda x: 'Própria' if any(u in str(x).upper() for u in UNIDADES_PROPRIAS) else 'Franqueada')

# --- ANÁLISE DE PERFIL ---
perfil_stats_list = []
for seg in modelos_dict.keys():
    subset = df[df['Segmento_ML'] == seg]
    if seg == 'Offline': subset = subset[subset['Número de atividades de vendas'] >= 3]
    sucesso = subset[subset['Etapa_Class'] == 'SUCESSO']
    perdido = subset[subset['Etapa_Class'] == 'PERDIDO']
    if len(sucesso) > 0 and len(perdido) > 0:
        perfil_stats_list.append({
            'Segmento': seg,
            'Ativ_Sucesso': sucesso['Número de atividades de vendas'].median(),
            'Ativ_Perdido': perdido['Número de atividades de vendas'].median(),
            'Dias_Sucesso': sucesso['Dias_Desde_Criacao'].median(),
            'Dias_Perdido': perdido['Dias_Desde_Criacao'].median(),
            'Dens_Sucesso': sucesso['Densidade_Atividade'].median(),
            'Dens_Perdido': perdido['Densidade_Atividade'].median(),
        })
df_perfil_stats = pd.DataFrame(perfil_stats_list)

# ==============================================================================
# 6.5 METRICAS DE FEEDBACK LOOP
# ==============================================================================
print("\n[FEEDBACK LOOP] Resumo do aprendizado com dados validados:")
print("="*60)

# Estatisticas do treino
total_treino = len(df_treino)
total_sucesso = (df_treino['Target'] == 1).sum()
total_perdido = (df_treino['Target'] == 0).sum()

print(f"  Base de Treino: {total_treino} leads")
print(f"    -> SUCESSO: {total_sucesso} ({total_sucesso/total_treino*100:.1f}%)")
print(f"    -> PERDIDO: {total_perdido} ({total_perdido/total_treino*100:.1f}%)")

# Contribuicao do feedback
n_mat_treino = df_treino['In_Matriculas_Finais'].sum()
n_fn_treino = df_treino['Is_Falso_Negativo'].sum()
n_vp_treino = df_treino['Is_Verdadeiro_Positivo'].sum()

print(f"\n  Contribuicao de Dados Validados:")
print(f"    -> Matriculas confirmadas: {n_mat_treino}")
print(f"    -> Falsos Negativos corrigidos: {n_fn_treino}")
print(f"    -> Verdadeiros Positivos reforçados: {n_vp_treino}")

# Impacto nos pesos
peso_total = df_treino['Sample_Weight'].sum()
peso_confirmados = df_treino[df_treino['Sample_Weight'] > 1.0]['Sample_Weight'].sum()
pct_peso = peso_confirmados / peso_total * 100 if peso_total > 0 else 0

print(f"\n  Impacto no Treino (via Sample Weights):")
print(f"    -> {pct_peso:.1f}% do 'peso' do treino vem de casos confirmados")
print(f"    -> Modelo agora prioriza aprender com dados reais")
print("="*60)

# ==============================================================================
# 7. EXPORTAÇÃO E GERAÇÃO DE EXCEL (COM DATA ATIVIDADE)
# ==============================================================================
print("\n[7/7] Gerando Relatórios com Data de Atividade...")

# Salvar Modelo
modelo_path = os.path.join(PASTA_MODELS, 'lead_scoring_model.pkl')
dados_export = {'encoders': encoders}
for seg, dados in modelos_dict.items():
    dados_export[f'model_{seg}'] = dados['model']
    dados_export[f'features_{seg}'] = dados['features']
with open(modelo_path, 'wb') as f: pickle.dump(dados_export, f)

# CSV Completo - AQUI ENTRA A DATA
csv_path = os.path.join(PASTA_DADOS, f'leads_scored_{DATA_HOJE}.csv')
cols_out = ['Record ID', 'Nome do negócio', 'Etapa do negócio', 'Unidade Desejada', 'Tipo_Unidade', 
            'Fonte_Cat', 'Segmento_ML', 'Faixa_Atividade', 'Número de atividades de vendas', 
            col_last_activity, # <--- ADICIONADO AQUI PARA EXPORTAR
            'Data de criação', 'Dias_Desde_Criacao', 'Probabilidade_Conversao', 'Nota_1a5', 'Prioridade_Atendimento']

# Garante que colunas existam
cols_existentes = [c for c in cols_out if c in df_score.columns]

df_score[cols_existentes].sort_values('Probabilidade_Conversao', ascending=False).to_csv(csv_path, index=False, sep=';', encoding='utf-8-sig')

# Log (agora com metricas de feedback)
log_path = os.path.join(PASTA_DADOS, 'Historico_Evolucao_IA.csv')
registros_log = []
for seg, dados in modelos_dict.items():
    registros_log.append({
        'Data_Execucao': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'Segmento': seg,
        'AUC': round(dados['auc'], 4),
        'Top_Features': ", ".join(dados['importance'].head(3)['Feature'].tolist()),
        'Matriculas_Confirmadas': n_mat_treino,
        'Falsos_Negativos_Corrigidos': n_fn_treino,
        'Pct_Peso_Confirmados': round(pct_peso, 1)
    })
df_log = pd.DataFrame(registros_log)
if os.path.exists(log_path): df_log = pd.concat([pd.read_csv(log_path, sep=';'), df_log])
df_log.to_csv(log_path, index=False, sep=';', encoding='utf-8-sig')

# --- FUNÇÕES DE EXCEL (LAYOUT) ---
HEADER_FILL = PatternFill(start_color='CE0E2D', end_color='CE0E2D', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def gerar_excel_segmentado(df_scored, df_historico, nome_arquivo, segmento_titulo, auc_modelo, stats_global, df_importance, df_perfil_stats, segmento_raw):
    wb = Workbook()
    source_text = f"Origem: {os.path.basename(CAMINHO_LEADS)}"

    def df_to_sheet(ws, df, start_row=1):
        # Seleciona colunas úteis para visualização, incluindo a data de atividade
        cols_visualizacao = ['Record ID', 'Nome do negócio', 'Unidade Desejada', 'Nota_1a5', 'Probabilidade_Conversao', 
                             'Prioridade_Atendimento', col_last_activity, 'Número de atividades de vendas', 'Etapa do negócio', 'Fonte_Cat']
        
        # Filtra apenas colunas que existem
        cols_finais = [c for c in cols_visualizacao if c in df.columns]
        df_vis = df[cols_finais].copy()
        
        # Formata data para string bonita
        if col_last_activity in df_vis.columns:
             df_vis[col_last_activity] = df_vis[col_last_activity].dt.strftime('%d/%m/%Y')

        rows = list(dataframe_to_rows(df_vis, index=False, header=True))
        for r_idx, row in enumerate(rows, start=start_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = THIN_BORDER
        
        # 2. Formata APENAS o cabeçalho
        for c_idx in range(1, len(rows[0]) + 1):
            cell = ws.cell(row=start_row, column=c_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

    def add_sheet_description(ws, lines, start_row=2):
        for i, text in enumerate(lines, start=start_row):
            cell = ws.cell(row=i, column=1, value=text)
            cell.font = Font(bold=False, size=9)
            cell.alignment = Alignment(wrap_text=True)

    def ajustar_largura(ws):
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # --- ABA 1: RESUMO ---
    ws1 = wb.active; ws1.title = '1_Resumo'
    ws1['A1'] = f'LEAD SCORING - {segmento_titulo}'; ws1['A1'].font = Font(bold=True, size=14)
    ws1['A2'] = source_text
    ws1['A3'] = f'Data: {DATA_HOJE}'
    ws1['A5'] = f'Precisão do Modelo (AUC): {auc_modelo:.4f}'
    
    add_sheet_description(ws1, [
        'Objetivo: Resumo executivo do Lead Scoring para tomada de decisão.',
        'Público: Gestores comerciais e operação. Interpretação: AUC indica qualidade do modelo; Notas 1-5 refletem probabilidade de matrícula.'
    ], start_row=6)

    # Contexto Global
    ws1['E3'] = "CONTEXTO GLOBAL"; ws1['E3'].font = Font(bold=True)
    ws1['E4'] = f"Total Leads Pontuados: {stats_global['total']:,}"
    
    # Tabela Resumo
    ws1['A7'] = "DISTRIBUIÇÃO POR NOTA:"
    ws1['A7'].font = Font(bold=True)
    if not df_scored.empty:
        resumo_notas = df_scored.groupby('Nota_1a5').agg(Qtd=('Nota_1a5', 'count'), Prob=('Probabilidade_Conversao', 'mean')).reset_index()
        resumo_notas['Share'] = resumo_notas['Qtd'] / resumo_notas['Qtd'].sum()
        
        # Escreve a tabela manual (pois df_to_sheet foi customizado acima para abas de lista)
        rows = list(dataframe_to_rows(resumo_notas, index=False, header=True))
        for r_idx, row in enumerate(rows, start=8):
             for c_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=r_idx, column=c_idx, value=value)
                cell.border = THIN_BORDER
                if r_idx == 8: cell.fill=HEADER_FILL; cell.font=HEADER_FONT

        # Formatar %
        for r in range(9, 9 + len(resumo_notas)):
            ws1.cell(row=r, column=3).number_format = '0.0%'
            ws1.cell(row=r, column=4).number_format = '0.0%'

    # --- ABA 2: TOP 500 (NOTA 5) ---
    ws2 = wb.create_sheet('2_Top500_Nota5')
    ws2['A1'] = 'TOP 500 LEADS - NOTA 5'; ws2['A1'].font = Font(bold=True)
    add_sheet_description(ws2, ['Objetivo: Lista dos 500 leads com maior probabilidade estimada (Nota 5).'], start_row=3)
    df_to_sheet(ws2, df_scored[df_scored['Nota_1a5'] == 5].head(500), start_row=6)
    ajustar_largura(ws2)

    # --- ABA 3: NOTA 4 ---
    ws3 = wb.create_sheet('3_Nota4')
    ws3['A1'] = 'TOP 500 LEADS - NOTA 4'; ws3['A1'].font = Font(bold=True)
    df_to_sheet(ws3, df_scored[df_scored['Nota_1a5'] == 4].head(500), start_row=6)
    ajustar_largura(ws3)
    
    # --- ABA 4: TOP 10 POR UNIDADE ---
    ws4 = wb.create_sheet('4_Top10_Por_Unidade')
    ws4['A1'] = 'TOP 10 POR UNIDADE'; ws4['A1'].font = Font(bold=True)
    if not df_scored.empty:
        top_unidade = df_scored.groupby('Unidade Desejada').apply(lambda x: x.nlargest(10, 'Probabilidade_Conversao')).reset_index(drop=True)
        df_to_sheet(ws4, top_unidade, start_row=3)
    ajustar_largura(ws4)

    # --- ABA 5: QUALIFICACAO ---
    ws5 = wb.create_sheet('5_Em_Qualificacao')
    ws5['A1'] = 'LEADS EM QUALIFICACAO'; ws5['A1'].font = Font(bold=True)
    df_to_sheet(ws5, df_scored[df_scored['Etapa do negócio'].str.contains('QUALIFICAÇÃO', case=False, na=False)].head(1000), start_row=6)
    ajustar_largura(ws5)
    
    # --- ABA 6: PAUSA ---
    ws6 = wb.create_sheet('6_Em_Pausa')
    ws6['A1'] = 'LEADS EM PAUSA'; ws6['A1'].font = Font(bold=True)
    df_to_sheet(ws6, df_scored[df_scored['Etapa do negócio'].str.contains('PAUSA', case=False, na=False)].head(1000), start_row=6)
    ajustar_largura(ws6)

    # --- ABA 8: FATORES DECISIVOS ---
    ws8 = wb.create_sheet('8_Fatores_Decisivos')
    ws8['A1'] = 'O QUE DEFINE A NOTA DO LEAD?'; ws8['A1'].font = Font(bold=True)
    if df_importance is not None:
        df_imp_final = df_importance.copy(); df_imp_final['Importance_Acum'] = df_imp_final['Importance'].cumsum()
        
        # Manual write for importance (different structure)
        rows = list(dataframe_to_rows(df_imp_final, index=False, header=True))
        for r_idx, row in enumerate(rows, start=5):
             for c_idx, value in enumerate(row, 1):
                cell = ws8.cell(row=r_idx, column=c_idx, value=value)
                cell.border = THIN_BORDER
                if r_idx == 5: cell.fill=HEADER_FILL; cell.font=HEADER_FONT

        for r in range(6, 6 + len(df_importance)):
            ws8.cell(row=r, column=2).number_format = '0.0%'
            ws8.cell(row=r, column=3).number_format = '0.0%'
    ajustar_largura(ws8)

    # --- ABA 9: PERFIL IDEAL ---
    ws9 = wb.create_sheet('9_Perfil_Ideal_Sucesso')
    ws9['A1'] = 'O DNA DO SUCESSO'; ws9['A1'].font = Font(bold=True, size=12, color='CE0E2D')
    if not df_perfil_stats.empty and segmento_raw is not None:
        stats_seg = df_perfil_stats[df_perfil_stats['Segmento'] == segmento_raw]
        if not stats_seg.empty:
            row = stats_seg.iloc[0]
            dados_perfil = [
                ['Característica', 'Leads de SUCESSO (Mediana)', 'Leads PERDIDOS (Mediana)', 'Conclusão'],
                ['Atividades de Vendas', row['Ativ_Sucesso'], row['Ativ_Perdido'], 'Meta sugerida'],
                ['Tempo até Resolução', row['Dias_Sucesso'], row['Dias_Perdido'], 'Ciclo esperado'],
                ['Frequência Contato', f"1 a cada {int(1/row['Dens_Sucesso'])} dias", f"1 a cada {int(1/row['Dens_Perdido'])} dias", 'Ritmo ideal']
            ]
            for i, linha in enumerate(dados_perfil, start=5):
                for j, val in enumerate(linha, start=1):
                    cell = ws9.cell(row=i, column=j, value=val)
                    cell.border = THIN_BORDER
                    if i == 5: 
                        cell.fill = HEADER_FILL
                        cell.font = HEADER_FONT
            ajustar_largura(ws9)

    # --- ABA 10: PROJECAO FINANCEIRA ---
    ws10 = wb.create_sheet('10_Projecao_Financeira')
    ws10['A1'] = 'PROJEÇÃO FINANCEIRA (PIPELINE REPRESADO)'; ws10['A1'].font = Font(bold=True, size=12, color='CE0E2D')
    
    if not df_scored.empty:
        df_fin = df_scored.copy()
        df_fin['Ticket'] = df_fin['Tipo_Unidade'].apply(lambda x: TICKET_PROPRIA if x=='Própria' else TICKET_FRANQUIA)
        df_fin['Esp_1M'] = df_fin['Ticket'] * df_fin['Probabilidade_Conversao']
        df_fin['Esp_12M'] = df_fin['Esp_1M'] * 12
        df_fin['LTV_Total'] = df_fin['Esp_1M'] * 66 # Simplificado

        def escrever_tabela_fin(ws, titulo, df_sub, start_r):
            ws.cell(row=start_r, column=1, value=titulo).font = Font(bold=True)
            if df_sub.empty: return start_r + 3
            fin_stats = df_sub.groupby('Nota_1a5').agg(Qtd=('Record ID', 'count'), Rec_1M=('Esp_1M', 'sum'), Rec_12M=('Esp_12M', 'sum'), LTV=('LTV_Total', 'sum')).reset_index().sort_values('Nota_1a5', ascending=False)
            
            # Header
            cols = ['Nota', 'Qtd', 'Rec. 1 Mês', 'Rec. 1 Ano', 'LTV Total']
            for c, val in enumerate(cols, 1):
                cell = ws.cell(row=start_r+1, column=c, value=val); cell.fill=HEADER_FILL; cell.font=HEADER_FONT; cell.border=THIN_BORDER
            
            # Rows
            for r, row in enumerate(dataframe_to_rows(fin_stats, index=False, header=False), start_r+2):
                ws.cell(r, 1, row[0]).border = THIN_BORDER
                ws.cell(r, 2, row[1]).border = THIN_BORDER
                for c in range(3): 
                    cell = ws.cell(r, 3+c, row[2+c]); cell.number_format='#,##0.00'; cell.border=THIN_BORDER
            return start_r + len(fin_stats) + 5

        row_next = escrever_tabela_fin(ws10, "1. UNIDADES PRÓPRIAS", df_fin[df_fin['Tipo_Unidade']=='Própria'], 4)
        escrever_tabela_fin(ws10, "2. UNIDADES FRANQUEADAS", df_fin[df_fin['Tipo_Unidade']=='Franqueada'], row_next)
        ajustar_largura(ws10)

    try: wb.save(nome_arquivo)
    except: wb.save(nome_arquivo.replace('.xlsx', '_backup.xlsx'))

# Gerar Arquivos
stats_global = {'total': len(df_score)}
for seg in modelos_dict.keys():
    df_out = df_score[df_score['Segmento_ML'] == seg]
    df_his = df[df['Segmento_ML'] == seg]
    path = os.path.join(PASTA_RELATORIOS, f'Lead_Scoring_{seg}_{DATA_HOJE}.xlsx')
    gerar_excel_segmentado(df_out, df_his, path, seg.upper(), modelos_dict[seg]['auc'], stats_global, modelos_dict[seg]['importance'], df_perfil_stats, seg)

print("\n[SUCESSO] Processo Finalizado. Vazamento Corrigido & Layout Restaurado.")
print(f"Arquivos em: {PASTA_RELATORIOS}")