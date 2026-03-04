"""
==============================================================================
GERADOR DE PLANILHAS DE FEEDBACK (FRANQUIAS E PRÓPRIAS)
==============================================================================
Script: 11.Gerar_Planilhas_Feedback.py
Objetivo: Gerar 2 arquivos Excel (Próprias e Franquias) com abas por unidade
          para coleta de feedback simples.
Colunas:  ID, Nome do Negócio, Nota, Feedback
Data:     2026-01-22
==============================================================================
"""

import pandas as pd
import os
import glob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASTA_DADOS = os.path.join(CAMINHO_BASE, 'Outputs', 'Dados_Scored')
PASTA_OUTPUT = os.path.join(CAMINHO_BASE, 'Outputs', 'Controle_Feedback')

os.makedirs(PASTA_OUTPUT, exist_ok=True)

UNIDADES_PROPRIAS = [
    'VILA LEOPOLDINA', 'SANTANA', 'PINHEIROS', 'PERDIZES', 
    'PACAEMBU', 'MORUMBI', 'JARDINS', 'ITAIM'
]

print("="*80)
print("GERADOR DE PLANILHAS DE FEEDBACK (SIMPLIFICADO)")
print("="*80)

# ==============================================================================
# 1. CARREGAR DADOS
# ==============================================================================
list_of_files = glob.glob(os.path.join(PASTA_DADOS, 'leads_scored_*.csv'))

if not list_of_files:
    print(f"[ERRO] Nenhum arquivo de leads encontrado em: {PASTA_DADOS}")
    exit()

latest_file = max(list_of_files, key=os.path.getmtime)
print(f"[LENDO] Arquivo base: {os.path.basename(latest_file)}")

try:
    df = pd.read_csv(latest_file, sep=';', encoding='utf-8-sig')
except:
    df = pd.read_csv(latest_file, sep=';', encoding='latin1')

# Garantir classificação de unidade
if 'Tipo_Unidade' not in df.columns:
    df['Tipo_Unidade'] = df['Unidade Desejada'].apply(
        lambda u: 'Própria' if any(p in str(u).upper() for p in UNIDADES_PROPRIAS) else 'Franqueada'
    )

# Selecionar e Renomear Colunas
cols_map = {
    'Record ID': 'ID',
    'Nome do negócio': 'Nome do Negócio',
    'Nota_1a5': 'Nota'
}

# Filtrar colunas existentes
cols_to_keep = [c for c in cols_map.keys() if c in df.columns] + ['Unidade Desejada', 'Tipo_Unidade']
df_final = df[cols_to_keep].rename(columns=cols_map)

# Adicionar coluna de Feedback vazia
df_final['Feedback'] = ''

# Ordenar por Nota (maior para menor)
if 'Nota' in df_final.columns:
    df_final = df_final.sort_values('Nota', ascending=False)

# ==============================================================================
# 2. FUNÇÃO GERADORA DE EXCEL
# ==============================================================================
def gerar_excel_por_tipo(df_subset, nome_arquivo):
    wb = Workbook()
    # Remover aba padrão
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    unidades = sorted(df_subset['Unidade Desejada'].dropna().unique())
    
    # Estilos
    HEADER_FILL = PatternFill(start_color='CE0E2D', end_color='CE0E2D', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF')
    THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for unidade in unidades:
        # Limpar nome da aba (Excel não aceita alguns caracteres e max 31 chars)
        nome_aba = str(unidade).replace('/', '-').replace('\\', '').replace('?', '').replace('*', '').replace('[', '').replace(']', '')[:30]
        ws = wb.create_sheet(title=nome_aba)
        
        # Filtrar dados da unidade
        df_unidade = df_subset[df_subset['Unidade Desejada'] == unidade].copy()
        
        # Selecionar colunas finais para exibição
        cols_view = ['ID', 'Nome do Negócio', 'Nota', 'Feedback']
        df_view = df_unidade[cols_view]
        
        # Escrever dados
        for r_idx, row in enumerate(dataframe_to_rows(df_view, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = THIN_BORDER
                
                if r_idx == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal='center')
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 15  # ID
        ws.column_dimensions['B'].width = 40  # Nome
        ws.column_dimensions['C'].width = 10  # Nota
        ws.column_dimensions['D'].width = 50  # Feedback
    
    caminho_final = os.path.join(PASTA_OUTPUT, nome_arquivo)
    wb.save(caminho_final)
    print(f"   -> Gerado: {nome_arquivo} ({len(unidades)} abas)")

# ==============================================================================
# 3. EXECUÇÃO
# ==============================================================================
DATA_HOJE = datetime.now().strftime('%Y-%m-%d')

gerar_excel_por_tipo(df_final[df_final['Tipo_Unidade'] == 'Própria'], f'Feedback_Proprias_{DATA_HOJE}.xlsx')
gerar_excel_por_tipo(df_final[df_final['Tipo_Unidade'] != 'Própria'], f'Feedback_Franquias_{DATA_HOJE}.xlsx')

print(f"\n[SUCESSO] Arquivos salvos em: {PASTA_OUTPUT}")