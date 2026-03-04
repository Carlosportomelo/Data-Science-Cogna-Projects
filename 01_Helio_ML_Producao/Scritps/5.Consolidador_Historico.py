"""
==============================================================================
CONSOLIDADOR DE HISTÓRICO - EVOLUÇÃO TEMPORAL
==============================================================================
Script: 3.Consolidador_Historico.py
Objetivo: Ler todos os CSVs diários (atuais e backup) e gerar um painel
          de evolução da qualidade dos leads por unidade ao longo do tempo.
Data: 2025-12-19
==============================================================================
"""

import pandas as pd
import os
import glob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side

# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASTA_OUTPUT = os.path.join(CAMINHO_BASE, 'Outputs')
PASTA_DADOS = os.path.join(PASTA_OUTPUT, 'Dados_Scored')
PASTA_BACKUP = os.path.join(PASTA_DADOS, 'Backup')
PASTA_DESTINO = os.path.join(PASTA_OUTPUT, 'DB_HELIO')
os.makedirs(PASTA_DESTINO, exist_ok=True)

# Lista para garantir classificação caso falte no CSV antigo
UNIDADES_PROPRIAS = [
    'VILA LEOPOLDINA', 'SANTANA', 'PINHEIROS', 'PERDIZES', 
    'PACAEMBU', 'MORUMBI', 'JARDINS', 'ITAIM'
]

print("="*80)
print("CONSOLIDADOR DE HISTÓRICO TEMPORAL")
print("="*80)

# ==============================================================================
# 1. COLETAR ARQUIVOS
# ==============================================================================
arquivos_atuais = glob.glob(os.path.join(PASTA_DADOS, 'leads_scored_*.csv'))
arquivos_backup = glob.glob(os.path.join(PASTA_BACKUP, 'leads_scored_*.csv'))
todos_arquivos = arquivos_atuais + arquivos_backup

if not todos_arquivos:
    print("❌ Nenhum arquivo de histórico encontrado.")
    exit()

print(f"📂 Encontrados {len(todos_arquivos)} arquivos de histórico (Atuais + Backup).")

# ==============================================================================
# 2. PROCESSAR DADOS
# ==============================================================================
dfs = []

for arq in todos_arquivos:
    try:
        # Extrair data do nome do arquivo (leads_scored_YYYY-MM-DD.csv)
        nome = os.path.basename(arq)
        data_str = nome.replace('leads_scored_', '').replace('.csv', '')
        
        # Ler CSV
        df_temp = pd.read_csv(arq, sep=';', encoding='utf-8-sig', usecols=lambda c: c in ['Unidade Desejada', 'Nota_1a5', 'Tipo_Unidade', 'Probabilidade_Conversao'])
        
        # Adicionar coluna de data
        df_temp['Data_Ref'] = data_str
        
        # Garantir Tipo_Unidade (para arquivos antigos que não tinham essa coluna)
        if 'Tipo_Unidade' not in df_temp.columns or df_temp['Tipo_Unidade'].isnull().any():
            def classificar(u):
                return 'Própria' if any(p in str(u).upper() for p in UNIDADES_PROPRIAS) else 'Franqueada'
            df_temp['Tipo_Unidade'] = df_temp['Unidade Desejada'].apply(classificar)
            
        dfs.append(df_temp)
        print(f"   → Lido: {data_str} ({len(df_temp)} leads)")
    except Exception as e:
        print(f"   ⚠️ Erro ao ler {nome}: {e}")

if not dfs:
    print("❌ Falha ao consolidar dados.")
    exit()

df_full = pd.concat(dfs, ignore_index=True)
df_full['Data_Ref'] = pd.to_datetime(df_full['Data_Ref']).dt.date
df_full = df_full.sort_values('Data_Ref')

# ==============================================================================
# 3. GERAR RELATÓRIO DE EVOLUÇÃO
# ==============================================================================
wb = Workbook()
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')

def criar_aba_evolucao(wb, titulo_aba, df_filtrado):
    ws = wb.create_sheet(titulo_aba)
    ws['A1'] = f"EVOLUÇÃO: {titulo_aba.upper()}"
    ws['A1'].font = Font(bold=True, size=14)
    
    if df_filtrado.empty:
        ws['A3'] = "Sem dados."
        return

    # Pivot: Data x Unidade (Média da Nota)
    pivot_nota = pd.pivot_table(
        df_filtrado, 
        index='Data_Ref', 
        columns='Unidade Desejada', 
        values='Nota_1a5', 
        aggfunc='mean'
    ).round(2)
    
    # Pivot: Data x Unidade (% Qualificados Nota 4+5)
    df_filtrado['Is_Qualificado'] = df_filtrado['Nota_1a5'].apply(lambda x: 1 if x >= 4 else 0)
    pivot_qual = pd.pivot_table(
        df_filtrado,
        index='Data_Ref',
        columns='Unidade Desejada',
        values='Is_Qualificado',
        aggfunc='mean'
    )

    # Escrever Tabela de Nota Média
    ws['A3'] = "NOTA MÉDIA (1 a 5) POR DIA"
    ws['A3'].font = Font(bold=True)
    
    rows = dataframe_to_rows(pivot_qual.reset_index(), index=False, header=True)
    
    # Cabeçalho
    ws.append([]) # Espaço
    ws.append(['DATA'] + list(pivot_qual.columns))
    
    for r_idx, row in enumerate(dataframe_to_rows(pivot_qual.reset_index(), index=False, header=False), start=6):
        ws.append(row)
        # Formatar como porcentagem
        for c_idx in range(2, len(row) + 2):
            ws.cell(row=r_idx, column=c_idx).number_format = '0%'

    # Estilizar Cabeçalho
    for cell in ws[5]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

# Remover aba padrão
wb.remove(wb.active)

# Gerar abas
criar_aba_evolucao(wb, "Proprias_Qualidade", df_full[df_full['Tipo_Unidade'] == 'Própria'])
criar_aba_evolucao(wb, "Franquias_Qualidade", df_full[df_full['Tipo_Unidade'] == 'Franqueada'])

# --- ABAS: HISTÓRICO DE RANKING (RELATÓRIOS CONSOLIDADOS) ---
print("   → Gerando histórico consolidado dos relatórios (Rankings)...")

def criar_aba_historico_ranking(wb, titulo_aba, df_subset):
    ws = wb.create_sheet(titulo_aba)
    ws['A1'] = f"HISTÓRICO DE RELATÓRIOS: {titulo_aba.upper()}"
    ws['A1'].font = Font(bold=True, size=14)
    
    if df_subset.empty:
        ws['A3'] = "Sem dados."
        return

    # Agrupar dados para recriar o Ranking (Unidade + Data)
    # Pivotar Nota_1a5 para colunas
    df_rank = df_subset.groupby(['Data_Ref', 'Unidade Desejada', 'Nota_1a5']).size().unstack(fill_value=0)
    
    # Garantir colunas 1 a 5
    for i in range(1, 6):
        if i not in df_rank.columns:
            df_rank[i] = 0
    df_rank = df_rank[[1, 2, 3, 4, 5]]
    
    # Calcular Totais e %
    df_rank['Total_Leads'] = df_rank.sum(axis=1)
    df_rank['Qualificados (4+5)'] = df_rank[4] + df_rank[5]
    df_rank['% Qualificados'] = (df_rank['Qualificados (4+5)'] / df_rank['Total_Leads'])
    
    # Resetar index para virar tabela plana
    df_final = df_rank.reset_index().sort_values(['Data_Ref', 'Total_Leads'], ascending=[False, False])
    
    # Escrever no Excel
    ws.append([])
    ws.append(list(df_final.columns)) # Header manual para controle
    
    for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=False), start=4):
        ws.append(row)
        # Formatar coluna de % (índice 9, pois começa em 1 e temos Data+Unidade+5Notas+Total+Quali = 9 colunas antes da %)
        ws.cell(row=r_idx, column=9).number_format = '0.0%'
        
    # Estilo Header
    for cell in ws[3]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        
    # Ajustar larguras
    ws.column_dimensions['A'].width = 12 # Data
    ws.column_dimensions['B'].width = 30 # Unidade

criar_aba_historico_ranking(wb, "Historico_Proprias", df_full[df_full['Tipo_Unidade'] == 'Própria'])
criar_aba_historico_ranking(wb, "Historico_Franqueadas", df_full[df_full['Tipo_Unidade'] == 'Franqueada'])

nome_arquivo = os.path.join(PASTA_DESTINO, f'Painel_Evolucao_Temporal_{datetime.now().strftime("%Y-%m-%d")}.xlsx')
wb.save(nome_arquivo)

print(f"\n✅ Painel gerado com sucesso: {nome_arquivo}")
print("   Abra este arquivo para ver a evolução dia a dia das unidades.")