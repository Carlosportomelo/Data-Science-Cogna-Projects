"""
==============================================================================
COMPARADOR DE RELATÓRIOS EXCEL (AUDITORIA DE IMPACTO)
==============================================================================
Objetivo: Comparar os relatórios Excel atuais (Relatorios_ML) com suas versões
          de backup (Relatorios_ML/Backup) para medir o impacto da correção.
==============================================================================
"""

import pandas as pd
import os
import glob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Configuração de Caminhos
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DIR_ATUAL = os.path.join(BASE_DIR, "Outputs", "Relatorios_ML")
# Aponta para a nova estrutura de Caixa Forte para garantir a comparação correta
DIR_BACKUP = os.path.join(DIR_ATUAL, "Backup", "Historico_Caixa_Forte")

print("="*80)
print("AUDITORIA DE IMPACTO - COMPARADOR DE RELATÓRIOS EXCEL")
print("="*80)

def encontrar_arquivo_mais_recente(diretorio, padrao):
    arquivos = glob.glob(os.path.join(diretorio, padrao))
    if not arquivos:
        return None
    return max(arquivos, key=os.path.getmtime)

# Segmentos para verificar
segmentos = ['Digital_Busca', 'Digital_Direto', 'Digital_Social', 'Offline']
data_hoje = pd.Timestamp.now().strftime('%Y-%m-%d')

# Tentar encontrar arquivos com a data de hoje
# Encontrar pares de arquivos (Mais recente na pasta Atual vs Mais recente na pasta Backup)
arquivos_para_comparar = []
for seg in segmentos:
    nome_arquivo = f"Lead_Scoring_{seg}_{data_hoje}.xlsx"
    caminho_atual = os.path.join(DIR_ATUAL, nome_arquivo)
    caminho_backup = os.path.join(DIR_BACKUP, nome_arquivo)
    padrao = f"Lead_Scoring_{seg}_*.xlsx"
    caminho_atual = encontrar_arquivo_mais_recente(DIR_ATUAL, padrao)
    caminho_backup = encontrar_arquivo_mais_recente(DIR_BACKUP, padrao)
    
    if caminho_atual and caminho_backup:
        arquivos_para_comparar.append((seg, caminho_atual, caminho_backup))

if not arquivos_para_comparar:
    print(f"❌ Nenhum par de arquivos (Atual vs Backup) encontrado para a data {data_hoje}.")
    print("   Verifique se o script de scoring rodou pelo menos 2 vezes hoje.")
    print(f"❌ Nenhum par de arquivos (Atual vs Backup) encontrado.")
    print("   Certifique-se de que existem relatórios gerados na pasta Relatorios_ML e em Relatorios_ML/Backup.")
    exit()

print(f"🔍 Encontrados {len(arquivos_para_comparar)} pares de arquivos para análise.\n")

# Listas para acumular dados do relatório Excel
audit_data = []
resumo_stats = []
transicoes_list = []
novos_leads_list = []

for seg, f_atual, f_backup in arquivos_para_comparar:
    # Obter datas dos arquivos para evidenciar o intervalo de tempo
    time_atual = datetime.fromtimestamp(os.path.getmtime(f_atual)).strftime('%d/%m %H:%M')
    time_backup = datetime.fromtimestamp(os.path.getmtime(f_backup)).strftime('%d/%m %H:%M')

    print(f"--- ANÁLISE: {seg.upper()} ---")
    print(f"   📂 Comparando: {time_backup} (Base) vs {time_atual} (Atual)")
    
    # Ler abas principais para capturar o máximo de leads
    # Expandido para capturar leads em todas as fases relevantes
    abas_interesse = ['2_Top500_Nota5', '3_Nota4', '5_Em_Qualificacao', '6_Em_Pausa']
    
    dfs_new = []
    dfs_old = []
    
    try:
        # Ler Atual
        xls_new = pd.ExcelFile(f_atual)
        for aba in abas_interesse:
            if aba in xls_new.sheet_names:
                # Pular cabeçalho descritivo (geralmente as primeiras linhas são texto)
                # O script gera header na linha 6 (index 5) para Top500 e Qualificacao
                df = pd.read_excel(xls_new, sheet_name=aba, header=5, dtype={'Record ID': str})
                if 'Record ID' in df.columns:
                    dfs_new.append(df)
        
        # Ler Backup
        xls_old = pd.ExcelFile(f_backup)
        for aba in abas_interesse:
            if aba in xls_old.sheet_names:
                df = pd.read_excel(xls_old, sheet_name=aba, header=5, dtype={'Record ID': str})
                if 'Record ID' in df.columns:
                    dfs_old.append(df)
                    
    except Exception as e:
        print(f"   ❌ Erro ao ler Excel: {e}")
        continue

    if not dfs_new or not dfs_old:
        print("   ⚠️  Abas de dados não encontradas ou vazias.")
        continue

    # Concatenar e remover duplicatas (um lead pode aparecer em abas diferentes se a lógica mudar, mas aqui garantimos unicidade por ID)
    df_new_full = pd.concat(dfs_new).drop_duplicates('Record ID')
    df_old_full = pd.concat(dfs_old).drop_duplicates('Record ID')
    
    # Identificar Novos Leads (Estão no New, mas não no Old)
    novos = df_new_full[~df_new_full['Record ID'].isin(df_old_full['Record ID'])].copy()
    print(f"   🆕 Novos Leads (Entraram): {len(novos):,}")
    
    if not novos.empty:
        novos['Segmento'] = seg
        novos['Status_Auditoria'] = 'Novo Lead'
        novos_leads_list.append(novos)

    # Merge
    merged = pd.merge(df_new_full, df_old_full, on='Record ID', how='inner', suffixes=('_new', '_old'))
    
    # Stats do segmento
    leads_corrigidos = 0
    notas_alteradas = 0

    # 1. Verificar Mudança de Data (Dias_Desde_Criacao)
    if 'Dias_Desde_Criacao_new' in merged.columns and 'Dias_Desde_Criacao_old' in merged.columns:
        dias_new = pd.to_numeric(merged['Dias_Desde_Criacao_new'], errors='coerce').fillna(-1)
        dias_old = pd.to_numeric(merged['Dias_Desde_Criacao_old'], errors='coerce').fillna(-1)
        
        diff = (dias_new - dias_old).abs()
        mudaram_data = merged[diff > 1] # Tolerância de 1 dia
        
        print(f"   📅 Leads com idade corrigida: {len(mudaram_data):,} (de {len(merged):,} analisados)")
        leads_corrigidos = len(mudaram_data)
        
        if not mudaram_data.empty:
            temp = mudaram_data.copy()
            temp['Segmento'] = seg
            temp['Tipo_Alteracao'] = 'Correção de Data (Idade)'
            temp['Detalhe_Antes'] = temp['Dias_Desde_Criacao_old'].astype(str) + ' dias'
            temp['Detalhe_Depois'] = temp['Dias_Desde_Criacao_new'].astype(str) + ' dias'
            audit_data.append(temp)

        if len(mudaram_data) > 0:
            ex = mudaram_data.iloc[0]
            print(f"      Exemplo (ID {ex['Record ID']}): Antes {ex['Dias_Desde_Criacao_old']} dias -> Agora {ex['Dias_Desde_Criacao_new']} dias")

    # 2. Verificar Mudança de Nota
    if 'Nota_1a5_new' in merged.columns and 'Nota_1a5_old' in merged.columns:
        nota_new = pd.to_numeric(merged['Nota_1a5_new'], errors='coerce').fillna(0)
        nota_old = pd.to_numeric(merged['Nota_1a5_old'], errors='coerce').fillna(0)
        
        caiu = merged[nota_new < nota_old]
        subiu = merged[nota_new > nota_old]
        notas_alteradas = len(caiu) + len(subiu)
        
        print(f"   ⭐ Notas alteradas: {len(caiu)+len(subiu):,}")
        print(f"      - Caíram (Correção de Falso Positivo): {len(caiu):,}")
        print(f"      - Subiram: {len(subiu):,}")
        
        # Detalhar Transições (Ex: 4 -> 5)
        mudancas = merged[nota_new != nota_old].copy()
        if not mudancas.empty:
            mudancas['N_Old'] = nota_old[mudancas.index].astype(int)
            mudancas['N_New'] = nota_new[mudancas.index].astype(int)
            
            # Calcular Delta de Probabilidade (A prova real da mudança)
            mudancas['Prob_Old'] = pd.to_numeric(mudancas['Probabilidade_Conversao_old'], errors='coerce').fillna(0)
            mudancas['Prob_New'] = pd.to_numeric(mudancas['Probabilidade_Conversao_new'], errors='coerce').fillna(0)
            mudancas['Delta_Prob'] = mudancas['Prob_New'] - mudancas['Prob_Old']
            
            mudancas['Transicao'] = mudancas['N_Old'].astype(str) + ' -> ' + mudancas['N_New'].astype(str)
            
            counts = mudancas['Transicao'].value_counts()
            print(f"      📊 Detalhe das Mudanças (Top 5):")
            for tr, qtd in counts.head(5).items():
                print(f"         - {tr}: {qtd} leads")
            
            df_tr = counts.reset_index()
            df_tr.columns = ['Transicao', 'Qtd']
            df_tr['Segmento'] = seg
            transicoes_list.append(df_tr)
            
            # --- DIAGNÓSTICO INTELIGENTE (POR QUE MUDOU?) ---
            # Prepara dados numéricos para comparação
            mudancas['Ativ_Old'] = pd.to_numeric(mudancas['Número de atividades de vendas_old'], errors='coerce').fillna(0)
            mudancas['Ativ_New'] = pd.to_numeric(mudancas['Número de atividades de vendas_new'], errors='coerce').fillna(0)
            
            def diagnosticar_motivo(row):
                delta_nota = row['N_New'] - row['N_Old']
                delta_ativ = row['Ativ_New'] - row['Ativ_Old']
                delta_prob_abs = abs(row['Delta_Prob'])
                
                # Se a probabilidade mudou muito pouco (< 3%), é apenas oscilação de fronteira
                if delta_prob_abs < 0.03:
                    return "⚠️ Oscilação Marginal (Fronteira de Nota)"
                
                if delta_nota < 0: # Caiu
                    if delta_ativ == 0: return "📉 Esfriamento (Tempo sem atividade)"
                    return "📉 Saturação (Muitas tentativas s/ sucesso)"
                else: # Subiu
                    if delta_ativ > 0: return "🚀 Reaquecimento (Nova Interação)"
                    return "🚀 Ajuste de Dados (Correção/Perfil)"

            mudancas['Motivo_Diagnostico'] = mudancas.apply(diagnosticar_motivo, axis=1)
            
            print(f"      🧐 Diagnóstico das Mudanças:")
            print(mudancas['Motivo_Diagnostico'].value_counts().to_string())
            
            # Preparar para Exportação
            mudancas['Segmento'] = seg
            mudancas['Tipo_Alteracao'] = mudancas.apply(lambda x: 'Nota Subiu' if x['N_New'] > x['N_Old'] else 'Nota Caiu', axis=1)
            
            # Colunas de Prova (Evidence)
            mudancas['Nota_Antes'] = mudancas['N_Old']
            mudancas['Nota_Depois'] = mudancas['N_New']
            mudancas['Prob_Antes'] = mudancas['Prob_Old']
            mudancas['Prob_Depois'] = mudancas['Prob_New']
            mudancas['Var_Probabilidade'] = mudancas['Delta_Prob']
            
            audit_data.append(mudancas)

    resumo_stats.append({
        'Segmento': seg,
        'Total Analisado': len(merged),
        'Novos Leads': len(novos),
        'Datas Corrigidas': leads_corrigidos,
        'Notas Alteradas': notas_alteradas
    })
    
    print("-" * 40)

# ==============================================================================
# GERAÇÃO DO RELATÓRIO EXCEL (LAYOUT RESTAURADO)
# ==============================================================================
print("\n[GERANDO EXCEL] Consolidando auditoria...")

wb = Workbook()

# Estilos
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') # Azul Corporativo
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def formatar_aba(ws, df):
    # Escrever dados
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if r_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center')
    
    # Ajustar largura
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

# Aba 1: Resumo
ws1 = wb.active
ws1.title = "Resumo_Impacto"
if resumo_stats:
    formatar_aba(ws1, pd.DataFrame(resumo_stats))

# Aba 2: Detalhamento
ws2 = wb.create_sheet("Leads_Afetados")
if audit_data:
    df_audit = pd.concat(audit_data)
    cols_view = ['Record ID', 'Segmento', 'Tipo_Alteracao', 'Motivo_Diagnostico', 'Nota_Antes', 'Nota_Depois', 'Var_Probabilidade', 'Prob_Antes', 'Prob_Depois', 'Nome do negócio_new', 'Unidade Desejada_new']
    # Filtrar colunas que existem
    cols_final = [c for c in cols_view if c in df_audit.columns]
    formatar_aba(ws2, df_audit[cols_final])

# Aba 3: Matriz de Mudanças
ws3 = wb.create_sheet("Matriz_Mudancas")
if transicoes_list:
    df_transicoes = pd.concat(transicoes_list)
    df_transicoes = df_transicoes[['Segmento', 'Transicao', 'Qtd']].sort_values(['Segmento', 'Qtd'], ascending=[True, False])
    formatar_aba(ws3, df_transicoes)

# Aba 4: Novos Leads
ws4 = wb.create_sheet("Novos_Leads")
if novos_leads_list:
    df_novos = pd.concat(novos_leads_list)
    cols_view_novos = ['Record ID', 'Segmento', 'Nome do negócio', 'Unidade Desejada', 'Nota_1a5', 'Probabilidade_Conversao', 'Etapa do negócio']
    cols_final_novos = [c for c in cols_view_novos if c in df_novos.columns]
    formatar_aba(ws4, df_novos[cols_final_novos])

nome_saida = f"Auditoria_Impacto_{data_hoje}.xlsx"
caminho_saida = os.path.join(DIR_ATUAL, nome_saida)
wb.save(caminho_saida)

print(f"✅ Relatório de auditoria salvo em: {caminho_saida}")
print("="*80)
