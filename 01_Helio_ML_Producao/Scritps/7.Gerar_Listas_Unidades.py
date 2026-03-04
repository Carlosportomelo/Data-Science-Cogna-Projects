"""
==============================================================================
GERADOR DE LISTAS POR UNIDADE
==============================================================================
Script: 2.Gerar_Listas_Unidades.py
Objetivo: Ler os dados pontuados (Scored) e gerar listas de Excel individuais
          para cada unidade, separando em pastas de Próprias e Franquias.
Data: 2026-01-08
==============================================================================
"""

import pandas as pd
import os
import glob
import unicodedata
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# ==============================================================================
# CONFIGURAÇÃO DE CAMINHOS
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASTA_DADOS_SCORED = os.path.join(CAMINHO_BASE, 'Outputs', 'Dados_Scored')
PASTA_OUTPUT_ROOT = os.path.join(CAMINHO_BASE, 'Outputs', 'Lista para Unidades')

print("="*80)
print("GERADOR DE LISTAS PARA UNIDADES")
print("="*80)

# ==============================================================================
# 1. CARREGAR DADOS MAIS RECENTES
# ==============================================================================
# Busca o arquivo leads_scored_*.csv mais recente
list_of_files = glob.glob(os.path.join(PASTA_DADOS_SCORED, 'leads_scored_*.csv'))

if not list_of_files:
    print(f"[ERRO] Nenhum arquivo de leads encontrado em: {PASTA_DADOS_SCORED}")
    print("Execute o script '1.ML_Lead_Scoring.py' primeiro.")
    exit()

latest_file = max(list_of_files, key=os.path.getmtime)
print(f"[LENDO] Arquivo base: {os.path.basename(latest_file)}")

# Extrair data do arquivo para usar nas pastas (Data do Dado e não Data de Hoje)
nome_arquivo = os.path.basename(latest_file)
DATA_HOJE = nome_arquivo.replace('leads_scored_', '').replace('.csv', '')

# Configurar pastas de saída com a data do arquivo
PASTA_PROPRIAS = os.path.join(PASTA_OUTPUT_ROOT, 'Proprias', DATA_HOJE)
PASTA_FRANQUIAS = os.path.join(PASTA_OUTPUT_ROOT, 'Franquias', DATA_HOJE)

# Criar estrutura de pastas
os.makedirs(PASTA_PROPRIAS, exist_ok=True)
os.makedirs(PASTA_FRANQUIAS, exist_ok=True)

try:
    df = pd.read_csv(latest_file, sep=';', encoding='utf-8-sig')
except Exception as e:
    print(f"[ERRO] Falha ao ler CSV: {e}")
    exit()

# --- Garantir e parsear coluna de Data de criação do lead (para export)
if 'Data de criação' in df.columns:
    try:
        df['Data de criação'] = pd.to_datetime(df['Data de criação'], dayfirst=True, errors='coerce')
    except Exception:
        pass

# Garantir classificação correta de Próprias vs Franquias
UNIDADES_PROPRIAS = [
    'VILA LEOPOLDINA', 'SANTANA', 'PINHEIROS', 'PERDIZES', 
    'PACAEMBU', 'MORUMBI', 'JARDINS', 'ITAIM'
]

def classificar_tipo(u):
    u_str = str(u).upper()
    for p in UNIDADES_PROPRIAS:
        if p in u_str: return 'Própria'
    return 'Franqueada'

df['Tipo_Unidade'] = df['Unidade Desejada'].apply(classificar_tipo)

# ==============================================================================
# 2. TRATAMENTO E HUMANIZAÇÃO DOS DADOS
# ==============================================================================
# Mapeamento de nomes técnicos para nomes amigáveis (Comercial)
DE_PARA_COLUNAS = {
    'Record ID': 'ID do Lead',
    'Nome do negócio': 'Nome do Negócio',
    'Etapa do negócio': 'Etapa Atual',
    'Unidade Desejada': 'Unidade',
    'Tipo_Unidade': 'Tipo de Unidade',
    'Número de atividades de vendas': 'Total de Interações',
    'Data de criação': 'Data de Criação',
    'Dias_Desde_Criacao': 'Dias na Base',
    'Proprietário do negócio': 'Responsável',
    'Probabilidade_Conversao': 'Chance de Fechamento (%)',
    'Nota_1a5': 'Qualidade (1-5)',
    'Prioridade_Atendimento': 'Recomendação de Ação'
}

# Colunas técnicas originais para filtrar antes de renomear
COLS_EXPORT = [
    'Record ID', 'Nome do negócio', 'Etapa do negócio', 'Unidade Desejada', 
    'Tipo_Unidade', 'Número de atividades de vendas', 'Data de criação', 
    'Dias_Desde_Criacao', 'Proprietário do negócio', 'Probabilidade_Conversao', 
    'Nota_1a5', 'Prioridade_Atendimento'
]

# Função para limpar termos técnicos da coluna de Prioridade
def limpar_recomendacao(texto):
    if pd.isna(texto): return ""
    txt = str(texto)
    # Remove prefixos numéricos (ex: "1_") e sufixos técnicos
    txt = txt.split('_', 1)[-1] if '_' in txt else txt
    txt = txt.replace('_', ' ').replace('TOP', '').replace('ALTA', '').replace('MEDIA', '')
    # Substitui termos negativos
    if 'Nutricao' in txt or 'Baixa' in txt:
        return 'Nutrição / Longo Prazo'
    return txt.strip()

if 'Prioridade_Atendimento' in df.columns:
    df['Prioridade_Atendimento'] = df['Prioridade_Atendimento'].apply(limpar_recomendacao)

# Preparar DataFrame Final (Humanizado)
df_final = df.copy()

# Garantir que colunas existam antes de selecionar
cols_existentes = [c for c in COLS_EXPORT if c in df_final.columns]
df_final = df_final[cols_existentes].rename(columns=DE_PARA_COLUNAS)

# ==============================================================================
# 3. GERAR ARQUIVOS INDIVIDUAIS POR UNIDADE
# ==============================================================================
unidades = df_final['Unidade'].dropna().unique()
print(f"[PROCESSANDO] Gerando listas para {len(unidades)} unidades encontradas...\n")

# Debug para verificar Paulínia
lista_debug = sorted([str(u) for u in unidades])
print(f"   [DEBUG] Unidades encontradas: {lista_debug}")
if any('PAULINIA' in str(u).upper().replace('Í', 'I') for u in lista_debug):
    print("   [DEBUG] ✅ Unidade Paulínia detectada na base.")
else:
    print("   [DEBUG] ❌ Unidade Paulínia NÃO detectada na base.")

# Estilos para o Excel
HEADER_FILL = PatternFill(start_color='CE0E2D', end_color='CE0E2D', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

count_proprias = 0
count_franquias = 0

for unidade in unidades:
    if unidade == '(Sem unidade)':
        continue

    # Filtrar dados da unidade
    df_unidade = df_final[df_final['Unidade'] == unidade].copy()
    
    # Ordenar por prioridade (Nota maior primeiro)
    col_prob = 'Chance de Fechamento (%)'
    if col_prob in df_unidade.columns:
        df_unidade = df_unidade.sort_values(col_prob, ascending=False)

    # Identificar Tipo (Própria ou Franquia)
    col_tipo = 'Tipo de Unidade'
    tipo = df_unidade[col_tipo].iloc[0] if col_tipo in df_unidade.columns else 'Franqueada'
    is_propria = str(tipo).strip().lower() in ['própria', 'propria']
    
    pasta_destino = PASTA_PROPRIAS if is_propria else PASTA_FRANQUIAS
    
    # Nome do arquivo seguro (remove caracteres inválidos)
    nome_safe = str(unidade).strip()
    # Remover acentos para evitar erro no Windows
    try:
        nome_safe = unicodedata.normalize('NFKD', nome_safe).encode('ASCII', 'ignore').decode('ASCII')
    except:
        pass
    for char in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
        nome_safe = nome_safe.replace(char, '-')
        
    filename = f"{nome_safe}_{DATA_HOJE}.xlsx"
    filepath = os.path.join(pasta_destino, filename)
    
    # Gerar Excel com formatação
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads Prioritários"
    
    # Escrever dados
    for r_idx, row in enumerate(dataframe_to_rows(df_unidade, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if r_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            
            # Formatar porcentagem na coluna de Chance de Fechamento
            header_val = ws.cell(row=1, column=c_idx).value
            if r_idx > 1 and header_val:
                if 'chance' in str(header_val).lower():
                    cell.number_format = '0%'
                # Formatar Data de criação
                if 'data' in str(header_val).lower() and 'cria' in str(header_val).lower():
                    try:
                        cell.number_format = 'DD/MM/YYYY'
                    except Exception:
                        pass
    
    # Ajustar largura das colunas
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

    wb.save(filepath)
    print(f"   -> Gerado ({'Própria' if is_propria else 'Franquia'}): {filename}")
    
    if is_propria:
        count_proprias += 1
    else:
        count_franquias += 1

print(f"[CONCLUÍDO] Arquivos gerados com sucesso!")
print(f"   → Próprias:  {count_proprias} arquivos em {PASTA_PROPRIAS}")
print(f"   → Franquias: {count_franquias} arquivos em {PASTA_FRANQUIAS}")

# ==============================================================================
# 4. GERAR RELATÓRIO MACRO GERENCIAL (VISÃO DO GESTOR)
# ==============================================================================
print("\n[GERENCIAL] Gerando Relatório Macro de Indicadores Comerciais...")

# Preparar dados agregados
# Usamos o df original (com colunas técnicas) para facilitar cálculos numéricos
df_macro = df.copy()

# Métricas por Unidade
indicadores = df_macro.groupby(['Unidade Desejada', 'Tipo_Unidade']).agg(
    Total_Leads=('Record ID', 'count'),
    Leads_Quentes=('Nota_1a5', lambda x: (x >= 4).sum()), # Nota 4 ou 5
    Media_Atividades=('Número de atividades de vendas', 'mean'),
    Leads_Trabalhados=('Número de atividades de vendas', lambda x: (x > 0).sum()),
    Dias_Medios_Base=('Dias_Desde_Criacao', 'mean'),
    Nota_Media_Qualidade=('Nota_1a5', 'mean')
).reset_index()

# Cálculos derivados
indicadores['% Leads Trabalhados'] = indicadores['Leads_Trabalhados'] / indicadores['Total_Leads']
indicadores['% Leads Quentes'] = indicadores['Leads_Quentes'] / indicadores['Total_Leads']

# Renomear para apresentação
cols_macro = {
    'Unidade Desejada': 'Unidade',
    'Tipo_Unidade': 'Tipo',
    'Total_Leads': 'Volume Total',
    'Leads_Quentes': 'Pipeline Quente (Nota 4-5)',
    '% Leads Quentes': '% Qualidade',
    'Media_Atividades': 'Média de Interações/Lead',
    '% Leads Trabalhados': 'Taxa de Cobertura (%)',
    'Dias_Medios_Base': 'Idade Média (Dias)',
    'Nota_Media_Qualidade': 'Score Médio (1-5)'
}

indicadores_final = indicadores.rename(columns=cols_macro)
# Reordenar colunas logicamente
ordem_cols = ['Unidade', 'Tipo', 'Volume Total', 'Pipeline Quente (Nota 4-5)', '% Qualidade', 
              'Taxa de Cobertura (%)', 'Média de Interações/Lead', 'Score Médio (1-5)', 'Idade Média (Dias)']
indicadores_final = indicadores_final[ordem_cols].sort_values(['Tipo', 'Volume Total'], ascending=[False, False])

# Gerar Excel Gerencial
wb_macro = Workbook()
ws_macro = wb_macro.active
ws_macro.title = "Visão Geral Unidades"

# Título
ws_macro['A1'] = "RELATÓRIO MACRO - PERFORMANCE COMERCIAL POR UNIDADE"
ws_macro['A1'].font = Font(bold=True, size=14)
ws_macro['A2'] = f"Data de Referência: {datetime.now().strftime('%d/%m/%Y')}"

# Escrever Tabela
start_row = 4
for r_idx, row in enumerate(dataframe_to_rows(indicadores_final, index=False, header=True), start_row):
    for c_idx, value in enumerate(row, 1):
        cell = ws_macro.cell(row=r_idx, column=c_idx, value=value)
        cell.border = THIN_BORDER
        
        # Cabeçalho
        if r_idx == start_row:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Formatação Condicional Básica (Cores alternadas ou negrito por tipo)
        if r_idx > start_row:
            # Formatar porcentagens
            if '%' in indicadores_final.columns[c_idx-1] or 'Taxa' in indicadores_final.columns[c_idx-1]:
                cell.number_format = '0.0%'
            # Formatar números decimais
            elif 'Média' in indicadores_final.columns[c_idx-1] or 'Score' in indicadores_final.columns[c_idx-1]:
                cell.number_format = '0.0'

# Ajustar larguras
for column_cells in ws_macro.columns:
    length = max(len(str(cell.value) or "") for cell in column_cells)
    ws_macro.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 25)

# ==============================================================================
# 5. NOVA ABA: QUALIDADE DO ATENDIMENTO (FREQUÊNCIA E SCORE)
# ==============================================================================
ws_qualidade = wb_macro.create_sheet("Qualidade Atendimento")
ws_qualidade['A1'] = "QUALIDADE DO ACOMPANHAMENTO (FREQUÊNCIA E SCORE)"
ws_qualidade['A1'].font = Font(bold=True, size=12)
ws_qualidade['A2'] = "Frequência: Média de dias entre contatos (Dias na Base / Nº Atividades). Quanto menor, mais intenso o acompanhamento."

# 1. Calcular Frequência (Apenas leads com atividade > 0 para evitar divisão por zero)
df_trab = df_macro[df_macro['Número de atividades de vendas'] > 0].copy()
df_trab['Frequencia_Dias'] = df_trab['Dias_Desde_Criacao'] / df_trab['Número de atividades de vendas']

# 2. Agrupar métricas de qualidade
metrics_qual = df_trab.groupby(['Unidade Desejada', 'Tipo_Unidade']).agg(
    Frequencia_Media=('Frequencia_Dias', 'mean'),
    Score_Medio=('Nota_1a5', 'mean')
).reset_index()

# 3. Calcular Conversão Real (Matrículas / Total) - Base Completa
df_macro['Is_Matricula'] = df_macro['Etapa do negócio'].astype(str).str.contains('Matrícula|Matricula', case=False, na=False).astype(int)
metrics_conv = df_macro.groupby('Unidade Desejada')['Is_Matricula'].mean().reset_index()

# 4. Consolidar e formatar
df_qualidade_final = pd.merge(metrics_qual, metrics_conv, on='Unidade Desejada', how='left')
df_qualidade_final.columns = ['Unidade', 'Tipo', 'Frequência (Dias entre contatos)', 'Score Médio dos Leads', 'Taxa Conversão Real']
df_qualidade_final = df_qualidade_final.sort_values('Frequência (Dias entre contatos)', ascending=True)

# 5. Escrever na aba
start_row_q = 4
for r_idx, row in enumerate(dataframe_to_rows(df_qualidade_final, index=False, header=True), start_row_q):
    for c_idx, value in enumerate(row, 1):
        cell = ws_qualidade.cell(row=r_idx, column=c_idx, value=value)
        cell.border = THIN_BORDER
        if r_idx == start_row_q:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            if c_idx == 3: cell.number_format = '0.0'  # Frequência
            if c_idx == 4: cell.number_format = '0.0'  # Score
            if c_idx == 5: cell.number_format = '0.0%' # Conversão

# Ajustar larguras
for column_cells in ws_qualidade.columns:
    length = max(len(str(cell.value) or "") for cell in column_cells)
    ws_qualidade.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 30)

# Salvar
nome_macro = f"Relatorio_Gerencial_Macro_{DATA_HOJE}.xlsx"
path_macro = os.path.join(PASTA_OUTPUT_ROOT, nome_macro)

try:
    wb_macro.save(path_macro)
    print(f"\n[SUCESSO] Relatório Gerencial salvo em: {path_macro}")
except Exception as e:
    print(f"\n[ERRO] Não foi possível salvar o relatório gerencial: {e}")