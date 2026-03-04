"""
==============================================================================
TESTE CEGO DE ACURÁCIA - VERSÃO AJUSTADA (CORRIGIDA E SEM VAZAMENTO)
==============================================================================
Script: 13_Blind_Test_Ajustado.py
Objetivo: Medir a acurácia REAL do modelo nos leads que ele PODE prever
          (excluindo vendas de balcão sem registro e CORRIGINDO O VAZAMENTO DE DATA)
Data: 2026-01-20
==============================================================================

JUSTIFICATIVA DO AJUSTE:
1. HIGIENE DE DADOS: 
   - Exclui vendas de balcão (<3 atividades) pois são imprevisíveis (ruído).
   - Foca nos leads rastreáveis (>=3 atividades) onde o modelo pode atuar.

2. CORREÇÃO DE VAZAMENTO (CRÍTICO):
   - A idade do lead agora é calculada em relação a HOJE (e não à data de fechamento).
   - Isso impede que o teste "roube" usando a informação de que o lead fechou rápido.
==============================================================================
"""

import pandas as pd
import numpy as np
import os
import pickle
from datetime import datetime
from sklearn.metrics import roc_auc_score, confusion_matrix, classification_report
from sklearn.preprocessing import LabelEncoder
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CAMINHO_LEADS = os.path.join(CAMINHO_BASE, 'Data', 'hubspot_leads.csv')
PASTA_OUTPUT = os.path.join(CAMINHO_BASE, 'Outputs')
PASTA_MODELS = os.path.join(CAMINHO_BASE, 'models')
MODELO_PATH = os.path.join(PASTA_MODELS, 'lead_scoring_model.pkl')
os.makedirs(PASTA_OUTPUT, exist_ok=True)
PASTA_BLIND_TEST = os.path.join(PASTA_OUTPUT, 'Relatorios_Blind_Test')
os.makedirs(PASTA_BLIND_TEST, exist_ok=True)

DATA_HOJE = datetime.now().strftime('%Y-%m-%d')
CAMINHO_SAIDA = os.path.join(PASTA_BLIND_TEST, f'Blind_Test_Realista_{DATA_HOJE}.xlsx')

# Configuração do corte de atividades
MIN_ATIVIDADES = 3  # Mínimo para considerar rastreável

# Estilos Excel
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
SUCCESS_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
WARNING_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
DANGER_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

def df_to_sheet(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if r_idx == start_row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center', wrap_text=True)

def ajustar_largura(ws):
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

print("="*80)
print("TESTE CEGO REALISTA - SEM VAZAMENTO DE DATA")
print("="*80)
print(f"📂 DIRETÓRIO DE EXECUÇÃO: {CAMINHO_BASE}")
print(f"📂 DADOS DE ENTRADA:      {CAMINHO_LEADS}")
print("-" * 80)

# ==============================================================================
# 1. CARREGAMENTO DOS DADOS
# ==============================================================================
print("\n[1/6] Carregando dados...")

if not os.path.exists(CAMINHO_LEADS):
    print(f"\n❌ [ERRO CRÍTICO] Arquivo de dados não encontrado: {CAMINHO_LEADS}")
    exit()

try:
    df = pd.read_csv(CAMINHO_LEADS, encoding='utf-8')
except:
    df = pd.read_csv(CAMINHO_LEADS, encoding='latin1')

df.columns = df.columns.str.strip()
print(f"  → Base carregada: {len(df):,} registros")

# ==============================================================================
# 2. PREPARAÇÃO DE FEATURES (CORRIGIDA)
# ==============================================================================
print("\n[2/6] Preparando features e CORRIGINDO VAZAMENTO...")

df['Data de criação'] = pd.to_datetime(df['Data de criação'], dayfirst=True, errors='coerce')
df['Data de fechamento'] = pd.to_datetime(df['Data de fechamento'], dayfirst=True, errors='coerce')

# --- CORREÇÃO DE VAZAMENTO DE DADOS (CRUCIAL) ---
# ANTES (ERRADO): Usava Data de Fechamento (O modelo adivinhava quem ganhava pelo tempo curto)
# data_referencia = df['Data de fechamento'].fillna(datetime.now())
# df['Dias_Desde_Criacao'] = (data_referencia - df['Data de criação']).dt.days

# AGORA (CORRETO): Usa SEMPRE a data de HOJE.
# O modelo deve julgar o lead como ele se apresenta HOJE no sistema, sem saber o futuro.
print("  >>> Aplicando cálculo de dias baseado em HOJE (Simulação Real)...")
df['Dias_Desde_Criacao'] = (pd.Timestamp.now() - df['Data de criação']).dt.days
df['Dias_Desde_Criacao'] = df['Dias_Desde_Criacao'].fillna(0).clip(lower=0)
# ------------------------------------------------

df['Ano'] = df['Data de criação'].dt.year
df['Mes'] = df['Data de criação'].dt.month
df['DiaSemana'] = df['Data de criação'].dt.dayofweek
df['Trimestre'] = df['Data de criação'].dt.quarter
df['Densidade_Atividade'] = df['Número de atividades de vendas'] / (df['Dias_Desde_Criacao'] + 1)

df['Fonte original do tráfego'] = df['Fonte original do tráfego'].fillna('(Sem fonte)')
df['Detalhamento da fonte original do tráfego 1'] = df['Detalhamento da fonte original do tráfego 1'].fillna('(Sem detalhe)')
df['Número de atividades de vendas'] = pd.to_numeric(df['Número de atividades de vendas'], errors='coerce').fillna(0)
df['Unidade Desejada'] = df['Unidade Desejada'].fillna('(Sem unidade)')

for col in ['Ano', 'Mes', 'DiaSemana', 'Trimestre', 'Dias_Desde_Criacao']:
    df[col] = df[col].fillna(0)

# Classificação de Atividade
def classificar_atividade_comercial(n):
    if n < 3: return '0_Outlier_Sem_Registro'
    elif n <= 6: return '1_Inicial_3_a_6'
    elif n <= 12: return '2_Engajamento_7_a_12'
    elif n <= 20: return '3_Persistencia_13_a_20'
    elif n <= 35: return '4_Alta_Persistencia_21_a_35'
    elif n <= 60: return '5_Recuperacao_Longa_36_a_60'
    else: return '6_Saturacao_Extrema_60+'

df['Faixa_Atividade'] = df['Número de atividades de vendas'].apply(classificar_atividade_comercial)

# Categorização de fonte
def categorizar_fonte(fonte):
    fonte = str(fonte).lower()
    if 'off-line' in fonte or 'offline' in fonte: return 'Offline'
    elif 'social pago' in fonte: return 'Meta_Ads'
    elif 'pesquisa paga' in fonte: return 'Google_Ads'
    elif 'pesquisa orgânica' in fonte: return 'Google_Organico'
    elif 'social orgânico' in fonte: return 'Social_Organico'
    elif 'tráfego direto' in fonte: return 'Direto'
    else: return 'Outros'

df['Fonte_Cat'] = df['Fonte original do tráfego'].apply(categorizar_fonte)

# Segmentação de ML
def definir_segmento_ml(fonte_cat):
    if fonte_cat == 'Offline': return 'Offline'
    elif fonte_cat in ['Google_Ads', 'Google_Organico']: return 'Digital_Busca'
    elif fonte_cat in ['Meta_Ads', 'Social_Organico']: return 'Digital_Social'
    else: return 'Digital_Direto'

df['Segmento_ML'] = df['Fonte_Cat'].apply(definir_segmento_ml)

def categorizar_detalhe_offline(row):
    if row['Fonte_Cat'] != 'Offline': return 'N/A'
    detalhe = str(row['Detalhamento da fonte original do tráfego 1']).upper()
    if 'CRM_UI' in detalhe: return 'CRM_UI'
    elif 'CONTACTS' in detalhe: return 'CONTACTS'
    elif 'IMPORT' in detalhe: return 'IMPORT'
    else: return 'Outros_Offline'

df['Detalhe_Offline'] = df.apply(categorizar_detalhe_offline, axis=1)

# Target
df['Sucesso'] = df['Etapa do negócio'].str.contains('MATRÍCULA|GANHO|WON', case=False, na=False).astype(int)
df['Perdido'] = df['Etapa do negócio'].str.contains('PERDIDO|LOST', case=False, na=False).astype(int)
df['Finalizado'] = df['Sucesso'] | df['Perdido']

# ==============================================================================
# 3. SEGMENTAÇÃO DA BASE
# ==============================================================================
print("\n[3/6] Segmentando a base...")

# Base completa de finalizados
df_final = df[df['Finalizado'] == 1].copy()

# Base RASTREÁVEL (>=3 atividades)
df_rastreavel = df_final[df_final['Número de atividades de vendas'] >= MIN_ATIVIDADES].copy()

# Base NÃO RASTREÁVEL (<3 atividades)
df_balcao = df_final[df_final['Número de atividades de vendas'] < MIN_ATIVIDADES].copy()

print(f"\n  POPULAÇÃO TOTAL FINALIZADA: {len(df_final):,}")
print(f"  ├── Rastreável (>={MIN_ATIVIDADES} ativ): {len(df_rastreavel):,} ({len(df_rastreavel)/len(df_final)*100:.1f}%)")
print(f"  │   ├── Matrículas: {df_rastreavel['Sucesso'].sum():,}")
print(f"  │   └── Perdidos: {df_rastreavel['Perdido'].sum():,}")
print(f"  └── Balcão (<{MIN_ATIVIDADES} ativ): {len(df_balcao):,} ({len(df_balcao)/len(df_final)*100:.1f}%)")

# ==============================================================================
# 4. CARREGAR MODELO E APLICAR SCORING
# ==============================================================================
print("\n[4/6] Carregando modelo e aplicando scoring...")

if not os.path.exists(MODELO_PATH):
    print(f"[ERRO] Modelo não encontrado: {MODELO_PATH}")
    exit()

with open(MODELO_PATH, 'rb') as f:
    modelo_data = pickle.load(f)

# Detectar formato do modelo (Segmentado vs Único)
is_segmented = False
if isinstance(modelo_data, dict) and any(k.startswith('model_') for k in modelo_data.keys()):
    print("  → Modelo segmentado detectado (Correto)")
    is_segmented = True
    encoders = modelo_data.get('encoders', {})
else:
    print("[ERRO] Formato do modelo inesperado (Esperado: dicionário segmentado)")
    exit()

# Aplicar encoders
FEATURES_CAT = ['Fonte_Cat', 'Detalhe_Offline', 'Unidade Desejada', 'Faixa_Atividade']
for col in FEATURES_CAT:
    if col in encoders:
        le = encoders[col]
        # Aplica transformação segura (trata valores novos como desconhecidos/primeira classe)
        df_final[f'{col}_enc'] = df_final[col].astype(str).apply(
            lambda x: le.transform([x])[0] if x in le.classes_ else -1
        )
        # Se -1 (desconhecido), substitui pelo primeiro valor válido ou moda (0)
        df_final[f'{col}_enc'] = df_final[f'{col}_enc'].replace(-1, 0)

# Aplicar Predição
proba = np.zeros(len(df_final))

if is_segmented:
    segmentos_disponiveis = [k.replace('model_', '') for k in modelo_data.keys() if k.startswith('model_')]
    
    for seg in segmentos_disponiveis:
        model_seg = modelo_data[f'model_{seg}']
        features_seg = modelo_data[f'features_{seg}']
        
        mask = df_final['Segmento_ML'] == seg
        if mask.sum() > 0:
            X_seg = df_final.loc[mask, features_seg].values
            proba[mask] = model_seg.predict_proba(X_seg)[:, 1]

df_final['Probabilidade'] = proba

# Criar nota 1-5
bins = [0, 0.2, 0.4, 0.6, 0.8, 1.0]
df_final['Nota'] = pd.cut(df_final['Probabilidade'], bins=bins, labels=[1, 2, 3, 4, 5], include_lowest=True).astype(int)

# Recriar os subsets agora com as notas calculadas
mask_balcao = (df_final['Número de atividades de vendas'] < MIN_ATIVIDADES)
df_rastreavel = df_final[~mask_balcao].copy()
df_balcao = df_final[mask_balcao].copy()

print(f"  → Scoring aplicado. {len(df_rastreavel):,} leads rastreáveis avaliados.")

# ==============================================================================
# 5. CALCULAR MÉTRICAS
# ==============================================================================
print("\n[5/6] Calculando métricas de acurácia...")

# Resumo por nota (AJUSTADO - Rastreável)
resumo_nota_adj = df_rastreavel.groupby('Nota').agg(
    Total=('Nota', 'count'),
    Matriculas=('Sucesso', 'sum'),
    Perdidos=('Perdido', 'sum')
).reset_index()
resumo_nota_adj['Taxa_Conversao'] = resumo_nota_adj['Matriculas'] / resumo_nota_adj['Total']

# Resumo por nota (GLOBAL - Tudo)
resumo_nota_global = df_final.groupby('Nota').agg(
    Total=('Nota', 'count'),
    Matriculas=('Sucesso', 'sum'),
    Perdido=('Perdido', 'sum')
).reset_index()
resumo_nota_global['Taxa_Conversao'] = resumo_nota_global['Matriculas'] / resumo_nota_global['Total']

print("\n  ACURÁCIA POR NOTA (BASE RASTREÁVEL - REALISTA):")
print(resumo_nota_adj.to_string(index=False))

# Métricas avançadas
df_rastreavel['Pred_Positiva'] = (df_rastreavel['Nota'] >= 4).astype(int)
y_true = df_rastreavel['Sucesso'].values
y_pred = df_rastreavel['Pred_Positiva'].values

matriz = confusion_matrix(y_true, y_pred, labels=[0, 1])
tn, fp, fn, tp = matriz.ravel()

precisao = tp / (tp + fp) if (tp + fp) > 0 else 0
recall = tp / (tp + fn) if (tp + fn) > 0 else 0
auc = roc_auc_score(y_true, df_rastreavel['Probabilidade'].values)

try:
    conv_nota5 = resumo_nota_adj[resumo_nota_adj['Nota'] == 5]['Taxa_Conversao'].values[0]
    conv_nota1 = resumo_nota_adj[resumo_nota_adj['Nota'] == 1]['Taxa_Conversao'].values[0]
    lift = conv_nota5 / conv_nota1 if conv_nota1 > 0 else 0
except:
    lift = 0

print(f"\n  MÉTRICAS REAIS (SEM COLAR):")
print(f"  ├── AUC-ROC: {auc:.4f}")
print(f"  ├── Precisão (Notas 4-5): {precisao:.1%}")
print(f"  ├── Recall (Notas 4-5): {recall:.1%}")
print(f"  └── LIFT (Nota5 / Nota1): {lift:.1f}x")

# --- SALVAR HISTÓRICO DE EVOLUÇÃO (NOVO) ---
CAMINHO_LOG_HISTORICO = os.path.join(PASTA_BLIND_TEST, 'Historico_Metricas_Blind_Test.csv')

dados_log = {
    'Data_Execucao': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    'AUC_ROC': round(auc, 4),
    'Precisao_Top_Leads': round(precisao, 4),
    'Recall_Top_Leads': round(recall, 4),
    'Lift_5vs1': round(lift, 2),
    'Total_Leads_Rastreaveis': len(df_rastreavel),
    'Matriculas_Identificadas': df_rastreavel['Sucesso'].sum()
}

df_log_novo = pd.DataFrame([dados_log])

if os.path.exists(CAMINHO_LOG_HISTORICO):
    df_log_antigo = pd.read_csv(CAMINHO_LOG_HISTORICO, sep=';')
    df_log_final = pd.concat([df_log_antigo, df_log_novo], ignore_index=True)
else:
    df_log_final = df_log_novo

df_log_final.to_csv(CAMINHO_LOG_HISTORICO, sep=';', index=False, encoding='utf-8-sig')
print(f"\n[HISTÓRICO] Métricas de evolução salvas em: {CAMINHO_LOG_HISTORICO}")

# ==============================================================================
# 6. EXPORTAR RESULTADOS
# ==============================================================================
print("\n[6/6] Exportando resultados...")

wb = Workbook()

# --- ABA 1: RESUMO ---
ws1 = wb.active
ws1.title = '1_Resumo_Executivo'

ws1['A1'] = 'TESTE CEGO DE ACURÁCIA (REALISTA - SEM VAZAMENTO)'
ws1['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws1['A3'] = f'Data: {DATA_HOJE}'

ws1['A5'] = 'MÉTRICAS CHAVE:'
ws1['A5'].font = Font(bold=True)
metricas = [
    ['AUC-ROC', auc],
    ['Precisão', precisao],
    ['Recall', recall],
    ['LIFT (5/1)', lift]
]
for i, (nome, val) in enumerate(metricas, 6):
    ws1.cell(i, 1, nome)
    cell = ws1.cell(i, 2, val)
    if nome != 'LIFT (5/1)': cell.number_format = '0.0%'
    else: cell.number_format = '0.0'

# --- ABA 2: POR NOTA ---
ws2 = wb.create_sheet('2_Acuracia_Por_Nota')
ws2['A1'] = 'ACURÁCIA POR NOTA (BASE RASTREÁVEL)'
ws2['A1'].font = Font(bold=True)

df_to_sheet(ws2, resumo_nota_adj, start_row=3)
for r in range(4, 4+len(resumo_nota_adj)):
    ws2.cell(r, 4).number_format = '0.0%' # Taxa Conv

# --- ABA 3: POR SEGMENTO ---
ws3 = wb.create_sheet('3_Por_Segmento')
ws3['A1'] = 'RECALL POR SEGMENTO'
ws3['A1'].font = Font(bold=True)

seg_stats = []
for seg in sorted(df_rastreavel['Segmento_ML'].unique()):
    sub = df_rastreavel[df_rastreavel['Segmento_ML'] == seg]
    total_mat = sub['Sucesso'].sum()
    if total_mat > 0:
        capt = sub[(sub['Sucesso']==1) & (sub['Nota']>=4)].shape[0]
        seg_stats.append({'Segmento': seg, 'Matrículas': total_mat, 'Capturadas (Nota 4-5)': capt, 'Recall': capt/total_mat})

df_seg_stats = pd.DataFrame(seg_stats)
df_to_sheet(ws3, df_seg_stats, start_row=3)
for r in range(4, 4+len(df_seg_stats)):
    ws3.cell(r, 4).number_format = '0.0%'

wb.save(CAMINHO_SAIDA)
print(f"\n✅ Arquivo gerado: {CAMINHO_SAIDA}")
print("="*80)