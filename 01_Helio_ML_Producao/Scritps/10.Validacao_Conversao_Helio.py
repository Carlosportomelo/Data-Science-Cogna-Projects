"""
==============================================================================
DASHBOARD DE ACOMPANHAMENTO - HELIO (EVOLUÇÃO PRÉ vs PÓS-LANÇAMENTO)
==============================================================================
Script: 10.Validacao_Conversao_Helio.py
Objetivo: Gerar Relatório de Evolução Temporal do Helio (Validação)

LINHA DO TEMPO:
- Pré-Lançamento (12/12/2025 e 05/01/2026): Baseline Orgânico (SEM intervenção)
- Pós-Lançamento (20/01/2026+): Período de Atuação Ativa do Helio

MÉTRICAS PRINCIPAIS:
1. Qualidade do Modelo (taxa conversão pós-lançamento)
2. Aderência Operacional (tempo de contato, negligência)
3. Feedback Loop (unidades validando qualidade dos leads)
4. Comparação Pré vs Pós Helio

Data: 2026-01-21
==============================================================================
"""

import pandas as pd
import os
import glob
import re
import unicodedata
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASTA_OUTPUT = os.path.join(CAMINHO_BASE, 'Outputs')
PASTA_RELATORIOS_ML = os.path.join(PASTA_OUTPUT, 'Relatorios_ML')
CAMINHO_LEADS_ATUAL = os.path.join(CAMINHO_BASE, 'Data', 'hubspot_leads.csv')
# ATUALIZAÇÃO: Usar base limpa de matrículas finais (Nome do Aluno = Nome do negócio)
CAMINHO_MATRICULAS = os.path.join(CAMINHO_BASE, 'Data', 'matriculas_finais_limpo.csv')
PASTA_VALIDACAO = os.path.join(PASTA_OUTPUT, 'DB_HELIO')

os.makedirs(PASTA_VALIDACAO, exist_ok=True)

DATA_HOJE = datetime.now().strftime('%Y-%m-%d')
CAMINHO_SAIDA = os.path.join(PASTA_VALIDACAO, f'Evolucao_Temporal_Helio_{DATA_HOJE}.xlsx')

# DATA DE LANÇAMENTO DO HELIO (marco divisor)
DATA_LANCAMENTO_HELIO = datetime(2026, 1, 20)

# Definição dos 3 Marcos
MARCOS = {
    'PreLancamento_Marco1': {
        'data': datetime(2025, 12, 12),
        'pasta': os.path.join(PASTA_RELATORIOS_ML, 'Relatório 12-12-25'),
        'descricao': '12/12/2025 - Baseline Organico 1',
        'periodo': 'Pre-Lancamento',
        'dias_apos_lancamento': -39  # 39 dias antes do lançamento
    },
    'PreLancamento_Marco2': {
        'data': datetime(2026, 1, 5),
        'pasta': os.path.join(PASTA_RELATORIOS_ML, 'Relatório 01-05-26'),
        'descricao': '05/01/2026 - Baseline Organico 2',
        'periodo': 'Pre-Lancamento',
        'dias_apos_lancamento': -15  # 15 dias antes do lançamento
    },
    'PosLancamento_Marco3': {
        'data': datetime(2026, 1, 20),
        'pasta': os.path.join(PASTA_RELATORIOS_ML, 'start'),
        'descricao': '20/01/2026 - LANCAMENTO DO HELIO',
        'periodo': 'Pos-Lancamento',
        'dias_apos_lancamento': 0  # Dia zero
    }
}

# Estilos Excel
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
FILL_PRE = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')  # Laranja claro
FILL_POS = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')  # Verde claro

# Regex para caracteres ilegais no Excel (XML)
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def limpar_dados_excel(df):
    """Remove caracteres ilegais que corrompem o Excel e trata NaNs"""
    if df.empty: return df
    df_clean = df.copy()

    # Preencher vazios
    df_clean = df_clean.fillna('')

    # --- AJUSTE DE TIPOS (SOLICITADO) ---
    # 1. Forçar colunas numéricas (preencher NaN com 0 para manter tipo numérico no Excel)
    for col in ['Numero_Atividades_Atual', 'Dias_Ociosidade']:
        if col in df_clean.columns:
            df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce').fillna(0)
            
    # 2. Forçar coluna de Nota como Texto
    if 'Nota_Original_1a5' in df_clean.columns:
        s_nota = pd.to_numeric(df_clean['Nota_Original_1a5'], errors='coerce').fillna(0)
        df_clean['Nota_Original_1a5'] = s_nota.astype(int).astype(str).replace('0', '')

    # Limpar strings
    for col in df_clean.select_dtypes(include=['object', 'string']):
        df_clean[col] = df_clean[col].astype(str).apply(
            lambda x: ILLEGAL_CHARACTERS_RE.sub("", x) if isinstance(x, str) else x
        )
    return df_clean

def df_to_sheet(ws, df, start_row=1, freeze_panes=False):
    """Adiciona DataFrame ao sheet com formatação"""
    df_safe = limpar_dados_excel(df)
    for r_idx, row in enumerate(dataframe_to_rows(df_safe, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if r_idx == start_row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center', wrap_text=True)

    if freeze_panes:
        ws.freeze_panes = ws.cell(row=start_row+1, column=1)

def ajustar_largura(ws):
    """Auto-ajusta largura das colunas"""
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

def normalizar_record_id(series):
    """
    Normaliza Record ID removendo problemas comuns:
    - Converte para string
    - Remove .0 de floats (53503877648.0 -> 53503877648)
    - Remove espacos
    - Remove aspas extras
    - Garante formato consistente
    """
    return (series
            .astype(str)
            .str.replace(r'\.0+$', '', regex=True)  # Remove .0 .00 etc
            .str.strip()
            .str.replace(r'["\'\s]', '', regex=True)  # Remove aspas e espacos
            .str.upper())  # Normaliza case

def carregar_marco(info_marco):
    """Carrega dados de um marco específico consolidando todos os arquivos Excel"""
    print(f"\n  >> Carregando: {info_marco['descricao']}")

    if not os.path.exists(info_marco['pasta']):
        print(f"     [!] Pasta nao encontrada: {info_marco['pasta']}")
        return pd.DataFrame()

    arquivos = glob.glob(os.path.join(info_marco['pasta'], 'Lead_Scoring_*.xlsx'))

    if not arquivos:
        print(f"     [!] Nenhum arquivo encontrado")
        return pd.DataFrame()

    dfs_leads = []

    for arquivo in arquivos:
        print(f"     - Lendo: {os.path.basename(arquivo)}")
        try:
            xl = pd.ExcelFile(arquivo)

            # Abas que contêm dados de leads com notas
            abas_relevantes = ['2_Top500_Nota5', '3_Nota4', '4_Top10_Por_Unidade', '5_Em_Qualificacao']

            for aba in abas_relevantes:
                if aba in xl.sheet_names:
                    try:
                        # Busca dinâmica do cabeçalho
                        df_preview = pd.read_excel(xl, sheet_name=aba, header=None, nrows=20)
                        header_idx = -1

                        for idx, row in df_preview.iterrows():
                            if row.astype(str).str.contains('Record ID', case=False, na=False).any():
                                header_idx = idx
                                break

                        if header_idx != -1:
                            df_temp = pd.read_excel(xl, sheet_name=aba, header=header_idx)

                            # Limpar nomes das colunas (remover espaços)
                            df_temp.columns = df_temp.columns.astype(str).str.strip()

                            # Normalizar Record ID imediatamente apos leitura
                            if 'Record ID' in df_temp.columns:
                                df_temp['Record ID'] = normalizar_record_id(df_temp['Record ID'])
                                dfs_leads.append(df_temp)
                                print(f"        [OK] {aba}: {len(df_temp)} leads")
                    except Exception as e:
                        print(f"        [!] Erro ao ler aba {aba}: {e}")

        except Exception as e:
            print(f"     [!] Erro ao ler arquivo {os.path.basename(arquivo)}: {e}")

    if dfs_leads:
        df_consolidado = pd.concat(dfs_leads, ignore_index=True)

        # Normalizar Record ID ANTES de remover duplicatas
        df_consolidado['Record ID'] = normalizar_record_id(df_consolidado['Record ID'])

        if 'Nota_1a5' in df_consolidado.columns:
            df_consolidado = df_consolidado.sort_values('Nota_1a5', ascending=False)

        df_consolidado = df_consolidado.drop_duplicates(subset=['Record ID'], keep='first')

        print(f"     [OK] Total consolidado: {len(df_consolidado)} leads unicos")
        return df_consolidado

    return pd.DataFrame()

print("="*80)
print("RELATÓRIO DE EVOLUÇÃO TEMPORAL - HELIO (VALIDAÇÃO)")
print("="*80)
print(f"Data de Lancamento do Helio: {DATA_LANCAMENTO_HELIO.strftime('%d/%m/%Y')}")
print(f"Dias desde o lancamento: {(datetime.now() - DATA_LANCAMENTO_HELIO).days}")

# ==============================================================================
# 1. CARREGAR OS 3 MARCOS HISTÓRICOS
# ==============================================================================
print("\n[1/5] Carregando os 3 marcos historicos...")

marcos_dados = {}

for marco_nome, marco_info in MARCOS.items():
    df_marco = carregar_marco(marco_info)
    if not df_marco.empty:
        df_marco['Marco_ID'] = marco_nome
        df_marco['Data_Marco'] = marco_info['data']
        df_marco['Descricao_Marco'] = marco_info['descricao']
        df_marco['Periodo'] = marco_info['periodo']
        df_marco['Dias_Apos_Lancamento_Marco'] = marco_info['dias_apos_lancamento']
        marcos_dados[marco_nome] = df_marco
    else:
        print(f"  [!] Marco {marco_nome} sem dados")

if not marcos_dados:
    print("\n[ERRO] Nenhum marco historico foi carregado. Verifique as pastas.")
    exit()

# ==============================================================================
# 2. CARREGAR SITUAÇÃO ATUAL (HUBSPOT)
# ==============================================================================
print("\n[2/5] Carregando situacao atual do HubSpot...")

if not os.path.exists(CAMINHO_LEADS_ATUAL):
    print(f"[ERRO] Arquivo nao encontrado: {CAMINHO_LEADS_ATUAL}")
    exit()

# Guardar amostra original antes da normalizacao (para diagnostico)
try:
    df_atual_original = pd.read_csv(CAMINHO_LEADS_ATUAL, encoding='utf-8', nrows=10)
except:
    df_atual_original = pd.read_csv(CAMINHO_LEADS_ATUAL, encoding='latin1', nrows=10)

try:
    df_atual = pd.read_csv(CAMINHO_LEADS_ATUAL, encoding='utf-8')
except:
    df_atual = pd.read_csv(CAMINHO_LEADS_ATUAL, encoding='latin1')

df_atual.columns = df_atual.columns.str.strip()
df_atual['Record ID'] = normalizar_record_id(df_atual['Record ID'])
df_atual = df_atual.drop_duplicates(subset=['Record ID'], keep='first')

print(f"  [OK] {len(df_atual)} leads atuais carregados")

# --- CARREGAR MATRÍCULAS FINAIS (SOURCE OF TRUTH) ---
# ATUALIZAÇÃO: matriculas_finais_limpo.csv com Nome do Aluno = Nome do negócio
print("  [INFO] Carregando base de Matrículas Finais para validação...")
set_matriculados_finais = set()
set_p1_paga = set()

def normalizar_nome_mat(nome):
    """Normaliza nome para matching."""
    if pd.isna(nome): return ""
    nome = str(nome).upper().strip()
    nome = unicodedata.normalize('NFKD', nome)
    nome = ''.join(c for c in nome if not unicodedata.combining(c))
    nome = ' '.join(nome.split())
    return nome

if os.path.exists(CAMINHO_MATRICULAS):
    try:
        df_mat = pd.read_csv(CAMINHO_MATRICULAS, sep=';', encoding='utf-8-sig')
    except:
        try:
            df_mat = pd.read_csv(CAMINHO_MATRICULAS, sep=';', encoding='latin1')
        except:
            df_mat = pd.read_csv(CAMINHO_MATRICULAS, sep=None, engine='python', encoding='utf-8')

    df_mat.columns = df_mat.columns.str.strip()
    print(f"    -> Colunas encontradas: {list(df_mat.columns)}")

    # Estrutura: Nome do Aluno, Status P1, Unidade, Tipo
    col_nome = next((c for c in df_mat.columns if 'nome' in c.lower() and 'aluno' in c.lower()), None)
    col_status_p1 = next((c for c in df_mat.columns if 'status' in c.lower() and 'p1' in c.lower()), None)
    col_tipo = next((c for c in df_mat.columns if 'tipo' in c.lower()), None)

    if col_nome:
        print(f"    -> Coluna de nome detectada: '{col_nome}'")

        # Filtrar apenas alunos NOVOS (Excluir REMA)
        if col_tipo:
            antes = len(df_mat)
            df_mat = df_mat[~df_mat[col_tipo].astype(str).str.contains('REMA', case=False, na=False)]
            print(f"    -> Filtrado REMA: {antes - len(df_mat)} removidos, {len(df_mat)} restantes.")

        df_mat['Nome_Normalizado'] = df_mat[col_nome].apply(normalizar_nome_mat)

        # Todos na base são matriculados
        set_matriculados_finais = set(df_mat['Nome_Normalizado'].unique())
        print(f"    -> {len(set_matriculados_finais)} matrículas efetivas carregadas.")

        # Filtrar P1 = PAGA ou BOLSA
        if col_status_p1:
            df_p1_paga = df_mat[df_mat[col_status_p1].astype(str).str.upper().isin(['PAGA', 'BOLSA'])]
            set_p1_paga = set(df_p1_paga['Nome_Normalizado'].unique())
            print(f"    -> {len(set_p1_paga)} com P1 paga/bolsa.")
    else:
        print("    [AVISO] Coluna 'Nome do Aluno' não identificada.")
else:
    print("    [AVISO] matriculas_finais_limpo.csv não encontrado.")

# Identificar coluna de data de última atividade
# Prioridade: 'Data da última atividade' > 'Data da última atividade de vendas' > Genérica
prioridades_atividade = ['Data da última atividade', 'Data da última atividade de vendas', 'Data do último contato']
col_atividade = None

for p in prioridades_atividade:
    match = next((c for c in df_atual.columns if c.strip().lower() == p.lower()), None)
    if match:
        col_atividade = match
        break

if not col_atividade:
    col_atividade = next((c for c in df_atual.columns if 'data' in c.lower() and ('atividade' in c.lower() or 'contato' in c.lower())), None)

print(f"  [CONFIG] Coluna de Atividade detectada: {col_atividade}")

# Identificar coluna de número de atividades (Esforço Comercial)
col_num_atividades = next((c for c in df_atual.columns if 'número' in c.lower() and 'atividades' in c.lower()), None)

# Colunas relevantes do atual
cols_atual_rel = ['Record ID', 'Etapa do negócio', 'Data de criação']
if col_atividade:
    cols_atual_rel.append(col_atividade)
if col_num_atividades:
    cols_atual_rel.append(col_num_atividades)

# Função para forçar formato brasileiro de data (DD/MM/YYYY)
def parse_data_br(series):
    # Tenta primeiro com hora
    d1 = pd.to_datetime(series, format='%d/%m/%Y %H:%M', errors='coerce')
    # Onde falhou, tenta sem hora
    mask = d1.isna()
    d1[mask] = pd.to_datetime(series[mask], format='%d/%m/%Y', errors='coerce')
    # Onde ainda falhou, tenta genérico com dayfirst=True
    mask = d1.isna()
    if mask.any():
        d1[mask] = pd.to_datetime(series[mask], dayfirst=True, errors='coerce', format=None)
    return d1

df_atual_slim = df_atual[cols_atual_rel].copy()
df_atual_slim.rename(columns={'Etapa do negócio': 'Etapa_Atual', 'Data de criação': 'Data_Criacao_Lead'}, inplace=True)

# Criar chave de join no atual para cruzamento com matriculas_finais
if 'Nome do negócio' in df_atual.columns and 'Unidade Desejada' in df_atual.columns:
    df_atual_slim['Key_Join_Mat'] = df_atual['Nome do negócio'].astype(str).str.upper().str.strip() + '_' + df_atual['Unidade Desejada'].astype(str).str.upper().str.strip()

if col_atividade:
    df_atual_slim.rename(columns={col_atividade: 'Data_Ultima_Atividade'}, inplace=True)
    # CORREÇÃO DE DATA: Forçar formato BR para evitar inversão (05/01 virar 01/05)
    df_atual_slim['Data_Ultima_Atividade'] = parse_data_br(df_atual_slim['Data_Ultima_Atividade'])
    
    # Fallback: Se data atividade vazia, usar data criacao para não quebrar cálculo de ociosidade
    df_atual_slim['Data_Ultima_Atividade'] = df_atual_slim['Data_Ultima_Atividade'].fillna(df_atual_slim['Data_Criacao_Lead'])

if col_num_atividades:
    df_atual_slim.rename(columns={col_num_atividades: 'Numero_Atividades_Atual'}, inplace=True)
    df_atual_slim['Numero_Atividades_Atual'] = pd.to_numeric(df_atual_slim['Numero_Atividades_Atual'], errors='coerce').fillna(0)

# Converter data de criação
df_atual_slim['Data_Criacao_Lead'] = parse_data_br(df_atual_slim['Data_Criacao_Lead'])

# ==============================================================================
# 3. CONSOLIDAR MARCOS - LÓGICA SIMPLIFICADA (CADA LEAD APARECE 1 VEZ)
# ==============================================================================
print("\n[3/5] Consolidando marcos (cada lead aparece 1 vez no marco original)...")

# Normalizar colunas essenciais de cada marco
colunas_essenciais = {
    'Record ID': 'Record ID',
    'Nota_1a5': 'Nota_Original_1a5',
    'Probabilidade_Conversao': 'Probabilidade_Original',
    'Unidade Desejada': 'Unidade_Desejada',
    'Nome do negócio': 'Nome_Negocio',
    'Data de criação': 'Data_Criacao'
}

for marco_nome, df_marco in marcos_dados.items():
    for col_original, col_nova in colunas_essenciais.items():
        if col_original in df_marco.columns:
            df_marco.rename(columns={col_original: col_nova}, inplace=True)

    if 'Nota_Original_1a5' in df_marco.columns:
        df_marco['Nota_Original_1a5'] = pd.to_numeric(df_marco['Nota_Original_1a5'], errors='coerce')

def normalizar_chave_dedup(nome):
    """Normalizacao agressiva para deduplicacao (remove acentos, espacos extras)"""
    if pd.isna(nome): return ""
    nome = str(nome).upper().strip()
    # Remove acentos
    nome = unicodedata.normalize('NFKD', nome)
    nome = ''.join(c for c in nome if not unicodedata.combining(c))
    nome = ' '.join(nome.split())
    return nome

# IMPORTANTE: Usar NOME como chave primaria (Record ID muda no HubSpot!)
# Normalizar nomes para garantir match correto
for marco_nome, df_marco in marcos_dados.items():
    if 'Nome_Negocio' in df_marco.columns:
        df_marco['Nome_Chave'] = df_marco['Nome_Negocio'].apply(normalizar_chave_dedup)

# Identificar marco ORIGINAL de cada lead por NOME (prioridade: Marco1 > Marco2 > Marco3)
print("\n  >> Regra Simples:")
print("     - Lead em 12/12? -> Reporta APENAS em 12/12")
print("     - Lead só em 05/01? -> Reporta em 05/01")
print("     - Lead só em 20/01? -> Reporta em 20/01")
print("     - Resultado: CADA LEAD APARECE UMA VEZ\n")

nomes_ja_vistos = set()
ids_ja_vistos = set()
df_historico_completo = pd.DataFrame()

for marco_nome in ['PreLancamento_Marco1', 'PreLancamento_Marco2', 'PosLancamento_Marco3']:
    if marco_nome in marcos_dados:
        df_marco = marcos_dados[marco_nome].copy()

        # Filtrar apenas leads NOVOS por NOME (nao vistos em marcos anteriores)
        if 'Nome_Chave' in df_marco.columns:
            # FIX: Deduplicar por nome dentro do marco para evitar duplicatas internas
            # Priorizar o mais recente (Data_Criacao)
            if 'Data_Criacao' in df_marco.columns:
                df_marco['Data_Criacao'] = pd.to_datetime(df_marco['Data_Criacao'], dayfirst=True, errors='coerce')
                df_marco = df_marco.sort_values('Data_Criacao', ascending=False)
            elif 'Nota_Original_1a5' in df_marco.columns:
                df_marco = df_marco.sort_values('Nota_Original_1a5', ascending=False)
            df_marco = df_marco.drop_duplicates(subset=['Nome_Chave'], keep='first')
            
            # [NOVO] Tambem deduplicar por Record ID dentro do marco (seguranca extra)
            if 'Record ID' in df_marco.columns:
                df_marco = df_marco.drop_duplicates(subset=['Record ID'], keep='first')

            # [CORRECAO] Verificar duplicidade por NOME ou ID
            mask_nome = ~df_marco['Nome_Chave'].isin(nomes_ja_vistos)
            if 'Record ID' in df_marco.columns:
                mask_id = ~df_marco['Record ID'].isin(ids_ja_vistos)
                mask_novos = mask_nome & mask_id
            else:
                mask_novos = mask_nome
            
            df_marco_novos = df_marco[mask_novos].copy()

            # Adicionar coluna identificando marco original
            df_marco_novos['Marco_Original'] = MARCOS[marco_nome]['descricao']

            # Adicionar ao historico
            if len(df_marco_novos) > 0:
                df_historico_completo = pd.concat([df_historico_completo, df_marco_novos], ignore_index=True)
                nomes_ja_vistos.update(df_marco_novos['Nome_Chave'].tolist())
                if 'Record ID' in df_marco_novos.columns:
                    ids_ja_vistos.update(df_marco_novos['Record ID'].tolist())

                print(f"  - {MARCOS[marco_nome]['descricao']}: {len(df_marco_novos)} leads NOVOS (de {len(df_marco)} total)")
        else:
            print(f"  [!] {MARCOS[marco_nome]['descricao']}: Coluna Nome_Negocio nao encontrada")

print(f"\n  [OK] {len(df_historico_completo)} leads unicos consolidados (sem duplicatas, por nome)")

# ==============================================================================
# 4. CRUZAR COM SITUAÇÃO ATUAL E CRIAR MÉTRICAS DE ACOMPANHAMENTO
# ==============================================================================
print("\n[4/5] Cruzando com situacao atual e criando metricas de acompanhamento...")
print(f"   [METODOLOGIA] Comparando Previsão (Relatórios Antigos) vs Realidade (Base HubSpot Hoje: {DATA_HOJE})")
print(f"   [CHAVE] Utilizando 'Nome do Negócio' para vincular o passado ao presente.")

# Preparar coluna de nome normalizado no HubSpot atual
if 'Nome do negócio' in df_atual.columns:
    df_atual['Nome_Chave_Atual'] = df_atual['Nome do negócio'].astype(str).str.strip().str.upper().str.replace(r'\s+', ' ', regex=True)
    df_atual_slim['Nome_Chave_Atual'] = df_atual['Nome_Chave_Atual']

# Preparar DataFrames para Match
df_atual_slim = df_atual_slim.sort_values('Data_Criacao_Lead', ascending=False) # Prioriza o mais recente

# Lookup secundário por Nome (apenas para fallback)
df_atual_nome_lookup = df_atual_slim.drop_duplicates(subset=['Nome_Chave_Atual'], keep='first')

# TENTATIVA 1: MERGE POR RECORD ID (Prioridade Máxima)
print("  - Tentando match por Record ID (Prioridade 1)...")
df_evolucao = pd.merge(
    df_historico_completo,
    df_atual_slim,
    on='Record ID',
    how='left',
    suffixes=('', '_Atual_ID')
)

# TENTATIVA 2: FALLBACK POR NOME (Para quem não deu match no ID)
mask_sem_match = df_evolucao['Etapa_Atual'].isna()
count_sem_match = mask_sem_match.sum()

if count_sem_match > 0:
    print(f"  - {count_sem_match} leads sem match por ID. Tentando fallback por Nome...")
    
    # Separar os não encontrados
    df_retry = df_evolucao[mask_sem_match].copy()
    df_ok = df_evolucao[~mask_sem_match].copy()
    
    # Remover colunas vazias vindas do merge falho (exceto chaves)
    cols_to_drop = [c for c in df_atual_slim.columns if c != 'Record ID' and c in df_retry.columns]
    df_retry = df_retry.drop(columns=cols_to_drop)
    
    # Merge por Nome
    df_retry_filled = pd.merge(
        df_retry,
        df_atual_nome_lookup,
        left_on='Nome_Chave',
        right_on='Nome_Chave_Atual',
        how='left',
        suffixes=('', '_Atual_Nome')
    )
    
    # Se encontrou por nome, atualiza o Record ID para o atual (pois o histórico pode estar obsoleto)
    if 'Record ID_Atual_Nome' in df_retry_filled.columns:
        df_retry_filled['Record ID'] = df_retry_filled['Record ID_Atual_Nome'].fillna(df_retry_filled['Record ID'])
    
    # Recombinar
    df_evolucao = pd.concat([df_ok, df_retry_filled], ignore_index=True)

mask_com_match = df_evolucao['Etapa_Atual'].notna()

# Garantir que Nota_Original_1a5 seja numérico (float) para comparações
if 'Nota_Original_1a5' in df_evolucao.columns:
    df_evolucao['Nota_Original_1a5'] = pd.to_numeric(df_evolucao['Nota_Original_1a5'], errors='coerce').fillna(0)

matches_por_nome = mask_com_match.sum()
print(f"  - Leads com match por nome: {matches_por_nome} ({matches_por_nome/len(df_evolucao)*100:.1f}%)")
print(f"  - Total de leads apos merge: {len(df_evolucao)}")

# ============================================================================
# DIAGNOSTICO: Verificar qualidade do merge (POR NOME)
# ============================================================================
print("\n  >> DIAGNOSTICO DE MERGE (por Nome do Negocio):")

# Leads sem match
leads_sem_match = df_evolucao[df_evolucao['Etapa_Atual'].isna()]
print(f"  - Leads historicos SEM match no HubSpot atual: {len(leads_sem_match)}")

if len(leads_sem_match) > 0:
    print(f"  - Representam {len(leads_sem_match) / len(df_evolucao) * 100:.1f}% do total historico")

    # Mostrar distribuicao por marco
    dist_marco = leads_sem_match.groupby('Marco_ID').size()
    print("\n  Distribuicao por marco:")
    for marco, count in dist_marco.items():
        print(f"     {marco}: {count} leads sem match")

    # Exemplos de nomes sem match
    if 'Nome_Negocio' in leads_sem_match.columns:
        exemplos_nomes = leads_sem_match[['Nome_Negocio', 'Record ID']].head(5)
        print(f"\n  Exemplos de leads sem match:")
        for idx, row in exemplos_nomes.iterrows():
            print(f"     - Nome: {row['Nome_Negocio'][:50]}, ID Historico: {row['Record ID']}")

# Estatisticas gerais
nomes_historicos = set(df_evolucao['Nome_Chave'].dropna().unique()) if 'Nome_Chave' in df_evolucao.columns else set()
nomes_atuais = set(df_atual['Nome_Chave_Atual'].dropna().unique()) if 'Nome_Chave_Atual' in df_atual.columns else set()
nomes_com_match = nomes_historicos.intersection(nomes_atuais)

print(f"\n  >> ESTATISTICAS DE MERGE (por Nome):")
print(f"  - Total Nomes historicos: {len(nomes_historicos)}")
print(f"  - Total Nomes HubSpot atual: {len(nomes_atuais)}")
print(f"  - Nomes com match: {len(nomes_com_match)} ({len(nomes_com_match)/len(nomes_historicos)*100:.1f}% se houver nomes)" if len(nomes_historicos) > 0 else "  - Nomes com match: 0")
print(f"  - Nomes sem match: {len(nomes_historicos - nomes_com_match)}")

# ============================================================================
# CLASSIFICAÇÃO SIMPLES DE STATUS E RESULTADO
# ============================================================================

# ATUALIZAÇÃO: Normalizar nomes no df_evolucao para matching com matrículas
def normalizar_nome_hubspot(nome):
    """Normaliza nome do HubSpot para matching."""
    if pd.isna(nome): return ""
    nome = str(nome).upper().strip()
    # Remove sufixos como " - 10 - UNIDADE"
    nome = re.sub(r'\s*-\s*\d+\s*-\s*.*$', '', nome)
    nome = unicodedata.normalize('NFKD', nome)
    nome = ''.join(c for c in nome if not unicodedata.combining(c))
    nome = ' '.join(nome.split())
    return nome

df_evolucao['Nome_Normalizado_Match'] = df_evolucao['Nome_Negocio'].apply(normalizar_nome_hubspot)

# Classificar Status Atual (Simplificado)
def classificar_status_simples(row):
    # 1. Checar Verdade Absoluta (Matriculas Finais por Nome Normalizado)
    nome_norm = row.get('Nome_Normalizado_Match', '')
    if nome_norm in set_matriculados_finais:
        return 'Matriculou'

    # 2. P1 Paga também é considerado próximo de matricular
    if nome_norm in set_p1_paga:
        return 'Matriculou'

    etapa = row['Etapa_Atual']
    if pd.isna(etapa):
        return 'Sem Informacao'
    etapa_lower = str(etapa).lower()
    if any(x in etapa_lower for x in ['matrícula', 'ganho', 'won', 'matricula']):
        return 'Matriculou'
    elif 'visita realizada' in etapa_lower:
        return 'Visita Realizada'
    elif any(x in etapa_lower for x in ['visita', 'agendada']):
        return 'Visita Agendada'
    elif any(x in etapa_lower for x in ['perdido', 'lost', 'arquivado']):
        return 'Perdido'
    elif any(x in etapa_lower for x in ['negociação', 'negociacao', 'proposta']):
        return 'Em Atendimento'
    else:
        return 'Em Atendimento'

df_evolucao['Status_Atual'] = df_evolucao.apply(classificar_status_simples, axis=1)

# Log de matches encontrados
n_match = df_evolucao['Nome_Normalizado_Match'].isin(set_matriculados_finais).sum()
n_p1 = df_evolucao['Nome_Normalizado_Match'].isin(set_p1_paga).sum()
print(f"  [MATCH] {n_match} leads encontrados na base de matrículas finais")
print(f"  [MATCH] {n_p1} leads com P1 paga")

# Calcular dias desde o marco
df_evolucao['Dias_Desde_Marco'] = (pd.Timestamp.now() - df_evolucao['Data_Marco']).dt.days

# Adicionar Data de Referência (Congela a data do relatório)
df_evolucao['Data_Relatorio'] = pd.to_datetime(DATA_HOJE)

# Calcular Ociosidade (Diferença entre Data do Relatório e Última Atividade)
df_evolucao['Dias_Ociosidade'] = (df_evolucao['Data_Relatorio'] - df_evolucao['Data_Ultima_Atividade']).dt.days

# [CORREÇÃO] Matriculados não devem ter ociosidade (zera contador para evitar falso positivo de SLA)
df_evolucao.loc[df_evolucao['Status_Atual'] == 'Matriculou', 'Dias_Ociosidade'] = 0

# Classificar Resultado de Conversão (Simples) - AGORA COM REGRAS DE SLA
def classificar_resultado_simples(row):
    status = row['Status_Atual']
    
    # Se já converteu ou perdeu, mantemos o status final
    if status in ['Matriculou', 'Visita Agendada', 'Visita Realizada', 'Perdido', 'Perdeu']:
        return status if status != 'Perdido' else 'Perdeu'
    
    # Se está em atendimento, aplicamos a regra de SLA do Helio
    dias = row['Dias_Ociosidade']
    nota = row.get('Nota_Original_1a5', 0)
    
    if pd.isna(dias):
        return "Não Atendido"
    
    dias = float(dias)
    
    if nota in [4, 5]:
        if dias <= 1: return "Normal"
        elif dias == 2: return "Mapear Status"
        elif dias <= 4: return "Ociosidade"
        else: return "ALTA OCIOSIDADE"
    elif nota == 3:
        if dias <= 3: return "Normal"
        elif dias == 4: return "Mapear Status"
        elif dias <= 6: return "Ociosidade"
        else: return "ALTA OCIOSIDADE"
    elif nota in [1, 2]:
        if dias <= 5: return "Normal"
        elif dias == 6: return "Mapear Status"
        elif dias <= 8: return "Ociosidade"
        else: return "ALTA OCIOSIDADE"
    else:
        return "Sem Atributo"

df_evolucao['Resultado_Conversao'] = df_evolucao.apply(classificar_resultado_simples, axis=1)

# ============================================================================
# NOVA LÓGICA DE SLA (QUALIDADE DO ATENDIMENTO)
# ============================================================================
def classificar_sla_atendimento(row):
    dias = row['Dias_Ociosidade']
    nota = row.get('Nota_Original_1a5', 0)
    
    # Se Dias_Ociosidade for nulo/NaN
    if pd.isna(dias):
        return "Não Atendido"
    
    dias = float(dias)
    
    if nota in [4, 5]:
        if dias <= 1: return "Normal"
        elif dias == 2: return "Mapear Status"
        elif dias <= 4: return "Ociosidade"
        else: return "ALTA OCIOSIDADE"
    elif nota == 3:
        if dias <= 3: return "Normal"
        elif dias == 4: return "Mapear Status"
        elif dias <= 6: return "Ociosidade"
        else: return "ALTA OCIOSIDADE"
    elif nota in [1, 2]:
        if dias <= 5: return "Normal"
        elif dias == 6: return "Mapear Status"
        elif dias <= 8: return "Ociosidade"
        else: return "ALTA OCIOSIDADE"
    else:
        return "Sem Atributo"

df_evolucao['Status_SLA'] = df_evolucao.apply(classificar_sla_atendimento, axis=1)

# ============================================================================
# RESUMO DE SLA NO CONSOLE (PARA CONFERÊNCIA IMEDIATA)
# ============================================================================
print("\n[DIAGNÓSTICO SLA] Distribuição para Leads Nota 4 e 5 (Elite):")
try:
    mask_elite = df_evolucao['Nota_Original_1a5'].isin([4, 5])
    print(df_evolucao[mask_elite]['Status_SLA'].value_counts())
except Exception as e:
    print(f"  (Não foi possível gerar resumo SLA: {e})")

# ============================================================================
# [NOVO] DIAGNÓSTICO DE LEAD ESPECÍFIC===============================
print("\n[DIAGNÓSTICO] Buscando lead exemplo: 'Patrick do Carmo Cordeiro'...")
try:
    # Busca parcial case-insensitive
    mask_patrick = df_evolucao['Nome_Negocio'].astype(str).str.contains('Patrick do Carmo', case=False, na=False)
    df_patrick = df_evolucao[mask_patrick]
    
    if not df_patrick.empty:
        for idx, row in df_patrick.iterrows():
            print(f"  > LEAD ENCONTRADO: {row['Nome_Negocio']}")
            print(f"    - Unidade: {row.get('Unidade_Desejada', 'N/A')}")
            print(f"    - Status Helio: {row.get('Status_Atual', 'N/A')} | Resultado: {row.get('Resultado_Conversao', 'N/A')}")
            print(f"    - Data Última Ativ.: {row.get('Data_Ultima_Atividade', 'N/A')} (Ociosidade: {row.get('Dias_Ociosidade', 'N/A')} dias)")
            print(f"    - SLA Calculado: {row.get('Status_SLA', 'N/A')}")
            print("-" * 40)
    else:
        print("  > Lead 'Patrick do Carmo' não encontrado na base consolidada.")
except Exception as e:
    print(f"  (Erro ao buscar lead: {e})")

print("\n[DIAGNÓSTICO] Buscando lead exemplo: 'Julia Coutinho Muniz'...")
try:
    # Busca parcial case-insensitive
    mask_julia = df_evolucao['Nome_Negocio'].astype(str).str.contains('Julia Coutinho Muniz', case=False, na=False)
    df_julia = df_evolucao[mask_julia]
    
    if not df_julia.empty:
        for idx, row in df_julia.iterrows():
            print(f"  > LEAD ENCONTRADO: {row['Nome_Negocio']}")
            print(f"    - Unidade: {row.get('Unidade_Desejada', 'N/A')}")
            print(f"    - Status Helio: {row.get('Status_Atual', 'N/A')} | Resultado: {row.get('Resultado_Conversao', 'N/A')}")
            print(f"    - Data Última Ativ.: {row.get('Data_Ultima_Atividade', 'N/A')} (Ociosidade: {row.get('Dias_Ociosidade', 'N/A')} dias)")
            print(f"    - SLA Calculado: {row.get('Status_SLA', 'N/A')}")
            print("-" * 40)
    else:
        print("  > Lead 'Julia Coutinho Muniz' não encontrado na base consolidada.")
except Exception as e:
    print(f"  (Erro ao buscar lead: {e})")

print(f"  [OK] {len(df_evolucao)} registros processados com metricas de acompanhamento")

# ==============================================================================
# 4.1. PERSISTÊNCIA DE HISTÓRICO (LOG DE EXECUÇÕES)
# ==============================================================================
CAMINHO_HISTORICO_LOG = os.path.join(PASTA_VALIDACAO, 'Historico_Execucoes_Helio.csv')
stats_hoje = pd.DataFrame([{
    'Data_Relatorio': DATA_HOJE,
    'Total_Leads_Monitorados': len(df_evolucao),
    'Total_Matriculas': len(df_evolucao[df_evolucao['Resultado_Conversao'] == 'Matriculou']),
    'Total_Perdidos': len(df_evolucao[df_evolucao['Resultado_Conversao'] == 'Perdeu'])
}])

if os.path.exists(CAMINHO_HISTORICO_LOG):
    try:
        df_hist_log = pd.read_csv(CAMINHO_HISTORICO_LOG, sep=';')
        df_hist_log = df_hist_log[df_hist_log['Data_Relatorio'] != DATA_HOJE]
        df_hist_log = pd.concat([df_hist_log, stats_hoje], ignore_index=True)
    except:
        df_hist_log = stats_hoje
else:
    df_hist_log = stats_hoje

# Salvar log atualizado
df_hist_log.to_csv(CAMINHO_HISTORICO_LOG, sep=';', index=False)
print(f"  [LOG] Histórico de execuções atualizado: {CAMINHO_HISTORICO_LOG}")

# ==============================================================================
# 5. GERAR EXCEL SIMPLIFICADO - FORMATO DIRETO
# ==============================================================================
print("\n[5/5] Gerando Excel simplificado...")

wb = Workbook()

# --- ABA 1: BASE COMPLETA (SIMPLES E DIRETA) ---
ws_base = wb.active
ws_base.title = 'Base_Completa'

# Colunas essenciais (formato simplificado)
cols_export = [
    'Record ID', 'Nome_Negocio', 'Marco_Original', 'Periodo',
    'Nota_Original_1a5', 'Probabilidade_Original',
    'Status_Atual', 'Etapa_Atual', 'Resultado_Conversao', 
    'Status_SLA', # Nova coluna adicionada
    'Data_Ultima_Atividade', 'Dias_Ociosidade', 'Data_Relatorio', 'Unidade_Desejada', 'Dias_Desde_Marco'
]

cols_disponiveis = [c for c in cols_export if c in df_evolucao.columns]
df_export = df_evolucao[cols_disponiveis].copy()

# Formatar datas
for col in ['Data_Ultima_Atividade']:
    if col in df_export.columns:
        df_export[col] = pd.to_datetime(df_export[col], errors='coerce').dt.strftime('%d/%m/%Y')

df_export = df_export.sort_values(['Periodo', 'Nota_Original_1a5'], ascending=[False, False])

df_to_sheet(ws_base, df_export, freeze_panes=True)
ajustar_largura(ws_base)

# --- ABA 2: PRÉ-LANÇAMENTO (Marco 12/12) ---
ws_pre_marco1 = wb.create_sheet('Pre_Lancamento_12_12')
df_pre_m1 = df_evolucao[df_evolucao['Marco_Original'] == '12/12/2025 - Baseline Organico 1'].copy()
if not df_pre_m1.empty:
    df_to_sheet(ws_pre_marco1, df_pre_m1[cols_disponiveis], freeze_panes=True)
ajustar_largura(ws_pre_marco1)

# --- ABA 3: PRÉ-LANÇAMENTO (Marco 05/01) ---
ws_pre_marco2 = wb.create_sheet('Pre_Lancamento_05_01')
df_pre_m2 = df_evolucao[df_evolucao['Marco_Original'] == '05/01/2026 - Baseline Organico 2'].copy()
if not df_pre_m2.empty:
    df_to_sheet(ws_pre_marco2, df_pre_m2[cols_disponiveis], freeze_panes=True)
ajustar_largura(ws_pre_marco2)

# --- ABA 4: PÓS-LANÇAMENTO (Marco 20/01) ---
ws_pos_marco3 = wb.create_sheet('Pos_Lancamento_20_01')
df_pos_m3 = df_evolucao[df_evolucao['Marco_Original'] == '20/01/2026 - LANCAMENTO DO HELIO'].copy()
if not df_pos_m3.empty:
    df_to_sheet(ws_pos_marco3, df_pos_m3[cols_disponiveis], freeze_panes=True)
ajustar_largura(ws_pos_marco3)

# --- ABA 5: ALTA PRIORIDADE (Nota >= 4) ---
ws_alta = wb.create_sheet('Alta_Prioridade')
df_alta = df_evolucao[df_evolucao['Nota_Original_1a5'] >= 4].copy()
if not df_alta.empty:
    df_to_sheet(ws_alta, df_alta[cols_disponiveis], freeze_panes=True)
ajustar_largura(ws_alta)

# --- ABA 6: MATRÍCULAS ---
ws_matriculas = wb.create_sheet('Matriculas')
df_matriculas = df_evolucao[df_evolucao['Resultado_Conversao'] == 'Matriculou'].copy()
if not df_matriculas.empty:
    df_to_sheet(ws_matriculas, df_matriculas[cols_disponiveis], freeze_panes=True)
ajustar_largura(ws_matriculas)

# --- ABA 7: VISITAS (AGENDADAS E REALIZADAS) ---
ws_visitas = wb.create_sheet('Visitas')
df_visitas = df_evolucao[df_evolucao['Resultado_Conversao'].isin(['Visita Agendada', 'Visita Realizada'])].copy()
if not df_visitas.empty:
    df_to_sheet(ws_visitas, df_visitas[cols_disponiveis], freeze_panes=True)
ajustar_largura(ws_visitas)

# --- ABA 8: PERDIDOS ---
ws_perdidos = wb.create_sheet('Perdidos')
df_perdidos = df_evolucao[df_evolucao['Resultado_Conversao'] == 'Perdeu'].copy()
if not df_perdidos.empty:
    df_to_sheet(ws_perdidos, df_perdidos[cols_disponiveis], freeze_panes=True)
ajustar_largura(ws_perdidos)

# --- ABA 8: ESTATÍSTICAS POR MARCO ---
ws_stats = wb.create_sheet('Estatisticas')
ws_stats['A1'] = 'ESTATISTICAS POR MARCO'
ws_stats['A1'].font = Font(bold=True, size=12)

# Calcular estatísticas simples por marco
stats_marco = df_evolucao.groupby('Marco_Original').agg(
    Total_Leads=('Record ID', 'nunique'),
    Nota_Media=('Nota_Original_1a5', 'mean'),
    Matriculou=('Resultado_Conversao', lambda x: (x == 'Matriculou').sum()),
    Visita_Agendada=('Resultado_Conversao', lambda x: (x == 'Visita Agendada').sum()),
    Visita_Realizada=('Resultado_Conversao', lambda x: (x == 'Visita Realizada').sum()),
    Perdeu=('Resultado_Conversao', lambda x: (x == 'Perdeu').sum()),
    Em_Atendimento=('Resultado_Conversao', lambda x: (x == 'Em Atendimento').sum()),
    Sem_Atendimento=('Resultado_Conversao', lambda x: (x == 'Sem Atendimento').sum()),
).reset_index()

stats_marco['Taxa_Conversao_%'] = ((stats_marco['Matriculou'] + stats_marco['Visita_Agendada'] + stats_marco['Visita_Realizada']) / stats_marco['Total_Leads'] * 100).round(1)
stats_marco['Nota_Media'] = stats_marco['Nota_Media'].round(2)

df_to_sheet(ws_stats, stats_marco, start_row=3)
ajustar_largura(ws_stats)

# --- ABA 9: DASHBOARD POR UNIDADE (FONTE DE DADOS) ---
ws_dash = wb.create_sheet('Dashboard_Unidades')
# Colunas otimizadas para BI
cols_dash = [
    'Unidade_Desejada', 'Record ID', 'Nome_Negocio', 'Marco_Original', 
    'Nota_Original_1a5', 'Status_Atual', 'Resultado_Conversao', 'Status_SLA',
    'Numero_Atividades_Atual', 'Data_Ultima_Atividade', 'Dias_Ociosidade', 'Data_Relatorio', 'Dias_Desde_Marco'
]
# Garantir que colunas existem
cols_dash_final = [c for c in cols_dash if c in df_evolucao.columns]

df_dash = df_evolucao[cols_dash_final].copy()
# Ordenar por Unidade e depois por Potencial (Nota)
if 'Unidade_Desejada' in df_dash.columns:
    df_dash.sort_values(['Unidade_Desejada', 'Nota_Original_1a5'], ascending=[True, False], inplace=True)

# Formatar data para string limpa
for col in ['Data_Ultima_Atividade', 'Data_Relatorio']:
    if col in df_dash.columns:
        df_dash[col] = pd.to_datetime(df_dash[col], errors='coerce').dt.strftime('%d/%m/%Y')

df_to_sheet(ws_dash, df_dash, freeze_panes=True)
ajustar_largura(ws_dash)

# --- ABA 10: KPIs POR UNIDADE (RESUMO EXECUTIVO) ---
ws_kpi = wb.create_sheet('KPIs_Unidades')
ws_kpi['A1'] = 'PERFORMANCE COMERCIAL POR UNIDADE (POTENCIAL vs REALIZADO)'
ws_kpi['A1'].font = Font(bold=True, size=12)

if 'Unidade_Desejada' in df_evolucao.columns:
    # Agregacao inteligente
    kpi_unidade = df_evolucao.groupby('Unidade_Desejada').agg(
        Leads_Recebidos=('Record ID', 'count'),
        Potencial_Alta_Prioridade=('Nota_Original_1a5', lambda x: (x >= 4).sum()),
        Matriculas=('Resultado_Conversao', lambda x: (x == 'Matriculou').sum()),
        Visitas=('Resultado_Conversao', lambda x: (x == 'Visita Agendada').sum() + (x == 'Visita Realizada').sum()),
        Perdas=('Resultado_Conversao', lambda x: (x == 'Perdeu').sum()),
        Atividades_Media=('Numero_Atividades_Atual', 'mean') if 'Numero_Atividades_Atual' in df_evolucao.columns else ('Record ID', lambda x: 0)
    ).reset_index()
    
    # Metricas calculadas
    kpi_unidade['Conversao_Total_%'] = ((kpi_unidade['Matriculas'] + kpi_unidade['Visitas']) / kpi_unidade['Leads_Recebidos'] * 100).round(1)
    if 'Atividades_Media' in kpi_unidade.columns:
        kpi_unidade['Atividades_Media'] = kpi_unidade['Atividades_Media'].round(1)
        
    kpi_unidade = kpi_unidade.sort_values('Leads_Recebidos', ascending=False)
    df_to_sheet(ws_kpi, kpi_unidade, start_row=3, freeze_panes=True)
    ajustar_largura(ws_kpi)

# --- ABA 11: ANALISE DE ACURACIA (PREVISAO vs REALIDADE) ---
ws_acc = wb.create_sheet('Analise_Acuracia')
ws_acc['A1'] = 'VALIDAÇÃO DE ACURÁCIA: O QUE O HELIO PREVIU vs O QUE ACONTECEU'
ws_acc['A1'].font = Font(bold=True, size=12)

dados_acuracia = []
periodos = ['Pre-Lancamento', 'Pos-Lancamento']

for periodo in periodos:
    df_p = df_evolucao[df_evolucao['Periodo'] == periodo].copy()
    if df_p.empty: continue
    
    # Filtrar apenas leads com desfecho claro
    df_fechados = df_p[df_p['Resultado_Conversao'].isin(['Matriculou', 'Visita Agendada', 'Visita Realizada', 'Perdeu'])].copy()
    total_fechados = len(df_fechados)
    
    if total_fechados > 0:
        # TP: Alta Prioridade (4-5) -> Sucesso (Matricula/Visita)
        tp_mat = len(df_fechados[(df_fechados['Nota_Original_1a5'] >= 4) & (df_fechados['Resultado_Conversao'] == 'Matriculou')])
        tp_vis = len(df_fechados[(df_fechados['Nota_Original_1a5'] >= 4) & (df_fechados['Resultado_Conversao'].isin(['Visita Agendada', 'Visita Realizada']))])
        tp = tp_mat + tp_vis
        
        # FP: Alta Prioridade (4-5) -> Perda
        fp = len(df_fechados[(df_fechados['Nota_Original_1a5'] >= 4) & (df_fechados['Resultado_Conversao'] == 'Perdeu')])
        
        # TN: Baixa Prioridade (1-3) -> Perda
        tn = len(df_fechados[(df_fechados['Nota_Original_1a5'] < 4) & (df_fechados['Resultado_Conversao'] == 'Perdeu')])
        
        # FN: Baixa Prioridade (1-3) -> Sucesso (Matricula/Visita)
        fn_mat = len(df_fechados[(df_fechados['Nota_Original_1a5'] < 4) & (df_fechados['Resultado_Conversao'] == 'Matriculou')])
        fn_vis = len(df_fechados[(df_fechados['Nota_Original_1a5'] < 4) & (df_fechados['Resultado_Conversao'].isin(['Visita Agendada', 'Visita Realizada']))])
        fn = fn_mat + fn_vis
        
        total_pred_pos = tp + fp
        total_pred_neg = tn + fn
        
        dados_acuracia.append({
            'Periodo': periodo,
            'Base_Fechada': total_fechados,
            'Leads_Alta_Prio_4_5': total_pred_pos,
            'Sucesso_Alta_Prio_Total': tp,
            '  > Matriculas': tp_mat,
            '  > Visitas': tp_vis,
            'Erro_Alta_Prio_Perda': fp,
            'Taxa_Acerto_Alta_Prio_%': (tp/total_pred_pos*100) if total_pred_pos else 0,
            'Leads_Baixa_Prio_1_3': total_pred_neg,
            'Acerto_Baixa_Prio_Perda': tn,
            'Erro_Baixa_Prio_Sucesso': fn,
            'Taxa_Acerto_Baixa_Prio_%': (tn/total_pred_neg*100) if total_pred_neg else 0
        })

if dados_acuracia:
    df_acc = pd.DataFrame(dados_acuracia)
    df_to_sheet(ws_acc, df_acc, start_row=3, freeze_panes=True)
    ajustar_largura(ws_acc)

# --- ABA 12: HISTÓRICO DE RELATÓRIOS (LOG) ---
ws_hist = wb.create_sheet('Historico_Relatorios')
ws_hist['A1'] = 'HISTÓRICO DE EXECUÇÕES E VOLUMETRIA'
ws_hist['A1'].font = Font(bold=True, size=12)
df_to_sheet(ws_hist, df_hist_log.sort_values('Data_Relatorio', ascending=False), start_row=3)
ajustar_largura(ws_hist)

# Salvar arquivo
try:
    wb.save(CAMINHO_SAIDA)
    print(f"\n{'='*80}")
    print(f"[OK] RELATORIO DE EVOLUÇÃO TEMPORAL GERADO COM SUCESSO!")
    print(f"{'='*80}")
    print(f"Arquivo: {CAMINHO_SAIDA}")
    print(f"\nRESUMO:")
    for idx, row in stats_marco.iterrows():
        marco = row['Marco_Original']
        total = row['Total_Leads']
        matriculas = row['Matriculou']
        visitas = row['Visita_Agendada']
        visitas_real = row['Visita_Realizada']
        taxa = row['Taxa_Conversao_%']
        print(f"\n   {marco}:")
        print(f"     - Total Leads: {total}")
        print(f"     - Matriculas: {matriculas}")
        print(f"     - Visitas (Agendadas + Realizadas): {visitas + visitas_real}")
        print(f"     - Taxa Conversao: {taxa}%")
    
    # --- NOVA SECAO: ANALISE DETALHADA ---
    print(f"\n{'='*80}")
    print(f"ANALISE DETALHADA DE PERFORMANCE (HELIO)")
    print(f"{'='*80}")

    # 1. Composicao da Base
    total_leads = len(df_evolucao)
    print(f"\n1. COMPOSICAO DA BASE CONSOLIDADA ({total_leads} leads unicos):")
    if total_leads > 0:
        comp = df_evolucao['Marco_Original'].value_counts()
        for marco, count in comp.items():
            pct = (count / total_leads) * 100
            print(f"   - {marco}: {count} leads ({pct:.1f}%)")

    # 2. Acuracia do Modelo (Comparativo Pre vs Pos)
    print(f"\n2. ACURACIA DO MODELO (PREVISAO vs REALIDADE) - PRE vs POS LANCAMENTO:")
    print(f"   [VALIDACAO] Verificando se a 'Nota' atribuída no passado se concretizou no Status Atual")
    print(f"{'='*80}")
    
    # Agrupar por Periodo
    periodos = ['Pre-Lancamento', 'Pos-Lancamento']
    
    for periodo in periodos:
        df_p = df_evolucao[df_evolucao['Periodo'] == periodo].copy()
        if df_p.empty: continue
        
        # Filtrar apenas leads com desfecho claro (Matriculou, Visita ou Perdeu)
        df_fechados = df_p[df_p['Resultado_Conversao'].isin(['Matriculou', 'Visita Agendada', 'Visita Realizada', 'Perdeu'])].copy()
        total_fechados = len(df_fechados)
        
        if total_fechados == 0:
            print(f"\n   >> {periodo.upper()}: Sem leads fechados (Matriculou/Visita/Perdeu) para analise.")
            continue
            
        # Definicoes: Nota >= 4 (Disse que ia matricular) | Nota < 4 (Disse que ia perder)
        # SUCESSO = Matriculou OU Visita Agendada
        tp = len(df_fechados[(df_fechados['Nota_Original_1a5'] >= 4) & (df_fechados['Resultado_Conversao'].isin(['Matriculou', 'Visita Agendada', 'Visita Realizada']))])
        # Detalhar TP (True Positives)
        tp_mat = len(df_fechados[(df_fechados['Nota_Original_1a5'] >= 4) & (df_fechados['Resultado_Conversao'] == 'Matriculou')])
        tp_vis = len(df_fechados[(df_fechados['Nota_Original_1a5'] >= 4) & (df_fechados['Resultado_Conversao'].isin(['Visita Agendada', 'Visita Realizada']))])
        tp = tp_mat + tp_vis

        tn = len(df_fechados[(df_fechados['Nota_Original_1a5'] < 4) & (df_fechados['Resultado_Conversao'] == 'Perdeu')])
        fp = len(df_fechados[(df_fechados['Nota_Original_1a5'] >= 4) & (df_fechados['Resultado_Conversao'] == 'Perdeu')])
        fn = len(df_fechados[(df_fechados['Nota_Original_1a5'] < 4) & (df_fechados['Resultado_Conversao'].isin(['Matriculou', 'Visita Agendada', 'Visita Realizada']))])
        
        # Detalhar FN (False Negatives)
        fn_mat = len(df_fechados[(df_fechados['Nota_Original_1a5'] < 4) & (df_fechados['Resultado_Conversao'] == 'Matriculou')])
        fn_vis = len(df_fechados[(df_fechados['Nota_Original_1a5'] < 4) & (df_fechados['Resultado_Conversao'].isin(['Visita Agendada', 'Visita Realizada']))])
        fn = fn_mat + fn_vis
        
        total_pred_pos = tp + fp
        total_pred_neg = tn + fn
        
        print(f"\n   >> {periodo.upper()} (Base Fechada: {total_fechados} leads):")
        
        if total_pred_pos > 0:
            print(f"      [1] DISSE QUE IA MATRICULAR (Nota 4-5): {total_pred_pos}")
            print(f"          [OK] CONVERSAO (Sucesso Total):   {tp} ({tp/total_pred_pos*100:.1f}%)")
            print(f"             - Matriculas: {tp_mat}")
            print(f"             - Visitas:    {tp_vis}")
            print(f"          [X]  ERROU (Perdeu):              {fp} ({fp/total_pred_pos*100:.1f}%)")

        if total_pred_neg > 0:
            print(f"      [2] DISSE QUE IA PERDER (Nota 1-3):     {total_pred_neg}")
            print(f"          [OK] Acertou (Perdeu):            {tn} ({tn/total_pred_neg*100:.1f}%)")
            print(f"          [X]  Errou (Matriculou/Visita):   {fn} ({fn/total_pred_neg*100:.1f}%)")
            print(f"             - Matriculas: {fn_mat}")
            print(f"             - Visitas:    {fn_vis}")

    print(f"\n[!] Cada lead aparece UMA VEZ no marco onde foi visto PRIMEIRO")
    print(f"[!] Total de leads unicos: {len(df_evolucao)}")
    print("="*80)

except Exception as e:
    print(f"\n[ERRO] Erro ao salvar Excel: {e}")
    import traceback
    traceback.print_exc()
