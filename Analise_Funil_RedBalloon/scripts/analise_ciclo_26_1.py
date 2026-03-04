"""
Análise Simples - Ciclo 26.1
Perguntas diretas sobre atividade comercial por canal e idade dos leads
"""

import pandas as pd
import numpy as np
import os
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================================
# 1. CARREGAR E PREPARAR DADOS
# ============================================================================

print("📊 Carregando dados...")
BASE_DIR = Path(__file__).parent.parent

# Candidate input files (in order of preference)

# Priorizar explicitamente a base atualizada fornecida pelo usuário
candidate_paths = [
    Path(r"C:\Users\a483650\Projetos\_DADOS_CENTRALIZADOS\hubspot\hubspot_leads_ATUAL.csv"),
    BASE_DIR / "data" / "hubspot_leads_redballoon.csv",
    Path(r"C:\Users\a483650\Projetos\_DADOS_CENTRALIZADOS\hubspot\historico\hubspot_leads_ATUAL.csv"),
    Path(r"C:\Users\a483650\Projetos\_ARQUIVO\projeto_helio_teste\Data\hubspot_leads_atual.csv"),
]

found_file = None
for p in candidate_paths:
    if Path(p).exists():
        found_file = p
        break

if found_file is None:
    raise FileNotFoundError(
        "Nenhum arquivo de leads encontrado. Verifique caminhos: " + ", ".join([str(p) for p in candidate_paths])
    )

print(f"📥 Usando base de dados: {found_file}")
df = pd.read_csv(found_file)

# Converter datas
df['Data de criação'] = pd.to_datetime(df['Data de criação'])
df['Data de fechamento'] = pd.to_datetime(df['Data de fechamento'], errors='coerce')
df['Data da última atividade'] = pd.to_datetime(df['Data da última atividade'], errors='coerce')

# Definir ciclos (Out/2025 até Mar/2026 = Ciclo 26.1)
def definir_ciclo(data):
    if pd.isna(data):
        return None
    ano = data.year
    mes = data.month
    
    if mes >= 10:  # Out, Nov, Dez → ciclo do próximo ano
        return f"{str(ano + 1)[-2:]}.1"
    elif mes <= 3:  # Jan, Fev, Mar → ciclo do ano atual
        return f"{str(ano)[-2:]}.1"
    else:
        return None

df['ciclo'] = df['Data de criação'].apply(definir_ciclo)

# Identificar conversões (matrícula concluída OU data de fechamento preenchida)
df['converteu'] = (df['Etapa do negócio'].str.upper().str.contains('MATRÍCULA CONCLUÍDA', na=False)) | (df['Data de fechamento'].notna())

# Canal (simplificado)
df['canal'] = df['Fonte original do tráfego'].fillna('Não informado')

# Classificação de unidades (Própria vs Franquia)
unidades_proprias = ['VILA LEOPOLDINA', 'MORUMBI', 'ITAIM BIBI', 'PACAEMBU', 'PINHEIROS', 'JARDINS', 'PERDIZES', 'SANTANA']

def classificar_unidade(unidade):
    if pd.isna(unidade):
        return 'N/A'
    unidade_upper = str(unidade).upper().strip()
    return 'Própria' if unidade_upper in unidades_proprias else 'Franquia'

df['tipo_unidade'] = df['Unidade Desejada'].apply(classificar_unidade)

# ============================================================================
# 2. ANÁLISE CICLO 26.1
# ============================================================================

print("\n🎯 Analisando Ciclo 26.1...")

# Calcular dias desde início do ciclo atual (1º Out 2025)
hoje = datetime.now()
inicio_ciclo_26_1 = datetime(2025, 10, 1)
dias_desde_inicio = (hoje - inicio_ciclo_26_1).days

print(f"   Dias desde início do ciclo: {dias_desde_inicio}")

# Filtrar ciclo 26.1
df_ciclo_26_1 = df[df['ciclo'] == '26.1'].copy()

# Calcular idade dos leads do ciclo 26.1 (em relação a hoje)
df_ciclo_26_1['idade_dias'] = (hoje - df_ciclo_26_1['Data de criação']).dt.days

# Classificar em faixas: D-1 separado para identificar erros
def classificar_idade(dias):
    if pd.isna(dias):
        return 'Sem data'
    elif dias <= 1:
        return 'D-1'
    elif dias <= 7:
        return '2 a 7 dias'
    elif dias <= 30:
        return '8 a 30 dias'
    elif dias <= 60:
        return '31 a 60 dias'
    elif dias <= 90:
        return '61 a 90 dias'
    else:
        return '> 90 dias'

df_ciclo_26_1['faixa_idade'] = df_ciclo_26_1['idade_dias'].apply(classificar_idade)

# Calcular recência da última atividade
df_ciclo_26_1['dias_desde_ultima_atividade'] = (hoje - df_ciclo_26_1['Data da última atividade']).dt.days

# Classificar recência da atividade
def classificar_recencia(dias):
    if pd.isna(dias):
        return 'Sem atividade'
    elif dias <= 1:
        return 'D-1'
    elif dias <= 7:
        return '≤ 7 dias'
    elif dias <= 30:
        return '≤ 30 dias'
    elif dias <= 60:
        return '≤ 60 dias'
    elif dias <= 90:
        return '≤ 90 dias'
    else:
        return '> 90 dias'

df_ciclo_26_1['recencia_atividade'] = df_ciclo_26_1['dias_desde_ultima_atividade'].apply(classificar_recencia)

# ============================================================================
# 3. CALCULAR MÉTRICAS
# ============================================================================

print("\n📈 Calculando métricas...")

# 3.1 Total de leads por canal
total_leads_por_canal = df_ciclo_26_1.groupby('canal').size()

# 3.2 Distribuição de idade no ciclo 26.1 por canal
distribuicao_idade = df_ciclo_26_1.groupby(['canal', 'faixa_idade']).size().unstack(fill_value=0)

# Garantir que todas as faixas existam
for faixa in ['D-1', '< 7 dias', '< 30 dias', '< 60 dias', '< 90 dias', '> 90 dias']:
    if faixa not in distribuicao_idade.columns:
        distribuicao_idade[faixa] = 0

# Reordenar colunas
ordem_faixas = ['D-1', '2 a 7 dias', '8 a 30 dias', '31 a 60 dias', '61 a 90 dias', '> 90 dias']
distribuicao_idade = distribuicao_idade[[col for col in ordem_faixas if col in distribuicao_idade.columns]]

# 3.3 Média de atividades para leads que converteram por faixa (HISTÓRICO)
# Primeiro, calcular idade no momento da conversão para todos os leads históricos
df_historico_convertidos = df[df['converteu'] == True].copy()
df_historico_convertidos['idade_conversao_dias'] = (
    df_historico_convertidos['Data de fechamento'] - df_historico_convertidos['Data de criação']
).dt.days

df_historico_convertidos['faixa_idade_conversao'] = df_historico_convertidos['idade_conversao_dias'].apply(classificar_idade)

# Média de atividades por faixa de idade (para leads que converteram)
media_atividades_convertidos = df_historico_convertidos.groupby('faixa_idade_conversao')['Número de atividades de vendas'].mean()

# 3.4 Total de leads por ETAPA x FAIXA DE IDADE (somando próprias e franquias)
print("\n🧩 Calculando total por etapa e faixa de idade...")

data_inicio_etapas = datetime(2025, 10, 1)
data_fim_etapas = datetime(2026, 3, 1)

df_etapas_periodo = df[
    (df['Data de criação'] >= data_inicio_etapas) &
    (df['Data de criação'] <= data_fim_etapas)
].copy()

# Idade calculada no ponto de corte solicitado (01/03/2026)
df_etapas_periodo['idade_dias_corte'] = (data_fim_etapas - df_etapas_periodo['Data de criação']).dt.days
df_etapas_periodo['faixa_idade_corte'] = df_etapas_periodo['idade_dias_corte'].apply(classificar_idade)

# Somar próprias e franquias: sem segmentação por tipo_unidade
df_etapas_periodo['Etapa'] = df_etapas_periodo['Etapa do negócio'].fillna('Sem etapa').astype(str)
# Limpar nome da etapa: remover texto entre parênteses
df_etapas_periodo['Etapa'] = df_etapas_periodo['Etapa'].str.replace(r'\s*\([^)]*\)', '', regex=True).str.strip().str.upper()

tabela_etapa_faixa = (
    df_etapas_periodo
    .groupby(['faixa_idade_corte', 'Etapa'])
    .size()
    .unstack(fill_value=0)
)

for faixa in ordem_faixas:
    if faixa not in tabela_etapa_faixa.index:
        tabela_etapa_faixa.loc[faixa] = 0

tabela_etapa_faixa = tabela_etapa_faixa.reindex(ordem_faixas, fill_value=0)

# Ordenar colunas na sequência desejada do funil
ordem_etapas_funil = [
    'NOVO NEGÓCIO',
    'EM QUALIFICAÇÃO',
    'VISITA AGENDADA',
    'VISITA REALIZADA',
    'LISTA DE ESPERA',
    'EM PAUSA',
    'MATRÍCULA REALIZADA',
    'NEGÓCIO PERDIDO'
]

# Ordenar: etapas na sequência + outras etapas alfabeticamente
colunas_ordenadas = []
for etapa in ordem_etapas_funil:
    if etapa in tabela_etapa_faixa.columns:
        colunas_ordenadas.append(etapa)

# Adicionar outras etapas que não estão na sequência (ordenadas alfabeticamente)
outras_etapas = sorted([col for col in tabela_etapa_faixa.columns if col not in ordem_etapas_funil])
colunas_ordenadas.extend(outras_etapas)

tabela_etapa_faixa = tabela_etapa_faixa[colunas_ordenadas]
tabela_etapa_faixa['Total Leads'] = tabela_etapa_faixa.sum(axis=1)
tabela_etapa_faixa = tabela_etapa_faixa.reset_index().rename(columns={'faixa_idade_corte': 'Faixa Idade'})

# Adicionar colunas de percentual para cada etapa
colunas_etapas_originais = [col for col in tabela_etapa_faixa.columns if col not in ['Faixa Idade', 'Total Leads']]
for col in colunas_etapas_originais:
    nome_pct = f'{col} %'
    tabela_etapa_faixa[nome_pct] = (tabela_etapa_faixa[col] / tabela_etapa_faixa['Total Leads'] * 100).fillna(0).round(1)

# Linha de total geral
colunas_etapas = [col for col in tabela_etapa_faixa.columns if col not in ['Faixa Idade']]
linha_total_etapas = {'Faixa Idade': 'TOTAL GERAL'}
for coluna in colunas_etapas:
    if coluna.endswith(' %'):
        # Para colunas %, recalcular com base nos totais
        col_base = coluna.replace(' %', '')
        if col_base in tabela_etapa_faixa.columns:
            total_col = tabela_etapa_faixa[col_base].sum()
            total_geral = tabela_etapa_faixa['Total Leads'].sum()
            linha_total_etapas[coluna] = round((total_col / total_geral * 100), 1) if total_geral > 0 else 0
        else:
            linha_total_etapas[coluna] = 0
    else:
        linha_total_etapas[coluna] = int(tabela_etapa_faixa[coluna].sum())
tabela_etapa_faixa = pd.concat([tabela_etapa_faixa, pd.DataFrame([linha_total_etapas])], ignore_index=True)

# ============================================================================
# 4. CRIAR TABELA CONSOLIDADA
# ============================================================================

print("\n📋 Criando tabela consolidada...")

# Criar dataframe consolidado - ANÁLISE CICLO 26.1
tabela = pd.DataFrame({
    'Canal': total_leads_por_canal.index,
    'Total Leads': total_leads_por_canal.values,
    '|': '',  # Divisória após Total Leads
})

# Definir ordem dos canais (usar a mesma ordem da tabela principal)
ordem_canais = total_leads_por_canal.index.tolist()

# Adicionar colunas intercaladas: Total, Atendidos, % + Divisória para cada faixa
for faixa in ordem_faixas:
    # Total de leads nessa faixa por canal
    if faixa in distribuicao_idade.columns:
        valores_total = []
        for canal in total_leads_por_canal.index:
            if canal in distribuicao_idade.index:
                valores_total.append(distribuicao_idade.loc[canal, faixa])
            else:
                valores_total.append(0)
        tabela[f'Total {faixa}'] = valores_total
    else:
        tabela[f'Total {faixa}'] = 0
    
    # Atendidos nessa faixa por canal
    leads_faixa = df_ciclo_26_1[df_ciclo_26_1['faixa_idade'] == faixa]
    atendidos_faixa = leads_faixa[leads_faixa['Número de atividades de vendas'] > 0].groupby('canal').size()
    
    valores_atendidos = []
    for canal in total_leads_por_canal.index:
        if canal in atendidos_faixa.index:
            valores_atendidos.append(atendidos_faixa.loc[canal])
        else:
            valores_atendidos.append(0)
    tabela[f'Atendidos {faixa}'] = valores_atendidos
    
    # Porcentagem de atendimento (como decimal para formatação correta no Excel)
    valores_percentual = []
    for i, canal in enumerate(total_leads_por_canal.index):
        total = tabela.loc[i, f'Total {faixa}']
        atendidos = tabela.loc[i, f'Atendidos {faixa}']
        if total > 0:
            valores_percentual.append(atendidos / total)  # Decimal: 0.386 = 38.6%
        else:
            valores_percentual.append(0)
    tabela[f'% {faixa}'] = valores_percentual
    
    # Adicionar divisória após cada grupo
    tabela[f'|.{faixa}'] = ''

# Adicionar linha de TOTAL
linha_total = {'Canal': 'TOTAL', 'Total Leads': total_leads_por_canal.sum(), '|': ''}

for faixa in ordem_faixas:
    # Somar totais da faixa
    if f'Total {faixa}' in tabela.columns:
        linha_total[f'Total {faixa}'] = tabela[f'Total {faixa}'].sum()
        linha_total[f'Atendidos {faixa}'] = tabela[f'Atendidos {faixa}'].sum()
        
        # Calcular % geral da faixa
        total_faixa = linha_total[f'Total {faixa}']
        atendidos_faixa = linha_total[f'Atendidos {faixa}']
        if total_faixa > 0:
            linha_total[f'% {faixa}'] = atendidos_faixa / total_faixa
        else:
            linha_total[f'% {faixa}'] = 0
    
    linha_total[f'|.{faixa}'] = ''

# Adicionar a linha de total ao dataframe
tabela = pd.concat([tabela, pd.DataFrame([linha_total])], ignore_index=True)

# ============================================================================
# SEGUNDA TABELA: RECÊNCIA DA ÚLTIMA ATIVIDADE POR FAIXA DE IDADE
# ============================================================================

print("\n📅 Criando tabela de recência de atividade...")

# Filtrar leads com data de atividade válida (não nula e não anterior à criação)
df_ciclo_26_1_valido = df_ciclo_26_1[
    (df_ciclo_26_1['Data da última atividade'].notna()) &
    (df_ciclo_26_1['Data da última atividade'] >= df_ciclo_26_1['Data de criação'])
].copy()

# Contar leads descartados por erro de preenchimento
total_leads_ciclo = len(df_ciclo_26_1)
leads_validos = len(df_ciclo_26_1_valido)
leads_descartados = total_leads_ciclo - leads_validos

print(f"   ⚠️  Leads com erro de preenchimento (data inválida): {leads_descartados:,}")
print(f"   ✅ Leads válidos para análise de recência: {leads_validos:,}")

# Calcular total de leads VÁLIDOS por canal
total_leads_validos_por_canal = df_ciclo_26_1_valido.groupby('canal').size()

# Criar tabela com mesma estrutura (usando apenas leads VÁLIDOS)
tabela_recencia = pd.DataFrame({
    'Canal': total_leads_por_canal.index,
    'Total Leads': [total_leads_validos_por_canal.get(canal, 0) for canal in total_leads_por_canal.index],
    '|': ''
})

# Para cada faixa de idade do lead
for faixa in ordem_faixas:
    # Total de leads nessa faixa de idade por canal (usando dados VÁLIDOS)
    leads_faixa = df_ciclo_26_1_valido[df_ciclo_26_1_valido['faixa_idade'] == faixa].groupby('canal').size()
    
    valores_total = []
    for canal in total_leads_por_canal.index:
        if canal in leads_faixa.index:
            valores_total.append(leads_faixa.loc[canal])
        else:
            valores_total.append(0)
    
    if len(valores_total) > 0:
        tabela_recencia[f'Total {faixa}'] = valores_total
    else:
        tabela_recencia[f'Total {faixa}'] = 0
    
    # Leads dessa faixa de idade que receberam atividade D-1 (dados VÁLIDOS)
    leads_faixa_ativ_d1 = df_ciclo_26_1_valido[
        (df_ciclo_26_1_valido['faixa_idade'] == faixa) & 
        (df_ciclo_26_1_valido['recencia_atividade'] == 'D-1')
    ].groupby('canal').size()
    
    valores_ativ_d1 = []
    for canal in total_leads_por_canal.index:
        if canal in leads_faixa_ativ_d1.index:
            valores_ativ_d1.append(leads_faixa_ativ_d1.loc[canal])
        else:
            valores_ativ_d1.append(0)
    tabela_recencia[f'Ativ D-1 {faixa}'] = valores_ativ_d1
    
    # Porcentagem com atividade D-1
    valores_percentual = []
    for i, canal in enumerate(total_leads_por_canal.index):
        total = tabela_recencia.loc[i, f'Total {faixa}']
        ativ_d1 = tabela_recencia.loc[i, f'Ativ D-1 {faixa}']
        if total > 0:
            valores_percentual.append(ativ_d1 / total)
        else:
            valores_percentual.append(0)
    tabela_recencia[f'% Ativ D-1 {faixa}'] = valores_percentual
    
    # Adicionar divisória
    tabela_recencia[f'|.{faixa}'] = ''

# Adicionar linha de TOTAL para tabela de recência (usando apenas leads VÁLIDOS)
linha_total_recencia = {'Canal': 'TOTAL', 'Total Leads': tabela_recencia['Total Leads'].sum(), '|': ''}

for faixa in ordem_faixas:
    if f'Total {faixa}' in tabela_recencia.columns:
        linha_total_recencia[f'Total {faixa}'] = tabela_recencia[f'Total {faixa}'].sum()
        linha_total_recencia[f'Ativ D-1 {faixa}'] = tabela_recencia[f'Ativ D-1 {faixa}'].sum()
        
        total_faixa = linha_total_recencia[f'Total {faixa}']
        ativ_d1_faixa = linha_total_recencia[f'Ativ D-1 {faixa}']
        if total_faixa > 0:
            linha_total_recencia[f'% Ativ D-1 {faixa}'] = ativ_d1_faixa / total_faixa
        else:
            linha_total_recencia[f'% Ativ D-1 {faixa}'] = 0
    
    linha_total_recencia[f'|.{faixa}'] = ''

tabela_recencia = pd.concat([tabela_recencia, pd.DataFrame([linha_total_recencia])], ignore_index=True)

# Adicionar linha de aviso sobre leads descartados
aviso_recencia = {
    'Canal': f'⚠️ ATENÇÃO: {leads_descartados:,} leads descartados por erro de preenchimento (data nula ou inválida)',
    'Total Leads': '',
    '|': ''
}
for faixa in ordem_faixas:
    aviso_recencia[f'Total {faixa}'] = ''
    aviso_recencia[f'Ativ D-1 {faixa}'] = ''
    aviso_recencia[f'% Ativ D-1 {faixa}'] = ''
    aviso_recencia[f'|.{faixa}'] = ''

tabela_recencia = pd.concat([tabela_recencia, pd.DataFrame([aviso_recencia])], ignore_index=True)

# Adicionar métricas de conversão por faixa
tabela_conversao = pd.DataFrame({
    'Faixa de Idade': ordem_faixas,
    'Média Atividades p/ Conversão (Histórico)': [
        media_atividades_convertidos.get(faixa, 0) for faixa in ordem_faixas
    ]
})
tabela_conversao['Média Atividades p/ Conversão (Histórico)'] = tabela_conversao['Média Atividades p/ Conversão (Histórico)'].round(1)

# ============================================================================
# TABELA DE MATRÍCULAS DO CICLO 26.1
# ============================================================================

print("\n🎓 Analisando matrículas efetivadas no ciclo 26.1...")

# Filtrar matrículas do ciclo 26.1
df_matriculas_26_1 = df_ciclo_26_1[df_ciclo_26_1['converteu'] == True].copy()

print(f"   Total de matrículas no ciclo 26.1: {len(df_matriculas_26_1):,}")

# Para matrículas, calcular idade no momento da conversão
df_matriculas_26_1['data_conversao'] = df_matriculas_26_1['Data de fechamento'].fillna(hoje)
df_matriculas_26_1['idade_na_conversao'] = (df_matriculas_26_1['data_conversao'] - df_matriculas_26_1['Data de criação']).dt.days

# NÃO filtrar dias negativos - eles evidenciam problemas na gestão do CRM!

# Diagnóstico: Analisar leads com idade negativa
matriculas_negativas = df_matriculas_26_1[df_matriculas_26_1['idade_na_conversao'] < 0].copy()
if len(matriculas_negativas) > 0:
    print(f"\n⚠️  ALERTA: {len(matriculas_negativas)} matrículas com idade negativa detectadas!")
    print("\n📊 Distribuição por canal:")
    print(matriculas_negativas['canal'].value_counts())
    print(f"\n📉 Média de dias negativos: {matriculas_negativas['idade_na_conversao'].mean():.1f} dias")
    print(f"   Pior caso: {matriculas_negativas['idade_na_conversao'].min()} dias")
    
    # Separar canais digitais (automáticos) vs manuais
    canais_digitais = ['Social pago', 'Social orgânico', 'Pesquisa paga', 'Pesquisa orgânica']
    digitais_negativos = matriculas_negativas[matriculas_negativas['canal'].isin(canais_digitais)]
    manuais_negativos = matriculas_negativas[~matriculas_negativas['canal'].isin(canais_digitais)]
    
    print(f"\n🤖 CANAIS DIGITAIS (automáticos) com idade negativa: {len(digitais_negativos)}")
    if len(digitais_negativos) > 0:
        print("   ⚠️  ISSO NÃO DEVERIA ACONTECER - evidência de retroatividade manual!")
        print(digitais_negativos['canal'].value_counts().to_string())
        
    print(f"\n👤 CANAIS MANUAIS com idade negativa: {len(manuais_negativos)}")
    if len(manuais_negativos) > 0:
        print(manuais_negativos['canal'].value_counts().to_string())
    
    # Mostrar exemplos de canais DIGITAIS com problema
    if len(digitais_negativos) > 0:
        print("\n🚨 Exemplos de CANAIS DIGITAIS com retroatividade:")
        exemplos_dig = digitais_negativos.nsmallest(5, 'idade_na_conversao')[['canal', 'Data de criação', 'data_conversao', 'idade_na_conversao']]
        for idx, row in exemplos_dig.iterrows():
            print(f"   • {row['canal']}: Lead criado {row['Data de criação'].strftime('%d/%m/%Y')}, mas convertido {row['data_conversao'].strftime('%d/%m/%Y')} → {row['idade_na_conversao']} dias")

# Classificar idade na conversão
df_matriculas_26_1['faixa_idade_conversao'] = df_matriculas_26_1['idade_na_conversao'].apply(classificar_idade)

# Criar tabela de matrículas por canal e faixa de idade
matriculas_por_canal = df_matriculas_26_1.groupby('canal').size()

# Calcular média de idade para matrículas com data válida
matriculas_com_idade = df_matriculas_26_1[
    (df_matriculas_26_1['Data da última atividade'].notna()) &
    (df_matriculas_26_1['Data da última atividade'] >= df_matriculas_26_1['Data de criação'])
].copy()

matriculas_com_idade['idade_real_dias'] = (
    matriculas_com_idade['Data da última atividade'] - 
    matriculas_com_idade['Data de criação']
).dt.days

# Para TABELA 1: usar idade oficial (data de conversão - data de criação)
# Idade na conversão já foi calculada anteriormente como 'idade_na_conversao'

tabela_matriculas = pd.DataFrame({
    'Canal': matriculas_por_canal.index,
    'Total Matrículas': matriculas_por_canal.values,
    'Média Geral Ativ': [df_matriculas_26_1[df_matriculas_26_1['canal'] == canal]['Número de atividades de vendas'].mean() for canal in matriculas_por_canal.index],
    'Média Idade (dias)': [df_matriculas_26_1[df_matriculas_26_1['canal'] == canal]['idade_na_conversao'].mean() for canal in matriculas_por_canal.index],
    '|': ''
})

# Para cada faixa de idade
for faixa in ordem_faixas:
    # Total de matrículas nessa faixa por canal
    matriculas_faixa = df_matriculas_26_1[df_matriculas_26_1['faixa_idade_conversao'] == faixa].groupby('canal').size()
    
    valores_total = []
    for canal in matriculas_por_canal.index:
        if canal in matriculas_faixa.index:
            valores_total.append(matriculas_faixa.loc[canal])
        else:
            valores_total.append(0)
    
    if len(valores_total) > 0:
        tabela_matriculas[f'Qtd {faixa}'] = valores_total
    else:
        tabela_matriculas[f'Qtd {faixa}'] = 0
    
    # Média de atividades para essa faixa
    matriculas_faixa_data = df_matriculas_26_1[df_matriculas_26_1['faixa_idade_conversao'] == faixa]
    
    valores_media_ativ = []
    for canal in matriculas_por_canal.index:
        matriculas_canal_faixa = matriculas_faixa_data[matriculas_faixa_data['canal'] == canal]
        if len(matriculas_canal_faixa) > 0:
            media = matriculas_canal_faixa['Número de atividades de vendas'].mean()
            valores_media_ativ.append(media)
        else:
            valores_media_ativ.append(0)
    
    tabela_matriculas[f'Média Ativ {faixa}'] = valores_media_ativ
    
    # Porcentagem de matrículas nessa faixa
    valores_percentual = []
    for i, canal in enumerate(matriculas_por_canal.index):
        total = tabela_matriculas.loc[i, 'Total Matrículas']
        qtd_faixa = tabela_matriculas.loc[i, f'Qtd {faixa}']
        if total > 0:
            valores_percentual.append(qtd_faixa / total)
        else:
            valores_percentual.append(0)
    tabela_matriculas[f'% {faixa}'] = valores_percentual
    
    # Adicionar divisória
    tabela_matriculas[f'|.{faixa}'] = ''

# Adicionar linha de TOTAL
linha_total_matriculas = {
    'Canal': 'TOTAL', 
    'Total Matrículas': matriculas_por_canal.sum(), 
    'Média Geral Ativ': df_matriculas_26_1['Número de atividades de vendas'].mean(),
    'Média Idade (dias)': df_matriculas_26_1['idade_na_conversao'].mean(),
    '|': ''
}

for faixa in ordem_faixas:
    if f'Qtd {faixa}' in tabela_matriculas.columns:
        linha_total_matriculas[f'Qtd {faixa}'] = tabela_matriculas[f'Qtd {faixa}'].sum()
        
        # Média geral de atividades para essa faixa
        matriculas_faixa_geral = df_matriculas_26_1[df_matriculas_26_1['faixa_idade_conversao'] == faixa]
        if len(matriculas_faixa_geral) > 0:
            linha_total_matriculas[f'Média Ativ {faixa}'] = matriculas_faixa_geral['Número de atividades de vendas'].mean()
        else:
            linha_total_matriculas[f'Média Ativ {faixa}'] = 0
        
        # Percentual geral
        total_mat = linha_total_matriculas['Total Matrículas']
        qtd_faixa = linha_total_matriculas[f'Qtd {faixa}']
        if total_mat > 0:
            linha_total_matriculas[f'% {faixa}'] = qtd_faixa / total_mat
        else:
            linha_total_matriculas[f'% {faixa}'] = 0
    
    linha_total_matriculas[f'|.{faixa}'] = ''

tabela_matriculas = pd.concat([tabela_matriculas, pd.DataFrame([linha_total_matriculas])], ignore_index=True)

# ============================================================================
# TABELA DE IDADE REAL (BASEADA NA ÚLTIMA ATIVIDADE)
# ============================================================================

print("\n🔍 Analisando idade REAL dos leads (baseada na última atividade)...")

# Filtrar matrículas com data de última atividade válida
df_matriculas_idade_real = df_matriculas_26_1[
    (df_matriculas_26_1['Data da última atividade'].notna()) &
    (df_matriculas_26_1['Data da última atividade'] >= df_matriculas_26_1['Data de criação'])
].copy()

matriculas_descartadas = len(df_matriculas_26_1) - len(df_matriculas_idade_real)

print(f"   ⚠️  Matrículas descartadas por erro de preenchimento: {matriculas_descartadas:,}")
print(f"   ✅ Matrículas válidas para análise: {len(df_matriculas_idade_real):,}")

# Calcular idade REAL (diferença entre última atividade e criação)
df_matriculas_idade_real['idade_real_dias'] = (
    df_matriculas_idade_real['Data da última atividade'] - 
    df_matriculas_idade_real['Data de criação']
).dt.days

# Classificar em faixas
df_matriculas_idade_real['faixa_idade_real'] = df_matriculas_idade_real['idade_real_dias'].apply(classificar_idade)

# Criar tabela
matriculas_por_canal_real = df_matriculas_idade_real.groupby('canal').size()

tabela_idade_real = pd.DataFrame({
    'Canal': matriculas_por_canal_real.index,
    'Total Matrículas': matriculas_por_canal_real.values,
    'Média Geral Ativ': [df_matriculas_idade_real[df_matriculas_idade_real['canal'] == canal]['Número de atividades de vendas'].mean() for canal in matriculas_por_canal_real.index],
    'Média Idade (dias)': [df_matriculas_idade_real[df_matriculas_idade_real['canal'] == canal]['idade_real_dias'].mean() for canal in matriculas_por_canal_real.index],
    '|': ''
})

# Para cada faixa de idade real
for faixa in ordem_faixas:
    # Total de matrículas nessa faixa por canal
    matriculas_faixa = df_matriculas_idade_real[df_matriculas_idade_real['faixa_idade_real'] == faixa].groupby('canal').size()
    
    valores_total = []
    for canal in matriculas_por_canal_real.index:
        if canal in matriculas_faixa.index:
            valores_total.append(matriculas_faixa.loc[canal])
        else:
            valores_total.append(0)
    
    if len(valores_total) > 0:
        tabela_idade_real[f'Qtd {faixa}'] = valores_total
    else:
        tabela_idade_real[f'Qtd {faixa}'] = 0
    
    # Média de atividades para essa faixa
    matriculas_faixa_data = df_matriculas_idade_real[df_matriculas_idade_real['faixa_idade_real'] == faixa]
    
    valores_media_ativ = []
    for canal in matriculas_por_canal_real.index:
        matriculas_canal_faixa = matriculas_faixa_data[matriculas_faixa_data['canal'] == canal]
        if len(matriculas_canal_faixa) > 0:
            media = matriculas_canal_faixa['Número de atividades de vendas'].mean()
            valores_media_ativ.append(media)
        else:
            valores_media_ativ.append(0)
    
    tabela_idade_real[f'Média Ativ {faixa}'] = valores_media_ativ
    
    # Porcentagem de matrículas nessa faixa
    valores_percentual = []
    for i, canal in enumerate(matriculas_por_canal_real.index):
        total = tabela_idade_real.loc[i, 'Total Matrículas']
        qtd_faixa = tabela_idade_real.loc[i, f'Qtd {faixa}']
        if total > 0:
            valores_percentual.append(qtd_faixa / total)
        else:
            valores_percentual.append(0)
    tabela_idade_real[f'% {faixa}'] = valores_percentual
    
    # Adicionar divisória
    tabela_idade_real[f'|.{faixa}'] = ''

# Adicionar linha de TOTAL
linha_total_idade_real = {
    'Canal': 'TOTAL', 
    'Total Matrículas': matriculas_por_canal_real.sum(), 
    'Média Geral Ativ': df_matriculas_idade_real['Número de atividades de vendas'].mean(),
    'Média Idade (dias)': df_matriculas_idade_real['idade_real_dias'].mean(),
    '|': ''
}

for faixa in ordem_faixas:
    if f'Qtd {faixa}' in tabela_idade_real.columns:
        linha_total_idade_real[f'Qtd {faixa}'] = tabela_idade_real[f'Qtd {faixa}'].sum()
        
        # Média geral de atividades para essa faixa
        matriculas_faixa_geral = df_matriculas_idade_real[df_matriculas_idade_real['faixa_idade_real'] == faixa]
        if len(matriculas_faixa_geral) > 0:
            linha_total_idade_real[f'Média Ativ {faixa}'] = matriculas_faixa_geral['Número de atividades de vendas'].mean()
        else:
            linha_total_idade_real[f'Média Ativ {faixa}'] = 0
        
        # Percentual geral
        total_mat = linha_total_idade_real['Total Matrículas']
        qtd_faixa = linha_total_idade_real[f'Qtd {faixa}']
        if total_mat > 0:
            linha_total_idade_real[f'% {faixa}'] = qtd_faixa / total_mat
        else:
            linha_total_idade_real[f'% {faixa}'] = 0
    
    linha_total_idade_real[f'|.{faixa}'] = ''

tabela_idade_real = pd.concat([tabela_idade_real, pd.DataFrame([linha_total_idade_real])], ignore_index=True)

# Adicionar linha de aviso
aviso_idade_real = {
    'Canal': f'⚠️ ATENÇÃO: {matriculas_descartadas:,} matrículas descartadas por erro de preenchimento (data nula ou inválida)',
    'Total Matrículas': '',
    '|': ''
}
for faixa in ordem_faixas:
    aviso_idade_real[f'Qtd {faixa}'] = ''
    aviso_idade_real[f'Média Ativ {faixa}'] = ''
    aviso_idade_real[f'% {faixa}'] = ''
    aviso_idade_real[f'|.{faixa}'] = ''

tabela_idade_real = pd.concat([tabela_idade_real, pd.DataFrame([aviso_idade_real])], ignore_index=True)

# ============================================================================
# CRIAR TABELAS POR UNIDADE + CANAL - VERSÃO BASE COMPLETA
# ============================================================================

print("\n📍 Preparando análise por unidade e canal (base completa)...")

# Análise de LEADS por UNIDADE + CANAL (BASE COMPLETA)
unidades_leads_completo_lista = []

# Obter todas unidades únicas
unidades_unicas = sorted(df_ciclo_26_1['Unidade Desejada'].dropna().unique())

for unidade in unidades_unicas:
    df_unidade = df_ciclo_26_1[df_ciclo_26_1['Unidade Desejada'] == unidade]
    
    # Agrupar por canal dentro da unidade - MOSTRAR TODOS OS CANAIS
    for canal in ordem_canais:
        df_canal = df_unidade[df_unidade['canal'] == canal]
        
        total_leads = len(df_canal)
        atendidos = (df_canal['Número de atividades de vendas'] > 0).sum()
        pct_atendimento = atendidos / total_leads if total_leads > 0 else 0
        media_ativ = df_canal['Número de atividades de vendas'].mean() if total_leads > 0 else 0
        
        # Tipo de unidade
        tipo = df_unidade['tipo_unidade'].iloc[0] if len(df_unidade) > 0 else 'N/A'
        
        linha = {
            'Unidade': unidade,
            'Tipo': tipo,
            'Canal': canal,
            'Total Leads': total_leads,
            'Atendidos': atendidos,
            '% Atendimento': pct_atendimento,
            '| ': '',
            'Média Ativ': media_ativ,
            '|  ': '',
        }
        
        # Adicionar faixas com total, atendidos e % + divisores
        for i, faixa in enumerate(ordem_faixas):
            df_faixa = df_canal[df_canal['faixa_idade'] == faixa]
            total_faixa = len(df_faixa)
            atendidos_faixa = (df_faixa['Número de atividades de vendas'] > 0).sum()
            pct_faixa = atendidos_faixa / total_faixa if total_faixa > 0 else 0
            
            linha[f'{faixa} Total'] = total_faixa
            linha[f'{faixa} Atend'] = atendidos_faixa
            linha[f'{faixa} %'] = pct_faixa
            if i < len(ordem_faixas) - 1:
                linha[f'|{i+3}'] = ''
        
        unidades_leads_completo_lista.append(linha)
    
    # Adicionar linha de TOTAL por unidade
    total_leads_unidade = len(df_unidade)
    atendidos_unidade = (df_unidade['Número de atividades de vendas'] > 0).sum()
    pct_atend_unidade = atendidos_unidade / total_leads_unidade if total_leads_unidade > 0 else 0
    media_ativ_unidade = df_unidade['Número de atividades de vendas'].mean()
    
    tipo_unidade_total = df_unidade['tipo_unidade'].iloc[0] if len(df_unidade) > 0 else 'N/A'
    
    linha_total = {
        'Unidade': unidade,
        'Tipo': tipo_unidade_total,
        'Canal': f'TOTAL {unidade}',
        'Total Leads': total_leads_unidade,
        'Atendidos': atendidos_unidade,
        '% Atendimento': pct_atend_unidade,
        '| ': '',
        'Média Ativ': media_ativ_unidade,
        '|  ': '',
    }
    
    for i, faixa in enumerate(ordem_faixas):
        df_faixa = df_unidade[df_unidade['faixa_idade'] == faixa]
        total_faixa = len(df_faixa)
        atendidos_faixa = (df_faixa['Número de atividades de vendas'] > 0).sum()
        pct_faixa = atendidos_faixa / total_faixa if total_faixa > 0 else 0
        media_ativ_faixa = df_faixa['Número de atividades de vendas'].mean() if total_faixa > 0 else 0
        
        linha_total[f'{faixa} Total'] = total_faixa
        linha_total[f'{faixa} Atend'] = atendidos_faixa
        linha_total[f'{faixa} %'] = pct_faixa
        if i < len(ordem_faixas) - 1:
            linha_total[f'|{i+3}'] = ''
    
    unidades_leads_completo_lista.append(linha_total)

unidades_leads_completo = pd.DataFrame(unidades_leads_completo_lista)

# Análise de MATRÍCULAS por UNIDADE + CANAL (BASE COMPLETA)
unidades_matriculas_completo_lista = []

for unidade in unidades_unicas:
    df_unidade_mat = df_matriculas_26_1[df_matriculas_26_1['Unidade Desejada'] == unidade]
    df_unidade_leads = df_ciclo_26_1[df_ciclo_26_1['Unidade Desejada'] == unidade]
    
    # Total de leads da unidade para taxa de conversão
    total_leads_unidade = len(df_unidade_leads)
    
    if total_leads_unidade == 0:
        continue
    
    # Tipo de unidade
    tipo = df_unidade_leads['tipo_unidade'].iloc[0] if len(df_unidade_leads) > 0 else 'N/A'
    
    # Agrupar por canal dentro da unidade - MOSTRAR TODOS OS CANAIS
    for canal in ordem_canais:
        df_canal = df_unidade_mat[df_unidade_mat['canal'] == canal]
        
        total_matriculas = len(df_canal)
        media_ativ = df_canal['Número de atividades de vendas'].mean() if total_matriculas > 0 else 0
        
        # Taxa de conversão do canal na unidade
        total_leads_canal_unidade = len(df_unidade_leads[df_unidade_leads['canal'] == canal])
        taxa_conv = (total_matriculas / total_leads_canal_unidade * 100) if total_leads_canal_unidade > 0 else 0
        
        # Média de idade (em dias) dos leads que converteram
        media_idade_dias = df_canal['idade_na_conversao'].mean() if total_matriculas > 0 else 0
        
        linha = {
            'Unidade': unidade,
            'Tipo': tipo,
            'Canal': canal,
            'Total Matriculas': total_matriculas,
            'Taxa Conversão %': taxa_conv / 100,  # Dividir por 100 para formato de %
            '| ': '',
            'Média Ativ': media_ativ,
            'Média Idade (dias)': media_idade_dias,
            '|  ': '',
        }
        
        # Adicionar faixas com Qtd, Média Ativ e % (estrutura igual tabela de matrículas)
        for i, faixa in enumerate(ordem_faixas):
            df_faixa = df_canal[df_canal['faixa_idade_conversao'] == faixa]
            total_faixa = len(df_faixa)
            media_ativ_faixa = df_faixa['Número de atividades de vendas'].mean() if total_faixa > 0 else 0
            pct_faixa = total_faixa / total_matriculas if total_matriculas > 0 else 0
            
            linha[f'Qtd {faixa}'] = total_faixa
            linha[f'Média Ativ {faixa}'] = media_ativ_faixa
            linha[f'% {faixa}'] = pct_faixa
            if i < len(ordem_faixas) - 1:
                linha[f'|{i+3}'] = ''
        
        unidades_matriculas_completo_lista.append(linha)
    
    # Adicionar linha de TOTAL por unidade
    total_matriculas_unidade = len(df_unidade_mat)
    media_ativ_unidade = df_unidade_mat['Número de atividades de vendas'].mean() if total_matriculas_unidade > 0 else 0
    taxa_conv_unidade = (total_matriculas_unidade / total_leads_unidade * 100) if total_leads_unidade > 0 else 0
    media_idade_unidade = df_unidade_mat['idade_na_conversao'].mean() if total_matriculas_unidade > 0 else 0
    
    linha_total = {
        'Unidade': unidade,
        'Tipo': tipo,
        'Canal': f'TOTAL {unidade}',
        'Total Matriculas': total_matriculas_unidade,
        'Taxa Conversão %': taxa_conv_unidade / 100,  # Dividir por 100 para formato de %
        '| ': '',
        'Média Ativ': media_ativ_unidade,
        'Média Idade (dias)': media_idade_unidade,
        '|  ': '',
    }
    
    for i, faixa in enumerate(ordem_faixas):
        df_faixa = df_unidade_mat[df_unidade_mat['faixa_idade_conversao'] == faixa]
        total_faixa = len(df_faixa)
        media_ativ_faixa = df_faixa['Número de atividades de vendas'].mean() if total_faixa > 0 else 0
        pct_faixa = total_faixa / total_matriculas_unidade if total_matriculas_unidade > 0 else 0
        
        linha_total[f'Qtd {faixa}'] = total_faixa
        linha_total[f'Média Ativ {faixa}'] = media_ativ_faixa
        linha_total[f'% {faixa}'] = pct_faixa
        if i < len(ordem_faixas) - 1:
            linha_total[f'|{i+3}'] = ''
    
    unidades_matriculas_completo_lista.append(linha_total)

unidades_matriculas_completo = pd.DataFrame(unidades_matriculas_completo_lista)

# ============================================================================
# CRIAR TABELAS POR UNIDADE + CANAL - VERSÃO BASE FILTRADA
# ============================================================================

print("\n📍 Preparando análise por unidade e canal (base filtrada)...")

# Filtrar leads com dados válidos para análise
df_ciclo_valido = df_ciclo_26_1[
    (df_ciclo_26_1['Data da última atividade'].notna()) &
    (df_ciclo_26_1['Data da última atividade'] >= df_ciclo_26_1['Data de criação'])
].copy()

# Filtrar matrículas com dados válidos
df_matriculas_valido = df_matriculas_26_1[
    (df_matriculas_26_1['Data da última atividade'].notna()) &
    (df_matriculas_26_1['Data da última atividade'] >= df_matriculas_26_1['Data de criação'])
].copy()

# Calcular idade REAL (diferença entre última atividade e criação) para matrículas filtradas
df_matriculas_valido['idade_real_dias'] = (
    df_matriculas_valido['Data da última atividade'] - 
    df_matriculas_valido['Data de criação']
).dt.days

# Classificar em faixas de idade REAL
df_matriculas_valido['faixa_idade_real'] = df_matriculas_valido['idade_real_dias'].apply(classificar_idade)

print(f"   Leads válidos: {len(df_ciclo_valido):,} de {len(df_ciclo_26_1):,} ({len(df_ciclo_valido)/len(df_ciclo_26_1)*100:.1f}%)")
print(f"   Matrículas válidas: {len(df_matriculas_valido):,} de {len(df_matriculas_26_1):,} ({len(df_matriculas_valido)/len(df_matriculas_26_1)*100:.1f}%)")

# Análise de LEADS por UNIDADE + CANAL (BASE FILTRADA)
unidades_leads_filtrado_lista = []

for unidade in unidades_unicas:
    df_unidade = df_ciclo_valido[df_ciclo_valido['Unidade Desejada'] == unidade]
    
    # Agrupar por canal dentro da unidade - MOSTRAR TODOS OS CANAIS
    for canal in ordem_canais:
        df_canal = df_unidade[df_unidade['canal'] == canal]
        
        total_leads = len(df_canal)
        atendidos = (df_canal['Número de atividades de vendas'] > 0).sum()
        pct_atendimento = atendidos / total_leads if total_leads > 0 else 0
        media_ativ = df_canal['Número de atividades de vendas'].mean() if total_leads > 0 else 0
        
        # Tipo de unidade
        tipo = df_unidade['tipo_unidade'].iloc[0] if len(df_unidade) > 0 else 'N/A'
        
        linha = {
            'Unidade': unidade,
            'Tipo': tipo,
            'Canal': canal,
            'Total Leads': total_leads,
            'Atendidos': atendidos,
            '% Atendimento': pct_atendimento,
            '| ': '',
            'Média Ativ': media_ativ,
            '|  ': '',
        }
        
        # Adicionar faixas com Qtd, Média Ativ e % (estrutura igual tabela de matrículas)
        for i, faixa in enumerate(ordem_faixas):
            df_faixa = df_canal[df_canal['faixa_idade'] == faixa]
            total_faixa = len(df_faixa)
            atendidos_faixa = (df_faixa['Número de atividades de vendas'] > 0).sum()
            pct_faixa = atendidos_faixa / total_faixa if total_faixa > 0 else 0
            media_ativ_faixa = df_faixa['Número de atividades de vendas'].mean() if total_faixa > 0 else 0
            
            linha[f'Qtd {faixa}'] = total_faixa
            linha[f'Média Ativ {faixa}'] = media_ativ_faixa
            linha[f'% {faixa}'] = pct_faixa
            if i < len(ordem_faixas) - 1:
                linha[f'|{i+3}'] = ''
        
        unidades_leads_filtrado_lista.append(linha)
    
    # Adicionar linha de TOTAL por unidade
    total_leads_unidade = len(df_unidade)
    if total_leads_unidade == 0:
        continue
        
    atendidos_unidade = (df_unidade['Número de atividades de vendas'] > 0).sum()
    pct_atend_unidade = atendidos_unidade / total_leads_unidade if total_leads_unidade > 0 else 0
    media_ativ_unidade = df_unidade['Número de atividades de vendas'].mean()
    
    tipo_unidade_total = df_unidade['tipo_unidade'].iloc[0] if len(df_unidade) > 0 else 'N/A'
    
    linha_total = {
        'Unidade': unidade,
        'Tipo': tipo_unidade_total,
        'Canal': f'TOTAL {unidade}',
        'Total Leads': total_leads_unidade,
        'Atendidos': atendidos_unidade,
        '% Atendimento': pct_atend_unidade,
        '| ': '',
        'Média Ativ': media_ativ_unidade,
        '|  ': '',
    }
    
    for i, faixa in enumerate(ordem_faixas):
        df_faixa = df_unidade[df_unidade['faixa_idade'] == faixa]
        total_faixa = len(df_faixa)
        atendidos_faixa = (df_faixa['Número de atividades de vendas'] > 0).sum()
        pct_faixa = atendidos_faixa / total_faixa if total_faixa > 0 else 0
        media_ativ_faixa = df_faixa['Número de atividades de vendas'].mean() if total_faixa > 0 else 0
        
        linha_total[f'Qtd {faixa}'] = total_faixa
        linha_total[f'Média Ativ {faixa}'] = media_ativ_faixa
        linha_total[f'% {faixa}'] = pct_faixa
        if i < len(ordem_faixas) - 1:
            linha_total[f'|{i+3}'] = ''
    
    unidades_leads_filtrado_lista.append(linha_total)

unidades_leads_filtrado = pd.DataFrame(unidades_leads_filtrado_lista)

# Análise de MATRÍCULAS por UNIDADE + CANAL (BASE FILTRADA)
unidades_matriculas_filtrado_lista = []

for unidade in unidades_unicas:
    df_unidade_mat = df_matriculas_valido[df_matriculas_valido['Unidade Desejada'] == unidade]
    df_unidade_leads = df_ciclo_valido[df_ciclo_valido['Unidade Desejada'] == unidade]
    
    # Total de leads da unidade para taxa de conversão
    total_leads_unidade = len(df_unidade_leads)
    
    if total_leads_unidade == 0:
        continue
    
    # Tipo de unidade
    tipo = df_unidade_leads['tipo_unidade'].iloc[0] if len(df_unidade_leads) > 0 else 'N/A'
    
    # Agrupar por canal dentro da unidade - MOSTRAR TODOS OS CANAIS
    for canal in ordem_canais:
        df_canal = df_unidade_mat[df_unidade_mat['canal'] == canal]
        
        total_matriculas = len(df_canal)
        media_ativ = df_canal['Número de atividades de vendas'].mean() if total_matriculas > 0 else 0
        
        # Taxa de conversão do canal na unidade
        total_leads_canal_unidade = len(df_unidade_leads[df_unidade_leads['canal'] == canal])
        taxa_conv = (total_matriculas / total_leads_canal_unidade * 100) if total_leads_canal_unidade > 0 else 0
        
        # Média de idade (em dias) dos leads que converteram (idade REAL)
        media_idade_dias = df_canal['idade_real_dias'].mean() if total_matriculas > 0 else 0
        
        linha = {
            'Unidade': unidade,
            'Tipo': tipo,
            'Canal': canal,
            'Total Matriculas': total_matriculas,
            'Taxa Conversão %': taxa_conv / 100,
            '| ': '',
            'Média Ativ': media_ativ,
            'Média Idade (dias)': media_idade_dias,
            '|  ': '',
        }
        
        # Adicionar faixas com Qtd, Média Ativ e % (usando faixa_idade_real)
        for i, faixa in enumerate(ordem_faixas):
            df_faixa = df_canal[df_canal['faixa_idade_real'] == faixa]
            total_faixa = len(df_faixa)
            media_ativ_faixa = df_faixa['Número de atividades de vendas'].mean() if total_faixa > 0 else 0
            pct_faixa = total_faixa / total_matriculas if total_matriculas > 0 else 0
            
            linha[f'Qtd {faixa}'] = total_faixa
            linha[f'Média Ativ {faixa}'] = media_ativ_faixa
            linha[f'% {faixa}'] = pct_faixa
            if i < len(ordem_faixas) - 1:
                linha[f'|{i+3}'] = ''
        
        unidades_matriculas_filtrado_lista.append(linha)
    
    # Adicionar linha de TOTAL por unidade
    total_matriculas_unidade = len(df_unidade_mat)
    media_ativ_unidade = df_unidade_mat['Número de atividades de vendas'].mean() if total_matriculas_unidade > 0 else 0
    taxa_conv_unidade = (total_matriculas_unidade / total_leads_unidade * 100) if total_leads_unidade > 0 else 0
    media_idade_unidade = df_unidade_mat['idade_real_dias'].mean() if total_matriculas_unidade > 0 else 0
    
    linha_total = {
        'Unidade': unidade,
        'Tipo': tipo,
        'Canal': f'TOTAL {unidade}',
        'Total Matriculas': total_matriculas_unidade,
        'Taxa Conversão %': taxa_conv_unidade / 100,
        '| ': '',
        'Média Ativ': media_ativ_unidade,
        'Média Idade (dias)': media_idade_unidade,
        '|  ': '',
    }
    
    for i, faixa in enumerate(ordem_faixas):
        df_faixa = df_unidade_mat[df_unidade_mat['faixa_idade_real'] == faixa]
        total_faixa = len(df_faixa)
        media_ativ_faixa = df_faixa['Número de atividades de vendas'].mean() if total_faixa > 0 else 0
        pct_faixa = total_faixa / total_matriculas_unidade if total_matriculas_unidade > 0 else 0
        
        linha_total[f'Qtd {faixa}'] = total_faixa
        linha_total[f'Média Ativ {faixa}'] = media_ativ_faixa
        linha_total[f'% {faixa}'] = pct_faixa
        if i < len(ordem_faixas) - 1:
            linha_total[f'|{i+3}'] = ''
    
    unidades_matriculas_filtrado_lista.append(linha_total)

unidades_matriculas_filtrado = pd.DataFrame(unidades_matriculas_filtrado_lista)


# ============================================================================
# 5. EXPORTAR PARA EXCEL COM FORMATAÇÃO BONITA
# ============================================================================

print("\n💾 Exportando para Excel...")

PASTA_OUTPUT = r'C:\Users\a483650\Projetos\06_Analise_Funil_RedBalloon\outputs'
arquivo_saida = os.path.join(PASTA_OUTPUT, 'Analise_Ciclo_26_1.xlsx')
arquivo_unidades = os.path.join(PASTA_OUTPUT, 'Analise_Ciclo_26_1_POR_UNIDADE.xlsx')
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)
arquivo_saida = OUTPUT_DIR / 'Analise_Ciclo_26_1.xlsx'
arquivo_unidades = OUTPUT_DIR / 'Analise_Ciclo_26_1_POR_UNIDADE.xlsx'
# Criar conteúdo da aba Resumo
resumo_conteudo = [
    ['ANÁLISE DO CICLO 26.1 - RED BALLOON', ''],
    ['Data de extração HubSpot: 12/Fev/2026', ''],
    ['', ''],
    ['DESCRIÇÃO DAS ABAS', ''],
    ['', ''],
    ['ABA: Ciclo 26.1', ''],
    ['  • TABELA 1 - ATENDIMENTO POR FAIXA DE IDADE', 'Mostra quantos leads de cada idade foram atendidos alguma vez'],
    ['    - Total: Total de leads naquela faixa de idade (por canal)', ''],
    ['    - Atendidos: Leads com Número de atividades > 0', ''],
    ['    - %: Percentual de atendimento', ''],
    ['', ''],
    ['  • TABELA 2 - RECÊNCIA DA ÚLTIMA ATIVIDADE (D-1)', 'Mostra quantos leads receberam atividade D-1 (ontem/hoje)'],
    ['    - Total: Total de leads válidos (com data de atividade preenchida corretamente)', ''],
    ['    - Ativ D-1: Leads cuja última atividade foi há ≤1 dia', ''],
    ['    - %: Percentual com atividade recente', ''],
    ['    - ATENÇÃO: Leads com erro de preenchimento foram descartados', ''],
    ['', ''],
    ['  • TABELA 3 - MÉDIA DE ATIVIDADES HISTÓRICA', 'Quantas atividades são necessárias para converter (dados históricos)'],
    ['', ''],
    ['', ''],
    ['ABA: Matrículas 26.1', ''],
    ['  • TABELA 1 - MATRÍCULAS POR IDADE NA CONVERSÃO', 'Distribuição das matrículas por idade do lead no momento da conversão'],
    ['    - Qtd: Quantidade de matrículas naquela faixa', ''],
    ['    - Média Ativ: Média de atividades comerciais registradas', ''],
    ['    - Média Idade (dias): Idade OFICIAL do lead até a conversão (todas as matrículas)', ''],
    ['    - %: Percentual das matrículas totais', ''],
    ['', ''],
    ['  • TABELA 2 - IDADE REAL (baseada na última atividade)', 'Diferença entre data da última atividade e data de criação'],
    ['    - Propósito: Descobrir a idade REAL trabalhada do lead', ''],
    ['    - Média Idade (dias): Idade até ÚLTIMA ATIVIDADE (apenas matrículas com dados válidos)', ''],
    ['    - Problema: Muitos leads são criados no HubSpot apenas no momento da matrícula', ''],
    ['    - Exemplo: Lead atendido 15 dias offline → criado dia 18 → matrícula dia 18 = D-1 "fantasma"', ''],
    ['    - ATENÇÃO: Matrículas sem data de atividade válida foram descartadas', ''],
    ['', ''],
    ['', ''],
    ['DIFERENÇAS IMPORTANTES', ''],
    ['', ''],
    ['ATENDIDOS vs ATIV D-1 (Aba Ciclo 26.1)', ''],
    ['', ''],
    ['ATENDIDOS D-1:', ''],
    ['  • Leads com idade ≤1 dia que TEM atividades registradas', ''],
    ['  • Baseado no contador "Número de atividades" (campo confiável)', ''],
    ['  • Não importa QUANDO a atividade aconteceu, apenas SE aconteceu', ''],
    ['', ''],
    ['ATIV D-1 D-1:', ''],
    ['  • Leads com idade ≤1 dia cuja ÚLTIMA atividade foi há ≤1 dia', ''],
    ['  • Baseado no campo "Data da última atividade" (tem erros de preenchimento)', ''],
    ['  • Mostra se o lead foi atendido RECENTEMENTE (D-1)', ''],
    ['', ''],
    ['Por que são diferentes?', ''],
    ['  • Campo "Data da última atividade" tem dados nulos ou inválidos', ''],
    ['  • Alguns leads têm atividades mas sem data registrada', ''],
    ['  • Alguns leads têm data de atividade ANTERIOR à criação (impossível)', ''],
    ['', ''],
    ['', ''],
    ['LIMITAÇÕES DA ANÁLISE', ''],
    ['', ''],
    ['Qualidade de Dados:', ''],
    ['  • ~2.200 leads com campo "Data da última atividade" inválido (19% dos leads)', ''],
    ['  • ~1.400 matrículas sem data de atividade válida (50% das matrículas)', ''],
    ['', ''],
    ['Leads "Fantasma":', ''],
    ['  • 66% das matrículas são classificadas como D-1 (idade oficial)', ''],
    ['  • Mas apenas 31% têm trabalho comercial real D-1', ''],
    ['  • Diferença = leads criados no sistema apenas no momento da matrícula', ''],
    ['  • Trabalho comercial OFFLINE não é capturado', ''],
    ['', ''],
    ['Trabalho Offline:', ''],
    ['  • Leads walk-in que matriculam na hora', ''],
    ['  • Indicações de alunos que chegam decididos', ''],
    ['  • Atendimentos por WhatsApp/telefone sem registro prévio no CRM', ''],
    ['', ''],
    ['', ''],
    ['ANÁLISE POR UNIDADE (ARQUIVO SEPARADO)', ''],
    ['', ''],
    ['Arquivo: Analise_Ciclo_26_1_POR_UNIDADE.xlsx', ''],
    ['  • 4 abas com análise desagregada por Unidade + Canal', ''],
    ['  • 2 abas BASE COMPLETA (headers azuis) + 2 abas BASE FILTRADA (headers laranjas)', ''],
    ['  • Todos os 10 canais exibidos para cada unidade (mesmo com valor 0)', ''],
    ['  • Distribuição por faixa etaria: Total / Atendidos / %', ''],
    ['', ''],
    ['ATENÇÃO - Base Filtrada:', ''],
    ['  • Abas filtradas têm viés: 100% dos leads foram atendidos', ''],
    ['  • Motivo: "Data da última atividade" só preenchida quando há atendimento', ''],
    ['  • Use para: validar qualidade de dados, NÃO para taxa de atendimento real', ''],
    ['', ''],
]

df_resumo = pd.DataFrame(resumo_conteudo, columns=['Item', 'Descrição'])
with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
    # Escrever aba Resumo primeiro
    df_resumo.to_excel(writer, sheet_name='Resumo', index=False, header=False, startrow=0)
    
    # Escrever a primeira tabela (atendimento)
    tabela.to_excel(writer, sheet_name='Ciclo 26.1 (Leads)', index=False, startrow=0)
    
    # Escrever a segunda tabela (recência) logo abaixo
    start_row_recencia = len(tabela) + 3
    tabela_recencia.to_excel(writer, sheet_name='Ciclo 26.1 (Leads)', index=False, startrow=start_row_recencia)
    
    # Escrever tabela de conversão mais abaixo
    start_row_conversao = start_row_recencia + len(tabela_recencia) + 4
    tabela_conversao.to_excel(writer, sheet_name='Ciclo 26.1 (Leads)', index=False, startrow=start_row_conversao)
    
    # Escrever tabela de matrículas em nova aba
    tabela_matriculas.to_excel(writer, sheet_name='Ciclo 26.1 (Matriculas)', index=False, startrow=0)
    
    # Escrever tabela de idade real logo abaixo
    start_row_idade_real = len(tabela_matriculas) + 3
    tabela_idade_real.to_excel(writer, sheet_name='Ciclo 26.1 (Matriculas)', index=False, startrow=start_row_idade_real)

    # Escrever tabela de Etapa x Faixa de Idade (Out/2025 até 01/03/2026)
    tabela_etapa_faixa.to_excel(writer, sheet_name='Etapas x Faixa Idade', index=False, startrow=0)
    
    # Pegar o workbook para formatação
    workbook = writer.book
    
    # ========================================================================
    # FORMATAR ABA RESUMO
    # ========================================================================
    ws_resumo = writer.sheets['Resumo']
    
    # Ajustar larguras
    ws_resumo.column_dimensions['A'].width = 70
    ws_resumo.column_dimensions['B'].width = 80
    
    # Formatar linha por linha
    for row_num in range(1, len(df_resumo) + 1):
        cell_a = ws_resumo.cell(row=row_num, column=1)
        cell_b = ws_resumo.cell(row=row_num, column=2)
        
        texto = str(cell_a.value) if cell_a.value else ''
        
        # Título principal
        if 'ANÁLISE DO CICLO' in texto:
            cell_a.font = Font(bold=True, size=16, color="1F4E78")
            ws_resumo.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        
        # Títulos de seção
        elif texto in ['DESCRIÇÃO DAS ABAS', 'DIFERENÇAS IMPORTANTES', 'LIMITAÇÕES DA ANÁLISE']:
            cell_a.font = Font(bold=True, size=14, color="2F5496")
            cell_a.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            ws_resumo.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        
        # Subtítulos (começam com ABA: ou nomes em maiúsculas seguidos de :)
        elif texto.startswith('ABA:') or (texto.endswith(':') and len(texto) < 50):
            cell_a.font = Font(bold=True, size=12, color="1F4E78")
        
        # Bullets
        elif '•' in texto:
            cell_a.font = Font(size=10)
            cell_b.font = Font(size=10, italic=True)
        
        # Texto normal
        else:
            cell_a.font = Font(size=10)
            cell_b.font = Font(size=10)
        
        # Alinhamento
        cell_a.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell_b.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # ========================================================================
    # FORMATAR ABA CICLO 26.1 (LEADS)
    # ========================================================================
    ws = writer.sheets['Ciclo 26.1 (Leads)']
    
    # Cores
    cor_header = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cor_header2 = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cor_divisoria = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    cor_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    font_header = Font(bold=True, color="FFFFFF", size=11)
    font_normal = Font(size=10)
    font_total = Font(bold=True, size=10)
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Formatar primeira tabela (por canal)
    # Header
    for col_num, column in enumerate(tabela.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        
        # Colunas divisórias
        if column.startswith('|'):
            cell.fill = cor_divisoria
            cell.font = Font(color="000000")
        else:
            cell.fill = cor_header
            cell.font = font_header
        
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    
    # Dados
    for row_num in range(2, len(tabela) + 2):
        # Verificar se é a linha de TOTAL (última linha)
        is_total_row = (row_num == len(tabela) + 1)
        
        for col_num in range(1, len(tabela.columns) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border_thin
            
            column_name = tabela.columns[col_num - 1]
            
            # Colunas divisórias
            if column_name.startswith('|'):
                cell.fill = cor_divisoria
            # Linha de TOTAL
            elif is_total_row:
                cell.fill = cor_total
                cell.font = font_total
                if col_num > 1:
                    cell.alignment = Alignment(horizontal='center')
            # Linhas normais
            else:
                cell.font = font_normal
                if col_num > 1:  # Numéricas centralizadas
                    cell.alignment = Alignment(horizontal='center')
            
            # Aplicar formato de porcentagem nas colunas %
            if column_name.startswith('%'):
                cell.number_format = '0.0%'
    
    # ========================================================================
    # FORMATAR SEGUNDA TABELA (RECÊNCIA DA ÚLTIMA ATIVIDADE)
    # ========================================================================
    
    start_row_rec = len(tabela) + 3
    
    # Adicionar título
    titulo_rec_cell = ws.cell(row=start_row_rec, column=1)
    titulo_rec_cell.value = f"RECÊNCIA DA ÚLTIMA ATIVIDADE (D-1) - {leads_descartados:,} leads descartados por erro"
    titulo_rec_cell.font = Font(bold=True, size=12, color="C55A11")
    ws.merge_cells(start_row=start_row_rec, start_column=1, end_row=start_row_rec, end_column=5)
    
    # Header da tabela de recência
    start_row_rec += 1
    for col_num, column in enumerate(tabela_recencia.columns, 1):
        cell = ws.cell(row=start_row_rec, column=col_num)
        
        # Colunas divisórias
        if column.startswith('|'):
            cell.fill = cor_divisoria
            cell.font = Font(color="000000")
        else:
            cell.fill = cor_header
            cell.font = font_header
        
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    
    # Dados da tabela de recência
    for row_num in range(start_row_rec + 1, start_row_rec + len(tabela_recencia) + 1):
        # Verificar se é a linha de AVISO (última linha)
        is_aviso_row = (row_num == start_row_rec + len(tabela_recencia))
        # Verificar se é a linha de TOTAL (penúltima linha)
        is_total_row = (row_num == start_row_rec + len(tabela_recencia) - 1)
        
        for col_num in range(1, len(tabela_recencia.columns) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border_thin
            
            column_name = tabela_recencia.columns[col_num - 1]
            
            # Colunas divisórias
            if column_name.startswith('|'):
                cell.fill = cor_divisoria
            # Linha de AVISO
            elif is_aviso_row:
                if col_num == 1:  # Primeira coluna com o texto do aviso
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    cell.font = Font(bold=True, size=9, italic=True, color="C55A11")
                    cell.alignment = Alignment(horizontal='left')
            # Linha de TOTAL
            elif is_total_row:
                cell.fill = cor_total
                cell.font = font_total
                if col_num > 1:
                    cell.alignment = Alignment(horizontal='center')
            # Linhas normais
            else:
                cell.font = font_normal
                if col_num > 1:
                    cell.alignment = Alignment(horizontal='center')
            
            # Aplicar formato de porcentagem nas colunas %
            if column_name.startswith('%'):
                cell.number_format = '0.0%'
    
    # ========================================================================
    # FORMATAR TERCEIRA TABELA (CONVERSÃO POR FAIXA)
    # ========================================================================
    
    start_row_2 = start_row_rec + len(tabela_recencia) + 4
    
    # Adicionar título
    titulo_cell = ws.cell(row=start_row_2, column=1)
    titulo_cell.value = "MÉDIA DE ATIVIDADES PARA CONVERSÃO (HISTÓRICO)"
    titulo_cell.font = Font(bold=True, size=12, color="2F5496")
    ws.merge_cells(start_row=start_row_2, start_column=1, end_row=start_row_2, end_column=2)
    
    # Header da segunda tabela
    start_row_2 += 1
    for col_num, column in enumerate(tabela_conversao.columns, 1):
        cell = ws.cell(row=start_row_2, column=col_num)
        cell.value = column
        cell.fill = cor_header2
        cell.font = font_header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
    
    # Dados da segunda tabela
    for idx, row in tabela_conversao.iterrows():
        row_num = start_row_2 + idx + 1
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border_thin
            cell.font = font_normal
            if col_num > 1:
                cell.alignment = Alignment(horizontal='center')
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 25  # Canal
    ws.column_dimensions['B'].width = 15  # Total Leads
    
    # Ajustar largura dinamicamente: divisórias finas, dados normais
    col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    for idx, column_name in enumerate(tabela.columns):
        if idx < len(col_letters):
            col_letter = col_letters[idx]
            if column_name.startswith('|'):
                ws.column_dimensions[col_letter].width = 2  # Divisória bem fina
            elif idx > 1:  # Pula Canal e Total Leads
                ws.column_dimensions[col_letter].width = 13  # Colunas de dados
    
    # ========================================================================
    # FORMATAR ABA DE MATRÍCULAS
    # ========================================================================
    ws_mat = writer.sheets['Ciclo 26.1 (Matriculas)']
    
    # Adicionar título
    ws_mat.insert_rows(1)
    titulo_mat = ws_mat.cell(row=1, column=1)
    titulo_mat.value = "MATRÍCULAS EFETIVADAS NO CICLO 26.1 - ANÁLISE POR FAIXA DE IDADE"
    titulo_mat.font = Font(bold=True, size=14, color="1F4E78")
    ws_mat.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    
    # Headers (agora na linha 2)
    for col_num, column in enumerate(tabela_matriculas.columns, 1):
        cell = ws_mat.cell(row=2, column=col_num)
        
        if column.startswith('|'):
            cell.fill = cor_divisoria
            cell.font = Font(color="000000")
        else:
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
        
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
    
    # Dados (começam na linha 3)
    for row_num in range(3, 3 + len(tabela_matriculas)):
        is_total_row = (row_num == 2 + len(tabela_matriculas))
        
        for col_num in range(1, len(tabela_matriculas.columns) + 1):
            cell = ws_mat.cell(row=row_num, column=col_num)
            cell.border = border_thin
            
            column_name = tabela_matriculas.columns[col_num - 1]
            
            # Colunas divisórias
            if column_name.startswith('|'):
                cell.fill = cor_divisoria
            # Linha de TOTAL
            elif is_total_row:
                cell.fill = cor_total
                cell.font = font_total
                if col_num > 1:
                    cell.alignment = Alignment(horizontal='center')
            # Linhas normais
            else:
                cell.font = font_normal
                if col_num > 1:
                    cell.alignment = Alignment(horizontal='center')
            
            # Aplicar formato de porcentagem nas colunas %
            if column_name.startswith('%'):
                cell.number_format = '0.0%'
            # Aplicar formato numérico com 1 casa decimal nas colunas de Média Ativ e Média Idade
            elif 'Média Ativ' in column_name or 'Média Idade' in column_name:
                cell.number_format = '0.0'
    
    # Ajustar larguras das colunas
    ws_mat.column_dimensions['A'].width = 25  # Canal
    ws_mat.column_dimensions['B'].width = 18  # Total Matrículas
    ws_mat.column_dimensions['C'].width = 16  # Média Geral Ativ
    ws_mat.column_dimensions['D'].width = 16  # Média Idade (dias)
    
    for idx, column_name in enumerate(tabela_matriculas.columns):
        if idx < len(col_letters):
            col_letter = col_letters[idx]
            if column_name.startswith('|'):
                ws_mat.column_dimensions[col_letter].width = 2
            elif idx > 3:  # Pula Canal, Total Matrículas, Média Geral e Média Idade
                ws_mat.column_dimensions[col_letter].width = 13
    
    # ========================================================================
    # FORMATAR SEGUNDA TABELA DA ABA DE MATRÍCULAS (IDADE REAL)
    # ========================================================================
    
    start_row_ir = start_row_idade_real + 1  # +1 porque começa com header
    
    # Adicionar título
    titulo_ir = ws_mat.cell(row=start_row_ir, column=1)
    titulo_ir.value = f"IDADE REAL DOS LEADS (baseada na última atividade) - {matriculas_descartadas:,} descartadas por erro"
    titulo_ir.font = Font(bold=True, size=12, color="C55A11")
    ws_mat.merge_cells(start_row=start_row_ir, start_column=1, end_row=start_row_ir, end_column=5)
    
    # Headers (linha seguinte)
    start_row_ir += 1
    for col_num, column in enumerate(tabela_idade_real.columns, 1):
        cell = ws_mat.cell(row=start_row_ir, column=col_num)
        
        if column.startswith('|'):
            cell.fill = cor_divisoria
            cell.font = Font(color="000000")
        else:
            cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
        
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
    
    # Dados
    for row_num in range(start_row_ir + 1, start_row_ir + len(tabela_idade_real) + 1):
        is_aviso_row = (row_num == start_row_ir + len(tabela_idade_real))
        is_total_row = (row_num == start_row_ir + len(tabela_idade_real) - 1)
        
        for col_num in range(1, len(tabela_idade_real.columns) + 1):
            cell = ws_mat.cell(row=row_num, column=col_num)
            cell.border = border_thin
            
            column_name = tabela_idade_real.columns[col_num - 1]
            
            # Colunas divisórias
            if column_name.startswith('|'):
                cell.fill = cor_divisoria
            # Linha de AVISO
            elif is_aviso_row:
                if col_num == 1:
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    cell.font = Font(bold=True, size=9, italic=True, color="C55A11")
                    cell.alignment = Alignment(horizontal='left')
            # Linha de TOTAL
            elif is_total_row:
                cell.fill = cor_total
                cell.font = font_total
                if col_num > 1:
                    cell.alignment = Alignment(horizontal='center')
            # Linhas normais
            else:
                cell.font = font_normal
                if col_num > 1:
                    cell.alignment = Alignment(horizontal='center')
            
            # Aplicar formato de porcentagem nas colunas %
            if column_name.startswith('%'):
                cell.number_format = '0.0%'
            # Aplicar formato numérico com 1 casa decimal nas colunas de Média Ativ e Média Idade
            elif 'Média Ativ' in column_name or 'Média Idade' in column_name:
                cell.number_format = '0.0'

print(f"\n✅ Análise por canal concluída!")
print(f"📁 Arquivo salvo: {arquivo_saida}")

# ============================================================================
# EXPORTAR ARQUIVO SEPARADO COM ANÁLISES POR UNIDADE
# ============================================================================

print(f"\n💾 Exportando análises por unidade...")

with pd.ExcelWriter(arquivo_unidades, engine='openpyxl') as writer_unidades:
    # Escrever análise por unidade - BASE COMPLETA
    unidades_leads_completo.to_excel(writer_unidades, sheet_name='Unidades (Leads)', index=False)
    unidades_matriculas_completo.to_excel(writer_unidades, sheet_name='Unidades (Matriculas)', index=False)
    
    # Escrever análise por unidade - BASE FILTRADA
    unidades_leads_filtrado.to_excel(writer_unidades, sheet_name='Unidades (Leads) FILTRADO', index=False)
    unidades_matriculas_filtrado.to_excel(writer_unidades, sheet_name='Unidades (Mat) FILTRADO', index=False)
    
    # Pegar o workbook para formatação
    workbook_unidades = writer_unidades.book
    
    # Definir estilos (reutilizar)
    cor_header = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cor_header_laranja = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    cor_divisoria = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    cor_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    font_header = Font(bold=True, color="FFFFFF", size=11)
    font_normal = Font(size=10)
    font_total = Font(bold=True, size=10)
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========================================================================
    # FORMATAR ABA UNIDADES (LEADS) - BASE COMPLETA
    # ========================================================================
    ws_un_leads = writer_unidades.sheets['Unidades (Leads)']
    
    # Header
    for col_num, column in enumerate(unidades_leads_completo.columns, 1):
        cell = ws_un_leads.cell(row=1, column=col_num)
        
        # Colunas divisórias
        if column.startswith('|'):
            cell.fill = cor_divisoria
            cell.font = Font(color="000000")
            ws_un_leads.column_dimensions[cell.column_letter].width = 2
        else:
            cell.fill = cor_header
            cell.font = font_header
        
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    
    # Ajustar largura das colunas principais
    ws_un_leads.column_dimensions['A'].width = 25  # Unidade
    ws_un_leads.column_dimensions['B'].width = 12  # Tipo
    ws_un_leads.column_dimensions['C'].width = 20  # Canal
    
    # Dados
    for row_num in range(2, len(unidades_leads_completo) + 2):
        # Verificar se é linha de TOTAL (agora o canal está na coluna 3)
        canal_value = ws_un_leads.cell(row=row_num, column=3).value
        is_total_row = (canal_value and 'TOTAL' in str(canal_value))
        
        for col_num in range(1, len(unidades_leads_completo.columns) + 1):
            cell = ws_un_leads.cell(row=row_num, column=col_num)
            cell.border = border_thin
            
            column_name = unidades_leads_completo.columns[col_num - 1]
            
            # Colunas divisórias
            if column_name.startswith('|'):
                cell.fill = cor_divisoria
            # Linha de TOTAL
            elif is_total_row:
                cell.fill = cor_total
                cell.font = font_total
                if col_num > 3:
                    cell.alignment = Alignment(horizontal='center')
            # Linhas normais
            else:
                cell.font = font_normal
                if col_num > 3:
                    cell.alignment = Alignment(horizontal='center')
            
            # Aplicar formato de porcentagem
            if '% Atendimento' in column_name or column_name.endswith(' %'):
                cell.number_format = '0.0%'
            # Aplicar formato numérico com 1 casa decimal
            elif 'Média Ativ' in column_name:
                cell.number_format = '0.0'
    
    # ========================================================================
    # FORMATAR ABA UNIDADES (MATRICULAS) - BASE COMPLETA
    # ========================================================================
    ws_un_mat = writer_unidades.sheets['Unidades (Matriculas)']
    
    # Header
    for col_num, column in enumerate(unidades_matriculas_completo.columns, 1):
        cell = ws_un_mat.cell(row=1, column=col_num)
        
        # Colunas divisórias
        if column.startswith('|'):
            cell.fill = cor_divisoria
            cell.font = Font(color="000000")
            ws_un_mat.column_dimensions[cell.column_letter].width = 2
        else:
            cell.fill = cor_header
            cell.font = font_header
        
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    
    # Ajustar largura das colunas principais
    ws_un_mat.column_dimensions['A'].width = 25  # Unidade
    ws_un_mat.column_dimensions['B'].width = 12  # Tipo
    ws_un_mat.column_dimensions['C'].width = 20  # Canal
    
    # Dados
    for row_num in range(2, len(unidades_matriculas_completo) + 2):
        # Verificar se é linha de TOTAL (agora o canal está na coluna 3)
        canal_value = ws_un_mat.cell(row=row_num, column=3).value
        is_total_row = (canal_value and 'TOTAL' in str(canal_value))
        
        for col_num in range(1, len(unidades_matriculas_completo.columns) + 1):
            cell = ws_un_mat.cell(row=row_num, column=col_num)
            cell.border = border_thin
            
            column_name = unidades_matriculas_completo.columns[col_num - 1]
            
            # Colunas divisórias
            if column_name.startswith('|'):
                cell.fill = cor_divisoria
            # Linha de TOTAL
            elif is_total_row:
                cell.fill = cor_total
                cell.font = font_total
                if col_num > 3:
                    cell.alignment = Alignment(horizontal='center')
            # Linhas normais
            else:
                cell.font = font_normal
                if col_num > 3:
                    cell.alignment = Alignment(horizontal='center')
            
            # Aplicar formato de porcentagem
            if 'Taxa Conversão %' in column_name or column_name.endswith(' %'):
                cell.number_format = '0.0%'
            # Aplicar formato numérico com 1 casa decimal
            elif 'Média Ativ' in column_name or 'Média Idade' in column_name:
                cell.number_format = '0.0'
    
    # ========================================================================
    # FORMATAR ABA UNIDADES (LEADS) - BASE FILTRADA (COR LARANJA)
    # ========================================================================
    ws_un_leads_filt = writer_unidades.sheets['Unidades (Leads) FILTRADO']
    
    # Cor laranja para alertas/filtros
    cor_header_laranja = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    
    # Header
    for col_num, column in enumerate(unidades_leads_filtrado.columns, 1):
        cell = ws_un_leads_filt.cell(row=1, column=col_num)
        
        # Colunas divisórias
        if column.startswith('|'):
            cell.fill = cor_divisoria
            cell.font = Font(color="000000")
            ws_un_leads_filt.column_dimensions[cell.column_letter].width = 2
        else:
            cell.fill = cor_header_laranja
            cell.font = font_header
        
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    
    # Ajustar largura das colunas principais
    ws_un_leads_filt.column_dimensions['A'].width = 25  # Unidade
    ws_un_leads_filt.column_dimensions['B'].width = 12  # Tipo
    ws_un_leads_filt.column_dimensions['C'].width = 20  # Canal
    
    # Dados
    for row_num in range(2, len(unidades_leads_filtrado) + 2):
        # Verificar se é linha de TOTAL
        canal_value = ws_un_leads_filt.cell(row=row_num, column=3).value
        is_total_row = (canal_value and 'TOTAL' in str(canal_value))
        
        for col_num in range(1, len(unidades_leads_filtrado.columns) + 1):
            cell = ws_un_leads_filt.cell(row=row_num, column=col_num)
            cell.border = border_thin
            
            column_name = unidades_leads_filtrado.columns[col_num - 1]
            
            # Colunas divisórias
            if column_name.startswith('|'):
                cell.fill = cor_divisoria
            # Linha de TOTAL
            elif is_total_row:
                cell.fill = cor_total
                cell.font = font_total
                if col_num > 3:
                    cell.alignment = Alignment(horizontal='center')
            # Linhas normais
            else:
                cell.font = font_normal
                if col_num > 3:
                    cell.alignment = Alignment(horizontal='center')
            
            # Aplicar formato de porcentagem
            if '% Atendimento' in column_name or column_name.endswith(' %'):
                cell.number_format = '0.0%'
            # Aplicar formato numérico com 1 casa decimal
            elif 'Média Ativ' in column_name:
                cell.number_format = '0.0'
    
    # ========================================================================
    # FORMATAR ABA UNIDADES (MATRICULAS) - BASE FILTRADA (COR LARANJA)
    # ========================================================================
    ws_un_mat_filt = writer_unidades.sheets['Unidades (Mat) FILTRADO']
    
    # Header
    for col_num, column in enumerate(unidades_matriculas_filtrado.columns, 1):
        cell = ws_un_mat_filt.cell(row=1, column=col_num)
        
        # Colunas divisórias
        if column.startswith('|'):
            cell.fill = cor_divisoria
            cell.font = Font(color="000000")
            ws_un_mat_filt.column_dimensions[cell.column_letter].width = 2
        else:
            cell.fill = cor_header_laranja
            cell.font = font_header
        
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    
    # Ajustar largura das colunas principais
    ws_un_mat_filt.column_dimensions['A'].width = 25  # Unidade
    ws_un_mat_filt.column_dimensions['B'].width = 12  # Tipo
    ws_un_mat_filt.column_dimensions['C'].width = 20  # Canal
    
    # Dados
    for row_num in range(2, len(unidades_matriculas_filtrado) + 2):
        # Verificar se é linha de TOTAL
        canal_value = ws_un_mat_filt.cell(row=row_num, column=3).value
        is_total_row = (canal_value and 'TOTAL' in str(canal_value))
        
        for col_num in range(1, len(unidades_matriculas_filtrado.columns) + 1):
            cell = ws_un_mat_filt.cell(row=row_num, column=col_num)
            cell.border = border_thin
            
            column_name = unidades_matriculas_filtrado.columns[col_num - 1]
            
            # Colunas divisórias
            if column_name.startswith('|'):
                cell.fill = cor_divisoria
            # Linha de TOTAL
            elif is_total_row:
                cell.fill = cor_total
                cell.font = font_total
                if col_num > 3:
                    cell.alignment = Alignment(horizontal='center')
            # Linhas normais
            else:
                cell.font = font_normal
                if col_num > 3:
                    cell.alignment = Alignment(horizontal='center')
            
            # Aplicar formato de porcentagem
            if 'Taxa Conversão %' in column_name or column_name.endswith(' %'):
                cell.number_format = '0.0%'
            # Aplicar formato numérico com 1 casa decimal
            elif 'Média Ativ' in column_name or 'Média Idade' in column_name:
                cell.number_format = '0.0'

print(f"\n✅ Análise por unidade concluída!")
print(f"📁 Arquivo salvo: {arquivo_unidades}")

# ============================================================================
# RESUMO FINAL DE ARQUIVOS GERADOS
# ============================================================================

print(f"\n" + "="*80)
print(f"🎉 PROCESSO CONCLUÍDO!")
print(f"="*80)
print(f"\n📂 Arquivos gerados:")
print(f"   1. {os.path.basename(arquivo_saida)}")
print(f"      → 4 abas: Resumo + Análises por Canal + Etapas x Faixa Idade")
print(f"   2. {os.path.basename(arquivo_unidades)}")
print(f"      → 4 abas: Análises por Unidade (Completa + Filtrada)")

# ============================================================================
# EXPORTAR ARQUIVO SEPARADO COM ERROS DE CRM
# ============================================================================

if len(matriculas_negativas) > 0:
    print(f"   3. Analise_Ciclo_26_1_ERROS_CRM.xlsx")
    print(f"      → 8 abas: Resumo por Unidade + Erros de CRM por Canal ({len(matriculas_negativas)} registros)")
    print(f"\n🚨 Gerando arquivo de erros...")
    
    arquivo_erros = os.path.join(PASTA_OUTPUT, 'Analise_Ciclo_26_1_ERROS_CRM.xlsx')
    arquivo_erros = OUTPUT_DIR / 'Analise_Ciclo_26_1_ERROS_CRM.xlsx'
    
    with pd.ExcelWriter(arquivo_erros, engine='openpyxl') as writer_erros:
        # Definir estilos
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        
        # ============================================================================
        # ABA 1: RESUMO POR UNIDADE
        # ============================================================================
        
        # Criar resumo por unidade
        resumo_unidades_lista = []
        
        for unidade in sorted(matriculas_negativas['Unidade Desejada'].dropna().unique()):
            # Erros desta unidade
            erros_unidade = matriculas_negativas[matriculas_negativas['Unidade Desejada'] == unidade]
            total_erros = len(erros_unidade)
            
            # Total de matrículas da unidade no ciclo
            total_matriculas_unidade = len(df_matriculas_26_1[df_matriculas_26_1['Unidade Desejada'] == unidade])
            
            # Tipo de unidade
            tipo = erros_unidade['tipo_unidade'].iloc[0] if len(erros_unidade) > 0 else 'N/A'
            
            # Percentual de erro
            pct_erro = (total_erros / total_matriculas_unidade * 100) if total_matriculas_unidade > 0 else 0
            
            resumo_unidades_lista.append({
                'Unidade': unidade,
                'Tipo': tipo,
                'Total Matrículas': total_matriculas_unidade,
                'Total Erros': total_erros,
                '% Erro': pct_erro / 100  # Para formatação de %
            })
        
        # Adicionar linha de TOTAL
        total_matriculas_geral = len(df_matriculas_26_1)
        total_erros_geral = len(matriculas_negativas)
        pct_erro_geral = (total_erros_geral / total_matriculas_geral * 100) if total_matriculas_geral > 0 else 0
        
        resumo_unidades_lista.append({
            'Unidade': 'TOTAL',
            'Tipo': '',
            'Total Matrículas': total_matriculas_geral,
            'Total Erros': total_erros_geral,
            '% Erro': pct_erro_geral / 100
        })
        
        df_resumo_unidades = pd.DataFrame(resumo_unidades_lista)
        
        # Exportar resumo
        df_resumo_unidades.to_excel(writer_erros, sheet_name='📊 Resumo por Unidade', index=False)
        
        # Formatar aba de resumo
        ws_resumo = writer_erros.sheets['📊 Resumo por Unidade']
        
        # Header
        cor_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font_header = Font(bold=True, color="FFFFFF", size=11)
        
        for col_num in range(1, 6):
            cell = ws_resumo.cell(row=1, column=col_num)
            cell.fill = cor_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
        
        # Formatar dados
        for row_num in range(2, len(df_resumo_unidades) + 2):
            for col_num in range(1, 6):
                cell = ws_resumo.cell(row=row_num, column=col_num)
                cell.border = border_thin
                cell.alignment = Alignment(horizontal='left' if col_num <= 2 else 'center')
                
                # Formatar números
                if col_num == 3 or col_num == 4:  # Total Matrículas / Total Erros
                    cell.number_format = '#,##0'
                elif col_num == 5:  # % Erro
                    cell.number_format = '0.0%'
                    # Colorir % alto de erro
                    if cell.value is not None and cell.value > 0.5:  # > 50%
                        cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                        cell.font = Font(bold=True, color="C55A11")
        
        # Destacar linha TOTAL
        ultima_linha = len(df_resumo_unidades) + 1
        for col_num in range(1, 6):
            cell = ws_resumo.cell(row=ultima_linha, column=col_num)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.font = Font(bold=True, size=11)
        
        # Ajustar largura das colunas
        ws_resumo.column_dimensions['A'].width = 30  # Unidade
        ws_resumo.column_dimensions['B'].width = 15  # Tipo
        ws_resumo.column_dimensions['C'].width = 18  # Total Matrículas
        ws_resumo.column_dimensions['D'].width = 15  # Total Erros
        ws_resumo.column_dimensions['E'].width = 12  # % Erro
        
        # ============================================================================
        # ABAS 2+: DETALHES POR CANAL
        # ============================================================================
        
        # Para cada canal com matrículas negativas, criar uma aba com dados completos
        canais_com_problema = matriculas_negativas['canal'].unique()
        
        for canal in sorted(canais_com_problema):
            # Filtrar matrículas negativas deste canal
            matriculas_canal = matriculas_negativas[matriculas_negativas['canal'] == canal].copy()
            
            # Buscar dados completos no dataframe original
            df_completo_canal = df[df.index.isin(matriculas_canal.index)].copy()
            
            # Adicionar colunas calculadas importantes
            df_completo_canal['idade_na_conversao'] = matriculas_canal['idade_na_conversao']
            df_completo_canal['data_conversao'] = matriculas_canal['data_conversao']
            
            # Ordenar por idade negativa (piores casos primeiro)
            df_completo_canal = df_completo_canal.sort_values('idade_na_conversao')
            
            # Nome da aba (limitado a 31 caracteres pelo Excel)
            nome_aba = f"⚠️ {canal}"[:31]
            
            # Exportar para Excel
            df_completo_canal.to_excel(writer_erros, sheet_name=nome_aba, index=False)
            
            # Formatar aba
            ws_diag = writer_erros.sheets[nome_aba]
            
            # Header com cor de alerta
            cor_alerta = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            font_alerta = Font(bold=True, color="FFFFFF", size=10)
            
            for col_num in range(1, len(df_completo_canal.columns) + 1):
                cell = ws_diag.cell(row=1, column=col_num)
                cell.fill = cor_alerta
                cell.font = font_alerta
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border_thin
            
            # Ajustar largura das colunas principais
            ws_diag.column_dimensions['A'].width = 20  # Primeira coluna
            
            # Destacar colunas importantes
            col_mapping = {col: idx + 1 for idx, col in enumerate(df_completo_canal.columns)}
            
            # Colorir células de idade_na_conversao negativa
            if 'idade_na_conversao' in col_mapping:
                col_idade = col_mapping['idade_na_conversao']
                for row_num in range(2, len(df_completo_canal) + 2):
                    cell = ws_diag.cell(row=row_num, column=col_idade)
                    if cell.value is not None and cell.value < 0:
                        cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                        cell.font = Font(bold=True, color="C55A11")
            
            # Destacar colunas de data importantes
            for col_name in ['Data de criação', 'Data de fechamento', 'data_conversao', 'Data da última atividade']:
                if col_name in col_mapping:
                    col_num = col_mapping[col_name]
                    ws_diag.column_dimensions[col_letters[col_num - 1] if col_num <= len(col_letters) else 'A'].width = 18
    
    print(f"   ✅ Arquivo de erros salvo!")
else:
    print(f"   (Nenhum erro de CRM encontrado)")

print(f"\n👍 Todos os arquivos estão em: {PASTA_OUTPUT}")
print(f"\n👍 Todos os arquivos estão em: {OUTPUT_DIR}")
print(f"="*80)

# ============================================================================
# 6. RESUMO NO CONSOLE
# ============================================================================

print("\n" + "="*80)
print("📊 RESUMO - CICLO 26.1")
print("="*80)

print(f"\n🎯 Total de Leads no Ciclo 26.1: {len(df_ciclo_26_1):,}")
print(f"📅 Período: 01/Out/2025 até {hoje.strftime('%d/%b/%Y')} ({dias_desde_inicio} dias)")
print(f"📥 Data de extração HubSpot: 12/Fev/2026")

print("\n📈 TAXA DE ATENDIMENTO POR FAIXA DE IDADE:")
print("-" * 80)
for faixa in ordem_faixas:
    leads_faixa = df_ciclo_26_1[df_ciclo_26_1['faixa_idade'] == faixa]
    total_faixa = len(leads_faixa)
    atendidos_faixa = len(leads_faixa[leads_faixa['Número de atividades de vendas'] > 0])
    if total_faixa > 0:
        taxa = 100 * atendidos_faixa / total_faixa
        print(f"{faixa:>12}: {atendidos_faixa:>4}/{total_faixa:<5} leads atendidos ({taxa:.1f}%)")
    else:
        print(f"{faixa:>12}: 0/0 leads")

print("\n📅 DISTRIBUIÇÃO DE IDADE DOS LEADS (Ciclo 26.1):")
print("-" * 80)
total_idade_26 = df_ciclo_26_1['faixa_idade'].value_counts()
print(total_idade_26.to_string())

print("\n🔄 RECÊNCIA DA ÚLTIMA ATIVIDADE (D-1) POR FAIXA DE IDADE:")
print("-" * 80)
print(f"⚠️  Análise baseada em {len(df_ciclo_26_1_valido):,} leads ({leads_descartados:,} descartados por erro de preenchimento)")
print("")
for faixa in ordem_faixas:
    leads_faixa = df_ciclo_26_1_valido[df_ciclo_26_1_valido['faixa_idade'] == faixa]
    total_faixa = len(leads_faixa)
    ativ_d1 = len(leads_faixa[leads_faixa['recencia_atividade'] == 'D-1'])
    if total_faixa > 0:
        taxa = 100 * ativ_d1 / total_faixa
        print(f"{faixa:>12}: {ativ_d1:>4}/{total_faixa:<5} leads com atividade D-1 ({taxa:.1f}%)")
    else:
        print(f"{faixa:>12}: 0/0 leads")

print("\n🎯 MÉDIA DE ATIVIDADES PARA CONVERSÃO (Histórico por Faixa de Idade):")
print("-" * 80)
print(tabela_conversao.to_string(index=False))

print("\n🎓 MATRÍCULAS EFETIVADAS NO CICLO 26.1:")
print("-" * 80)
print(f"Total de matrículas: {len(df_matriculas_26_1):,}")
print(f"Taxa de conversão: {100*len(df_matriculas_26_1)/len(df_ciclo_26_1):.1f}%")
print("")
print("Distribuição por faixa de idade na conversão:")
for faixa in ordem_faixas:
    matriculas_faixa = df_matriculas_26_1[df_matriculas_26_1['faixa_idade_conversao'] == faixa]
    qtd = len(matriculas_faixa)
    if qtd > 0:
        taxa = 100 * qtd / len(df_matriculas_26_1)
        media_ativ = matriculas_faixa['Número de atividades de vendas'].mean()
        print(f"{faixa:>12}: {qtd:>3} matrículas ({taxa:>5.1f}%) - Média de {media_ativ:.1f} atividades")
    else:
        print(f"{faixa:>12}: 0 matrículas")

print("\n🔍 IDADE REAL DOS LEADS (baseada na última atividade comercial):")
print("-" * 80)
print(f"⚠️  Análise baseada em {len(df_matriculas_idade_real):,} matrículas ({matriculas_descartadas:,} descartadas por erro)")
print("")
for faixa in ordem_faixas:
    matriculas_faixa = df_matriculas_idade_real[df_matriculas_idade_real['faixa_idade_real'] == faixa]
    qtd = len(matriculas_faixa)
    if qtd > 0:
        taxa = 100 * qtd / len(df_matriculas_idade_real)
        media_ativ = matriculas_faixa['Número de atividades de vendas'].mean()
        print(f"{faixa:>12}: {qtd:>3} matrículas ({taxa:>5.1f}%) - Média de {media_ativ:.1f} atividades")
    else:
        print(f"{faixa:>12}: 0 matrículas")

print("\n" + "="*80)
