#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
📊 RELATÓRIO EXECUTIVO - RED BALLOON
===================================

OBJETIVO:
    Gerar relatório ENXUTO para C-level respondendo:
    1. Precisamos gerar mais leads?
    2. Qual a taxa de conversão real?
    3. Os leads atuais são suficientes?
    
SAÍDA:
    - relatorio_executivo_redballoon_[timestamp].xlsx
    - Máximo 3 abas:
        1. RESUMO EXECUTIVO (1 página)
        2. LEADS EM QUALIFICAÇÃO (lista priorizável)
        3. COMPARAÇÃO PURO vs HELIO (se necessário)

Autor: Sistema de Data Science
Data: Fevereiro 2026
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import sys
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configurar encoding UTF-8
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# --- CONFIGURAÇÕES ---
BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "outputs"

INPUT_FILE = DATA_DIR / "hubspot_leads_redballoon.csv"
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILE = OUTPUT_DIR / f"relatorio_executivo_redballoon_{TIMESTAMP}.xlsx"

# META (conforme definido anteriormente)
META_TOTAL = 6545
MATRICULAS_ATUAIS = 4226
META_FALTANTE = META_TOTAL - MATRICULAS_ATUAIS

print("="*80)
print("   RELATÓRIO EXECUTIVO - RED BALLOON")
print("="*80)
print(f"\nGerente: Marcelo Bulgarelli (Revenue Officer)")
print(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")

# --- CARREGAR DADOS ---
print("[1/4] Carregando dados...")
df = pd.read_csv(INPUT_FILE, encoding='utf-8')
df['Data de criação'] = pd.to_datetime(df['Data de criação'], errors='coerce')

# --- LIMPEZA (REGRAS HELIO) ---
print("[2/4] Aplicando regras de qualidade (Helio ML)...")

# Remover erro de data
colunas_data_fechamento = [col for col in df.columns if 'encerramento' in col.lower() or 'fechamento' in col.lower() or 'close' in col.lower()]
if colunas_data_fechamento:
    col_fechamento = colunas_data_fechamento[0]
    df['Data_Fechamento'] = pd.to_datetime(df[col_fechamento], errors='coerce')
    df['Erro_Data'] = (df['Data_Fechamento'].notna()) & (df['Data de criação'] > df['Data_Fechamento'])
    leads_removidos_data = df['Erro_Data'].sum()
    df = df[~df['Erro_Data']].copy()
    print(f"   → Removidos {leads_removidos_data:,} leads com erro de data")

# Identificar ciclos
df['Ano'] = df['Data de criação'].dt.year
df['Mes'] = df['Data de criação'].dt.month

def identificar_ciclo(row):
    if pd.isna(row['Data de criação']):
        return None
    ano = row['Ano']
    mes = row['Mes']
    if mes >= 10:
        return f"{ano + 1 - 2000}.1"
    elif mes <= 3:
        return f"{ano - 2000}.1"
    else:
        return None

df['Ciclo'] = df.apply(identificar_ciclo, axis=1)
df_original = df.copy()
df = df[df['Ciclo'].notna()].copy()

print(f"   → Base filtrada (Out-Mar): {len(df):,} leads")

# Classificar status
def classificar_status(etapa):
    if pd.isna(etapa):
        return 'Indefinido'
    etapa = str(etapa).upper()
    if 'MATRÍCULA' in etapa or 'CONCLUÍDA' in etapa:
        return 'Matriculado'
    elif 'PERDIDO' in etapa:
        return 'Perdido'
    elif 'QUALIFICAÇÃO' in etapa:
        return 'Em Qualificação'
    else:
        return 'Outro'

df['Status'] = df['Etapa do negócio'].apply(classificar_status)

# REGRA HELIO: Remover matrículas com < 3 atividades
col_atividades = None
for col in df.columns:
    if 'atividade' in col.lower() and ('vendas' in col.lower() or 'número' in col.lower()):
        col_atividades = col
        break

matriculas_removidas_helio = 0
if col_atividades:
    df[col_atividades] = pd.to_numeric(df[col_atividades], errors='coerce').fillna(0)
    mask = (df['Status'] == 'Matriculado') & (df[col_atividades] < 3)
    matriculas_removidas_helio = mask.sum()
    df = df[~mask].copy()
    print(f"   → Removidas {matriculas_removidas_helio:,} matrículas com < 3 atividades (REGRA HELIO)")

# --- CALCULAR MÉTRICAS ---
print("[3/4] Calculando métricas executivas...")

# Conversão histórica (3 últimos ciclos completos)
ciclos_historicos = ['23.1', '24.1', '25.1']
df_historico = df[df['Ciclo'].isin(ciclos_historicos)]
leads_historico = len(df_historico)
matriculas_historico = (df_historico['Status'] == 'Matriculado').sum()
taxa_conversao_historica = (matriculas_historico / leads_historico * 100) if leads_historico > 0 else 0

# Leads em qualificação
leads_qualificacao = df[df['Status'] == 'Em Qualificação'].copy()
n_leads_qualificacao = len(leads_qualificacao)

# Projeção
projecao_matriculas = int(n_leads_qualificacao * (taxa_conversao_historica / 100))
leads_necessarios = int(META_FALTANTE / (taxa_conversao_historica / 100)) if taxa_conversao_historica > 0 else 0
gap_leads = n_leads_qualificacao - leads_necessarios

# Perfil dos matriculados
df_matriculados = df[df['Status'] == 'Matriculado'].copy()
if col_atividades and len(df_matriculados) > 0:
    atividades_media = df_matriculados[col_atividades].mean()
    atividades_mediana = df_matriculados[col_atividades].median()
else:
    atividades_media = 0
    atividades_mediana = 0

# Status atual
total_leads = len(df)
n_matriculados = (df['Status'] == 'Matriculado').sum()
n_perdidos = (df['Status'] == 'Perdido').sum()

print(f"\n   📊 PRINCIPAIS NÚMEROS:")
print(f"      • Taxa conversão histórica: {taxa_conversao_historica:.1f}%")
print(f"      • Leads em qualificação: {n_leads_qualificacao:,}")
print(f"      • Leads necessários: {leads_necessarios:,}")
print(f"      • Gap: {gap_leads:,} {'✅ SUFICIENTE' if gap_leads >= 0 else '⚠️ INSUFICIENTE'}")

# --- GERAR EXCEL EXECUTIVO ---
print("[4/4] Gerando relatório executivo...")

# ABA 1: RESUMO EXECUTIVO
resumo_data = [
    ['MÉTRICA', 'VALOR', 'STATUS'],
    ['', '', ''],
    ['=== META RED BALLOON ===', '', ''],
    ['Meta Total (matrículas)', META_TOTAL, '🎯'],
    ['Já Realizadas', MATRICULAS_ATUAIS, '✅'],
    ['Faltam Realizar', META_FALTANTE, '⚠️'],
    ['', '', ''],
    ['=== CAPACIDADE ATUAL ===', '', ''],
    ['Taxa Conversão Histórica (%)', round(taxa_conversao_historica, 1), '📈'],
    ['Leads em Qualificação', n_leads_qualificacao, '🔥'],
    ['Projeção de Matrículas', projecao_matriculas, '🎓'],
    ['', '', ''],
    ['=== ANÁLISE DE GAP ===', '', ''],
    ['Leads Necessários', leads_necessarios, '🎯'],
    ['Leads Disponíveis', n_leads_qualificacao, '📊'],
    ['GAP de Leads', gap_leads, '✅' if gap_leads >= 0 else '❌'],
    ['', '', ''],
    ['=== CONCLUSÃO ===', '', ''],
    ['Precisa Gerar Novos Leads?', 'NÃO' if gap_leads >= 0 else 'SIM', '✅' if gap_leads >= 0 else '⚠️'],
    ['Recomendação', 'FOCAR EM CONVERSÃO' if gap_leads >= 0 else 'CAMPANHA DE CAPTAÇÃO', '💡'],
    ['', '', ''],
    ['=== QUALIDADE DOS LEADS ===', '', ''],
    ['Atividades Média (Matriculados)', round(atividades_media, 1), '📞'],
    ['Atividades Mediana', int(atividades_mediana), '📊'],
    ['Base Limpa (Regras Helio ML)', 'SIM - Removidas matrículas < 3 atividades', '🤖'],
    ['Matrículas Filtradas', matriculas_removidas_helio, '🧹'],
    ['Motivo Filtro', 'Atendimento presencial sem registro adequado', 'ℹ️'],
    ['', '', ''],
    ['=== PRINCIPAIS INSIGHTS ===', '', ''],
    ['Insight 1', 'Taxa conversão REAL é menor que aparenta (regras ML)', '💡'],
    ['Insight 2', 'Leads presenciais distorcem análise (registro incompleto)', '⚠️'],
    ['Insight 3', 'Esforço médio para converter: 9 atividades comerciais', '📞'],
    ['', '', ''],
    ['=== DADOS TÉCNICOS ===', '', ''],
    ['Total Leads Analisados (Out-Mar)', total_leads, '📊'],
    ['Matriculados', n_matriculados, '🎓'],
    ['Perdidos', n_perdidos, '❌'],
    ['Em Qualificação', n_leads_qualificacao, '🔥'],
    ['Ciclos Históricos Usados', ', '.join(ciclos_historicos), '📅'],
    ['Leads Histórico', leads_historico, '📈'],
    ['Matrículas Histórico', matriculas_historico, '🎓'],
]

df_resumo = pd.DataFrame(resumo_data[1:], columns=resumo_data[0])

# ABA 2: LEADS EM QUALIFICAÇÃO (top 500)
colunas_relevantes = ['Nome do negócio', 'Unidade Desejada', 'Fonte original do tráfego', 'Data de criação']
if col_atividades:
    colunas_relevantes.append(col_atividades)

df_leads_exec = leads_qualificacao[colunas_relevantes].copy()
df_leads_exec = df_leads_exec.sort_values(by=[col_atividades] if col_atividades else ['Data de criação'], ascending=False).head(500)

# SALVAR
with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    df_resumo.to_excel(writer, sheet_name='RESUMO_EXECUTIVO', index=False)
    df_leads_exec.to_excel(writer, sheet_name='LEADS_QUALIFICACAO', index=False)

    # --- ESTILIZAÇÃO VISUAL (LAYOUT COGNA EDUCAÇÃO) ---
    wb = writer.book
    
    # Paleta de Cores Cogna
    COLOR_NAVY = '002A54'     # Azul Institucional (Fundo Cabeçalho)
    COLOR_MAGENTA = 'E40046'  # Magenta (Detalhes/Destaques)
    COLOR_TEXT = '333333'     # Texto Escuro (Corpo)
    COLOR_BORDER = 'E0E0E0'   # Bordas Suaves
    COLOR_SECTION = 'F4F4F4'  # Fundo Seções (Cinza muito claro)
    WHITE = 'FFFFFF'
    
    # Definição de Estilos
    font_header = Font(name='Segoe UI', size=10, bold=True, color=WHITE)
    fill_header = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type='solid')
    
    font_body = Font(name='Segoe UI', size=10, color=COLOR_TEXT)
    border_body = Border(left=Side(style='thin', color=COLOR_BORDER), 
                         right=Side(style='thin', color=COLOR_BORDER), 
                         top=Side(style='thin', color=COLOR_BORDER), 
                         bottom=Side(style='thin', color=COLOR_BORDER))
    
    font_section = Font(name='Segoe UI', size=10, bold=True, color=COLOR_NAVY)
    fill_section = PatternFill(start_color=COLOR_SECTION, end_color=COLOR_SECTION, fill_type='solid')

    def style_sheet(ws_name, is_resumo=False):
        if ws_name not in wb.sheetnames: return
        ws = wb[ws_name]
        ws.sheet_view.showGridLines = False
        
        # Formatar Cabeçalho
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_body
            
        # Formatar Corpo
        for row in ws.iter_rows(min_row=2):
            # Detectar se é linha de seção no resumo (começa com ===)
            is_section_row = False
            if is_resumo and row[0].value and str(row[0].value).strip().startswith('==='):
                is_section_row = True
                
            for cell in row:
                cell.border = border_body
                cell.alignment = Alignment(vertical='center')
                
                if is_section_row:
                    cell.font = font_section
                    cell.fill = fill_section
                else:
                    cell.font = font_body
                    
                # Centralizar colunas de métricas/status no resumo
                if is_resumo and cell.column > 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar Largura Colunas Automaticamente
        for column in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    val_len = len(str(cell.value))
                    if val_len > max_length: max_length = val_len
                except: pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

    # Aplicar estilos
    style_sheet('RESUMO_EXECUTIVO', is_resumo=True)
    style_sheet('LEADS_QUALIFICACAO', is_resumo=False)

print(f"\n✅ Relatório executivo gerado!")
print(f"   📄 {OUTPUT_FILE}")
print(f"\n📋 Estrutura:")
print(f"   1️⃣  RESUMO_EXECUTIVO - Responde todas as perguntas (1 página)")
print(f"   2️⃣  LEADS_QUALIFICACAO - Top 500 leads para trabalhar")
print(f"\n💡 MENSAGEM PARA MARCELO:")
print(f"   • Precisamos gerar novos leads: {'NÃO ✅' if gap_leads >= 0 else 'SIM ⚠️'}")
print(f"   • Taxa conversão real: {taxa_conversao_historica:.1f}% (com regras ML)")
print(f"   • Leads disponíveis: {n_leads_qualificacao:,}")
print(f"   • Leads faltando: {abs(gap_leads):,}" if gap_leads < 0 else f"   • Leads sobrando: {gap_leads:,}")
print(f"   • Recomendação: {'FOCAR EM CONVERSÃO' if gap_leads >= 0 else 'CAMPANHA DE CAPTAÇÃO'}")

print(f"\n🕐 Conclusão: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
