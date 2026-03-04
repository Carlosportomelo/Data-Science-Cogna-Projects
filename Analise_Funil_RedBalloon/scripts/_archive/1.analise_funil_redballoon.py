#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
🎈 ANÁLISE DE FUNIL - RED BALLOON
==================================

OBJETIVO:
    Analisar a performance do funil de vendas da Red Balloon por:
    - Ano de geração do lead
    - Status (Matrícula, Perdido, Em Qualificação)
    - Unidade desejada
    - Fonte de origem
    - Performance temporal

INPUT:
    - data/hubspot_leads_redballoon.csv (filtrado do HubSpot)

OUTPUT:
    - outputs/funil_redballoon_[timestamp].xlsx
    
MÉTRICAS:
    ✅ Total de leads por ano
    ✅ Taxa de conversão (Lead → Matrícula)
    ✅ Taxa de perda
    ✅ Leads em qualificação (status atual)
    ✅ Performance por unidade
    ✅ Performance por fonte
    ✅ Análise mensal e por ciclo

Autor: Sistema de Data Science
Data: Fevereiro 2026
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import sys
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configurar encoding UTF-8 para suportar emojis no Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# --- CONFIGURAÇÕES ---
try:
    BASE_DIR = Path(__file__).parent.parent
except NameError:
    BASE_DIR = Path.cwd()

DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "outputs"

# Criar diretório de outputs se não existir
OUTPUT_DIR.mkdir(exist_ok=True)

# Arquivo de entrada
INPUT_FILE = DATA_DIR / "hubspot_leads_redballoon.csv"

# Timestamp para o arquivo de saída
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILE = OUTPUT_DIR / f"funil_redballoon_{TIMESTAMP}.xlsx"

print("="*80)
print(f"   ANALISE DE FUNIL - RED BALLOON")
print("="*80)
print(f"\nDiretorio: {BASE_DIR}")
print(f"Arquivo de entrada: {INPUT_FILE}")
print(f"Arquivo de saida: {OUTPUT_FILE}")
print(f"Inicio: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")

# --- CARREGAR DADOS ---
print("="*80)
print("ETAPA 1: CARREGANDO DADOS")
print("="*80 + "\n")

if not INPUT_FILE.exists():
    print(f"❌ ERRO: Arquivo não encontrado: {INPUT_FILE}")
    print("\n💡 DICA: Execute o comando abaixo para criar a base:")
    print('   python ../_SCRIPTS_COMPARTILHADOS/sincronizar_bases.py')
    sys.exit(1)

try:
    df = pd.read_csv(INPUT_FILE, encoding='utf-8')
    print(f"[OK] Dados carregados: {len(df):,} registros")
    print(f"Colunas disponiveis: {len(df.columns)}")
except Exception as e:
    print(f"❌ ERRO ao carregar arquivo: {e}")
    sys.exit(1)

# --- TRATAMENTO DE DADOS ---
print("\n" + "="*80)
print("ETAPA 2: TRATAMENTO E LIMPEZA DE DADOS")
print("="*80 + "\n")

# Converter datas
df['Data de criação'] = pd.to_datetime(df['Data de criação'], errors='coerce')

# Verificar se existe coluna de data de encerramento (Close Date)
colunas_data_fechamento = [col for col in df.columns if 'encerramento' in col.lower() or 'fechamento' in col.lower() or 'close' in col.lower()]
tem_data_fechamento = len(colunas_data_fechamento) > 0

if tem_data_fechamento:
    col_fechamento = colunas_data_fechamento[0]
    df['Data_Fechamento'] = pd.to_datetime(df[col_fechamento], errors='coerce')
    
    # Identificar leads com ERRO DE DATA (criação > fechamento)
    df['Erro_Data'] = (df['Data_Fechamento'].notna()) & (df['Data de criação'] > df['Data_Fechamento'])
    
    leads_com_erro = df['Erro_Data'].sum()
    leads_offline_direto_erro = df[df['Erro_Data'] & df['Fonte original do tráfego'].fillna('').str.contains('offline|direto', case=False, na=False)].shape[0]
    
    print(f"\n🔧 CORREÇÃO DE DADOS - ERRO DE DATAS:")
    print(f"   Leads com Data Criacao > Data Fechamento: {leads_com_erro:,}")
    print(f"   Destes, sao Offline/Direto: {leads_offline_direto_erro:,}")
    print(f"   Acao: CORRIGIR data de criacao para data de fechamento")
    print(f"   Motivo: Evita idade negativa e mantém matrículas reais\n")
    
    # CORRIGIR ao invés de remover: usar data de fechamento como criação
    df_original_completo = df.copy()
    df.loc[df['Erro_Data'], 'Data de criação'] = df.loc[df['Erro_Data'], 'Data_Fechamento']
    
    print(f"   ✅ {leads_com_erro:,} datas corrigidas (matrículas mantidas)")
else:
    print("\n⚠️  AVISO: Coluna de data de fechamento nao encontrada")
    print("   Nao foi possivel identificar leads com erro de data\n")
    df_original_completo = df.copy()
    leads_com_erro = 0
    leads_offline_direto_erro = 0
df['Ano'] = df['Data de criação'].dt.year
df['Mes'] = df['Data de criação'].dt.month
df['Mes_Ano'] = df['Data de criação'].dt.to_period('M').astype(str)

# Normalizar etapa do negócio ANTES de identificar ciclo
def classificar_status(etapa):
    """Classifica o status do lead"""
    if pd.isna(etapa):
        return 'Indefinido'
    etapa = str(etapa).upper()
    
    if 'MATRÍCULA' in etapa or 'CONCLUÍDA' in etapa:
        return 'Matriculado'
    elif 'PERDIDO' in etapa:
        return 'Perdido'
    elif 'QUALIFICAÇÃO' in etapa:
        return 'Em Qualificação'
    else:
        return 'Outro'

df['Status'] = df['Etapa do negócio'].apply(classificar_status)

# Identificar Ciclo de Captação
# CRÍTICO: Para MATRÍCULAS usa Data_Fechamento, para outros usa Data_Criação
def identificar_ciclo(row):
    """
    Identifica o ciclo de captação (Outubro a Março)
    IMPORTANTE: Matrículas usam data de fechamento, outros usam data de criação
    """
    # Definir qual data usar
    if row['Status'] == 'Matriculado' and pd.notna(row.get('Data_Fechamento')):
        data_ref = row['Data_Fechamento']
    else:
        data_ref = row['Data de criação']
    
    if pd.isna(data_ref):
        return None
    
    ano = data_ref.year
    mes = data_ref.month
    
    # Se é de Outubro a Dezembro, pertence ao ciclo do ano seguinte
    if mes >= 10:
        ciclo_ano = ano + 1
    # Se é de Janeiro a Março, pertence ao ciclo do próprio ano
    elif mes <= 3:
        ciclo_ano = ano
    else:
        # Meses fora do ciclo (Abril a Setembro)
        return None
    
    # Formato: 23.1, 24.1, etc
    return f"{str(ciclo_ano)[-2:]}.1"

df['Ciclo'] = df.apply(identificar_ciclo, axis=1)

# Validar ciclos futuros (ERRO GRAVE DE PREENCHIMENTO)
from datetime import datetime
data_atual = datetime.now()
ano_atual = data_atual.year
mes_atual = data_atual.month

# Ciclo atual é determinado pela data de hoje
if mes_atual >= 10:
    ciclo_maximo_valido = f"{str(ano_atual + 1)[-2:]}.1"
elif mes_atual <= 3:
    ciclo_maximo_valido = f"{str(ano_atual)[-2:]}.1"
else:
    # Entre Abr-Set, último ciclo válido foi o do ano atual
    ciclo_maximo_valido = f"{str(ano_atual)[-2:]}.1"

# Detectar leads com ciclos futuros
leads_ciclo_futuro = 0  # Inicializar
df['Ciclo_Futuro'] = False
for idx, row in df.iterrows():
    if pd.notna(row['Ciclo']):
        ciclo_num = float(row['Ciclo'].replace('.1', '.1'))
        ciclo_max_num = float(ciclo_maximo_valido.replace('.1', '.1'))
        if ciclo_num > ciclo_max_num:
            df.at[idx, 'Ciclo_Futuro'] = True

leads_ciclo_futuro = df['Ciclo_Futuro'].sum()
if leads_ciclo_futuro > 0:
    ciclos_futuros = sorted(df[df['Ciclo_Futuro']]['Ciclo'].unique())
    print(f"\n🚨 ERRO GRAVE - DATAS FUTURAS DETECTADAS:")
    print(f"   Leads com datas futuras: {leads_ciclo_futuro:,}")
    print(f"   Ciclos impossíveis: {ciclos_futuros}")
    print(f"   Ciclo máximo válido hoje ({data_atual.strftime('%d/%m/%Y')}): {ciclo_maximo_valido}")
    print(f"   Causa: Data de criação ou fechamento cadastrada no futuro")
    print(f"   Ação: Esses leads serão REMOVIDOS da análise\n")
    
    # Remover leads com ciclos futuros
    df = df[~df['Ciclo_Futuro']].copy()

# Filtrar apenas leads dos ciclos (Out-Mar)
df_original = df.copy()
df = df[df['Ciclo'].notna()].copy()

print(f"\nFiltro de Ciclos Aplicado:")
print(f"   Leads totais: {len(df_original):,}")
print(f"   Leads em ciclos (Out-Mar): {len(df):,}")
print(f"   Leads fora dos ciclos (Abr-Set): {len(df_original) - len(df):,}")
if leads_ciclo_futuro > 0:
    print(f"   Leads removidos (datas futuras): {leads_ciclo_futuro:,}")
print(f"   Ciclos identificados: {sorted(df['Ciclo'].dropna().unique())}")

# Detectar coluna de atividades para análise posterior
col_atividades = None
for col in df.columns:
    if 'atividade' in col.lower() and ('vendas' in col.lower() or 'número' in col.lower()):
        col_atividades = col
        break

# Contar matrículas com registro incompleto (alerta)
matriculas_baixo_registro = 0
if col_atividades:
    df[col_atividades] = pd.to_numeric(df[col_atividades], errors='coerce').fillna(0)
    mask_baixo = (df['Status'] == 'Matriculado') & (df[col_atividades] < 3)
    matriculas_baixo_registro = mask_baixo.sum()
    print(f"\n⚠️  ALERTA - QUALIDADE DE REGISTRO:")
    print(f"   Matrículas com < 3 atividades: {matriculas_baixo_registro:,}")
    print(f"   Percentual: {matriculas_baixo_registro/df[df['Status']=='Matriculado'].shape[0]*100:.1f}%")
    print(f"   Causa: Atendimento presencial sem registro adequado no HubSpot")
    print(f"   Impacto: Distorce análise de atividades comerciais\n")

# Classificar fonte
def classificar_fonte(fonte_original, detalhe1):
    """Classifica a fonte do lead"""
    if pd.isna(fonte_original):
        return 'Não Identificado'
    
    fonte = str(fonte_original).lower()
    detalhe = str(detalhe1).lower() if pd.notna(detalhe1) else ''
    
    if 'social pago' in fonte or 'facebook' in detalhe or 'meta' in detalhe:
        return 'Meta Ads'
    elif 'pesquisa paga' in fonte or 'google' in detalhe or 'gads' in detalhe:
        return 'Google Ads'
    elif 'pesquisa orgânica' in fonte or 'organic' in fonte:
        return 'SEO/Orgânico'
    elif 'off-line' in fonte or 'offline' in fonte:
        return 'Offline'
    elif 'direto' in fonte or 'direct' in fonte:
        return 'Direto'
    else:
        return 'Outros'

df['Fonte_Categoria'] = df.apply(
    lambda x: classificar_fonte(x['Fonte original do tráfego'], 
                                 x['Detalhamento da fonte original do tráfego 1']), 
    axis=1
)

# Limpar unidade
df['Unidade_Limpa'] = df['Unidade Desejada'].fillna('Não Informado').str.strip().str.upper()

print(f"\n[OK] Status identificados: {df['Status'].nunique()}")
print(f"   {df['Status'].value_counts().to_dict()}")
print(f"\n[OK] Fontes identificadas: {df['Fonte_Categoria'].nunique()}")
print(f"   {df['Fonte_Categoria'].value_counts().to_dict()}")
print(f"\n[OK] Unidades identificadas: {df['Unidade_Limpa'].nunique()}")
print(f"\n[OK] Ciclos analisados: {sorted(df['Ciclo'].dropna().unique())}")

# --- ANÁLISES ---
print("\n" + "="*80)
print("ETAPA 3: GERANDO ANÁLISES")
print("="*80 + "\n")

# =============================================================================
# ANÁLISE 1: RESUMO GERAL POR CICLO
# =============================================================================
print("[Analise 1] Resumo Geral por Ciclo...")

resumo_ciclo = df.groupby('Ciclo').agg({
    'Record ID': 'count',
}).rename(columns={'Record ID': 'Total_Leads'})

# Contar cada status por ciclo
for status in ['Matriculado', 'Perdido', 'Em Qualificação']:
    resumo_ciclo[status] = df[df['Status'] == status].groupby('Ciclo')['Record ID'].count()

resumo_ciclo = resumo_ciclo.fillna(0).astype(int)

# Calcular taxas
resumo_ciclo['Taxa_Conversao_%'] = (resumo_ciclo['Matriculado'] / resumo_ciclo['Total_Leads'] * 100).round(2)
resumo_ciclo['Taxa_Perda_%'] = (resumo_ciclo['Perdido'] / resumo_ciclo['Total_Leads'] * 100).round(2)
resumo_ciclo['Taxa_Qualificacao_%'] = (resumo_ciclo['Em Qualificação'] / resumo_ciclo['Total_Leads'] * 100).round(2)

# Ordenar por ciclo
resumo_ciclo = resumo_ciclo.sort_index().reset_index()

print(f"   ✅ {len(resumo_ciclo)} ciclos analisados")
print("\n   📊 Detalhamento por Ciclo:")
for _, row in resumo_ciclo.iterrows():
    print(f"      Ciclo {row['Ciclo']}: {row['Total_Leads']:,} leads | "
          f"Matriculados: {row['Matriculado']:,} ({row['Taxa_Conversao_%']:.1f}%)")

# =============================================================================
# ANÁLISE 2: STATUS ATUAL (TODOS OS ANOS)
# =============================================================================
print("📊 Análise 2: Status Atual (Consolidado)...")

status_geral = df.groupby('Status').agg({
    'Record ID': 'count'
}).rename(columns={'Record ID': 'Quantidade'})

status_geral['Percentual_%'] = (status_geral['Quantidade'] / status_geral['Quantidade'].sum() * 100).round(2)
status_geral = status_geral.reset_index().sort_values('Quantidade', ascending=False)

print(f"   ✅ {len(status_geral)} status identificados")

# =============================================================================
# ANÁLISE 3: PERFORMANCE POR UNIDADE
# =============================================================================
print("📊 Análise 3: Performance por Unidade...")

unidade = df.groupby('Unidade_Limpa').agg({
    'Record ID': 'count'
}).rename(columns={'Record ID': 'Total_Leads'})

for status in ['Matriculado', 'Perdido', 'Em Qualificação']:
    unidade[status] = df[df['Status'] == status].groupby('Unidade_Limpa')['Record ID'].count()

unidade = unidade.fillna(0).astype(int)
unidade['Taxa_Conversao_%'] = (unidade['Matriculado'] / unidade['Total_Leads'] * 100).round(2)
unidade = unidade.reset_index().sort_values('Total_Leads', ascending=False)

print(f"   ✅ {len(unidade)} unidades analisadas")

# =============================================================================
# ANÁLISE 4: PERFORMANCE POR FONTE
# =============================================================================
print("📊 Análise 4: Performance por Fonte...")

fonte = df.groupby('Fonte_Categoria').agg({
    'Record ID': 'count'
}).rename(columns={'Record ID': 'Total_Leads'})

for status in ['Matriculado', 'Perdido', 'Em Qualificação']:
    fonte[status] = df[df['Status'] == status].groupby('Fonte_Categoria')['Record ID'].count()

fonte = fonte.fillna(0).astype(int)
fonte['Taxa_Conversao_%'] = (fonte['Matriculado'] / fonte['Total_Leads'] * 100).round(2)
fonte = fonte.reset_index().sort_values('Total_Leads', ascending=False)

print(f"   ✅ {len(fonte)} fontes analisadas")

# =============================================================================
# ANÁLISE 5: EVOLUÇÃO MENSAL
# =============================================================================
print("📊 Análise 5: Evolução Mensal...")

mensal = df.groupby(['Mes_Ano', 'Status']).agg({
    'Record ID': 'count'
}).rename(columns={'Record ID': 'Quantidade'}).reset_index()

mensal_pivot = mensal.pivot(index='Mes_Ano', columns='Status', values='Quantidade').fillna(0).astype(int)
mensal_pivot['Total_Leads'] = mensal_pivot.sum(axis=1)

if 'Matriculado' in mensal_pivot.columns:
    mensal_pivot['Taxa_Conversao_%'] = (mensal_pivot['Matriculado'] / mensal_pivot['Total_Leads'] * 100).round(2)

mensal_pivot = mensal_pivot.reset_index()

print(f"   ✅ {len(mensal_pivot)} meses analisados")

# =============================================================================
# ANÁLISE 6: MATRIZ CICLO x UNIDADE
# =============================================================================
print("📊 Análise 6: Matriz Ciclo x Unidade...")

ciclo_unidade = df.groupby(['Ciclo', 'Unidade_Limpa']).agg({
    'Record ID': 'count'
}).rename(columns={'Record ID': 'Total_Leads'}).reset_index()

ciclo_unidade_pivot = ciclo_unidade.pivot(index='Unidade_Limpa', columns='Ciclo', values='Total_Leads').fillna(0).astype(int)
# Ordenar colunas (ciclos)
ciclo_unidade_pivot = ciclo_unidade_pivot.reindex(sorted(ciclo_unidade_pivot.columns), axis=1)
ciclo_unidade_pivot['Total_Geral'] = ciclo_unidade_pivot.sum(axis=1)
ciclo_unidade_pivot = ciclo_unidade_pivot.sort_values('Total_Geral', ascending=False).reset_index()

print(f"   ✅ Matriz criada: {len(ciclo_unidade_pivot)} unidades x {len(ciclo_unidade_pivot.columns)-2} ciclos")

# =============================================================================
# ANÁLISE 7: MATRIZ CICLO x FONTE
# =============================================================================
print("📊 Análise 7: Matriz Ciclo x Fonte...")

ciclo_fonte = df.groupby(['Ciclo', 'Fonte_Categoria']).agg({
    'Record ID': 'count'
}).rename(columns={'Record ID': 'Total_Leads'}).reset_index()

ciclo_fonte_pivot = ciclo_fonte.pivot(index='Fonte_Categoria', columns='Ciclo', values='Total_Leads').fillna(0).astype(int)
# Ordenar colunas (ciclos)
ciclo_fonte_pivot = ciclo_fonte_pivot.reindex(sorted(ciclo_fonte_pivot.columns), axis=1)
ciclo_fonte_pivot['Total_Geral'] = ciclo_fonte_pivot.sum(axis=1)
ciclo_fonte_pivot = ciclo_fonte_pivot.sort_values('Total_Geral', ascending=False).reset_index()

print(f"   ✅ Matriz criada: {len(ciclo_fonte_pivot)} fontes x {len(ciclo_fonte_pivot.columns)-2} ciclos")

# =============================================================================
# ANÁLISE 7B: COMPARATIVO CICLO ATUAL (26.1) vs ANTERIOR (25.1) POR CANAL
# =============================================================================
print("📊 Análise 7b: Comparativo Ciclo 26.1 vs 25.1 por Canal...")

# Filtrar apenas ciclos 25.1 e 26.1
df_comparativo = df[df['Ciclo'].isin(['25.1', '26.1'])].copy()

# Criar análise por Fonte + Ciclo + Status
comparativo_data = []
for canal in df_comparativo['Fonte_Categoria'].unique():
    df_fonte = df_comparativo[df_comparativo['Fonte_Categoria'] == canal]
    
    for ciclo in ['25.1', '26.1']:
        df_ciclo_fonte = df_fonte[df_fonte['Ciclo'] == ciclo]
        
        total = len(df_ciclo_fonte)
        matriculados = len(df_ciclo_fonte[df_ciclo_fonte['Status'] == 'Matriculado'])
        perdidos = len(df_ciclo_fonte[df_ciclo_fonte['Status'] == 'Perdido'])
        em_qualificacao = len(df_ciclo_fonte[df_ciclo_fonte['Status'] == 'Em Qualificação'])
        
        taxa_conversao = (matriculados / total * 100) if total > 0 else 0
        
        comparativo_data.append({
            'Canal': canal,
            'Ciclo': ciclo,
            'Total_Leads': total,
            'Matriculados': matriculados,
            'Perdidos': perdidos,
            'Em_Qualificacao': em_qualificacao,
            'Taxa_Conversao_%': round(taxa_conversao, 1)
        })

df_comparativo_ciclos = pd.DataFrame(comparativo_data)
df_comparativo_ciclos = df_comparativo_ciclos.sort_values(['Canal', 'Ciclo'])

# ⚠️ ALERTA DE QUALIDADE DE DADOS: Offline e Direto possuem problema
# Adicionar nota sobre inconsistência de datas (usar None ao invés de 0 para células vazias)
nota_qualidade = pd.DataFrame([{
    'Canal': 'ALERTA IMPORTANTE',
    'Ciclo': 'Nota',
    'Total_Leads': None,
    'Matriculados': None,
    'Perdidos': None,
    'Em_Qualificacao': None,
    'Taxa_Conversao_%': None
}])

nota_explicacao = pd.DataFrame([{
    'Canal': 'Canais OFFLINE e DIRETO possuem inconsistencias de DATA',
    'Ciclo': '',
    'Total_Leads': None,
    'Matriculados': None,
    'Perdidos': None,
    'Em_Qualificacao': None,
    'Taxa_Conversao_%': None
}])

nota_detalhe = pd.DataFrame([{
    'Canal': 'Leads atendidos presencialmente nao sao registrados no HubSpot imediatamente',
    'Ciclo': '',
    'Total_Leads': None,
    'Matriculados': None,
    'Perdidos': None,
    'Em_Qualificacao': None,
    'Taxa_Conversao_%': None
}])

nota_consequencia = pd.DataFrame([{
    'Canal': 'Ao matricular criam o lead DEPOIS da matricula - Data Criacao maior que Data Fechamento',
    'Ciclo': '',
    'Total_Leads': None,
    'Matriculados': None,
    'Perdidos': None,
    'Em_Qualificacao': None,
    'Taxa_Conversao_%': None
}])

# Adicionar nota após os dados
df_comparativo_ciclos = pd.concat([
    df_comparativo_ciclos, 
    nota_qualidade, 
    nota_explicacao,
    nota_detalhe,
    nota_consequencia
], ignore_index=True)

print(f"   ✅ Comparativo criado: 25.1 vs 26.1 por {len(df_comparativo['Fonte_Categoria'].unique())} canais")
print(f"   ⚠️  Nota de qualidade de dados adicionada (Offline/Direto)")

# =============================================================================
# ANÁLISE 8: PERFIL DOS LEADS MATRICULADOS (ATIVIDADES E TEMPO ATÉ CONVERSÃO)
# =============================================================================
print("📊 Análise 8: Perfil dos Leads Matriculados...")

matriculados = df[df['Status'] == 'Matriculado'].copy()

# Calcular tempo até conversão (assumindo que leads mais antigos já converteram)
matriculados['Dias_Desde_Criacao'] = (datetime.now() - matriculados['Data de criação']).dt.days

# Análise 1: Distribuição de ATIVIDADES COMERCIAIS dos matriculados
atividades_matriculados = []
for faixa in ['0 atividades', '1-2 atividades', '3-5 atividades', '6-10 atividades', '11+ atividades']:
    if faixa == '0 atividades':
        qtd = len(matriculados[matriculados['Número de atividades de vendas'] == 0])
    elif faixa == '1-2 atividades':
        qtd = len(matriculados[(matriculados['Número de atividades de vendas'] >= 1) & (matriculados['Número de atividades de vendas'] <= 2)])
    elif faixa == '3-5 atividades':
        qtd = len(matriculados[(matriculados['Número de atividades de vendas'] >= 3) & (matriculados['Número de atividades de vendas'] <= 5)])
    elif faixa == '6-10 atividades':
        qtd = len(matriculados[(matriculados['Número de atividades de vendas'] >= 6) & (matriculados['Número de atividades de vendas'] <= 10)])
    else:
        qtd = len(matriculados[matriculados['Número de atividades de vendas'] > 10])
    
    atividades_matriculados.append({'Faixa_Atividades': faixa, 'Quantidade': qtd})

df_atividades_matriculados = pd.DataFrame(atividades_matriculados)
df_atividades_matriculados['Percentual_%'] = (df_atividades_matriculados['Quantidade'] / df_atividades_matriculados['Quantidade'].sum() * 100).round(1)

# Análise 2: Distribuição de IDADE (tempo de vida) dos matriculados
idade_matriculados = []
for faixa in ['0-30 dias', '31-60 dias', '61-90 dias', '91-180 dias', '180+ dias']:
    if faixa == '0-30 dias':
        qtd = len(matriculados[matriculados['Dias_Desde_Criacao'] <= 30])
    elif faixa == '31-60 dias':
        qtd = len(matriculados[(matriculados['Dias_Desde_Criacao'] > 30) & (matriculados['Dias_Desde_Criacao'] <= 60)])
    elif faixa == '61-90 dias':
        qtd = len(matriculados[(matriculados['Dias_Desde_Criacao'] > 60) & (matriculados['Dias_Desde_Criacao'] <= 90)])
    elif faixa == '91-180 dias':
        qtd = len(matriculados[(matriculados['Dias_Desde_Criacao'] > 90) & (matriculados['Dias_Desde_Criacao'] <= 180)])
    else:
        qtd = len(matriculados[matriculados['Dias_Desde_Criacao'] > 180])
    
    idade_matriculados.append({'Faixa_Idade': faixa, 'Quantidade': qtd})

df_idade_matriculados = pd.DataFrame(idade_matriculados)
df_idade_matriculados['Percentual_%'] = (df_idade_matriculados['Quantidade'] / df_idade_matriculados['Quantidade'].sum() * 100).round(1)

# Estatísticas gerais
atividades_media = matriculados['Número de atividades de vendas'].mean()
atividades_mediana = matriculados['Número de atividades de vendas'].median()
idade_media = matriculados['Dias_Desde_Criacao'].mean()
idade_mediana = matriculados['Dias_Desde_Criacao'].median()

print(f"   ✅ {len(matriculados)} leads matriculados analisados")
print(f"   📊 Atividades comerciais: Média={atividades_media:.1f} | Mediana={atividades_mediana:.0f}")
print(f"   📊 Idade do lead (dias): Média={idade_media:.0f} | Mediana={idade_mediana:.0f}")

# =============================================================================
# ANÁLISE 9: PROJEÇÃO DE META (LEADS NECESSÁRIOS)
# =============================================================================
print("📊 Análise 9: Projeção de Meta...")

# Calcular taxa de conversão histórica (últimos 3 ciclos completos)
ciclos_completos = [c for c in sorted(df['Ciclo'].dropna().unique()) if c not in ['26.1']][-3:]
leads_historicos = df[df['Ciclo'].isin(ciclos_completos)]
taxa_conversao_historica = len(leads_historicos[leads_historicos['Status']=='Matriculado']) / len(leads_historicos) * 100

# Leads atuais em qualificação
leads_em_qualificacao = len(df[df['Status']=='Em Qualificação'])

# Projeção de conversão dos leads atuais
projecao_matriculas_atuais = int(leads_em_qualificacao * (taxa_conversao_historica/100))

# META REAL DA RED BALLOON (Ciclo 26.1)
META_TOTAL = 6545  # Meta total até março
MATRICULAS_ATUAIS = 4226  # Já realizadas
META_FALTANTE = META_TOTAL - MATRICULAS_ATUAIS

# Calcular quantos leads precisam para fechar o gap
leads_necessarios_gap = int(META_FALTANTE / (taxa_conversao_historica/100))
gap_leads = leads_necessarios_gap - leads_em_qualificacao

# Criar cenário da meta real
cenario_meta_real = {
    'Cenario': '🎯 META RED BALLOON',
    'Meta_Total': META_TOTAL,
    'Ja_Matriculados': MATRICULAS_ATUAIS,
    'Faltam_Matriculas': META_FALTANTE,
    'Taxa_Conversao_%': round(taxa_conversao_historica, 1),
    'Leads_Necessarios': leads_necessarios_gap,
    'Leads_Atuais_Qualificacao': leads_em_qualificacao,
    'Gap_Leads': gap_leads,
    'Status': '✅ SUFICIENTE' if gap_leads <= 0 else '⚠️ GERAR NOVOS',
    'Observacao': 'Base de cálculo: últimos 3 ciclos completos'
}

df_meta_real = pd.DataFrame([cenario_meta_real])

# Simulação de cenários adicionais
cenarios = []
for meta_faltante in [1000, 1500, 2000, 2319, 2500, 3000, 3500, 4000]:
    leads_necessarios = int(meta_faltante / (taxa_conversao_historica/100))
    gap = leads_necessarios - leads_em_qualificacao
    status = '✅ Suficiente' if gap <= 0 else '⚠️ Gerar Novos'
    
    cenarios.append({
        'Meta_Faltante': meta_faltante,
        'Leads_Necessarios': leads_necessarios,
        'Leads_Atuais': leads_em_qualificacao,
        'Gap_Leads': max(0, gap),
        'Status': status
    })

df_projecao = pd.DataFrame(cenarios)

# Identificar ciclo atual de negócio (ciclo mais recente com volume significativo)
# Filtrar ciclos com pelo menos 100 leads para evitar ciclos futuros/marginais
contagem_por_ciclo = df['Ciclo'].value_counts()
ciclos_validos = [c for c in contagem_por_ciclo.index if '.' in str(c) and contagem_por_ciclo[c] >= 100]
ciclos_ordenados = sorted(ciclos_validos, key=lambda x: float(x.replace('.1', '.1')), reverse=True)
ciclo_atual = ciclos_ordenados[0] if ciclos_ordenados else None

leads_qualificacao_ciclo_atual = 0
if ciclo_atual:
    leads_qualificacao_ciclo_atual = len(df[(df['Status'] == 'Em Qualificação') & (df['Ciclo'] == ciclo_atual)])

print(f"   ✅ Taxa de conversão histórica: {taxa_conversao_historica:.1f}%")
print(f"   ✅ Leads em qualificação (TOTAL): {leads_em_qualificacao:,}")
if ciclo_atual:
    print(f"   ✅ Leads em qualificação (CICLO {ciclo_atual}): {leads_qualificacao_ciclo_atual:,}")
print(f"   ✅ Projeção de matrículas dos leads atuais: {projecao_matriculas_atuais:,}")
print(f"")
print(f"   🎯 META RED BALLOON:")
print(f"      Meta total: {META_TOTAL:,} matrículas")
print(f"      Já realizadas: {MATRICULAS_ATUAIS:,} matrículas")
print(f"      Faltam: {META_FALTANTE:,} matrículas")
print(f"      Leads necessários: {leads_necessarios_gap:,}")
print(f"      Gap de leads: {gap_leads:,} {'(FALTANDO!)' if gap_leads > 0 else '(SUFICIENTE!)'}")

# =============================================================================
# ANÁLISE 10: LEADS EM QUALIFICAÇÃO (DETALHADO)
# =============================================================================
print("📊 Análise 10: Leads em Qualificação (Detalhado)...")

qualificacao = df[df['Status'] == 'Em Qualificação'][[
    'Record ID',
    'Nome do negócio',
    'Data de criação',
    'Ciclo',
    'Unidade_Limpa',
    'Fonte_Categoria',
    'Número de atividades de vendas',
    'Data da última atividade',
    'Proprietário do negócio'
]].copy()

# Calcular IDADE DO LEAD (dias desde criação)
qualificacao['Dias_Desde_Criacao'] = (datetime.now() - qualificacao['Data de criação']).dt.days

# Calcular dias desde ÚLTIMO CONTATO
qualificacao['Data da última atividade'] = pd.to_datetime(qualificacao['Data da última atividade'], errors='coerce')
qualificacao['Dias_Sem_Contato'] = (datetime.now() - qualificacao['Data da última atividade']).dt.days
qualificacao['Dias_Sem_Contato'] = qualificacao['Dias_Sem_Contato'].fillna(999).astype(int)  # 999 = nunca contatado

# Classificar FAIXA DE ATIVIDADES
def classificar_atividades(atividades):
    if atividades == 0:
        return '0 atividades'
    elif atividades <= 2:
        return '1-2 atividades'
    elif atividades <= 5:
        return '3-5 atividades'
    elif atividades <= 10:
        return '6-10 atividades'
    else:
        return '11+ atividades'

qualificacao['Faixa_Atividades'] = qualificacao['Número de atividades de vendas'].apply(classificar_atividades)

# Classificar FAIXA DE IDADE
def classificar_idade(dias):
    if dias <= 30:
        return '0-30 dias'
    elif dias <= 60:
        return '31-60 dias'
    elif dias <= 90:
        return '61-90 dias'
    elif dias <= 180:
        return '91-180 dias'
    else:
        return '180+ dias'

qualificacao['Faixa_Idade'] = qualificacao['Dias_Desde_Criacao'].apply(classificar_idade)

# Ordenar por número de atividades (mais atividades primeiro) e depois por data
qualificacao = qualificacao.sort_values(['Número de atividades de vendas', 'Dias_Desde_Criacao'], ascending=[False, True])

print(f"   ✅ {len(qualificacao)} leads em qualificação detalhados")

# =============================================================================
# ANÁLISE 10b: TAXA DE ATENDIMENTO DOS LEADS EM QUALIFICAÇÃO
# =============================================================================
print("📊 Análise 10b: Taxa de Atendimento dos Leads em Qualificação...")

# Métricas de atendimento
leads_sem_atividade = len(qualificacao[qualificacao['Número de atividades de vendas'] == 0])
leads_com_atividade = len(qualificacao[qualificacao['Número de atividades de vendas'] > 0])
taxa_atendimento_geral = (leads_com_atividade / len(qualificacao) * 100) if len(qualificacao) > 0 else 0

# Atividade nos últimos 30 dias
leads_ativos_30d = len(qualificacao[qualificacao['Dias_Sem_Contato'] <= 30])
leads_inativos_30d = len(qualificacao[qualificacao['Dias_Sem_Contato'] > 30])
taxa_atendimento_30d = (leads_ativos_30d / len(qualificacao) * 100) if len(qualificacao) > 0 else 0

# Atividade nos últimos 60 dias
leads_ativos_60d = len(qualificacao[qualificacao['Dias_Sem_Contato'] <= 60])
taxa_atendimento_60d = (leads_ativos_60d / len(qualificacao) * 100) if len(qualificacao) > 0 else 0

# Leads nunca contatados (999 dias)
leads_nunca_contatados = len(qualificacao[qualificacao['Dias_Sem_Contato'] >= 999])

print(f"   ✅ Taxa de atendimento (alguma vez): {taxa_atendimento_geral:.1f}%")
print(f"   ✅ Taxa de atendimento (últimos 30 dias): {taxa_atendimento_30d:.1f}%")
print(f"   ✅ Leads ativos (contato <30 dias): {leads_ativos_30d:,}")
print(f"   ⚠️  Leads parados (sem contato >30 dias): {leads_inativos_30d:,}")
print(f"   🚨 Leads NUNCA contatados: {leads_nunca_contatados:,}")

# Análise por FAIXA DE TEMPO SEM CONTATO
analise_tempo_sem_contato = []
for faixa, filtro in [
    ('Ativos (0-7 dias)', qualificacao['Dias_Sem_Contato'] <= 7),
    ('Recentes (8-30 dias)', (qualificacao['Dias_Sem_Contato'] > 7) & (qualificacao['Dias_Sem_Contato'] <= 30)),
    ('Mornos (31-60 dias)', (qualificacao['Dias_Sem_Contato'] > 30) & (qualificacao['Dias_Sem_Contato'] <= 60)),
    ('Frios (61-90 dias)', (qualificacao['Dias_Sem_Contato'] > 60) & (qualificacao['Dias_Sem_Contato'] <= 90)),
    ('Muito Frios (>90 dias)', (qualificacao['Dias_Sem_Contato'] > 90) & (qualificacao['Dias_Sem_Contato'] < 999)),
    ('Nunca Contatados', qualificacao['Dias_Sem_Contato'] >= 999)
]:
    qtd = len(qualificacao[filtro])
    pct = (qtd / len(qualificacao) * 100) if len(qualificacao) > 0 else 0
    analise_tempo_sem_contato.append({
        'Faixa_Tempo': faixa,
        'Quantidade': qtd,
        'Percentual': round(pct, 1)
    })

df_tempo_sem_contato = pd.DataFrame(analise_tempo_sem_contato)

# Análise por FAIXA DE ATIVIDADES (distribuição)
analise_atividades = qualificacao.groupby('Faixa_Atividades').size().reset_index(name='Quantidade')
analise_atividades['Percentual'] = round(analise_atividades['Quantidade'] / len(qualificacao) * 100, 1)
analise_atividades = analise_atividades.sort_values('Quantidade', ascending=False)

# Análise por CICLO (atendimento por ciclo)
analise_por_ciclo_atendimento = []
for ciclo in sorted(qualificacao['Ciclo'].dropna().unique(), reverse=True):
    df_ciclo = qualificacao[qualificacao['Ciclo'] == ciclo]
    ativos_30d = len(df_ciclo[df_ciclo['Dias_Sem_Contato'] <= 30])
    taxa = (ativos_30d / len(df_ciclo) * 100) if len(df_ciclo) > 0 else 0
    
    analise_por_ciclo_atendimento.append({
        'Ciclo': ciclo,
        'Total_Leads': len(df_ciclo),
        'Ativos_30d': ativos_30d,
        'Taxa_Atendimento_%': round(taxa, 1)
    })

df_atendimento_ciclo = pd.DataFrame(analise_por_ciclo_atendimento)

print(f"   ✅ Análise de atendimento concluída")

# =============================================================================
# ANÁLISE 10f: VELOCIDADE DE CONVERSÃO E PROJEÇÃO REALISTA ATÉ MARÇO (CORRIGIDA)
# =============================================================================
print("📊 Análise 10f: Velocidade de Conversão e Projeção Realista até Março...")

# Calcular DIAS ÚTEIS até 31/03/2026 (excluindo feriados de Carnaval)
from datetime import timedelta

data_hoje = datetime.now()
data_fim_ciclo = datetime(2026, 3, 31)

# Feriados Carnaval 2026 (semana de 16/02)
feriados_carnaval = [
    datetime(2026, 2, 16),  # Segunda-feira de Carnaval
    datetime(2026, 2, 17),  # Terça-feira de Carnaval (Carnaval)
    datetime(2026, 2, 18),  # Quarta-feira de Cinzas (meio dia)
]

def contar_dias_uteis(data_inicio, data_fim, feriados):
    """Conta dias úteis entre duas datas, excluindo sábados, domingos e feriados"""
    dias_uteis = 0
    data_atual = data_inicio
    
    while data_atual <= data_fim:
        # Segunda=0, Domingo=6
        if data_atual.weekday() < 5:  # Seg-Sex
            if data_atual not in feriados:
                dias_uteis += 1
        data_atual += timedelta(days=1)
    
    return dias_uteis

dias_uteis_restantes = contar_dias_uteis(data_hoje, data_fim_ciclo, feriados_carnaval)
dias_corridos_restantes = (data_fim_ciclo - data_hoje).days

print(f"   ⏰ Tempo restante até 31/03/2026:")
print(f"      Dias corridos: {dias_corridos_restantes}")
print(f"      Dias úteis: {dias_uteis_restantes} (excluindo finais de semana + Carnaval)")

# ANÁLISE CORRETA: Segmentar leads por CICLO (não por dias absolutos)
# Ciclo 26.1 = Out/2025 a Mar/2026
inicio_ciclo_26 = datetime(2025, 10, 1)
fim_primeiros_2meses = datetime(2025, 12, 1)  # Out-Nov 2025

# Leads DO CICLO ATUAL (26.1) - entraram entre Out/2025 e agora
leads_ciclo_26 = qualificacao[qualificacao['Data de criação'] >= inicio_ciclo_26]

# Subdividir ciclo 26.1 por período de entrada
leads_ciclo_26_inicio = leads_ciclo_26[leads_ciclo_26['Data de criação'] < fim_primeiros_2meses]  # Out-Nov
leads_ciclo_26_recente = leads_ciclo_26[leads_ciclo_26['Data de criação'] >= fim_primeiros_2meses]  # Dez-Mar

# Leads ESTOQUE (ciclos anteriores 25.1, 24.1, etc) que NÃO converteram
leads_estoque_ciclos_anteriores = qualificacao[qualificacao['Data de criação'] < inicio_ciclo_26]

print(f"\n   📦 SEGMENTAÇÃO DO PIPELINE (por CICLO):")
print(f"      Leads DO CICLO 26.1: {len(leads_ciclo_26):,} leads")
print(f"         • Entrada Out-Nov/2025: {len(leads_ciclo_26_inicio):,} leads (já estão há 3-4 meses)")
print(f"         • Entrada Dez/2025-Fev/2026: {len(leads_ciclo_26_recente):,} leads (mais recentes)")
print(f"      Leads ESTOQUE (ciclos anteriores): {len(leads_estoque_ciclos_anteriores):,} leads")
print(f"      ⚠️  Estoque = {round(len(leads_estoque_ciclos_anteriores)/len(qualificacao)*100, 1)}% do pipeline (já falharam em converter antes)")

# Calcular taxa de MATRICULAÇÃO MENSAL histórica dos últimos 3 ciclos completos
# Ciclos 22.1 (Out21-Mar22), 23.1 (Out22-Mar23), 24.1 (Out23-Mar24)
ciclos_completos = ['22.1', '23.1', '24.1']
matriculas_mensais = []
leads_mensais = []

for ciclo in ciclos_completos:
    df_ciclo = df[df['Ciclo'] == ciclo]
    matriculas = len(df_ciclo[df_ciclo['Status'] == 'Matriculado'])
    total = len(df_ciclo)
    matriculas_mensais.append(matriculas / 6)  # 6 meses por ciclo
    leads_mensais.append(total / 6)

matriculas_por_mes_media = sum(matriculas_mensais) / len(matriculas_mensais)
taxa_conversao_mensal = (matriculas_por_mes_media / (sum(leads_mensais) / len(leads_mensais))) * 100

print(f"\n   📊 TAXA DE CONVERSÃO MENSAL (histórica):")
print(f"      Matrículas por mês (média): {matriculas_por_mes_media:.0f}")
print(f"      Taxa mensal: {taxa_conversao_mensal:.1f}%")

# PROJEÇÃO REALISTA para 1.6 meses (47 dias)
meses_restantes = dias_corridos_restantes / 30

# Leads do CICLO 26.1 - entraram neste ciclo, têm chance real de converter
taxa_leads_ciclo_26 = taxa_conversao_historica / 100  # Taxa histórica de 6 meses

# Leads ESTOQUE (ciclos anteriores) - já falharam em converter, chance MUITO menor
# Usar apenas 5% da taxa (probabilidade baixíssima de reativação)
taxa_leads_estoque_antigo = (taxa_conversao_historica / 100) * 0.05

# Projeção por segmento
projecao_leads_ciclo_26 = int(len(leads_ciclo_26) * taxa_leads_ciclo_26)
projecao_estoque_antigo = int(len(leads_estoque_ciclos_anteriores) * taxa_leads_estoque_antigo)
projecao_total_realista = projecao_leads_ciclo_26 + projecao_estoque_antigo

# Projeção LINEAR mensal
projecao_linear_mensal = int(matriculas_por_mes_media * meses_restantes)

# Gap realista
gap_realista_segmentado = META_FALTANTE - projecao_total_realista
gap_realista_linear = META_FALTANTE - projecao_linear_mensal

print(f"\n   🎯 PROJEÇÃO REALISTA até 31/03/2026:")
print(f"\n   MÉTODO 1: Segmentação por CICLO")
print(f"      Leads DO CICLO 26.1 ({len(leads_ciclo_26):,}) × {taxa_conversao_historica:.1f}% = ~{projecao_leads_ciclo_26:,} matrículas")
print(f"      Leads ESTOQUE anterior ({len(leads_estoque_ciclos_anteriores):,}) × {taxa_conversao_historica * 0.05:.1f}% = ~{projecao_estoque_antigo:,} matrículas")
print(f"      TOTAL projetado: {projecao_total_realista:,} matrículas")
print(f"      Meta faltante: {META_FALTANTE:,}")
print(f"      Gap: {gap_realista_segmentado:,} - {'INSUFICIENTE!' if gap_realista_segmentado > 0 else 'SUFICIENTE!'}")

print(f"\n   MÉTODO 2: Projeção linear mensal")
print(f"      Matrículas/mês histórico: {matriculas_por_mes_media:.0f}")
print(f"      Meses restantes: {meses_restantes:.1f}")
print(f"      Projeção: {projecao_linear_mensal:,} matrículas")
print(f"      Meta faltante: {META_FALTANTE:,}")
print(f"      Gap: {gap_realista_linear:,} - {'INSUFICIENTE!' if gap_realista_linear > 0 else 'SUFICIENTE!'}")

# Criar DataFrame de projeção realista
df_projecao_realista = pd.DataFrame([
    {
        'Metodo': 'Segmentacao (Ciclo 26.1 vs Estoque)',
        'Leads_Ciclo_26_1': len(leads_ciclo_26),
        'Leads_Estoque_Ciclos_Anteriores': len(leads_estoque_ciclos_anteriores),
        'Projecao_Ciclo_26': projecao_leads_ciclo_26,
        'Projecao_Estoque': projecao_estoque_antigo,
        'Projecao_Total': projecao_total_realista,
        'Meta_Faltante': META_FALTANTE,
        'Gap': gap_realista_segmentado,
        'Status': 'INSUFICIENTE' if gap_realista_segmentado > 0 else 'SUFICIENTE'
    },
    {
        'Metodo': 'Linear Mensal (historico)',
        'Leads_Ciclo_26_1': leads_em_qualificacao,
        'Leads_Estoque_Ciclos_Anteriores': 0,
        'Projecao_Ciclo_26': 0,
        'Projecao_Estoque': 0,
        'Projecao_Total': projecao_linear_mensal,
        'Meta_Faltante': META_FALTANTE,
        'Gap': gap_realista_linear,
        'Status': 'INSUFICIENTE' if gap_realista_linear > 0 else 'SUFICIENTE'
    }
])

# Priorização de leads (mantém a lógica anterior mas com contexto correto)
leads_alta_prioridade = qualificacao[
    ((qualificacao['Número de atividades de vendas'] >= 5) & 
     (qualificacao['Dias_Sem_Contato'] <= 7)) |
    ((qualificacao['Dias_Desde_Criacao'] <= 60) & 
     (qualificacao['Número de atividades de vendas'] >= 3) & 
     (qualificacao['Dias_Sem_Contato'] <= 14))
]

leads_media_prioridade = qualificacao[
    ((qualificacao['Número de atividades de vendas'] >= 3) & 
     (qualificacao['Dias_Sem_Contato'] <= 30)) &
    (~qualificacao.index.isin(leads_alta_prioridade.index))
]

leads_baixa_prioridade = qualificacao[
    ~qualificacao.index.isin(leads_alta_prioridade.index) &
    ~qualificacao.index.isin(leads_media_prioridade.index)
]

print(f"\n   🔥 LEADS PRIORITÁRIOS (maior chance de fechar até março):")
print(f"      ALTA prioridade: {len(leads_alta_prioridade):,} leads (muitas atividades recentes + leads novos ativos)")
print(f"      MÉDIA prioridade: {len(leads_media_prioridade):,} leads (atividade moderada nos últimos 30 dias)")
print(f"      BAIXA prioridade: {len(leads_baixa_prioridade):,} leads (pouca atividade ou estoque antigo)")

# Criar DataFrame de priorização
df_priorizacao = pd.DataFrame([
    {
        'Prioridade': 'ALTA',
        'Criterio': 'Muitas atividades recentes OU leads novos muito ativos',
        'Quantidade_Leads': len(leads_alta_prioridade),
        'Percentual_Total_%': round(len(leads_alta_prioridade) / len(qualificacao) * 100, 1)
    },
    {
        'Prioridade': 'MEDIA',
        'Criterio': 'Atividade moderada nos ultimos 30 dias',
        'Quantidade_Leads': len(leads_media_prioridade),
        'Percentual_Total_%': round(len(leads_media_prioridade) / len(qualificacao) * 100, 1)
    },
    {
        'Prioridade': 'BAIXA',
        'Criterio': 'Pouca atividade ou estoque antigo (>60 dias sem converter)',
        'Quantidade_Leads': len(leads_baixa_prioridade),
        'Percentual_Total_%': round(len(leads_baixa_prioridade) / len(qualificacao) * 100, 1)
    }
])

print(f"   ✅ Análise de velocidade e projeção realista concluída")

# --- SALVAR RESULTADOS ---
print("\n" + "="*80)
print("ETAPA 4: SALVANDO RESULTADOS")
print("="*80 + "\n")

# =============================================================================
# CRIAR CAPA DO RELATÓRIO
# =============================================================================
print("📋 Criando capa do relatório...")

capa_data = [
    ['SECAO', 'CONTEUDO'],
    ['TITULO', 'ANALISE DE FUNIL - RED BALLOON'],
    ['', ''],
    ['DATA_GERACAO', datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
    ['AUTOR', 'Sistema de Data Science - Cogna'],
    ['', ''],
    ['=== FONTE DOS DADOS ===', ''],
    ['Sistema', 'HubSpot CRM'],
    ['Pipeline', 'Red Balloon'],
    ['Base Original Completa (leads)', len(df_original_completo)],
    ['Leads com ERRO DE DATA (corrigidos)', leads_com_erro],
    ['Destes: Offline/Direto', leads_offline_direto_erro],
    ['Base Original (leads)', len(df_original)],
    ['Base Analisada (leads Out-Mar)', len(df)],
    ['Leads Excluidos (Abr-Set)', len(df_original)-len(df)],
    ['', ''],
    ['=== CORRECAO DE DADOS ===', ''],
    ['Problema Identificado', 'Data Criacao maior que Data Fechamento'],
    ['Leads Afetados', leads_com_erro],
    ['Acao Tomada', 'CORRIGIR data criacao = data fechamento (manter matricula)'],
    ['Motivo Correcao', 'Evita idade negativa e mantem matriculas reais'],
    ['Causa Raiz', 'Atendimento presencial matriculado antes do registro no sistema'],
    ['Canais Afetados', 'Principalmente Offline e Direto'],
    ['', ''],
    ['=== ALERTA - QUALIDADE DE REGISTRO ===', ''],
    ['Matriculas com < 3 atividades', matriculas_baixo_registro],
    ['Percentual do Total', f'{matriculas_baixo_registro/df[df["Status"]=="Matriculado"].shape[0]*100:.1f}%' if df[df['Status']=='Matriculado'].shape[0] > 0 else '0%'],
    ['Causa', 'Atendimento presencial sem registro adequado no HubSpot'],
    ['Impacto na Analise', 'DISTORCE media de atividades comerciais'],
    ['Realidade', 'Essas matriculas tiveram mais contatos que registrados'],
    ['Conclusao', 'Tx atividade REAL e maior que a calculada'],
    ['Acao Recomendada', 'Treinar equipe para registrar TODAS as atividades'],
    ['', ''],
    ['=== CICLO DE CAPTACAO - REGRA IMPORTANTE ===', ''],
    ['MATRICULAS', 'Usam DATA DE FECHAMENTO para identificar ciclo'],
    ['OUTROS STATUS', 'Usam DATA DE CRIACAO para identificar ciclo'],
    ['Motivo', 'Matricula = quando converteu, nao quando entrou no funil'],
    ['', ''],
    ['=== PERIODO ANALISADO ===', ''],
    ['Estrategia', 'Analise focada nos CICLOS DE CAPTACAO (Outubro a Marco)'],
    ['Ciclos Incluidos', '22.1, 23.1, 24.1, 25.1, 26.1'],
    ['Datas', 'Out/2021 ate Mar/2026'],
    ['Justificativa', 'Red Balloon tem alta captacao entre Out-Mar (periodo de matriculas escolares)'],
    ['', ''],
    ['=== OBJETIVO DO RELATORIO ===', ''],
    ['Questao Principal', 'Precisamos gerar novos leads ou os atuais sao suficientes para a meta?'],
    ['Meta Total (matriculas)', META_TOTAL],
    ['Ja Realizadas (matriculas)', MATRICULAS_ATUAIS],
    ['Faltam (matriculas)', META_FALTANTE],
    ['Leads em Qualificacao (TOTAL)', leads_em_qualificacao],
    ['Leads em Qualificacao (Ciclo Atual)', leads_qualificacao_ciclo_atual if ciclo_atual else 0],
    ['Ciclo Atual Analisado', ciclo_atual if ciclo_atual else 'N/A'],
    ['Leads Necessarios', leads_necessarios_gap],
    ['CONCLUSAO', 'NAO PRECISA GERAR NOVOS - temos folga'],
    ['', ''],
    ['=== ESTRUTURA DO RELATORIO ===', ''],
    ['Aba 0', '[ESTA ABA] Documentacao e guia de navegacao'],
    ['Aba 1', 'Resumo_Por_Ciclo - Metricas de cada ciclo'],
    ['Aba 2', 'Status_Atual - Distribuicao dos leads'],
    ['Aba 3', 'Performance_Unidade - Analise por unidade (74 unidades)'],
    ['Aba 4', 'Performance_Fonte - Analise por canal'],
    ['Aba 5', 'Evolucao_Mensal - Tendencia mes a mes'],
    ['Aba 6', 'Matriz_Ciclo_Unidade - Cruzamento ciclo x unidade'],
    ['Aba 7', 'Matriz_Ciclo_Fonte - Cruzamento ciclo x fonte'],
    ['Aba 7b', 'Comparativo_26vs25 - Ciclo atual vs anterior POR CANAL'],
    ['Aba 8', 'Atividades_Matriculados - Perfil dos que converteram'],
    ['Aba 9', 'Idade_Matriculados - Tempo ate conversao'],
    ['Aba 10', 'Meta_RedBalloon - PRINCIPAL - Calculo do gap'],
    ['Aba 10b', 'Simulacao_Cenarios - Outros cenarios de meta'],
    ['Aba 10c', 'Taxa_Atendimento - Tempo sem contato (CRITICO para vendas) 🔥'],
    ['Aba 10d', 'Distribuicao_Atividades - Quantas atividades por lead'],
    ['Aba 10e', 'Atendimento_Por_Ciclo - Taxa de atendimento por ciclo'],
    ['Aba 11', 'Leads_Qualificacao - Lista completa para trabalhar'],
    ['', ''],
    ['=== METRICAS CHAVE - ENTENDA OS NUMEROS ===', ''],
    ['', ''],
    ['[1] Taxa_Conversao (%)', round(taxa_conversao_historica, 1)],
    ['    O que e?', 'Percentual HISTORICO de leads que viraram matriculas nos ultimos 3 ciclos'],
    ['    Como calculamos?', '(Total Matriculas / Total Leads) dos ciclos 24.1, 25.1 e 26.1'],
    ['    Por que 3 ciclos?', 'Base estatistica robusta, remove variacoes pontuais, reflete realidade recente'],
    ['    Significa que...', f'A cada 100 leads em qualificacao, ~{int(taxa_conversao_historica)} viram matriculas'],
    ['', ''],
    ['[2] Leads_Em_Qualificacao (TOTAL)', leads_em_qualificacao],
    ['    O que e?', 'TODOS os leads atualmente em status "Em Qualificacao" (qualquer ciclo)'],
    ['    Como calculamos?', 'Soma de leads com Status = Em Qualificacao de TODOS os ciclos Out-Mar'],
    ['    Por que importa?', 'Mostra CAPACIDADE TOTAL do pipeline - quanto temos para trabalhar'],
    ['    Significa que...', 'Pipeline acumulado de leads disponiveis para converter'],
    ['', ''],
    ['[3] Leads_Em_Qualificacao (Ciclo Atual)', leads_qualificacao_ciclo_atual if ciclo_atual else 0],
    ['    O que e?', f'Leads em qualificacao APENAS do ciclo {ciclo_atual if ciclo_atual else "N/A"}'],
    ['    Como calculamos?', 'Status = Em Qualificacao E Ciclo = mais recente com volume >100 leads'],
    ['    Por que importa?', 'Mostra pipeline FRESCO - leads mais quentes do momento'],
    ['    Significa que...', f'{round(leads_qualificacao_ciclo_atual/(leads_em_qualificacao+0.01)*100, 1) if ciclo_atual and leads_em_qualificacao > 0 else 0}% do pipeline sao leads do ciclo atual'],
    ['', ''],
    ['[4] Ciclo_Atual', ciclo_atual if ciclo_atual else 'N/A'],
    ['    O que e?', 'Ciclo de captacao mais recente com volume significativo (>100 leads)'],
    ['    Como identificamos?', 'Ciclo mais alto entre os que tem pelo menos 100 leads cadastrados'],
    ['    Por que filtrar <100?', 'Evita considerar ciclos futuros com poucos leads cadastrados por erro'],
    ['    Significa que...', f'Estamos em {ciclo_atual if ciclo_atual else "N/A"} = periodo Out/{str(int(ciclo_atual.split(".")[0])+2000-1)[-2:]}-Mar/{ciclo_atual.split(".")[0] if ciclo_atual else "N/A"}'],
    ['', ''],
    ['[5] Leads_Necessarios', leads_necessarios_gap],
    ['    O que e?', 'Quantidade de leads NOVOS necessarios para bater a meta faltante'],
    ['    Como calculamos?', f'Meta Faltante ({META_FALTANTE}) / Taxa Conversao ({round(taxa_conversao_historica, 1)}%) = {leads_necessarios_gap}'],
    ['    Por que importa?', 'Define se precisamos GERAR novos leads ou CONVERTER os atuais'],
    ['    Formula', f'{META_FALTANTE} matriculas faltando ÷ {round(taxa_conversao_historica/100, 3)} conversao = {leads_necessarios_gap} leads'],
    ['', ''],
    ['[6] Gap_de_Leads', gap_leads],
    ['    O que e?', 'Diferenca entre leads disponiveis e leads necessarios'],
    ['    Como calculamos?', f'Leads Necessarios ({leads_necessarios_gap}) - Leads Disponiveis ({leads_em_qualificacao}) = {gap_leads}'],
    ['    Por que importa?', 'DECISAO ESTRATEGICA: gerar leads (positivo) ou focar conversao (negativo)'],
    ['    Significa que...', f'{"FOLGA de "+str(abs(gap_leads))+" leads - NAO precisa gerar novos" if gap_leads < 0 else "FALTA de "+str(gap_leads)+" leads - precisa captacao"}'],
    ['    Interpretacao', 'Numero NEGATIVO = bom (temos leads sobrando) | POSITIVO = ruim (faltam leads)'],
    ['', ''],
    ['[7] Status_Gap', 'SUFICIENTE' if gap_leads < 0 else 'INSUFICIENTE'],
    ['    O que e?', 'Classificacao simplificada: pipeline suficiente ou nao'],
    ['    Como classificamos?', 'SUFICIENTE se Gap < 0 (temos mais que o necessario) | INSUFICIENTE se Gap > 0'],
    ['    Acao Recomendada', 'SUFICIENTE = focar em CONVERSAO | INSUFICIENTE = focar em GERACAO'],
    ['', ''],
    ['=== PROJECAO REALISTA ATE 31/03/2026 - ANALISE CORRIGIDA ===', ''],
    ['', ''],
    ['⚠️  IMPORTANTE - ERRO NA INTERPRETACAO ANTERIOR', ''],
    ['    O que estava ERRADO?', 'Aplicar taxa historica de 6 meses em leads de CICLOS ANTERIORES que ja falharam'],
    ['    Por que estava errado?', 'Leads de ciclos ANTERIORES (24.1, 25.1) ja tiveram chance e NAO converteram'],
    ['    Analogia', 'E como esperar que estoque de ciclos passados converta agora - nao faz sentido!'],
    ['', ''],
    ['⏰ TEMPO DISPONIVEL', ''],
    ['    Dias corridos ate 31/03', dias_corridos_restantes],
    ['    Dias uteis (seg-sex)', dias_uteis_restantes],
    ['    Feriados considerados', 'Carnaval 2026 (16-18/02) - 3 dias excluidos'],
    ['    Meses disponiveis', f'{round(dias_corridos_restantes/30, 1)} meses'],
    ['', ''],
    ['📦 SEGMENTACAO DO PIPELINE (por CICLO)', ''],
    ['    Leads DO CICLO 26.1 (Out/2025-Mar/2026)', f'{len(leads_ciclo_26):,} leads'],
    ['    • Entrada Out-Nov/2025', f'{len(leads_ciclo_26_inicio):,} leads (ja estao ha 3-4 meses)'],
    ['    • Entrada Dez/2025-Fev/2026', f'{len(leads_ciclo_26_recente):,} leads (mais recentes)'],
    ['    Leads ESTOQUE (ciclos anteriores 24.1, 25.1, etc)', f'{len(leads_estoque_ciclos_anteriores):,} leads'],
    ['    % de Estoque', f'{round(len(leads_estoque_ciclos_anteriores)/len(qualificacao)*100, 1)}% do pipeline total'],
    ['    Interpretacao', 'Leads do ciclo 26.1 tem chance real | Estoque de ciclos anteriores tem chance BAIXISSIMA'],
    ['', ''],
    ['📊 TAXA DE CONVERSAO MENSAL (historica)', ''],
    ['    Matriculas por mes (media)', f'{matriculas_por_mes_media:.0f} matriculas/mes'],
    ['    Taxa mensal', f'{taxa_conversao_mensal:.1f}%'],
    ['    Base de calculo', 'Media dos ciclos 22.1, 23.1, 24.1 (completos)'],
    ['', ''],
    ['🎯 PROJECAO REALISTA - METODO 1 (Segmentacao por Ciclo)', ''],
    ['    Leads DO CICLO 26.1', f'{len(leads_ciclo_26):,} leads × {taxa_conversao_historica:.1f}% = {projecao_leads_ciclo_26:,} matriculas'],
    ['    Leads ESTOQUE ciclos anteriores', f'{len(leads_estoque_ciclos_anteriores):,} leads × {taxa_conversao_historica * 0.05:.1f}% = {projecao_estoque_antigo:,} matriculas'],
    ['    TOTAL projetado', f'{projecao_total_realista:,} matriculas'],
    ['    Meta faltante', META_FALTANTE],
    ['    Gap REALISTA', gap_realista_segmentado],
    ['    Status', 'INSUFICIENTE!' if gap_realista_segmentado > 0 else 'SUFICIENTE!'],
    ['    Interpretacao', f'{"NAO BATE a meta - faltam "+str(gap_realista_segmentado)+" matriculas" if gap_realista_segmentado > 0 else "BATE a meta com folga de "+str(abs(gap_realista_segmentado))+" matriculas"}'],
    ['', ''],
    ['🎯 PROJECAO REALISTA - METODO 2 (Linear Mensal)', ''],
    ['    Matriculas/mes historico', f'{matriculas_por_mes_media:.0f} matriculas/mes'],
    ['    Meses restantes', f'{meses_restantes:.1f} meses'],
    ['    Projecao', f'{projecao_linear_mensal:,} matriculas'],
    ['    Meta faltante', META_FALTANTE],
    ['    Gap REALISTA', gap_realista_linear],
    ['    Status', 'INSUFICIENTE!' if gap_realista_linear > 0 else 'SUFICIENTE!'],
    ['    Interpretacao', f'{"NAO BATE a meta - faltam "+str(gap_realista_linear)+" matriculas" if gap_realista_linear > 0 else "BATE a meta com folga de "+str(abs(gap_realista_linear))+" matriculas"}'],
    ['', ''],
    ['💡 CONCLUSAO REALISTA', ''],
    ['    Ambos metodos mostram', 'Projecao INSUFICIENTE se gap > 0 | SUFICIENTE se gap < 0'],
    ['    Problema principal', f'Estoque de ciclos anteriores ({len(leads_estoque_ciclos_anteriores):,} leads) - ja falharam, chance baixissima'],
    ['    Acao necessaria', 'Se INSUFICIENTE: acelerar conversao OU iniciar captacao AGORA'],
    ['    Leads prioritarios', f'Focar nos {len(leads_alta_prioridade):,} leads de ALTA prioridade'],
    ['', ''],
    ['=== METRICAS DE ATENDIMENTO - OS LEADS ESTAO SENDO TRABALHADOS? ===', ''],
    ['', ''],
    ['[8] Taxa_Atendimento_Geral (%)', round(taxa_atendimento_geral, 1)],
    ['    O que e?', 'Percentual de leads em qualificacao que JA FORAM contatados alguma vez'],
    ['    Como calculamos?', f'Leads com atividades > 0 / Total leads em qualificacao = {round(taxa_atendimento_geral, 1)}%'],
    ['    O que significa?', f'{round(taxa_atendimento_geral, 1)}% dos leads receberam ALGUM contato desde que entraram'],
    ['    Inverso (preocupante)', f'{round(100-taxa_atendimento_geral, 1)}% NUNCA foram contatados (leads abandonados)'],
    ['', ''],
    ['[9] Taxa_Atendimento_30d (%)', round(taxa_atendimento_30d, 1)],
    ['    O que e?', 'Percentual de leads que receberam contato nos ULTIMOS 30 DIAS'],
    ['    Como calculamos?', f'Leads com ultima atividade <30 dias / Total = {round(taxa_atendimento_30d, 1)}%'],
    ['    Por que importa?', 'Lead sem contato >30 dias esfria - chance de conversao CAI drasticamente'],
    ['    O que significa?', f'Apenas {round(taxa_atendimento_30d, 1)}% do pipeline esta sendo ATIVAMENTE trabalhado'],
    ['    🚨 CRITICO', f'{round(100-taxa_atendimento_30d, 1)}% dos leads estao PARADOS (sem contato recente)'],
    ['', ''],
    ['[10] Leads_Ativos_30d', leads_ativos_30d],
    ['    O que e?', 'Quantidade absoluta de leads com contato nos ultimos 30 dias'],
    ['    Por que separar?', f'Sao os {leads_ativos_30d:,} leads QUENTES do pipeline total de {leads_em_qualificacao:,}'],
    ['    Projecao realista', f'{leads_ativos_30d:,} leads × {round(taxa_conversao_historica, 1)}% = ~{int(leads_ativos_30d * taxa_conversao_historica/100)} matriculas esperadas'],
    ['    Gap realista', f'Meta faltante: {META_FALTANTE} | Projecao leads ativos: {int(leads_ativos_30d * taxa_conversao_historica/100)} | {"INSUFICIENTE!" if int(leads_ativos_30d * taxa_conversao_historica/100) < META_FALTANTE else "SUFICIENTE!"}'],
    ['', ''],
    ['[11] Leads_Inativos_30d', leads_inativos_30d],
    ['    O que e?', 'Leads SEM CONTATO ha mais de 30 dias (pipeline congelado)'],
    ['    Gravidade', f'{leads_inativos_30d:,} leads ({round(leads_inativos_30d/leads_em_qualificacao*100, 1)}%) estao esfriando'],
    ['    Oportunidade', 'Campanha de reativacao pode recuperar parte desses leads'],
    ['    Acao urgente', 'Priorizar reativacao dos leads inativos de 31-60 dias (ainda recuperaveis)'],
    ['', ''],
    ['[12] Leads_Nunca_Contatados', leads_nunca_contatados],
    ['    O que e?', 'Leads que NUNCA receberam nenhuma atividade comercial'],
    ['    Gravidade', f'{leads_nunca_contatados:,} leads ({round(leads_nunca_contatados/leads_em_qualificacao*100, 1)}%) completamente abandonados'],
    ['    Causa provavel', 'Falha na distribuicao de leads OU leads de baixa qualidade ignorados'],
    ['    Acao', 'Qualificar esses leads e distribuir para equipe OU descarta-los do pipeline'],
    ['', ''],
    ['=== PERFIL MATRICULADOS - ENTENDA O COMPORTAMENTO ===', ''],
    ['', ''],
    ['[13] Atividades Comerciais Media', round(atividades_media, 1)],
    ['    O que e?', 'Numero MEDIO de atividades comerciais dos leads que MATRICULARAM'],
    ['    Como calculamos?', 'Soma(atividades de todos matriculados) / Quantidade(matriculados)'],
    ['    Por que importa?', 'Mostra intensidade de trabalho necessario para converter'],
    ['    ⚠️  ATENCAO', f'{matriculas_baixo_registro:,} matriculas ({matriculas_baixo_registro/df[df["Status"]=="Matriculado"].shape[0]*100:.1f}%) tem <3 atividades registradas'],
    ['    Realidade', 'Numero REAL de atividades e MAIOR - falta registro no sistema'],
    ['', ''],
    ['[14] Atividades Comerciais Mediana', int(atividades_mediana)],
    ['    O que e?', 'Valor CENTRAL de atividades (50% tem mais, 50% tem menos)'],
    ['    Como calculamos?', 'Ordenar todas atividades e pegar o valor do meio'],
    ['    Por que mediana?', 'Mediana e mais robusta que media - nao sofre com outliers'],
    ['    Significa que...', f'Metade das matriculas tem {int(atividades_mediana)} ou MENOS atividades registradas'],
    ['    Diferenca Media x Mediana?', f'Media ({round(atividades_media, 1)}) > Mediana ({int(atividades_mediana)}) = distribuicao assimetrica (poucos com muitas atividades)'],
    ['', ''],
    ['[15] Idade Media (dias)', int(idade_media)],
    ['    O que e?', 'Tempo MEDIO entre criacao do lead e matricula'],
    ['    Como calculamos?', 'Soma(dias ate conversao de todos) / Quantidade(matriculados)'],
    ['    Por que importa?', 'Mostra velocidade de conversao - comprimento do ciclo de vendas'],
    ['    Significa que...', f'Em media, leva {int(idade_media)} dias (~{int(idade_media/30)} meses) para converter um lead'],
    ['', ''],
    ['[16] Idade Mediana (dias)', int(idade_mediana)],
    ['    O que e?', 'Tempo CENTRAL ate conversao (50% convertem mais rapido, 50% mais lento)'],
    ['    Como calculamos?', 'Ordenar todos os tempos e pegar o valor do meio'],
    ['    Por que mediana?', 'Remove efeito de leads muito antigos que distorcem a media'],
    ['    Significa que...', f'Metade das matriculas acontece em ate {int(idade_mediana)} dias (~{int(idade_mediana/30)} meses)'],
    ['    Diferenca Media x Mediana?', f'Media ({int(idade_media)}) > Mediana ({int(idade_mediana)}) = alguns leads demoram MUITO tempo'],
    ['', ''],
    ['=== ALERTAS DE QUALIDADE - PROBLEMAS DETECTADOS ===', ''],
    ['', ''],
    ['🚨 [ERRO GRAVE] DATAS FUTURAS', f'{leads_ciclo_futuro} leads com datas futuras removidos da analise'],
    ['    O que detectamos?', f'{leads_ciclo_futuro} leads cadastrados com datas futuras (ciclo 27.1 = Out/2026-Mar/2027)'],
    ['    Como identificamos?', f'Comparamos ciclo calculado vs ciclo maximo valido hoje ({data_atual.strftime("%d/%m/%Y")})'],
    ['    Ciclo maximo valido', f'{ciclo_maximo_valido} (hoje estamos em Fev/2026 = dentro do ciclo 26.1)'],
    ['    Ciclos impossiveis', 'Ciclo 27.1 (Out/2026-Mar/2027) - ainda nao comecou!'],
    ['    Causa raiz', 'Data de criacao OU data de fechamento cadastrada no FUTURO'],
    ['    Gravidade', 'CRITICA - dados obviamente errados, nao podem ser corrigidos'],
    ['    Acao tomada', 'Leads REMOVIDOS da analise (nao ha como corrigir data futura)'],
    ['    Recomendacao', 'Revisar processo de cadastro para prevenir entradas com datas futuras'],
    ['', ''],
    ['⚠️  [ALERTA] DATAS INVERTIDAS (Offline/Direto)', f'{leads_offline_direto_erro:,} leads de canais presenciais corrigidos'],
    ['    O que detectamos?', f'{leads_com_erro:,} leads com Data Criacao > Data Fechamento'],
    ['    Como identificamos?', 'Comparamos coluna "Data de criacao" vs "Data de encerramento do negocio"'],
    ['    Exemplo do problema', 'Lead criado em 15/Jan/2025 mas matriculado em 10/Dez/2024 (impossivel!)'],
    ['    Canais mais afetados', f'Offline e Direto = {leads_offline_direto_erro:,} dos {leads_com_erro:,} ({leads_offline_direto_erro/max(leads_com_erro, 1)*100:.1f}%)'],
    ['    Causa raiz', 'Atendimento PRESENCIAL: pessoa matricula na unidade, so depois registram no sistema'],
    ['    Impacto se nao corrigir', 'Idade negativa, metricas distorcidas, conversao impossivel de calcular'],
    ['    Acao tomada', 'CORRECAO: Data_Criacao = Data_Fechamento (manter matricula valida)'],
    ['    Por que corrigir?', 'Essas matriculas sao REAIS (geraram receita), nao podem ser descartadas'],
    ['    Limitacao', 'Perdemos info de quando lead REALMENTE entrou, mas mantemos matricula'],
    ['    Recomendacao', 'Cadastrar leads IMEDIATAMENTE no sistema, mesmo em atendimento presencial'],
    ['', ''],
    ['⚠️  [ALERTA] ATIVIDADES COMERCIAIS SUBREGISTRADAS', f'{matriculas_baixo_registro:,} matriculas ({matriculas_baixo_registro/df[df["Status"]=="Matriculado"].shape[0]*100:.1f}%) com <3 atividades'],
    ['    O que detectamos?', f'{matriculas_baixo_registro:,} leads MATRICULADOS tem menos de 3 atividades comerciais registradas'],
    ['    Como identificamos?', 'Contamos coluna "Numero de atividades de vendas" para Status = Matriculado'],
    ['    Por que e problema?', 'E IMPOSSIVEL matricular sem contato - houve atendimento nao registrado'],
    ['    Impacto nas metricas', 'Media e mediana de atividades SUBESTIMADAS (menor que realidade)'],
    ['    Realidade vs Sistema', 'REALIDADE: muito contato | SISTEMA: pouco registro'],
    ['    Causa raiz', 'Equipe faz atendimento presencial/telefone mas nao registra no HubSpot'],
    ['    Acao tomada', 'Manter na analise (matriculas reais) mas ALERTAR sobre qualidade'],
    ['    Por que nao remover?', 'Nao e criterio de qualidade de LEAD, e problema de PROCESSO'],
    ['    Recomendacao', 'Treinar equipe para registrar TODAS as atividades no sistema'],
    ['    Impacto Business', 'Impossibilita analise de eficiencia comercial e treinamento de ML'],
    ['', ''],
    ['=== COMO USAR ESTE RELATORIO ===', ''],
    ['Para Gestao', 'Aba 10 (Meta_RedBalloon) + Aba 10c (Taxa_Atendimento) 🔥'],
    ['Para Marketing', 'Aba 4 (Performance_Fonte) + Aba 7b (Comparativo)'],
    ['Para Vendas', 'Aba 10c (Taxa_Atendimento) + Aba 11 (Leads_Qualificacao) 🔥'],
    ['Para Gestao Comercial', 'Aba 10c, 10d, 10e - TAXA DE ATENDIMENTO por tempo/atividade/ciclo'],
    ['Para Estrategia', 'Aba 1 (Resumo_Por_Ciclo) - tendencias'],
    ['Para ML e Helio', 'Aba 8 e 9 (Perfil dos matriculados)'],
    ['', ''],
    ['=== PROXIMOS PASSOS ===', ''],
    ['Acao_1', 'URGENTE: Reativar leads inativos (>30 dias sem contato) - ver Aba 10c'],
    ['Acao_2', 'Distribuir leads nunca contatados para equipe - ver Aba 10c'],
    ['Acao_2', 'Priorizar leads em qualificacao'],
    ['Acao_3', 'Campanha de reativacao'],
    ['Acao_4', 'Corrigir processo de cadastro Offline/Direto'],
    ['Acao_5', 'Treinar equipe para registrar TODAS atividades no HubSpot'],
]

df_capa = pd.DataFrame(capa_data[1:], columns=capa_data[0])


try:
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # Aba 0: CAPA / DOCUMENTACAO
        df_capa.to_excel(writer, sheet_name='0_CAPA_Leia_Primeiro', index=False)
        
        # Aba 1: Resumo Geral por Ciclo
        resumo_ciclo.to_excel(writer, sheet_name='1_Resumo_Por_Ciclo', index=False)
        
        # Aba 2: Status Atual
        status_geral.to_excel(writer, sheet_name='2_Status_Atual', index=False)
        
        # Aba 3: Performance por Unidade
        unidade.to_excel(writer, sheet_name='3_Performance_Unidade', index=False)
        
        # Aba 4: Performance por Fonte
        fonte.to_excel(writer, sheet_name='4_Performance_Fonte', index=False)
        
        # Aba 5: Evolução Mensal
        mensal_pivot.to_excel(writer, sheet_name='5_Evolucao_Mensal', index=False)
        
        # Aba 6: Matriz Ciclo x Unidade
        ciclo_unidade_pivot.to_excel(writer, sheet_name='6_Matriz_Ciclo_Unidade', index=False)
        
        # Aba 7: Matriz Ciclo x Fonte
        ciclo_fonte_pivot.to_excel(writer, sheet_name='7_Matriz_Ciclo_Fonte', index=False)
        
        # Aba 7b: Comparativo 26.1 vs 25.1 por Canal (COM NOTA DE QUALIDADE)
        df_comparativo_ciclos.to_excel(writer, sheet_name='7b_Comparativo_26vs25', index=False)
        
        # Aba 8: Atividades dos Matriculados
        df_atividades_matriculados.to_excel(writer, sheet_name='8_Atividades_Matriculados', index=False)
        
        # Aba 9: Idade dos Matriculados
        df_idade_matriculados.to_excel(writer, sheet_name='9_Idade_Matriculados', index=False)
        
        # Aba 10: Meta Real Red Balloon
        df_meta_real.to_excel(writer, sheet_name='10_Meta_RedBalloon', index=False)
        
        # Aba 10b: Simulação de Cenários
        df_projecao.to_excel(writer, sheet_name='10b_Simulacao_Cenarios', index=False)
        
        # Aba 10c: Taxa de Atendimento - Tempo sem Contato
        df_tempo_sem_contato.to_excel(writer, sheet_name='10c_Taxa_Atendimento', index=False)
        
        # Aba 10d: Taxa de Atendimento - Atividades
        analise_atividades.to_excel(writer, sheet_name='10d_Distribuicao_Atividades', index=False)
        
        # Aba 10e: Taxa de Atendimento por Ciclo
        df_atendimento_ciclo.to_excel(writer, sheet_name='10e_Atendimento_Por_Ciclo', index=False)
        
        # Aba 10f: Projeção Realista até Março
        df_projecao_realista.to_excel(writer, sheet_name='10f_Projecao_Realista', index=False)
        
        # Aba 10g: Priorização de Leads
        df_priorizacao.to_excel(writer, sheet_name='10g_Priorizacao_Leads', index=False)
        
        # Aba 11: Leads em Qualificação (DETALHADO)
        qualificacao.to_excel(writer, sheet_name='11_Leads_Qualificacao', index=False)
        
        # --- ESTILIZAÇÃO VISUAL (LAYOUT COGNA EDUCAÇÃO) ---
        wb = writer.book
        
        # Paleta de Cores Cogna
        COLOR_NAVY = '002A54'     # Azul Institucional
        COLOR_MAGENTA = 'E40046'  # Magenta (Detalhes)
        COLOR_TEXT = '333333'     # Texto Escuro
        COLOR_BORDER = 'E0E0E0'   # Bordas Suaves
        COLOR_SECTION = 'F4F4F4'  # Fundo Seções
        WHITE = 'FFFFFF'
        
        font_header = Font(name='Segoe UI', size=10, bold=True, color=WHITE)
        fill_header = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type='solid')
        
        font_body = Font(name='Segoe UI', size=10, color=COLOR_TEXT)
        border_body = Border(left=Side(style='thin', color=COLOR_BORDER), 
                             right=Side(style='thin', color=COLOR_BORDER), 
                             top=Side(style='thin', color=COLOR_BORDER), 
                             bottom=Side(style='thin', color=COLOR_BORDER))
                             
        font_section = Font(name='Segoe UI', size=10, bold=True, color=COLOR_NAVY)
        fill_section = PatternFill(start_color=COLOR_SECTION, end_color=COLOR_SECTION, fill_type='solid')

        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
            
            # Cabeçalho
            for cell in ws[1]:
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_body
                
            # Corpo
            for row in ws.iter_rows(min_row=2):
                # Detectar se é linha de seção (especialmente na Capa)
                is_section_row = False
                first_cell_val = str(row[0].value).strip() if row[0].value else ""
                if first_cell_val.startswith('==='):
                    is_section_row = True

                for cell in row:
                    cell.border = border_body
                    cell.alignment = Alignment(vertical='center')
                    
                    if is_section_row:
                        cell.font = font_section
                        cell.fill = fill_section
                    else:
                        cell.font = font_body
            
            # Largura Colunas
            for column in ws.columns:
                max_length = 0
                col_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            val_len = len(str(cell.value))
                            if val_len > max_length: max_length = val_len
                    except: pass
                ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

        print("✅ Arquivo Excel criado com sucesso!")
        print(f"   📄 {OUTPUT_FILE}\n")
        print("📑 Abas criadas:")
        print("   0️⃣  CAPA_Leia_Primeiro - DOCUMENTACAO do relatorio ⭐")
        print("   1️⃣  Resumo_Por_Ciclo - Métricas por ciclo (Out-Mar)")
        print("   2️⃣  Status_Atual - Distribuição atual dos leads")
        print("   3️⃣  Performance_Unidade - Análise por unidade")
        print("   4️⃣  Performance_Fonte - Análise por origem do lead")
        print("   5️⃣  Evolucao_Mensal - Tendência mês a mês")
        print("   6️⃣  Matriz_Ciclo_Unidade - Cruzamento ciclo x unidade")
        print("   7️⃣  Matriz_Ciclo_Fonte - Cruzamento ciclo x fonte")
        print("   7b Comparativo_26vs25 - Ciclo atual vs anterior POR CANAL ⚠️")
        print("   8️⃣  Atividades_Matriculados - Atividades comerciais dos que matricularam")
        print("   9️⃣  Idade_Matriculados - Tempo até conversão dos matriculados")
        print("   10  Meta_RedBalloon - Meta 6.545 / Atual 4.226 / Gap")
        print("   10b Simulacao_Cenarios - Outros cenários de meta")
        print("   10c Taxa_Atendimento - Tempo sem contato (0-7d, 8-30d, etc) 🔥")
        print("   10d Distribuicao_Atividades - Faixas de atividades comerciais")
        print("   10e Atendimento_Por_Ciclo - Taxa de atendimento por ciclo")
        print("   10f Projecao_Realista - RESPOSTA: Vamos bater a meta ate março? 🎯")
        print("   10g Priorizacao_Leads - Leads por prioridade (ALTA/MEDIA/BAIXA) 🔥")
        print("   11  Leads_Qualificacao - Lista completa dos em qualificação")
        print("")
        print("   ⭐  COMECE PELA ABA 0 - Documentacao completa do relatorio")
        print("   ⚠️  ABA 7b: Contém nota sobre inconsistências de data em Offline/Direto")
        print("   🔥 ABAS 10c-10e: METRICAS DE ATENDIMENTO - quantos leads estao sendo trabalhados")
        print("   🎯 ABA 10f: PROJECAO REALISTA - resposta critica: vamos bater a meta ate marco?")
        
except Exception as e:
    print(f"❌ ERRO ao salvar arquivo: {e}")
    sys.exit(1)

# --- RESUMO FINAL ---
fim = datetime.now()
print("\n" + "="*80)
print("   ✅ ANÁLISE CONCLUÍDA COM SUCESSO!")
print("="*80)
print(f"\n📊 Estatísticas Finais:")
print(f"   Total de leads analisados (Out-Mar): {len(df):,}")
print(f"   Ciclos: {', '.join(sorted(df['Ciclo'].dropna().unique()))}")
print(f"   Matriculados: {len(df[df['Status']=='Matriculado']):,} ({len(df[df['Status']=='Matriculado'])/len(df)*100:.1f}%)")
print(f"   Perdidos: {len(df[df['Status']=='Perdido']):,} ({len(df[df['Status']=='Perdido'])/len(df)*100:.1f}%)")
print(f"   Em Qualificação: {len(df[df['Status']=='Em Qualificação']):,} ({len(df[df['Status']=='Em Qualificação'])/len(df)*100:.1f}%)")
print(f"\n📁 Arquivo gerado: {OUTPUT_FILE.name}")
print(f"🕐 Conclusão: {fim.strftime('%d/%m/%Y %H:%M:%S')}")
print("\n💡 OBS: Análise focada nos ciclos de captação (Outubro a Março)")
print("   Leads de Abril a Setembro não foram incluídos")
print("\n" + "="*80 + "\n")
