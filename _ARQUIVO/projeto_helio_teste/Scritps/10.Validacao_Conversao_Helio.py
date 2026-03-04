"""
==============================================================================
VALIDAÇÃO DE EVOLUÇÃO - HÉLIO (PREVISTO vs REALIZADO)
==============================================================================
Script: 10.Validacao_Conversao_Helio.py
Objetivo: Comparar o scoring passado (Nota 4 e 5) com o status ATUAL da base.
          Responde:
          1. Dos leads Nota 5 antigos, quantos viraram Matrícula ou Visita?
          2. Dos leads Nota 1/2 antigos, quantos foram Perdidos de fato?
Data: 2026-01-19
==============================================================================
"""

import pandas as pd
import os
import glob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASTA_OUTPUT = os.path.join(CAMINHO_BASE, 'Outputs')
PASTA_DADOS_SCORED = os.path.join(PASTA_OUTPUT, 'Dados_Scored')
PASTA_BACKUP_SCORED = os.path.join(PASTA_DADOS_SCORED, 'Backup')
CAMINHO_LEADS_ATUAL = os.path.join(CAMINHO_BASE, 'Data', 'hubspot_leads.csv')
PASTA_VALIDACAO = os.path.join(PASTA_OUTPUT, 'Relatorios_Validacao')

os.makedirs(PASTA_VALIDACAO, exist_ok=True)

DATA_HOJE = datetime.now().strftime('%Y-%m-%d')
CAMINHO_SAIDA = os.path.join(PASTA_VALIDACAO, f'Validacao_Evolucao_Helio_{DATA_HOJE}.xlsx')

# Estilos Excel
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
SUCCESS_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') # Verde
DANGER_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Vermelho
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def df_to_sheet(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if r_idx == start_row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center', wrap_text=True)

def ajustar_largura(ws):
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

print("="*80)
print("VALIDAÇÃO DE EVOLUÇÃO - O QUE ACONTECEU COM OS LEADS?")
print("="*80)

# ==============================================================================
# 1. CARREGAR DADOS (FOTO ANTIGA vs. FOTO ATUAL)
# ==============================================================================
print("\n[1/4] Carregando bases...")

# --- FOTO ANTIGA (Leads Pontuados no Passado) ---
arquivos_scored = glob.glob(os.path.join(PASTA_DADOS_SCORED, 'leads_scored_*.csv'))
arquivos_backup = glob.glob(os.path.join(PASTA_BACKUP_SCORED, 'leads_scored_*.csv'))
todos_arquivos = arquivos_scored + arquivos_backup

if not todos_arquivos:
    print(f"❌ [ERRO] Nenhum histórico de scoring encontrado.")
    print("   Rode o script '1.ML_Lead_Scoring.py' primeiro para gerar um histórico.")
    exit()

# Função para extrair data do nome do arquivo (mais seguro que data de criação do Windows)
def extrair_data_nome(caminho):
    try:
        nome = os.path.basename(caminho)
        data_str = nome.replace('leads_scored_', '').replace('.csv', '')
        return datetime.strptime(data_str, '%Y-%m-%d')
    except:
        return datetime.fromtimestamp(os.path.getctime(caminho))

# Pegar o arquivo MAIS ANTIGO disponível com base na data do nome
arquivo_antigo = min(todos_arquivos, key=extrair_data_nome)
data_arquivo_antigo = extrair_data_nome(arquivo_antigo).strftime('%Y-%m-%d')
print(f"  → Foto do Passado (Previsão): {os.path.basename(arquivo_antigo)} ({data_arquivo_antigo})")

try:
    df_antigo = pd.read_csv(arquivo_antigo, sep=';', encoding='utf-8-sig')
except:
    print("   Erro ao ler arquivo antigo. Verifique o formato.")
    exit()

# --- FOTO ATUAL (Realidade Hoje) ---
if not os.path.exists(CAMINHO_LEADS_ATUAL):
    print(f"❌ [ERRO] Arquivo atual não encontrado: {CAMINHO_LEADS_ATUAL}")
    exit()

print(f"  → Foto Atual (Realidade):     {os.path.basename(CAMINHO_LEADS_ATUAL)}")
try:
    df_atual = pd.read_csv(CAMINHO_LEADS_ATUAL, encoding='utf-8')
except:
    df_atual = pd.read_csv(CAMINHO_LEADS_ATUAL, encoding='latin1')

df_atual.columns = df_atual.columns.str.strip()

# Padronizar ID para cruzamento
df_antigo['Record ID'] = df_antigo['Record ID'].astype(str).str.strip()
df_atual['Record ID'] = df_atual['Record ID'].astype(str).str.strip()

# ==============================================================================
# 2. CRUZAMENTO E ANÁLISE
# ==============================================================================
print("\n[2/4] Cruzando previsões com resultados reais...")

# Preparar colunas
cols_antigo = ['Record ID', 'Nota_1a5', 'Probabilidade_Conversao', 'Etapa do negócio', 'Nome do negócio', 'Unidade Desejada']
cols_atual = ['Record ID', 'Etapa do negócio']

# Filtrar colunas existentes
cols_antigo = [c for c in cols_antigo if c in df_antigo.columns]

df_antigo_cut = df_antigo[cols_antigo].rename(columns={'Etapa do negócio': 'Etapa_Passada'})
df_atual_cut = df_atual[cols_atual].rename(columns={'Etapa do negócio': 'Etapa_Atual'})

# Merge (Left Join no Antigo para ver o destino de quem foi pontuado)
df_validacao = pd.merge(df_antigo_cut, df_atual_cut, on='Record ID', how='inner')

print(f"  → Leads analisados (Interseção): {len(df_validacao):,}")

# --- DEFINIÇÃO DE SUCESSO E PERDA ---
# Ajuste aqui as palavras-chave do seu funil
df_validacao['Virou_Matricula'] = df_validacao['Etapa_Atual'].str.contains('MATRÍCULA|MATRICULA|GANHO|WON', case=False, na=False)
df_validacao['Virou_Visita'] = df_validacao['Etapa_Atual'].str.contains('VISITA|AGENDADA', case=False, na=False)
df_validacao['Virou_Perdido'] = df_validacao['Etapa_Atual'].str.contains('PERDIDO|LOST|DESQUALIFICADO|ARQUIVADO', case=False, na=False)

# ==============================================================================
# 3. CÁLCULO DE MÉTRICAS
# ==============================================================================
print("\n[3/4] Calculando taxas de evolução...")

resumo_list = []

for nota in sorted(df_validacao['Nota_1a5'].unique()):
    subset = df_validacao[df_validacao['Nota_1a5'] == nota]
    total = len(subset)
    
    matriculas = subset['Virou_Matricula'].sum()
    visitas = subset['Virou_Visita'].sum()
    perdidos = subset['Virou_Perdido'].sum()
    
    # Visita inclui matrícula (pois quem matriculou passou por visita ou é sucesso igual)
    # Mas aqui vamos contar Visita como "Visita ou Matrícula" para ver avanço positivo
    avanco_positivo = matriculas + visitas 
    
    resumo_list.append({
        'Nota (Previsão)': nota,
        'Total Leads': total,
        'Evoluiu p/ Matrícula': matriculas,
        '% Matrícula': matriculas / total if total > 0 else 0,
        'Evoluiu p/ Visita': visitas,
        '% Visita': visitas / total if total > 0 else 0,
        'Foi Perdido (Correto p/ Nota Baixa)': perdidos,
        '% Perdido': perdidos / total if total > 0 else 0
    })

df_resumo = pd.DataFrame(resumo_list)

print("\n--- RESULTADO DA VALIDAÇÃO ---")
print(df_resumo.to_string(index=False, formatters={
    '% Matrícula': '{:.1%}'.format,
    '% Visita': '{:.1%}'.format,
    '% Perdido': '{:.1%}'.format
}))

# --- RESPOSTA DIRETA NO TERMINAL ---
print("\n" + "="*80)
print("RESPOSTA DIRETA: QUANTAS MATRÍCULAS O HÉLIO PREVIU COMO NOTA 5?")
print("="*80)
try:
    row_n5 = df_resumo[df_resumo['Nota (Previsão)'] == 5].iloc[0]
    print(f"🎯 Leads classificados como Nota 5: {int(row_n5['Total Leads'])}")
    print(f"✅ Desses, viraram MATRÍCULA:      {int(row_n5['Evoluiu p/ Matrícula'])}")
    print(f"📊 Taxa de Conversão Real:         {row_n5['% Matrícula']:.1%}")
except IndexError:
    print("⚠️ Nenhum lead Nota 5 encontrado no histórico analisado.")
print("="*80)

# ==============================================================================
# 4. GERAR RELATÓRIO EXCEL
# ==============================================================================
print("\n[4/4] Gerando Excel detalhado...")

wb = Workbook()

# --- ABA 1: RESUMO ---
ws1 = wb.active
ws1.title = '1_Resumo_Validacao'
ws1['A1'] = 'VALIDAÇÃO DE PREVISÃO DO HÉLIO'
ws1['A1'].font = Font(bold=True, size=14)
ws1['A2'] = f"Comparando previsão de: {os.path.basename(arquivo_antigo)}"
ws1['A3'] = f"Com status real em: {os.path.basename(CAMINHO_LEADS_ATUAL)}"

df_to_sheet(ws1, df_resumo, start_row=5)

# Formatar %
for row in range(6, 6 + len(df_resumo)):
    ws1.cell(row=row, column=4).number_format = '0.0%' # % Matrícula
    ws1.cell(row=row, column=6).number_format = '0.0%' # % Visita
    ws1.cell(row=row, column=8).number_format = '0.0%' # % Perdido

ajustar_largura(ws1)

# --- ABA 2: NOTA 4 e 5 QUE EVOLUÍRAM (PROVA DE SUCESSO) ---
ws2 = wb.create_sheet('2_Sucesso_Nota_4_5')
ws2['A1'] = 'LEADS NOTA 4 E 5 QUE VIRARAM VISITA OU MATRÍCULA'
ws2['A1'].font = Font(bold=True)

mask_sucesso = (df_validacao['Nota_1a5'].isin([4, 5])) & (df_validacao['Virou_Matricula'] | df_validacao['Virou_Visita'])
df_sucesso = df_validacao[mask_sucesso].sort_values('Virou_Matricula', ascending=False)

df_to_sheet(ws2, df_sucesso, start_row=3)
ajustar_largura(ws2)

# --- ABA 3: NOTA 1 e 2 QUE FORAM PERDIDOS (PROVA DE PRECISÃO NEGATIVA) ---
ws3 = wb.create_sheet('3_Perdidos_Nota_1_2')
ws3['A1'] = 'LEADS NOTA 1 E 2 QUE FORAM CONFIRMADOS COMO PERDIDOS'
ws3['A1'].font = Font(bold=True)

mask_perda = (df_validacao['Nota_1a5'].isin([1, 2])) & (df_validacao['Virou_Perdido'])
df_perda = df_validacao[mask_perda]

df_to_sheet(ws3, df_perda, start_row=3)
ajustar_largura(ws3)

try:
    wb.save(CAMINHO_SAIDA)
    print(f"\n✅ Relatório salvo em: {CAMINHO_SAIDA}")
except Exception as e:
    print(f"\n❌ Erro ao salvar Excel: {e}")

print("="*80)