"""
==============================================================================
MODELO DE ML - SCORING DE PROBABILIDADE DE CONVERSÃO
==============================================================================
Script: 12_ML_Lead_Scoring.py
Objetivo: Treinar modelo para prever conversão e pontuar leads em qualificação/pausa
Data: 2025-12-12
==============================================================================

METODOLOGIA:
1. TARGET: Matrícula Concluída (Y=1) vs Negócio Perdido (Y=0)
2. FEATURES: Fonte, Unidade, Atividades de Vendas, Mês, Dia da Semana, etc.
3. MODELO: Random Forest (robusto, interpretável)
4. OUTPUT: Nota de 1 a 5 baseada na probabilidade de conversão
==============================================================================
"""

import pandas as pd
import numpy as np
import os
import shutil
import glob
import pickle
from datetime import datetime
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import classification_report, roc_auc_score, confusion_matrix
import warnings
warnings.filterwarnings('ignore')

# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASTA_DATA = os.path.join(CAMINHO_BASE, 'Data')

# Função para encontrar o arquivo mais recente (suporta versionamento ex: hubspot_leads_v2.csv)
def encontrar_arquivo_recente(pasta, padrao):
    arquivos = glob.glob(os.path.join(pasta, padrao))
    if not arquivos:
        return None
    return max(arquivos, key=os.path.getmtime)

# Busca dinâmica dos arquivos (prioriza o mais novo na pasta Data)
CAMINHO_LEADS = encontrar_arquivo_recente(PASTA_DATA, 'hubspot_leads*.csv') or os.path.join(PASTA_DATA, 'hubspot_leads.csv')
CAMINHO_PERDIDOS = encontrar_arquivo_recente(PASTA_DATA, 'hubspot_negocios_perdidos*.csv') or os.path.join(PASTA_DATA, 'hubspot_negocios_perdidos.csv')

PASTA_OUTPUT = os.path.join(CAMINHO_BASE, 'Outputs')
PASTA_MODELS = os.path.join(CAMINHO_BASE, 'models')

# Novas pastas organizadas
PASTA_DADOS = os.path.join(PASTA_OUTPUT, 'Dados_Scored')
PASTA_RELATORIOS = os.path.join(PASTA_OUTPUT, 'Relatorios_ML')

os.makedirs(PASTA_DADOS, exist_ok=True)
os.makedirs(PASTA_RELATORIOS, exist_ok=True)
os.makedirs(PASTA_MODELS, exist_ok=True)

# --- CONSTANTES FINANCEIRAS ---
TICKET_PROPRIA = 1222.25
TICKET_FRANQUIA = 907.49

# Função de Backup
def gerenciar_backup(pasta, ignorar=None):
    if ignorar is None: ignorar = []
    pasta_backup = os.path.join(pasta, 'Backup')
    os.makedirs(pasta_backup, exist_ok=True)
    
    for item in os.listdir(pasta):
        caminho_item = os.path.join(pasta, item)
        if item == 'Backup' or item in ignorar:
            continue
        
        if os.path.isfile(caminho_item):
            destino = os.path.join(pasta_backup, item)
            # Se já existir no backup (ex: rodou 2x no mesmo dia), remove o antigo do backup
            if os.path.exists(destino):
                try: os.remove(destino)
                except: pass
            try: shutil.move(caminho_item, destino)
            except: pass

# Executar limpeza/backup antes de começar
print("[BACKUP] Organizando arquivos antigos em Backup...")
gerenciar_backup(PASTA_DADOS, ignorar=['Historico_Evolucao_IA.csv']) # Não move o log histórico
gerenciar_backup(PASTA_RELATORIOS)

DATA_HOJE = datetime.now().strftime('%Y-%m-%d')

# ==============================================================================
# CONFIGURAÇÃO DE UNIDADES
# ==============================================================================
UNIDADES_PROPRIAS = [
    'VILA LEOPOLDINA', 'SANTANA', 'PINHEIROS', 'PERDIZES', 
    'PACAEMBU', 'MORUMBI', 'JARDINS', 'ITAIM'
]

print("="*80)
print("MODELO DE ML - LEAD SCORING")
print("="*80)
print(f"[DIR] DIRETÓRIO DE EXECUÇÃO: {CAMINHO_BASE}")
print(f"[DADOS] DADOS DE ENTRADA:      {CAMINHO_LEADS}")
print("-" * 80)

# ==============================================================================
# 1. CARREGAMENTO DOS DADOS
# ==============================================================================
print("\n[1/7] Carregando dados...")

if not os.path.exists(CAMINHO_LEADS):
    print(f"\n[AVISO] Arquivo de dados não encontrado: {CAMINHO_LEADS}")
    print("   O Script 11 serve para TREINAR o modelo (criar o cérebro).")
    print("   Como você já tem o modelo pronto (.pkl) copiado, você NÃO precisa rodar este script.")
    print("   (A menos que queira retreinar do zero: nesse caso, coloque o 'hubspot_leads.csv' na pasta Data)")
    exit()

try:
    df = pd.read_csv(CAMINHO_LEADS, encoding='utf-8')
except:
    df = pd.read_csv(CAMINHO_LEADS, encoding='latin1')

# GARANTIA: Limpar nomes das colunas imediatamente para evitar erro de chave
df.columns = df.columns.str.strip()
print(f"  → Base principal: {len(df):,} registros")

# Carregar base de perdidos se existir (para enriquecer com motivo)
if os.path.exists(CAMINHO_PERDIDOS):
    try:
        df_perdidos = pd.read_csv(CAMINHO_PERDIDOS, encoding='utf-8')
    except:
        df_perdidos = pd.read_csv(CAMINHO_PERDIDOS, encoding='latin1')
    
    # Limpar nomes das colunas
    df_perdidos.columns = df_perdidos.columns.str.strip()
    
    # Padronizar Record ID para string em ambas as bases para garantir o match exato
    if 'Record ID' in df.columns:
        df['Record ID'] = df['Record ID'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()

    if 'Record ID' in df_perdidos.columns:
        df_perdidos['Record ID'] = df_perdidos['Record ID'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
        
        qtd_antes_drop = len(df_perdidos)
        # Remover duplicatas de ID na base de perdidos para evitar explosão do join
        df_perdidos = df_perdidos.drop_duplicates('Record ID')
        qtd_apos_drop = len(df_perdidos)
        duplicatas = qtd_antes_drop - qtd_apos_drop
        
        print(f"  → Limpeza de Duplicatas na base de Perdidos: {duplicatas:,} registros removidos (de {qtd_antes_drop:,} para {qtd_apos_drop:,})")

        # --- DIAGNÓSTICO DE PERDIDOS (RESPOSTA AO USUÁRIO) ---
        print(f"  → Base de Negócios Perdidos carregada: {len(df_perdidos):,} registros")
        
        ids_leads = set(df['Record ID'])
        ids_perdidos = set(df_perdidos['Record ID'])
        
        perdidos_nao_encontrados = len(ids_perdidos - ids_leads)
        match_perdidos = len(ids_perdidos.intersection(ids_leads))
        
        print(f"  → Cruzamento de IDs:")
        print(f"    - Encontrados na base de Leads: {match_perdidos:,}")
        print(f"    - NÃO encontrados na base de Leads: {perdidos_nao_encontrados:,}")
        if len(df_perdidos) > 0:
            match_rate = (match_perdidos / len(df_perdidos)) * 100
            print(f"    - Taxa de Cobertura do Match: {match_rate:.1f}% (Negócios Perdidos Únicos que estão na Base Principal)")
        if perdidos_nao_encontrados > 0:
            print(f"    [ATENCAO]: {perdidos_nao_encontrados:,} negócios perdidos foram ignorados pois seus IDs não estão na exportação principal de leads.")
            print(f"       (Isso explica a diferença entre o total do HubSpot e o total analisado)")
        # -----------------------------------------------------
        
        # Selecionar colunas para enriquecimento
        cols_enrich = ['Record ID']
        motivo_col = [c for c in df_perdidos.columns if 'motivo' in c.lower()]
        
        if motivo_col:
            cols_enrich.append(motivo_col[0])
            
        df_enrich = df_perdidos[cols_enrich].copy()
        if len(cols_enrich) > 1:
            df_enrich.columns = ['Record ID', 'Motivo_Perda']
        
        # Flag para identificar que veio da base de perdidos (Fundamental para o Target)
        df_enrich['In_Lost_Base'] = True
        
        # Merge (Left Join) para não duplicar linhas da base principal
        df = df.merge(df_enrich, on='Record ID', how='left')
        df['In_Lost_Base'] = df['In_Lost_Base'].fillna(False)
        
        print(f"  → Enriquecimento concluído: {df['In_Lost_Base'].sum():,} leads identificados na base de perdidos")
else:
    print(f"  → Base de perdidos não encontrada (usando apenas base principal)")

# ==============================================================================
# 2. PREPARAÇÃO DOS DADOS
# ==============================================================================
print("\n[2/7] Preparando features...")

df['Data de criação'] = pd.to_datetime(df['Data de criação'], errors='coerce')
df['Data de fechamento'] = pd.to_datetime(df['Data de fechamento'], errors='coerce') # Carregar data de fechamento

# SANITY CHECK: Data de fechamento (Crucial para evitar "Efeito Zumbi")
if df['Data de fechamento'].notna().sum() == 0:
    print("\n[AVISO CRITICO] A coluna 'Data de fechamento' parece vazia ou não foi lida corretamente.")
    print("   Isso fará com que leads históricos de sucesso pareçam 'velhos' (Efeito Zumbi).")
    print("   Verifique se o nome da coluna no CSV mudou (ex: 'Close Date').\n")

# Criar features temporais
df['Ano'] = df['Data de criação'].dt.year
df['Mes'] = df['Data de criação'].dt.month
df['DiaSemana'] = df['Data de criação'].dt.dayofweek
df['Trimestre'] = df['Data de criação'].dt.quarter

# CORREÇÃO CRONOLÓGICA (VIAGEM NO TEMPO):
# Para leads fechados, a idade é (Fechamento - Criação). Para abertos, é (Hoje - Criação).
# Isso evita que leads antigos de sucesso pareçam "velhos/zumbis" para o modelo.
data_referencia = df['Data de fechamento'].fillna(datetime.now())
df['Dias_Desde_Criacao'] = (data_referencia - df['Data de criação']).dt.days
# Garantir que não haja dias negativos (erro de base)
df['Dias_Desde_Criacao'] = df['Dias_Desde_Criacao'].clip(lower=0)

# Tratar nulos
df['Fonte original do tráfego'] = df['Fonte original do tráfego'].fillna('(Sem fonte)')
df['Detalhamento da fonte original do tráfego 1'] = df['Detalhamento da fonte original do tráfego 1'].fillna('(Sem detalhe)')
df['Número de atividades de vendas'] = pd.to_numeric(df['Número de atividades de vendas'], errors='coerce').fillna(0)
print(f"  → Máximo de atividades encontrado na base: {df['Número de atividades de vendas'].max()}")
df['Unidade Desejada'] = df['Unidade Desejada'].fillna('(Sem unidade)')

# Tratar nulos nas features temporais (preencher com 0 ou média para não quebrar o modelo)
for col in ['Ano', 'Mes', 'DiaSemana', 'Trimestre', 'Dias_Desde_Criacao']:
    df[col] = df[col].fillna(0)

# Criar feature de "Densidade de Atividade" (Velocidade de Vendas)
# Objetivo: Diferenciar lead "velho e trabalhado" (bom) de lead "velho e parado" (ruim)
df['Densidade_Atividade'] = df['Número de atividades de vendas'] / (df['Dias_Desde_Criacao'] + 1)

# Criar feature de "Sweet Spot" de atividades (Solicitação: X=Quente, Y=Excesso)
# Lógica: 0 (Novo), 1-3 (Inicial), 4-12 (Ideal/Quente), >12 (Risco/Stalling)
# Criar feature de "Ondas de Atividade" (Ajuste: 0-2 é outlier operacional/balcão)
def classificar_atividade_comercial(n):
    # < 3: Histórico não registrado / Balcão Imediato (Outlier Operacional)
    # O modelo deve ignorar o sucesso "fácil" desses casos para não se iludir
    if n < 3:
        return '0_Outlier_Sem_Registro'
    # Ondas de Conversão Baseadas em Esforço Real
    # Faixas neutras para permitir que o ML identifique onde está a tração real
    elif n <= 6:
        return '1_Inicial_3_a_6'
    elif n <= 12:
        return '2_Engajamento_7_a_12'
    elif n <= 20:
        return '3_Persistencia_13_a_20'
    elif n <= 35:
        return '4_Alta_Persistencia_21_a_35'
    elif n <= 60:
        return '5_Recuperacao_Longa_36_a_60'
    else:
        return '6_Saturacao_Extrema_60+'

df['Faixa_Atividade'] = df['Número de atividades de vendas'].apply(classificar_atividade_comercial)

# Criar feature de categoria de fonte simplificada
def categorizar_fonte(fonte):
    fonte = str(fonte).lower()
    if 'off-line' in fonte or 'offline' in fonte:
        return 'Offline'
    elif 'social pago' in fonte:
        return 'Meta_Ads'
    elif 'pesquisa paga' in fonte:
        return 'Google_Ads'
    elif 'pesquisa orgânica' in fonte:
        return 'Google_Organico'
    elif 'social orgânico' in fonte:
        return 'Social_Organico'
    elif 'tráfego direto' in fonte:
        return 'Direto'
    else:
        return 'Outros'

df['Fonte_Cat'] = df['Fonte original do tráfego'].apply(categorizar_fonte)

# --- NOVA SEGMENTAÇÃO DE MODELOS (SOLICITAÇÃO: SEPARAR VISÕES DIGITAIS) ---
def definir_segmento_ml(fonte_cat):
    if fonte_cat == 'Offline':
        return 'Offline'
    elif fonte_cat in ['Google_Ads', 'Google_Organico']:
        return 'Digital_Busca' # Alta Intenção
    elif fonte_cat in ['Meta_Ads', 'Social_Organico']:
        return 'Digital_Social' # Descoberta/Atenção
    else:
        return 'Digital_Direto' # Direto + Outros (Indicação/Link)

df['Segmento_ML'] = df['Fonte_Cat'].apply(definir_segmento_ml)
print(f"\n[INFO] Segmentação de Modelos definida:")
print(df['Segmento_ML'].value_counts())

# --- VERIFICAÇÃO DE VOLUMETRIA (RESPOSTA AO USUÁRIO) ---
print("\n" + "="*50)
print("[GRAFICO] DISTRIBUIÇÃO TOTAL DE LEADS POR CANAL (BASE COMPLETA)")
print("="*50)
counts = df['Fonte_Cat'].value_counts()
shares = df['Fonte_Cat'].value_counts(normalize=True) * 100
summary_df = pd.DataFrame({'Qtd': counts, 'Share (%)': shares})
print(summary_df.round(1))
print("="*50 + "\n")

# Criar feature de detalhe simplificado para offline
def categorizar_detalhe_offline(row):
    if row['Fonte_Cat'] != 'Offline':
        return 'N/A'
    detalhe = str(row['Detalhamento da fonte original do tráfego 1']).upper()
    if 'CRM_UI' in detalhe:
        return 'CRM_UI'
    elif 'CONTACTS' in detalhe:
        return 'CONTACTS'
    elif 'IMPORT' in detalhe:
        return 'IMPORT'
    else:
        return 'Outros_Offline'

df['Detalhe_Offline'] = df.apply(categorizar_detalhe_offline, axis=1)

# ==============================================================================
# 3. DEFINIÇÃO DO TARGET
# ==============================================================================
print("\n[3/7] Definindo target para treinamento...")

# Classificar etapas
def classificar_etapa(row):
    etapa = str(row['Etapa do negócio']).lower()
    # Se o lead está na base de perdidos, consideramos PERDIDO (exceto se já virou matrícula)
    in_lost = row.get('In_Lost_Base', False)

    if 'matrícula' in etapa:
        return 'SUCESSO'
    elif 'perdido' in etapa:
        return 'PERDIDO'
    elif in_lost:
        return 'PERDIDO' # CRUCIAL: O modelo aprende com o histórico real de falhas (mesmo se o CRM estiver desatualizado)
    elif 'qualificação' in etapa:
        return 'QUALIFICACAO'
    elif 'pausa' in etapa:
        return 'PAUSA'
    else:
        return 'ANDAMENTO'

df['Etapa_Class'] = df.apply(classificar_etapa, axis=1)

# Estatísticas
print("\nDistribuição das etapas (Classificação para o Modelo):")
dist_etapas = df['Etapa_Class'].value_counts()
print(dist_etapas)

total_analisado = dist_etapas.drop('ANDAMENTO', errors='ignore').sum()
total_ignorado = dist_etapas.get('ANDAMENTO', 0)

print(f"\n  → Total Analisado (Treino + Score): {total_analisado:,}")
print(f"  → Total Ignorado (Etapa 'ANDAMENTO' ou indefinida): {total_ignorado:,}")
print(f"    (Leads que não são Sucesso, Perdido, Qualificação ou Pausa não entram no modelo)")

# --- ANÁLISE DE ONDAS DE CONVERSÃO (OFFLINE > 2 ATIVIDADES) ---
print("\n" + "="*60)
print("[ANALISE] ANÁLISE DE ONDAS DE CONVERSÃO (OFFLINE - SEM OUTLIERS < 3)")
print("="*60)

# Filtrar apenas Offline e remover o outlier de balcão (< 3)
df_off_analysis = df[(df['Fonte_Cat'] == 'Offline') & (df['Número de atividades de vendas'] >= 3)].copy()

# Criar flag de conversão para análise
df_off_analysis['Convertido'] = (df_off_analysis['Etapa_Class'] == 'SUCESSO').astype(int)

# Agrupar por número de atividades
activity_analysis = df_off_analysis.groupby('Número de atividades de vendas').agg(
    Total=('Record ID', 'count'),
    Matriculas=('Convertido', 'sum')
).reset_index()

activity_analysis['Taxa_Conv'] = (activity_analysis['Matriculas'] / activity_analysis['Total']).fillna(0)

# Mostrar as ondas (até 25 atividades)
print(f"\nTotal de Leads Offline Trabalhados (>2 ativ): {len(df_off_analysis):,}")
print(f"\nTotal de Leads Offline Trabalhados (>=3 ativ): {len(df_off_analysis):,}")
print("Isso remove o ruído de matrículas imediatas de balcão.")
print("\nOndas de Conversão (Atividade 3 a 60):")
print(activity_analysis[activity_analysis['Número de atividades de vendas'] <= 60].to_string(index=False))
print("="*60 + "\n")

# --- ANÁLISE DE CONVERSÃO POR SEGMENTO DIGITAL (NOVA VISÃO SEPARADA) ---
for seg in ['Digital_Busca', 'Digital_Social', 'Digital_Direto']:
    print("\n" + "="*60)
    print(f"[ANALISE] ANÁLISE DE CONVERSÃO: {seg.upper()}")
    print("="*60)

    df_seg_analysis = df[df['Segmento_ML'] == seg].copy()
    if len(df_seg_analysis) == 0:
        print("  (Sem dados para este segmento)")
        continue
        
    df_seg_analysis['Convertido'] = (df_seg_analysis['Etapa_Class'] == 'SUCESSO').astype(int)

    activity_analysis_seg = df_seg_analysis.groupby('Número de atividades de vendas').agg(
        Total=('Record ID', 'count'),
        Matriculas=('Convertido', 'sum')
    ).reset_index()

    activity_analysis_seg['Taxa_Conv'] = (activity_analysis_seg['Matriculas'] / activity_analysis_seg['Total']).fillna(0)

    print(f"\nTotal de Leads {seg}: {len(df_seg_analysis):,}")
    print(f"\nConversão por Atividade ({seg} - Top 15 faixas):")
    print(activity_analysis_seg.head(15).to_string(index=False))
    print("="*60 + "\n")

# Dataset de treino: apenas SUCESSO e PERDIDO
df_treino = df[df['Etapa_Class'].isin(['SUCESSO', 'PERDIDO'])].copy()
df_treino['Target'] = (df_treino['Etapa_Class'] == 'SUCESSO').astype(int)

print(f"\n  → Dataset de treino: {len(df_treino):,} registros")
print(f"  → Sucesso (1): {df_treino['Target'].sum():,} ({df_treino['Target'].mean()*100:.1f}%)")
print(f"  → Perdido (0): {(1-df_treino['Target']).sum():,} ({(1-df_treino['Target'].mean())*100:.1f}%)")

# Dataset para scoring: QUALIFICAÇÃO e PAUSA
df_score = df[df['Etapa_Class'].isin(['QUALIFICACAO', 'PAUSA'])].copy()
print(f"\n  → Dataset para scoring: {len(df_score):,} registros")

# ==============================================================================
# 4. ENCODING DE FEATURES
# ==============================================================================
print("\n[4/7] Encoding de features categóricas...")

# Features para o modelo
FEATURES_CAT = ['Fonte_Cat', 'Detalhe_Offline', 'Unidade Desejada', 'Faixa_Atividade']
FEATURES_NUM = ['Número de atividades de vendas', 'Mes', 'DiaSemana', 'Trimestre', 'Dias_Desde_Criacao', 'Densidade_Atividade']

# DEFINIÇÃO DE FEATURES ESPECÍFICAS POR MODELO
# Digital: Não usa atividades para não penalizar leads novos
# Offline: Usa atividades pois já nascem com interação (balcão)
FEATURES_NUM_BASE = ['Mes', 'DiaSemana', 'Trimestre', 'Dias_Desde_Criacao', 'Densidade_Atividade']

# Label Encoders
encoders = {}
for col in FEATURES_CAT:
    le = LabelEncoder()
    # Fit em todos os dados para garantir consistência
    le.fit(df[col].astype(str))
    encoders[col] = le
    df_treino[f'{col}_enc'] = le.transform(df_treino[col].astype(str))
    df_score[f'{col}_enc'] = le.transform(df_score[col].astype(str))

# Features codificadas
FEATURES_CAT_ENC = [f'{col}_enc' for col in FEATURES_CAT]

# ==============================================================================
# 5. TREINAMENTO DO MODELO
# ==============================================================================
print("\n[5/7] Treinando modelos ESPECIALISTAS (Offline, Busca, Social, Direto)...")

# Função auxiliar para treinar e avaliar
def treinar_modelo_segmentado(nome, df_segmento, features_segmento):
    print(f"\n  >>> Treinando Modelo: {nome}")
    if len(df_segmento) < 50:
        print(f"      [AVISO] Poucos dados para {nome} ({len(df_segmento)}). Pulando.")
        return None, 0, None
    
    X = df_segmento[features_segmento].values
    y = df_segmento['Target'].values
    
    # Split
    X_tr, X_te, y_tr, y_te = train_test_split(X, y, test_size=0.2, random_state=42, stratify=y)
    
    clf = RandomForestClassifier(
        n_estimators=100, max_depth=12, min_samples_split=20,
        class_weight='balanced', # RESTAURADO: Para equilibrar o acerto entre QUEM MATRICULOU e QUEM PERDEU.
        random_state=42, n_jobs=-1
    )
    clf.fit(X_tr, y_tr)
    
    # Métricas
    auc = roc_auc_score(y_te, clf.predict_proba(X_te)[:, 1])
    print(f"      Registros: {len(df_segmento):,} | AUC-ROC: {auc:.4f}")
    
    # Feature Importance
    imp = pd.DataFrame({'Feature': features_segmento, 'Importance': clf.feature_importances_}).sort_values('Importance', ascending=False)
    print(f"      Top Features: {imp.head(3)['Feature'].tolist()}")
    
    # Mostrar peso específico da Fonte para validar a tese do usuário
    if 'Fonte_Cat_enc' in features_segmento:
        peso_fonte = imp.loc[imp['Feature'] == 'Fonte_Cat_enc', 'Importance'].values[0]
        print(f"      → Peso da Distinção (Pago vs Orgânico): {peso_fonte:.1%}")
    
    return clf, auc, imp

# Dicionário para armazenar modelos e metadados
modelos_dict = {}

# Loop para treinar um modelo por segmento
segmentos_unicos = df['Segmento_ML'].unique()

for segmento in segmentos_unicos:
    print(f"\n--- Processando Segmento: {segmento} ---")
    
    # Filtro de dados
    if segmento == 'Offline':
        # Offline usa regra de >= 3 atividades e INCLUI atividades nas features
        # Offline:
        # - Removemos 'Fonte_Cat' pois é constante ("Offline") e não agrega valor.
        # - Mantemos 'Detalhe_Offline' (CRM_UI vs Import) e 'Unidade'.
        # - Usamos Atividades.
        cols_cat_seg = ['Detalhe_Offline', 'Unidade Desejada', 'Faixa_Atividade']
        cols_cat_seg_enc = [f'{c}_enc' for c in cols_cat_seg]
        
        df_tr_seg = df_treino[(df_treino['Segmento_ML'] == segmento) & (df_treino['Número de atividades de vendas'] >= 3)].copy()
        feats_seg = cols_cat_seg_enc + FEATURES_NUM_BASE + ['Número de atividades de vendas']
    else:
        # Digitais (Busca, Social, Direto) NÃO usam atividades (para não punir leads novos)
        # Digitais:
        # - Mantemos 'Fonte_Cat' (Diferencia Pago vs Orgânico dentro do segmento).
        # - Removemos 'Detalhe_Offline' (é sempre N/A).
        # - NÃO usamos atividades.
        cols_cat_seg = ['Fonte_Cat', 'Unidade Desejada', 'Faixa_Atividade']
        cols_cat_seg_enc = [f'{c}_enc' for c in cols_cat_seg]
        
        df_tr_seg = df_treino[df_treino['Segmento_ML'] == segmento].copy()
        feats_seg = cols_cat_seg_enc + FEATURES_NUM_BASE
    
    modelo, auc, imp = treinar_modelo_segmentado(segmento.upper(), df_tr_seg, feats_seg)
    
    if modelo:
        modelos_dict[segmento] = {
            'model': modelo,
            'auc': auc,
            'features': feats_seg,
            'importance': imp
        }

# ==============================================================================
# 6. SCORING DOS LEADS EM QUALIFICAÇÃO/PAUSA
# ==============================================================================
print("\n[6/7] Aplicando modelos segmentados...")

df_score['Probabilidade_Conversao'] = 0.0

# Aplicar scoring usando o modelo correto para cada segmento
for segmento, dados_modelo in modelos_dict.items():
    mask = df_score['Segmento_ML'] == segmento
    if mask.any():
        feats = dados_modelo['features']
        model = dados_modelo['model']
        
        X_score = df_score.loc[mask, feats].values
        probs = model.predict_proba(X_score)[:, 1]
        
        df_score.loc[mask, 'Probabilidade_Conversao'] = probs
        print(f"  → {segmento} pontuados: {mask.sum():,}")

# Criar nota 1-5 baseada em PROBABILIDADE BALANCEADA (0-100% distribuído)
# Com class_weight='balanced', as probabilidades se espalham melhor, permitindo faixas padrão.
bins = [0, 0.2, 0.4, 0.6, 0.8, 1.0]
df_score['Nota_1a5'] = pd.cut(
    df_score['Probabilidade_Conversao'],
    bins=bins,
    labels=[1, 2, 3, 4, 5],
    include_lowest=True
)
df_score['Nota_1a5'] = df_score['Nota_1a5'].astype(int)

# --- NOVA LÓGICA: SEPARAÇÃO DE SUPER QUALIFICADOS (SOLICITAÇÃO DE URGÊNCIA) ---
# Objetivo: Separar a "Nata" (Probabilidade > 90%) dos leads Nota 5 padrão.
# Ajustado para refletir a nova régua de Nota 5 (>80% no modelo balanceado)
corte_super = 0.90 # Super Qualificado agora é >90% (Elite da Elite)
print(f"\n  → Corte para Super Qualificado (Probabilidade Fixa > {corte_super*100:.0f}%): {corte_super:.2f}")

def definir_prioridade(row):
    # Se for Nota 5 E estiver no Top 10% -> Super Qualificado
    if row['Probabilidade_Conversao'] >= corte_super and row['Nota_1a5'] == 5:
        return '1_Super_Qualificado_TOP'
    elif row['Nota_1a5'] == 5:
        return '2_Alta_Prioridade_ALTA'
    elif row['Nota_1a5'] == 4:
        return '3_Media_Prioridade_MEDIA'
    else:
        return '4_Nutricao_Baixa'

df_score['Prioridade_Atendimento'] = df_score.apply(definir_prioridade, axis=1)

print(f"\n  → {len(df_score):,} leads pontuados")
print("\nDistribuição das notas:")
print(df_score['Nota_1a5'].value_counts().sort_index())
print("\nDistribuição de Prioridade:")
print(df_score['Prioridade_Atendimento'].value_counts().sort_index())

# Classificação de Tipo de Unidade (Própria vs Franqueada)
def classificar_tipo_unidade(unidade):
    unidade_upper = str(unidade).upper()
    for propria in UNIDADES_PROPRIAS:
        if propria in unidade_upper:
            return 'Própria'
    return 'Franqueada'

df_score['Tipo_Unidade'] = df_score['Unidade Desejada'].apply(classificar_tipo_unidade)

# Estatísticas por nota
print("\nProbabilidade média por nota:")
stats_nota = df_score.groupby('Nota_1a5')['Probabilidade_Conversao'].agg(['mean', 'min', 'max', 'count'])
print(stats_nota.round(3))

# --- ANÁLISE DE PERFIL DE SUCESSO (DNA DA MATRÍCULA) ---
print("\n[ANÁLISE] Gerando Perfil de Sucesso (Características de quem Matricula)...")
perfil_stats_list = []

for seg in segmentos_unicos:
    subset = df[df['Segmento_ML'] == seg]
    
    # AJUSTE DE COERÊNCIA: Para Offline, analisar apenas leads TRABALHADOS (>=3 atividades)
    # Isso garante que a Aba 9 (Perfil) não seja contaminada por balcão (0 ativ) e contradiga o Modelo (Aba 8).
    if seg == 'Offline':
        subset = subset[subset['Número de atividades de vendas'] >= 3]
        
    sucesso = subset[subset['Etapa_Class'] == 'SUCESSO']
    perdido = subset[subset['Etapa_Class'] == 'PERDIDO']
    
    if len(sucesso) > 0 and len(perdido) > 0:
        stats = {
            'Segmento': seg,
            'Ativ_Sucesso': sucesso['Número de atividades de vendas'].median(),
            'Ativ_Perdido': perdido['Número de atividades de vendas'].median(),
            'Dias_Sucesso': sucesso['Dias_Desde_Criacao'].median(),
            'Dias_Perdido': perdido['Dias_Desde_Criacao'].median(),
            'Dens_Sucesso': sucesso['Densidade_Atividade'].median(),
            'Dens_Perdido': perdido['Densidade_Atividade'].median(),
            'Conv_Rate': len(sucesso) / len(subset)
        }
        perfil_stats_list.append(stats)

df_perfil_stats = pd.DataFrame(perfil_stats_list)
print(df_perfil_stats.to_string(index=False))

# ==============================================================================
# 7. EXPORTAÇÃO
# ==============================================================================
print("\n[7/7] Exportando resultados...")

# Salvar modelo
# Salvando dicionário com ambos os modelos
modelo_path = os.path.join(PASTA_MODELS, 'lead_scoring_model.pkl')
dados_exportacao_modelo = {'encoders': encoders}
for seg, dados in modelos_dict.items():
    dados_exportacao_modelo[f'model_{seg}'] = dados['model']
    dados_exportacao_modelo[f'features_{seg}'] = dados['features']

with open(modelo_path, 'wb') as f:
    pickle.dump(dados_exportacao_modelo, f)
print(f"  → Modelo salvo: {modelo_path}")

# Preparar output
output_cols = [
    'Record ID', 'Nome do negócio', 'Etapa do negócio', 'Unidade Desejada', 'Tipo_Unidade',
    'Fonte original do tráfego', 'Fonte_Cat', 'Segmento_ML', 'Faixa_Atividade', 'Número de atividades de vendas',
    'Data de criação', 'Dias_Desde_Criacao', 'Proprietário do negócio',
    'Probabilidade_Conversao', 'Nota_1a5', 'Prioridade_Atendimento'
]

# Filtrar colunas existentes
output_cols_exist = [c for c in output_cols if c in df_score.columns]
df_output = df_score[output_cols_exist].copy()

# Ordenar por probabilidade (maior primeiro)
df_output = df_output.sort_values('Probabilidade_Conversao', ascending=False)

# Exportar CSV completo
csv_path = os.path.join(PASTA_DADOS, f'leads_scored_{DATA_HOJE}.csv')
df_output.to_csv(csv_path, index=False, encoding='utf-8-sig', sep=';')
print(f"  → CSV completo: {csv_path}")

# --- NOVO: DIÁRIO DE BORDO (HISTÓRICO DE EVOLUÇÃO DA IA) ---
# Objetivo: Registrar como o "cérebro" da IA muda ao longo do tempo para contextualizar mudanças de comportamento.
log_path = os.path.join(PASTA_DADOS, 'Historico_Evolucao_IA.csv')
print(f"\n[LOG] Atualizando histórico de evolução (Diário de Bordo) em: {log_path}")

registros_log = []
for seg, dados in modelos_dict.items():
    # Capturar os 3 principais fatores de decisão deste modelo hoje
    top_feats = ", ".join(dados['importance'].head(3)['Feature'].tolist())
    
    registros_log.append({
        'Data_Execucao': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'Segmento': seg,
        'AUC_Performance': round(dados['auc'], 4),
        'Fatores_Decisivos_Top3': top_feats,
        'Volume_Treino': len(df_treino[df_treino['Segmento_ML'] == seg]) if 'Segmento_ML' in df_treino.columns else 0
    })

df_log_novo = pd.DataFrame(registros_log)

if os.path.exists(log_path):
    try:
        df_log_antigo = pd.read_csv(log_path, sep=';', encoding='utf-8-sig')
        df_log_final = pd.concat([df_log_antigo, df_log_novo], ignore_index=True)
    except:
        df_log_final = df_log_novo
else:
    df_log_final = df_log_novo

df_log_final.to_csv(log_path, index=False, sep=';', encoding='utf-8-sig')

# ==============================================================================
# GERAÇÃO DE EXCEL SEGMENTADO (OFFLINE vs DIGITAL)
# ==============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Estilos Globais
HEADER_FILL = PatternFill(start_color='CE0E2D', end_color='CE0E2D', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def gerar_excel_segmentado(df_scored, df_historico, nome_arquivo, segmento_titulo, auc_modelo, stats_global, df_importance, df_perfil_stats=None, segmento_raw=None):
    wb = Workbook()
    
    # Texto de fonte detalhado (Solicitação: Lugar e Colunas)
    leads_fn = f"Data/{os.path.basename(CAMINHO_LEADS)}"
    lost_fn = f"Data/{os.path.basename(CAMINHO_PERDIDOS)}"
    source_text = f"Origem: {leads_fn} & {lost_fn} | Colunas-Chave: 'Etapa do negócio', 'Fonte original', 'Atividades'"

    def df_to_sheet(ws, df, start_row=1):
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = THIN_BORDER
                if r_idx == start_row:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)

    def add_sheet_description(ws, lines, start_row=2):
        """Adiciona um bloco descritivo (em linguagem de negócio) no topo da aba.
        lines: lista de strings. start_row: linha onde começa (1-indexed).
        """
        for i, text in enumerate(lines, start=start_row):
            cell = ws.cell(row=i, column=1, value=text)
            cell.font = Font(bold=False, size=9)
            cell.alignment = Alignment(wrap_text=True)
            # Mesclar algumas colunas para melhor visual
            try:
                ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=min(6, ws.max_column if ws.max_column>0 else 6))
            except Exception:
                pass

    def ajustar_largura(ws):
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # --- ABA 1: RESUMO ---
    ws1 = wb.active
    ws1.title = '1_Resumo'
    ws1['A1'] = f'LEAD SCORING - {segmento_titulo}'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A2'] = source_text
    ws1['A3'] = f'Data: {DATA_HOJE}'
    ws1['A4'] = f'Leads Pontuados: {len(df_scored):,}'
    ws1['A5'] = f'Precisão do Modelo (AUC): {auc_modelo:.4f}'

    add_sheet_description(ws1, [
        'Objetivo: Resumo executivo do Lead Scoring para tomada de decisão.',
        'Público: Gestores comerciais e operação. Interpretação: AUC indica qualidade do modelo; Notas 1-5 refletem probabilidade de matrícula.',
        "Uso prático: priorizar contatos pela coluna 'Prioridade_Atendimento'. Leads sem modelo serão indicados como 'Sem modelo'.",
        "Caveats: Segmento 'Offline' pode apresentar maior ruído — ver aba '8_Fatores_Decisivos' e '9_Perfil_Ideal_Sucesso' para contexto."
    ], start_row=6)

    # Contexto Global (Solicitado)
    ws1['E3'] = "CONTEXTO GLOBAL (Todos os Canais)"
    ws1['E3'].font = Font(bold=True)
    ws1['E4'] = f"Total Leads Pontuados: {stats_global['total']:,}"
    if stats_global['total'] > 0:
        ws1['E5'] = f"Offline: {stats_global['offline']:,} ({stats_global['offline']/stats_global['total']:.1%})"
        ws1['E6'] = f"Digital: {stats_global['digital']:,} ({stats_global['digital']/stats_global['total']:.1%})"

    ws1['A7'] = 'DISTRIBUIÇÃO POR NOTA:'
    ws1['A7'].font = Font(bold=True)
    
    if not df_scored.empty:
        resumo_notas = df_scored.groupby('Nota_1a5').agg(
            Quantidade=('Nota_1a5', 'count'),
            Prob_Media=('Probabilidade_Conversao', 'mean')
        ).reset_index()
        
        # Adicionar Share de cada nota
        resumo_notas['Share'] = resumo_notas['Quantidade'] / resumo_notas['Quantidade'].sum()
        
        df_to_sheet(ws1, resumo_notas, start_row=8)
        
        # Formatar porcentagens (Prob_Media e Share)
        for r in range(9, 9 + len(resumo_notas)):
            ws1.cell(row=r, column=3).number_format = '0.0%' # Prob
            ws1.cell(row=r, column=4).number_format = '0.0%' # Share

    # --- ABA 2: TOP 500 (NOTA 5) ---
    ws2 = wb.create_sheet('2_Top500_Nota5')
    ws2['A1'] = 'TOP 500 LEADS - NOTA 5'
    ws2['A1'].font = Font(bold=True)
    ws2['A2'] = source_text
    add_sheet_description(ws2, [
        'Objetivo: Lista dos 500 leads com maior probabilidade estimada (Nota 5).',
        'Público: Equipe de atendimento — contactar imediatamente. Verificar campo Prioridade_Atendimento antes do contato.'
    ], start_row=3)
    df_to_sheet(ws2, df_scored[df_scored['Nota_1a5'] == 5].head(500), start_row=6)
    ajustar_largura(ws2)

    # --- ABA 3: NOTA 4 ---
    ws3 = wb.create_sheet('3_Nota4')
    ws3['A1'] = 'TOP 500 LEADS - NOTA 4'
    ws3['A1'].font = Font(bold=True)
    ws3['A2'] = source_text
    add_sheet_description(ws3, [
        'Objetivo: Leads com alta prioridade (Nota 4).',
        'Público: Equipe de atendimento — priorizar dentro da fila após Nota 5.'
    ], start_row=3)
    df_to_sheet(ws3, df_scored[df_scored['Nota_1a5'] == 4].head(500), start_row=6)
    ajustar_largura(ws3)

    # --- ABA 4: TOP 10 POR UNIDADE ---
    ws4 = wb.create_sheet('4_Top10_Por_Unidade')
    ws4['A1'] = 'TOP 10 POR UNIDADE'
    ws4['A1'].font = Font(bold=True)
    ws4['A2'] = source_text
    add_sheet_description(ws4, [
        'Objetivo: Exibir até 10 leads prioritários por unidade (melhor alocação de esforço).',
        'Público: Gerentes de unidade — usar para definir roteiros de contato locais.'
    ], start_row=3)
    if not df_scored.empty:
        top_unidade = df_scored.groupby('Unidade Desejada').apply(
            lambda x: x.nlargest(10, 'Probabilidade_Conversao')
        ).reset_index(drop=True)
        df_to_sheet(ws4, top_unidade, start_row=3)
    ajustar_largura(ws4)

    # --- ABA 5: EM QUALIFICAÇÃO ---
    ws5 = wb.create_sheet('5_Em_Qualificacao')
    ws5['A1'] = 'LEADS EM QUALIFICAÇÃO'
    ws5['A1'].font = Font(bold=True)
    ws5['A2'] = source_text
    add_sheet_description(ws5, [
        'Objetivo: Listar leads em etapa de Qualificação (não finalizados).',
        'Público: Equipe comercial — revisar e priorizar ações conforme nota.'
    ], start_row=3)
    df_to_sheet(ws5, df_scored[df_scored['Etapa do negócio'].str.contains('QUALIFICAÇÃO', case=False, na=False)].head(1000), start_row=6)
    ajustar_largura(ws5)

    # --- ABA 6: EM PAUSA ---
    ws6 = wb.create_sheet('6_Em_Pausa')
    ws6['A1'] = 'LEADS EM PAUSA'
    ws6['A1'].font = Font(bold=True)
    ws6['A2'] = source_text
    add_sheet_description(ws6, [
        'Objetivo: Leads colocados em Pausa — reengajar quando apropriado.',
        'Público: Equipe de relacionamento — usar notas para priorização de reengajamento.'
    ], start_row=3)
    df_to_sheet(ws6, df_scored[df_scored['Etapa do negócio'].str.contains('PAUSA', case=False, na=False)].head(1000), start_row=6)
    ajustar_largura(ws6)

    # --- ABA 7: RESUMO CONVERSÃO vs ATIVIDADE (NOVA) ---
    ws7 = wb.create_sheet('7_Resumo_Conv_vs_Atividade')
    ws7['A1'] = 'HISTÓRICO: TAXA DE CONVERSÃO POR NÚMERO DE ATIVIDADES'
    ws7['A1'].font = Font(bold=True)
    ws7['A2'] = source_text
    add_sheet_description(ws7, [
        'Objetivo: Comparar taxa histórica de conversão por número de atividades e cruzar com a nota média do modelo.',
        'Público: Analistas e gestores — usar para verificar se esforço correlaciona com conversão.'
    ], start_row=3)
    
    # Filtrar apenas leads finalizados (Sucesso ou Perdido) para cálculo real de conversão
    df_closed = df_historico[df_historico['Etapa_Class'].isin(['SUCESSO', 'PERDIDO'])].copy()
    df_closed['Is_Won'] = (df_closed['Etapa_Class'] == 'SUCESSO').astype(int)
    
    if not df_closed.empty:
        resumo_ativ = df_closed.groupby('Número de atividades de vendas').agg(
            Total_Finalizados=('Record ID', 'count'),
            Sucesso=('Is_Won', 'sum')
        ).reset_index()
        
        resumo_ativ['Perdido'] = resumo_ativ['Total_Finalizados'] - resumo_ativ['Sucesso']
        resumo_ativ['Taxa_Conversao'] = resumo_ativ['Sucesso'] / resumo_ativ['Total_Finalizados']
        
        # Calcular totais antes de renomear colunas (para evitar KeyError)
        total_finalizados = resumo_ativ['Total_Finalizados'].sum()
        total_sucesso = resumo_ativ['Sucesso'].sum()
        total_perdido = resumo_ativ['Perdido'].sum()

        # Limitar para não ficar gigante (ex: até 30 atividades)
        # resumo_ativ = resumo_ativ[resumo_ativ['Número de atividades de vendas'] <= 30] # REMOVIDO LIMITADOR
        
        # Adicionar Nota Média do Modelo para essa faixa (Cruzamento Histórico vs Modelo Atual)
        if not df_scored.empty:
            # Converter para int para cálculo (usando copy para segurança)
            temp_score = df_scored.copy()
            temp_score['Nota_Calc'] = temp_score['Nota_1a5'].astype(int)
            
            # Calcular Média da Nota e Média de Dias (Recência) para explicar notas baixas
            stats_atual = temp_score.groupby('Número de atividades de vendas').agg(
                Nota_Media_Atual=('Nota_Calc', 'mean'),
                Dias_Media_Atual=('Dias_Desde_Criacao', 'mean')
            ).reset_index()
            
            resumo_ativ = pd.merge(resumo_ativ, stats_atual, on='Número de atividades de vendas', how='left')
            
            # Aplicar a regra solicitada:
            # Para < 3 atividades, a nota é distorcida e não deve ser mostrada.
            # Para >= 3, mostramos a nota média calculada.
            resumo_ativ['Nota_Media_Atual'] = resumo_ativ['Nota_Media_Atual'].fillna(0).round(1) # Calcula para todos primeiro
            resumo_ativ['Dias_Media_Atual'] = resumo_ativ['Dias_Media_Atual'].fillna(0).round(0) # Idade média do lead
            resumo_ativ.loc[resumo_ativ['Número de atividades de vendas'] < 3, 'Nota_Media_Atual'] = "Sem rastreio de atividade"
            
            # Renomear colunas para apresentação final mais amigável
            resumo_ativ.columns = [
                'Nº Atividades', 'Total Finalizados', 'Matrículas', 'Perdidos', 
                'Taxa Conv. Histórica', 'Nota Média (Modelo)', 'Idade Média (Dias)'
            ]

        # Formatar colunas
        df_to_sheet(ws7, resumo_ativ, start_row=3)
        
        # Aplicar formato de porcentagem na coluna de taxa (coluna 5)
        for row in range(4, 4 + len(resumo_ativ)):
            ws7.cell(row=row, column=5).number_format = '0.0%'
            
        # Adicionar linha de total para clareza
        total_row_num = 4 + len(resumo_ativ)

        # Célula de Título
        total_cell = ws7.cell(row=total_row_num, column=1, value='TOTAL (do Segmento)')
        total_cell.font = Font(bold=True)

        # Células de Valores
        ws7.cell(row=total_row_num, column=2, value=total_finalizados).font = Font(bold=True)
        ws7.cell(row=total_row_num, column=3, value=total_sucesso).font = Font(bold=True)
        ws7.cell(row=total_row_num, column=4, value=total_perdido).font = Font(bold=True)

        # Célula de Taxa de Conversão Total
        if total_finalizados > 0:
            taxa_total = total_sucesso / total_finalizados
            taxa_cell = ws7.cell(row=total_row_num, column=5, value=taxa_total)
            taxa_cell.font = Font(bold=True)
            taxa_cell.number_format = '0.0%'

        # Adicionar Média Geral da Nota (Apenas para leads com rastreio >= 3)
        if not df_scored.empty and 'Nota_Media_Atual' in resumo_ativ.columns:
            # Filtrar apenas leads válidos para a média geral (ignora < 3)
            valid_leads_score = df_scored[df_scored['Número de atividades de vendas'] >= 3]
            if not valid_leads_score.empty:
                media_geral_validos = valid_leads_score['Nota_1a5'].astype(float).mean()
                media_dias_validos = valid_leads_score['Dias_Desde_Criacao'].mean()
                
                # Coluna 6 é a Nota_Media_Atual
                ws7.cell(row=total_row_num, column=6, value=media_geral_validos).font = Font(bold=True)
                ws7.cell(row=total_row_num, column=6).number_format = '0.0'
                
                # Coluna 7 é a Dias_Media_Atual
                ws7.cell(row=total_row_num, column=7, value=media_dias_validos).font = Font(bold=True)
                ws7.cell(row=total_row_num, column=7).number_format = '0'

        # Aplicar estilo na linha inteira
        total_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        for col_idx in range(1, ws7.max_column + 1):
            ws7.cell(row=total_row_num, column=col_idx).fill = total_fill
            ws7.cell(row=total_row_num, column=col_idx).border = THIN_BORDER

    # --- ABA 8: FATORES DECISIVOS (FEATURE IMPORTANCE) ---
    ws8 = wb.create_sheet('8_Fatores_Decisivos')
    ws8['A1'] = 'O QUE DEFINE A NOTA DO LEAD? (PESO DAS VARIÁVEIS)'
    ws8['A1'].font = Font(bold=True)
    ws8['A2'] = 'Interpretação: Quanto maior a %, mais essa variável influencia o robô a dar uma nota alta ou baixa.'
    ws8['A3'] = source_text
    add_sheet_description(ws8, [
        'Objetivo: Explicar quais variáveis mais influenciam a nota (interpretação do modelo).',
        'Público: Gestores e equipe técnica; usar para entender prioridades de ação e possíveis vieses.'
    ], start_row=4)
    if df_importance is not None:
        # Adicionar coluna de % acumulada para Pareto
        df_importance['Importance_Acum'] = df_importance['Importance'].cumsum()
        
        # Renomear para ficar bonito
        df_imp_final = df_importance.rename(columns={'Importance': 'Peso (%)', 'Importance_Acum': 'Peso Acumulado'})
        df_to_sheet(ws8, df_imp_final, start_row=5)
        
        # Formatar %
        for r in range(6, 6 + len(df_importance)):
            ws8.cell(row=r, column=2).number_format = '0.0%'
            ws8.cell(row=r, column=3).number_format = '0.0%'
            
    ajustar_largura(ws8)

    # --- ABA 9: PERFIL IDEAL DE SUCESSO (CARACTERÍSTICAS) ---
    ws9 = wb.create_sheet('9_Perfil_Ideal_Sucesso')
    ws9['A1'] = 'O DNA DO SUCESSO: O QUE DIFERENCIA QUEM MATRICULA?'
    ws9['A1'].font = Font(bold=True, size=12, color='CE0E2D')
    ws9['A2'] = 'Comparativo REAL (Histórico): Como se comportam os leads que compram vs os que desistem.'
    ws9['A3'] = 'Use isso para definir metas: Se o Sucesso tem 10 atividades e o Perdido tem 2, sua meta deve ser 10.'
    add_sheet_description(ws9, [
        'Objetivo: Perfil comparativo entre leads que convertem e que perdem — insight operacional.',
        'Público: Operação e gestão — usar para definir metas de atividades e ritmo de contato.'
    ], start_row=4)
    
    if df_perfil_stats is not None and segmento_raw is not None:
        # Filtrar estatísticas deste segmento
        stats_seg = df_perfil_stats[df_perfil_stats['Segmento'] == segmento_raw]
        
        if not stats_seg.empty:
            row = stats_seg.iloc[0]
            
            # Calcular frequência legível (1 contato a cada X dias)
            freq_sucesso = int(1 / row['Dens_Sucesso']) if row['Dens_Sucesso'] > 0 else 0
            freq_perdido = int(1 / row['Dens_Perdido']) if row['Dens_Perdido'] > 0 else 0
            
            dados_perfil = [
                ['Característica', 'Leads de SUCESSO (Mediana)', 'Leads PERDIDOS (Mediana)', 'Conclusão / Regra'],
                ['Atividades de Vendas', row['Ativ_Sucesso'], row['Ativ_Perdido'], 
                 'Sucesso exige mais esforço' if row['Ativ_Sucesso'] > row['Ativ_Perdido'] else 'Volume não é diferencial'],
                ['Tempo até Resolução (Dias)', row['Dias_Sucesso'], row['Dias_Perdido'], 
                 'Sucesso é mais rápido' if row['Dias_Sucesso'] < row['Dias_Perdido'] else 'Sucesso demora mais'],
                ['Frequência de Contato', f"1 a cada {freq_sucesso} dias", f"1 a cada {freq_perdido} dias", 
                 'Sucesso não deixa o lead esfriar' if freq_sucesso < freq_perdido else 'Ritmo igual'],
            ]
            
            for i, linha in enumerate(dados_perfil, start=5):
                for j, val in enumerate(linha, start=1):
                    cell = ws9.cell(row=i, column=j, value=val)
                    if i == 5:
                        cell.fill = HEADER_FILL
                        cell.font = HEADER_FONT
            
            ajustar_largura(ws9)

    # --- ABA 10: PROJEÇÃO FINANCEIRA (NOVO) ---
    ws10 = wb.create_sheet('10_Projecao_Financeira')
    ws10['A1'] = 'PROJEÇÃO FINANCEIRA (PIPELINE REPRESADO)'
    ws10['A1'].font = Font(bold=True, size=12, color='CE0E2D')
    ws10['A2'] = f'Premissas: Ticket Própria R$ {TICKET_PROPRIA:,.2f} | Ticket Franquia R$ {TICKET_FRANQUIA:,.2f} | LTV: 66 Meses (Com discriminação de Ajuste 7% a.a.)'
    add_sheet_description(ws10, [
        'Objetivo: Projeção financeira esperada por nota do lead, considerando premissas de ticket e LTV.',
        'Público: Gestão financeira e comercial — atenção: valores são estimativas baseadas em probabilidade (não garantias).'
    ], start_row=3)
    
    if not df_scored.empty:
        # Calcular valor por lead
        def get_ticket(tipo):
            return TICKET_PROPRIA if 'Própria' in str(tipo) else TICKET_FRANQUIA
            
        df_fin = df_scored.copy()
        df_fin['Ticket_Medio'] = df_fin['Tipo_Unidade'].apply(get_ticket)
        
        # Cálculo de LTV Ponderado (Esperado) com Reajuste Anual de 7%
        taxa_anual = 1.07
        fator_ano1 = 12
        fator_ano2 = 12 * taxa_anual
        fator_ano3 = 12 * (taxa_anual**2)
        fator_ano4 = 12 * (taxa_anual**3)
        fator_ano5 = 12 * (taxa_anual**4)
        fator_ano6_semestral = 6 * (taxa_anual**5)
        
        mult_24m = fator_ano1 + fator_ano2
        mult_66m_com_ajuste = fator_ano1 + fator_ano2 + fator_ano3 + fator_ano4 + fator_ano5 + fator_ano6_semestral
        mult_66m_sem_ajuste = 66

        df_fin['Esp_1M'] = df_fin['Ticket_Medio'] * df_fin['Probabilidade_Conversao']
        df_fin['Esp_12M'] = df_fin['Esp_1M'] * 12
        df_fin['Esp_24M'] = df_fin['Esp_1M'] * mult_24m
        df_fin['LTV_Base'] = df_fin['Esp_1M'] * mult_66m_sem_ajuste
        df_fin['LTV_Total'] = df_fin['Esp_1M'] * mult_66m_com_ajuste
        df_fin['LTV_Ajuste'] = df_fin['LTV_Total'] - df_fin['LTV_Base']
        
        # Função auxiliar para gerar tabela por tipo
        def escrever_tabela_financeira(ws, titulo, df_sub, linha_inicial):
            ws.cell(row=linha_inicial, column=1, value=titulo).font = Font(bold=True, size=11)
            
            if df_sub.empty:
                ws.cell(row=linha_inicial+1, column=1, value="(Sem dados para este perfil)")
                return linha_inicial + 3

            fin_stats = df_sub.groupby('Nota_1a5').agg(
                Qtd_Leads=('Record ID', 'count'),
                Rec_1M=('Esp_1M', 'sum'),
                Rec_12M=('Esp_12M', 'sum'),
                Rec_24M=('Esp_24M', 'sum'),
                LTV_Base=('LTV_Base', 'sum'),
                LTV_Ajuste=('LTV_Ajuste', 'sum'),
                LTV_Total=('LTV_Total', 'sum')
            ).reset_index().sort_values('Nota_1a5', ascending=False)
            
            # Totais
            total_qtd = fin_stats['Qtd_Leads'].sum()
            totais_vals = [
                fin_stats['Rec_1M'].sum(), 
                fin_stats['Rec_12M'].sum(), fin_stats['Rec_24M'].sum(), 
                fin_stats['LTV_Base'].sum(), fin_stats['LTV_Ajuste'].sum(), fin_stats['LTV_Total'].sum()
            ]
            
            cols_fin = ['Nota (1-5)', 'Qtd Leads', 'Rec. 1 Mês', 'Rec. 1 Ano', 'Rec. 2 Anos', 'LTV Base (66m)', 'Ajuste (7% a.a.)', 'LTV Total']
            
            # Header
            header_row = linha_inicial + 1
            for c_idx, col_name in enumerate(cols_fin, 1):
                cell = ws.cell(row=header_row, column=c_idx, value=col_name)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
            
            current_row = header_row + 1
            for r_idx, row in enumerate(dataframe_to_rows(fin_stats, index=False, header=False), start=current_row):
                ws.cell(row=r_idx, column=1, value=row[0]).border = THIN_BORDER
                ws.cell(row=r_idx, column=2, value=row[1]).border = THIN_BORDER
                
                # Colunas de Valores (3 a 8)
                for c_offset in range(6):
                    val = row[2 + c_offset]
                    cell = ws.cell(row=r_idx, column=3 + c_offset, value=val)
                    cell.number_format = '#,##0.00'
                    cell.border = THIN_BORDER
                
                current_row = r_idx
            
            # Linha de Total
            total_row = current_row + 1
            ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
            ws.cell(row=total_row, column=2, value=total_qtd).font = Font(bold=True)
            
            for i, val in enumerate(totais_vals):
                cell = ws.cell(row=total_row, column=3 + i, value=val)
                cell.number_format = '#,##0.00'
                cell.font = Font(bold=True)
            
            # Bordas do total
            for c in range(1, 9):
                ws.cell(row=total_row, column=c).border = THIN_BORDER
                ws.cell(row=total_row, column=c).fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

            return total_row + 3 # Espaço para a próxima tabela

        # 1. Tabela Próprias
        df_proprias = df_fin[df_fin['Tipo_Unidade'] == 'Própria']
        prox_linha = escrever_tabela_financeira(ws10, "1. UNIDADES PRÓPRIAS", df_proprias, 4)
        
        # 2. Tabela Franquias
        df_franquias = df_fin[df_fin['Tipo_Unidade'] == 'Franqueada']
        escrever_tabela_financeira(ws10, "2. UNIDADES FRANQUEADAS", df_franquias, prox_linha)
            
        ajustar_largura(ws10)

    try:
        wb.save(nome_arquivo)
        print(f"  → Arquivo gerado: {nome_arquivo}")
    except PermissionError:
        print(f"\n[ERRO] O arquivo parece estar ABERTO no Excel: {nome_arquivo}")
        print("       Salvando cópia de segurança para não perder o processamento...")
        novo_nome = nome_arquivo.replace('.xlsx', f'_Copia_{datetime.now().strftime("%H%M%S")}.xlsx')
        wb.save(novo_nome)
        print(f"  → Salvo como: {novo_nome}")

# Estatísticas Globais para o Resumo
stats_global = {
    'total': len(df_output),
    'offline': len(df_output[df_output['Segmento_ML'] == 'Offline']),
    'digital': len(df_output[df_output['Segmento_ML'] != 'Offline'])
}

# Gerar Arquivos Separados por Segmento
for segmento in modelos_dict.keys():
    df_out_seg = df_output[df_output['Segmento_ML'] == segmento].copy()
    df_hist_seg = df[df['Segmento_ML'] == segmento].copy()
    
    dados_modelo = modelos_dict[segmento]
    auc_seg = dados_modelo['auc']
    imp_seg = dados_modelo['importance']
    
    nome_arquivo = f'Lead_Scoring_{segmento}_{DATA_HOJE}.xlsx'
    path_seg = os.path.join(PASTA_RELATORIOS, nome_arquivo)
    
    gerar_excel_segmentado(df_out_seg, df_hist_seg, path_seg, segmento.upper(), auc_seg, stats_global, imp_seg, df_perfil_stats, segmento)

# ==============================================================================
# RESUMO FINAL
# ==============================================================================
print("\n" + "="*80)
print("RESUMO FINAL")
print("="*80)
print(f"\n[OK] Modelos treinados separadamente por segmento:")
for seg, dados in modelos_dict.items():
    print(f"   - {seg}: AUC {dados['auc']:.4f}")

print(f"[OK] {len(df_score):,} leads pontuados")
print(f"\n[GRAFICO] DISTRIBUIÇÃO DAS NOTAS:")
for nota in sorted(df_score['Nota_1a5'].dropna().unique()):
    qtd = len(df_score[df_score['Nota_1a5'] == nota])
    prob_media = df_score[df_score['Nota_1a5'] == nota]['Probabilidade_Conversao'].mean()
    print(f"   Nota {int(nota)}: {qtd:,} leads (prob. média: {prob_media*100:.1f}%)")

print(f"\n[ARQUIVOS] ARQUIVOS GERADOS:")
print(f"   → {modelo_path}")
for segmento in modelos_dict.keys():
    print(f"   → Lead_Scoring_{segmento}_{DATA_HOJE}.xlsx")
print("="*80)
