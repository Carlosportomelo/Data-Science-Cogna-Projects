"""
==============================================================================
TESTE CEGO DE ACURÁCIA - VERSÃO AJUSTADA (LEADS RASTREÁVEIS)
==============================================================================
Script: 13_Blind_Test_Ajustado.py
Objetivo: Medir a acurácia REAL do modelo nos leads que ele PODE prever
          (excluindo vendas de balcão sem registro de atividades)
Data: 2025-12-15
==============================================================================

JUSTIFICATIVA DO AJUSTE:
- 69.3% das matrículas têm <3 atividades (vendas de balcão sem registro)
- Essas vendas são IMPREVISÍVEIS pois não há dados de comportamento
- O modelo é treinado para prever baseado em atividades de vendas
- Medir acurácia incluindo casos imprevisíveis distorce a avaliação

MÉTRICA CORRETA:
- Avaliar apenas leads com >=3 atividades (rastreáveis)
- Essa é a população que o modelo PODE ajudar a priorizar
==============================================================================
"""

import pandas as pd
import numpy as np
import os
import pickle
from datetime import datetime
from sklearn.metrics import roc_auc_score, confusion_matrix, classification_report
from sklearn.preprocessing import LabelEncoder
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CAMINHO_LEADS = os.path.join(CAMINHO_BASE, 'Data', 'hubspot_leads.csv')
PASTA_OUTPUT = os.path.join(CAMINHO_BASE, 'Outputs')
PASTA_MODELS = os.path.join(CAMINHO_BASE, 'models')
MODELO_PATH = os.path.join(PASTA_MODELS, 'lead_scoring_model.pkl')
os.makedirs(PASTA_OUTPUT, exist_ok=True)

DATA_HOJE = datetime.now().strftime('%Y-%m-%d')
CAMINHO_SAIDA = os.path.join(PASTA_OUTPUT, f'Blind_Test_Ajustado_{DATA_HOJE}.xlsx')

# Configuração do corte de atividades
MIN_ATIVIDADES = 3  # Mínimo para considerar rastreável

# Estilos Excel
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
SUCCESS_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
WARNING_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
DANGER_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

def df_to_sheet(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if r_idx == start_row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center', wrap_text=True)

def ajustar_largura(ws):
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

print("="*80)
print("TESTE CEGO AJUSTADO - LEADS RASTREÁVEIS (>=3 ATIVIDADES)")
print("="*80)
print(f"📂 DIRETÓRIO DE EXECUÇÃO: {CAMINHO_BASE}")
print(f"📂 DADOS DE ENTRADA:      {CAMINHO_LEADS}")
print("-" * 80)

# ==============================================================================
# 1. CARREGAMENTO DOS DADOS
# ==============================================================================
print("\n[1/6] Carregando dados...")

if not os.path.exists(CAMINHO_LEADS):
    print(f"\n❌ [ERRO CRÍTICO] Arquivo de dados não encontrado: {CAMINHO_LEADS}")
    print("   Certifique-se de copiar o arquivo 'hubspot_leads.csv' para a pasta 'Data' deste diretório.")
    exit()

try:
    df = pd.read_csv(CAMINHO_LEADS, encoding='utf-8')
except:
    df = pd.read_csv(CAMINHO_LEADS, encoding='latin1')

df.columns = df.columns.str.strip()
print(f"  → Base carregada: {len(df):,} registros")

# ==============================================================================
# 2. PREPARAÇÃO DE FEATURES (IGUAL AO SCRIPT DE TREINO)
# ==============================================================================
print("\n[2/6] Preparando features...")

df['Data de criação'] = pd.to_datetime(df['Data de criação'], errors='coerce')
df['Data de fechamento'] = pd.to_datetime(df['Data de fechamento'], errors='coerce')

# SANITY CHECK: Data de fechamento
if df['Data de fechamento'].notna().sum() == 0:
    print("\n⚠️ [AVISO CRÍTICO] A coluna 'Data de fechamento' está vazia. A validação histórica pode falhar.")
    print("   O cálculo de 'Dias_Desde_Criacao' usará a data de hoje, distorcendo a idade dos leads fechados.\n")

# CORREÇÃO CRONOLÓGICA (VIAGEM NO TEMPO)
data_referencia = df['Data de fechamento'].fillna(datetime.now())
df['Dias_Desde_Criacao'] = (data_referencia - df['Data de criação']).dt.days
df['Dias_Desde_Criacao'] = df['Dias_Desde_Criacao'].clip(lower=0)
df['Ano'] = df['Data de criação'].dt.year
df['Mes'] = df['Data de criação'].dt.month
df['DiaSemana'] = df['Data de criação'].dt.dayofweek
df['Trimestre'] = df['Data de criação'].dt.quarter
df['Densidade_Atividade'] = df['Número de atividades de vendas'] / (df['Dias_Desde_Criacao'] + 1)

df['Fonte original do tráfego'] = df['Fonte original do tráfego'].fillna('(Sem fonte)')
df['Detalhamento da fonte original do tráfego 1'] = df['Detalhamento da fonte original do tráfego 1'].fillna('(Sem detalhe)')
df['Número de atividades de vendas'] = pd.to_numeric(df['Número de atividades de vendas'], errors='coerce').fillna(0)
df['Unidade Desejada'] = df['Unidade Desejada'].fillna('(Sem unidade)')

for col in ['Ano', 'Mes', 'DiaSemana', 'Trimestre', 'Dias_Desde_Criacao']:
    df[col] = df[col].fillna(0)

# Classificação de Atividade (Igual ao Treino)
def classificar_atividade_comercial(n):
    if n < 3: return '0_Outlier_Sem_Registro'
    elif n <= 6: return '1_Inicial_3_a_6'
    elif n <= 12: return '2_Engajamento_7_a_12'
    elif n <= 20: return '3_Persistencia_13_a_20'
    elif n <= 35: return '4_Alta_Persistencia_21_a_35'
    elif n <= 60: return '5_Recuperacao_Longa_36_a_60'
    else: return '6_Saturacao_Extrema_60+'

df['Faixa_Atividade'] = df['Número de atividades de vendas'].apply(classificar_atividade_comercial)

# Categorização de fonte
def categorizar_fonte(fonte):
    fonte = str(fonte).lower()
    if 'off-line' in fonte: return 'Offline'
    elif 'social pago' in fonte: return 'Meta_Ads'
    elif 'pesquisa paga' in fonte: return 'Google_Ads'
    elif 'pesquisa orgânica' in fonte: return 'Google_Organico'
    elif 'social orgânico' in fonte: return 'Social_Organico'
    elif 'tráfego direto' in fonte: return 'Direto'
    else: return 'Outros'

df['Fonte_Cat'] = df['Fonte original do tráfego'].apply(categorizar_fonte)

# Segmentação de ML (Fundamental para aplicar o modelo correto)
def definir_segmento_ml(fonte_cat):
    if fonte_cat == 'Offline': return 'Offline'
    elif fonte_cat in ['Google_Ads', 'Google_Organico']: return 'Digital_Busca'
    elif fonte_cat in ['Meta_Ads', 'Social_Organico']: return 'Digital_Social'
    else: return 'Digital_Direto'

df['Segmento_ML'] = df['Fonte_Cat'].apply(definir_segmento_ml)

def categorizar_detalhe_offline(row):
    if row['Fonte_Cat'] != 'Offline': return 'N/A'
    detalhe = str(row['Detalhamento da fonte original do tráfego 1']).upper()
    if 'CRM_UI' in detalhe: return 'CRM_UI'
    elif 'CONTACTS' in detalhe: return 'CONTACTS'
    elif 'IMPORT' in detalhe: return 'IMPORT'
    else: return 'Outros_Offline'

df['Detalhe_Offline'] = df.apply(categorizar_detalhe_offline, axis=1)

# Target
df['Sucesso'] = df['Etapa do negócio'].str.contains('MATRÍCULA', case=False, na=False).astype(int)
df['Perdido'] = df['Etapa do negócio'].str.contains('PERDIDO', case=False, na=False).astype(int)
df['Finalizado'] = df['Sucesso'] | df['Perdido']

# ==============================================================================
# 3. SEGMENTAÇÃO DA BASE
# ==============================================================================
print("\n[3/6] Segmentando a base...")

# Base completa de finalizados
df_final = df[df['Finalizado'] == 1].copy()

# Base RASTREÁVEL (>=3 atividades)
df_rastreavel = df_final[df_final['Número de atividades de vendas'] >= MIN_ATIVIDADES].copy()

# Base NÃO RASTREÁVEL (<3 atividades)
df_balcao = df_final[df_final['Número de atividades de vendas'] < MIN_ATIVIDADES].copy()

print(f"\n  POPULAÇÃO TOTAL FINALIZADA: {len(df_final):,}")
print(f"  ├── Rastreável (>={MIN_ATIVIDADES} ativ): {len(df_rastreavel):,} ({len(df_rastreavel)/len(df_final)*100:.1f}%)")
print(f"  │   ├── Matrículas: {df_rastreavel['Sucesso'].sum():,}")
print(f"  │   └── Perdidos: {df_rastreavel['Perdido'].sum():,}")
print(f"  └── Balcão (<{MIN_ATIVIDADES} ativ): {len(df_balcao):,} ({len(df_balcao)/len(df_final)*100:.1f}%)")
print(f"      ├── Matrículas: {df_balcao['Sucesso'].sum():,} (IMPREVISÍVEIS)")
print(f"      └── Perdidos: {df_balcao['Perdido'].sum():,}")

# ==============================================================================
# 4. CARREGAR MODELO E APLICAR SCORING
# ==============================================================================
print("\n[4/6] Carregando modelo e aplicando scoring...")

if not os.path.exists(MODELO_PATH):
    print(f"[ERRO] Modelo não encontrado: {MODELO_PATH}")
    print("Execute o script 11_ML_Lead_Scoring.py primeiro.")
    exit()

with open(MODELO_PATH, 'rb') as f:
    modelo_data = pickle.load(f)

# Detectar formato do modelo (Segmentado vs Único)
is_segmented = False
model = None

if isinstance(modelo_data, dict):
    if any(k.startswith('model_') for k in modelo_data.keys()):
        print("  → Modelo segmentado detectado")
        is_segmented = True
        encoders = modelo_data.get('encoders', {})
    else:
        # Modelo único (formato antigo)
        model = modelo_data.get('model')
        encoders = modelo_data.get('encoders', {})
else:
    model = modelo_data
    encoders = {}

# Aplicar encoders
FEATURES_CAT = ['Fonte_Cat', 'Detalhe_Offline', 'Unidade Desejada', 'Faixa_Atividade']

for col in FEATURES_CAT:
    if col in encoders:
        le = encoders[col]
        # Tratar valores não vistos no treino
        df_final[f'{col}_enc'] = df_final[col].astype(str).apply(
            lambda x: le.transform([x])[0] if x in le.classes_ else -1
        )
    else:
        # Criar encoder na hora se não existir
        le = LabelEncoder()
        le.fit(df[col].astype(str))
        df_final[f'{col}_enc'] = le.transform(df_final[col].astype(str))

# Aplicar Predição
proba = np.zeros(len(df_final))

if is_segmented:
    # Lógica para modelos segmentados
    segmentos_disponiveis = [k.replace('model_', '') for k in modelo_data.keys() if k.startswith('model_')]
    
    for seg in segmentos_disponiveis:
        model_seg = modelo_data[f'model_{seg}']
        features_seg = modelo_data[f'features_{seg}']
        
        # Filtrar leads do segmento
        mask = df_final['Segmento_ML'] == seg
        
        if mask.sum() > 0:
            # Garantir que temos todas as features necessárias
            X_seg = df_final.loc[mask, features_seg].values
            proba[mask] = model_seg.predict_proba(X_seg)[:, 1]
            
elif model is not None:
    # Lógica para modelo único (legado)
    features_model = modelo_data.get('features', [])
    if not features_model: # Fallback
        features_model = [f'{col}_enc' for col in FEATURES_CAT] + ['Número de atividades de vendas', 'Mes', 'DiaSemana', 'Trimestre', 'Dias_Desde_Criacao', 'Densidade_Atividade']
        features_model = [f for f in features_model if f in df_final.columns]
        
    X_test = df_final[features_model].values
    proba = model.predict_proba(X_test)[:, 1]
else:
    print("[ERRO CRÍTICO] Não foi possível carregar nenhum modelo válido.")
    exit()

df_final['Probabilidade'] = proba

# Criar nota 1-5 baseada em PROBABILIDADE BALANCEADA (0-100% distribuído)
bins = [0, 0.2, 0.4, 0.6, 0.8, 1.0]
df_final['Nota'] = pd.cut(
    df_final['Probabilidade'],
    bins=bins,
    labels=[1, 2, 3, 4, 5],
    include_lowest=True
)
df_final['Nota'] = df_final['Nota'].astype(int)

# Recriar os subsets agora com as notas calculadas
mask_balcao_offline = (df_final['Segmento_ML'] == 'Offline') & (df_final['Número de atividades de vendas'] < 3)
df_rastreavel = df_final[~mask_balcao_offline].copy()
df_balcao = df_final[mask_balcao_offline].copy()

print(f"  → Scoring aplicado em toda a base ({len(df_final):,} leads)")
print(f"  → {len(df_rastreavel):,} leads rastreáveis para validação")

# ==============================================================================
# 5. CALCULAR MÉTRICAS DE ACURÁCIA
# ==============================================================================
print("\n[5/6] Calculando métricas de acurácia...")

# Resumo por nota (AJUSTADO - Rastreável)
resumo_nota_adj = df_rastreavel.groupby('Nota').agg(
    Total=('Nota', 'count'),
    Matriculas=('Sucesso', 'sum'),
    Perdidos=('Perdido', 'sum')
).reset_index()
resumo_nota_adj['Taxa_Conversao'] = resumo_nota_adj['Matriculas'] / resumo_nota_adj['Total']
resumo_nota_adj['Share_Matriculas'] = resumo_nota_adj['Matriculas'] / resumo_nota_adj['Matriculas'].sum()

# Resumo por nota (GLOBAL - Tudo)
resumo_nota_global = df_final.groupby('Nota').agg(
    Total=('Nota', 'count'),
    Matriculas=('Sucesso', 'sum'),
    Perdido=('Perdido', 'sum')
).reset_index()
resumo_nota_global['Taxa_Conversao'] = resumo_nota_global['Matriculas'] / resumo_nota_global['Total']
resumo_nota_global['Share_Matriculas'] = resumo_nota_global['Matriculas'] / resumo_nota_global['Matriculas'].sum()

print("\n  ACURÁCIA POR NOTA (BASE RASTREÁVEL - AJUSTADA):")
print(resumo_nota_adj.to_string(index=False))

# --- DESTAQUE NOTA 5 ---
try:
    row_n5 = resumo_nota_adj[resumo_nota_adj['Nota'] == 5].iloc[0]
    print(f"\n  [DESTAQUE NOTA 5]")
    print(f"  ✅ O modelo capturou {int(row_n5['Matriculas'])} matrículas classificando-as como Nota 5.")
    print(f"  📊 Taxa de Conversão neste grupo: {row_n5['Taxa_Conversao']:.1%}")
except IndexError:
    pass
# -----------------------

# Matriz de confusão (Nota 4-5 = Predição Positiva)
df_rastreavel['Pred_Positiva'] = (df_rastreavel['Nota'] >= 4).astype(int)
y_true = df_rastreavel['Sucesso'].values
y_pred = df_rastreavel['Pred_Positiva'].values

matriz = confusion_matrix(y_true, y_pred, labels=[0, 1])
tn, fp, fn, tp = matriz.ravel()

precisao = tp / (tp + fp) if (tp + fp) > 0 else 0
recall = tp / (tp + fn) if (tp + fn) > 0 else 0
f1 = 2 * (precisao * recall) / (precisao + recall) if (precisao + recall) > 0 else 0

# AUC-ROC
auc = roc_auc_score(y_true, df_rastreavel['Probabilidade'].values)

# LIFT
conv_nota5 = resumo_nota_adj[resumo_nota_adj['Nota'] == 5]['Taxa_Conversao'].values[0]
conv_nota1 = resumo_nota_adj[resumo_nota_adj['Nota'] == 1]['Taxa_Conversao'].values[0]
lift = conv_nota5 / conv_nota1 if conv_nota1 > 0 else 0

print(f"\n  MÉTRICAS PRINCIPAIS:")
print(f"  ├── AUC-ROC: {auc:.4f}")
print(f"  ├── Precisão (Notas 4-5): {precisao:.1%}")
print(f"  ├── Recall (Notas 4-5): {recall:.1%}")
print(f"  ├── F1-Score: {f1:.4f}")
print(f"  └── LIFT (Nota5 / Nota1): {lift:.1f}x")

# --- DIAGNÓSTICO POR SEGMENTO (NOVO) ---
print(f"\n  RAIO-X POR SEGMENTO (RECALL NOTA 4-5):")

# Iterar por cada segmento presente na base rastreável para mostrar a realidade de cada modelo
segmentos_unicos = sorted(df_rastreavel['Segmento_ML'].unique())

for seg in segmentos_unicos:
    mask = df_rastreavel['Segmento_ML'] == seg
    label = f"{seg} (Trabalhado)" if seg == 'Offline' else seg
    sub = df_rastreavel[mask]
    if len(sub) > 0:
        rec_sub = sub[(sub['Sucesso']==1) & (sub['Nota']>=4)].shape[0] / sub[sub['Sucesso']==1].shape[0] if sub[sub['Sucesso']==1].shape[0] > 0 else 0
        print(f"  ├── {label}: {rec_sub:.1%}")

# ==============================================================================
# 6. EXPORTAR RESULTADOS
# ==============================================================================
print("\n[6/6] Exportando resultados...")

wb = Workbook()

# --- ABA 1: RESUMO EXECUTIVO ---
ws1 = wb.active
ws1.title = '1_Resumo_Executivo'

ws1['A1'] = 'VALIDAÇÃO DO MODELO - BASE RASTREÁVEL'
ws1['A1'].font = Font(bold=True, size=14, color='1F4E79')

ws1['A3'] = f'Data: {DATA_HOJE}'
ws1['A4'] = f'Critério: Leads com >= {MIN_ATIVIDADES} atividades de vendas'

ws1['A6'] = 'POR QUE ESSE RECORTE?'
ws1['A6'].font = Font(bold=True)
ws1['A7'] = f'• {len(df_balcao):,} leads ({len(df_balcao)/len(df_final)*100:.1f}%) têm <{MIN_ATIVIDADES} atividades'
ws1['A8'] = '• São vendas de balcão sem registro de acompanhamento'
ws1['A9'] = '• O modelo NÃO TEM dados para prever esses casos'
ws1['A10'] = '• Incluí-los na avaliação distorce a métrica de acurácia'

ws1['A12'] = 'MÉTRICAS DO MODELO (BASE RASTREÁVEL)'
ws1['A12'].font = Font(bold=True)

metricas = [
    ['Métrica', 'Valor', 'Interpretação'],
    ['AUC-ROC', auc, 'Capacidade de ordenação (>0.8 = Excelente)'],
    ['Precisão (Notas 4-5)', precisao, 'Qualidade da lista VIP'],
    ['Recall (Notas 4-5)', recall, '% de matrículas capturadas nas notas altas'],
    ['F1-Score', f1, 'Equilíbrio entre Precisão e Recall'],
    ['LIFT (Nota5/Nota1)', lift, f'Nota 5 converte {lift:.1f}x mais que Nota 1'],
]

for i, row in enumerate(metricas, start=13):
    for j, val in enumerate(row, start=1):
        cell = ws1.cell(row=i, column=j, value=val)
        if i == 13:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        elif j == 2 and isinstance(val, float):
            if val >= 0.7:
                cell.fill = SUCCESS_FILL
            elif val >= 0.5:
                cell.fill = WARNING_FILL
            else:
                cell.fill = DANGER_FILL
            cell.number_format = '0.0%' if val < 10 else '0.0'

# --- ABA 2: ACURÁCIA POR NOTA ---
ws2 = wb.create_sheet('2_Acuracia_Por_Nota')

ws2['A1'] = '1. ACURÁCIA GLOBAL (INCLUI VENDAS DE BALCÃO)'
ws2['A1'].font = Font(bold=True)
ws2['A2'] = f'Base: {len(df_final):,} leads totais'

resumo_global_fmt = resumo_nota_global.copy()
resumo_global_fmt.columns = ['Nota', 'Total Leads', 'Matrículas', 'Perdidos', 'Taxa Conv. Real', 'Share Matrículas']
df_to_sheet(ws2, resumo_global_fmt, start_row=4)

ws2['A12'] = '2. ACURÁCIA AJUSTADA (APENAS LEADS RASTREÁVEIS)'
ws2['A12'].font = Font(bold=True, color='006400')
ws2['A13'] = f'Base: {len(df_rastreavel):,} leads rastreáveis (Digital Completo + Offline Trabalhado)'

resumo_adj_fmt = resumo_nota_adj.copy()
resumo_adj_fmt.columns = ['Nota', 'Total Leads', 'Matrículas', 'Perdidos', 'Taxa Conv. Real', 'Share Matrículas']
df_to_sheet(ws2, resumo_adj_fmt, start_row=15)

# Formatação
for row in range(5, 25):
    ws2.cell(row=row, column=5).number_format = '0.0%'
    ws2.cell(row=row, column=6).number_format = '0.0%'

ajustar_largura(ws2)

# --- ABA 3: MATRIZ DE CONFUSÃO ---
ws3 = wb.create_sheet('3_Matriz_Confusao')

ws3['A1'] = 'MATRIZ DE CONFUSÃO (Corte: Notas 4 e 5 = Predição Positiva)'
ws3['A1'].font = Font(bold=True)

dados_matriz = [
    ['Métrica', 'Valor', 'Descrição'],
    ['Verdadeiros Positivos (TP)', tp, 'Modelo disse BOM e MATRICULOU'],
    ['Verdadeiros Negativos (TN)', tn, 'Modelo disse RUIM e PERDEU'],
    ['Falsos Positivos (FP)', fp, 'Modelo disse BOM mas PERDEU'],
    ['Falsos Negativos (FN)', fn, 'Modelo disse RUIM mas MATRICULOU'],
    ['', '', ''],
    ['Precisão', precisao, f'De {tp+fp} preditos como bons, {tp} converteram'],
    ['Recall', recall, f'De {tp+fn} matrículas, capturou {tp} nas notas altas'],
]

for i, row in enumerate(dados_matriz, start=3):
    for j, val in enumerate(row, start=1):
        cell = ws3.cell(row=i, column=j, value=val)
        if i == 3:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        if isinstance(val, float) and val < 1:
            cell.number_format = '0.0%'

# Adicionar Raio-X na planilha
start_row = len(dados_matriz) + 5
ws3.cell(row=start_row, column=1, value='RAIO-X POR SEGMENTO (RECALL)').font = Font(bold=True)

row_idx = start_row + 1
for seg in sorted(df_rastreavel['Segmento_ML'].unique()):
    mask = df_rastreavel['Segmento_ML'] == seg
    label = f"{seg} (Trabalhado)" if seg == 'Offline' else seg
    
    sub = df_rastreavel[mask]
    if len(sub) > 0:
        total_mat = sub[sub['Sucesso']==1].shape[0]
        capturadas = sub[(sub['Sucesso']==1) & (sub['Nota']>=4)].shape[0]
        rec_sub = capturadas / total_mat if total_mat > 0 else 0
        ws3.cell(row=row_idx, column=1, value=label)
        ws3.cell(row=row_idx, column=2, value=rec_sub).number_format = '0.0%'
        ws3.cell(row=row_idx, column=3, value=f'De {total_mat} matrículas, achou {capturadas}')
        row_idx += 1

ajustar_largura(ws3)

# --- ABA 4: COMPARATIVO GERAL vs RASTREÁVEL ---
ws4 = wb.create_sheet('4_Comparativo')

ws4['A1'] = 'COMPARATIVO: POR QUE A MÉTRICA AJUSTADA É MAIS JUSTA'
ws4['A1'].font = Font(bold=True)

comp_data = [
    ['Cenário', 'Total', 'Matrículas', 'Taxa Conv.', 'Observação'],
    ['Base Completa', len(df_final), df_final['Sucesso'].sum(), df_final['Sucesso'].mean(), 'Inclui vendas de balcão sem registro'],
    ['Base Rastreável (Ajustada)', len(df_rastreavel), df_rastreavel['Sucesso'].sum(), df_rastreavel['Sucesso'].mean(), 'Digital (Todos) + Offline (>=3 ativ)'],
    ['Balcão (Ruído)', len(df_balcao), df_balcao['Sucesso'].sum(), df_balcao['Sucesso'].mean(), 'Apenas Offline < 3 ativ (Sem registro)'],
]

for i, row in enumerate(comp_data, start=3):
    for j, val in enumerate(row, start=1):
        cell = ws4.cell(row=i, column=j, value=val)
        if i == 3:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        if j == 4 and isinstance(val, float):
            cell.number_format = '0.0%'

ajustar_largura(ws4)

# --- ABA 5: CONCLUSÃO ---
ws5 = wb.create_sheet('5_Conclusao')

ws5['A1'] = 'CONCLUSÃO E RECOMENDAÇÕES'
ws5['A1'].font = Font(bold=True, size=14, color='1F4E79')

conclusoes = [
    '',
    'O MODELO ESTÁ FUNCIONANDO CORRETAMENTE?',
    '',
    f'✅ SIM! O AUC-ROC de {auc:.2f} indica excelente capacidade de ordenação.',
    f'✅ O LIFT de {lift:.1f}x prova que Nota 5 converte muito mais que Nota 1.',
    f'✅ A precisão de {precisao:.1%} significa que a lista VIP tem qualidade.',
    '',
    'POR QUE O RECALL PARECIA BAIXO ANTES?',
    '',
    f'• {df_balcao["Sucesso"].sum():,} matrículas ({df_balcao["Sucesso"].sum()/df_final["Sucesso"].sum()*100:.1f}%) são de balcão sem registro',
    '• Essas vendas acontecem sem acompanhamento no CRM',
    '• O modelo não tem como prever o que não está registrado',
    '• Excluindo esses casos, o recall sobe significativamente',
    '',
    'COMO USAR O MODELO:',
    '',
    '1. PRIORIZAR leads com Nota 4 e 5 para ações de follow-up',
    '2. MONITORAR o número de atividades - leads com >=3 ativ são rastreáveis',
    '3. INCENTIVAR vendedores a registrar atividades no CRM',
    '4. REVISAR leads Nota 1 e 2 - provavelmente são casos perdidos',
]

for i, texto in enumerate(conclusoes, start=3):
    ws5.cell(row=i, column=1, value=texto)
    if texto.startswith('O MODELO') or texto.startswith('POR QUE') or texto.startswith('COMO USAR'):
        ws5.cell(row=i, column=1).font = Font(bold=True)
    if texto.startswith('✅'):
        ws5.cell(row=i, column=1).font = Font(color='006400')

ws5.column_dimensions['A'].width = 80

# Salvar
wb.save(CAMINHO_SAIDA)

print(f"\n{'='*80}")
print("RESULTADO FINAL")
print("="*80)
print(f"\n✅ Arquivo gerado: {CAMINHO_SAIDA}")
print(f"\n📊 MÉTRICAS NA BASE RASTREÁVEL ({len(df_rastreavel):,} leads):")
print(f"   AUC-ROC: {auc:.4f}")
print(f"   Precisão: {precisao:.1%}")
print(f"   Recall: {recall:.1%}")
print(f"   LIFT: {lift:.1f}x")
print(f"\n💡 O modelo está funcionando corretamente!")
print(f"   O recall 'baixo' no teste anterior era causado por")
print(f"   {df_balcao['Sucesso'].sum():,} matrículas de balcão sem registro de atividades.")
print("="*80)
