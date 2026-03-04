"""
==============================================================================
SCORING DE ARQUIVO EXTERNO (PROVA DE FOGO)
==============================================================================
Script: 15.Score_External_File.py
Objetivo: Pontuar uma base aleatória fornecida por terceiros (Gestores)
          para validar a assertividade do modelo "no escuro".
Input:    Data/base_teste.csv (Deve ter as mesmas colunas do HubSpot)
Output:   Outputs/Resultado_Teste_Cego_[DATA].xlsx
==============================================================================
"""

import pandas as pd
import numpy as np
import os
import pickle
from datetime import datetime
from sklearn.preprocessing import LabelEncoder
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CAMINHO_INPUT_CSV = os.path.join(CAMINHO_BASE, 'Data', 'base_teste.csv')
CAMINHO_INPUT_CSV_DOUBLE = os.path.join(CAMINHO_BASE, 'Data', 'base_teste.csv.csv') # Correção para erro comum
CAMINHO_INPUT_XLSX = os.path.join(CAMINHO_BASE, 'Data', 'base_teste.xlsx')
PASTA_OUTPUT = os.path.join(CAMINHO_BASE, 'Outputs')
PASTA_MODELS = os.path.join(CAMINHO_BASE, 'models')
MODELO_PATH = os.path.join(PASTA_MODELS, 'lead_scoring_model.pkl')

DATA_HOJE = datetime.now().strftime('%Y-%m-%d')

# LÓGICA DE SELEÇÃO DE ARQUIVO (BLINDADA)
arquivo_selecionado = None
if os.path.exists(CAMINHO_INPUT_CSV):
    arquivo_selecionado = CAMINHO_INPUT_CSV
elif os.path.exists(CAMINHO_INPUT_CSV_DOUBLE):
    arquivo_selecionado = CAMINHO_INPUT_CSV_DOUBLE
elif os.path.exists(CAMINHO_INPUT_XLSX):
    arquivo_selecionado = CAMINHO_INPUT_XLSX

print("="*80)
print("SCORING DE ARQUIVO EXTERNO - PROVA DE FOGO")
print("Objetivo: Testar o modelo em uma base 'cega' (base_teste.csv ou .xlsx)")
print("="*80)
print(f"📂 DIRETÓRIO DE EXECUÇÃO: {CAMINHO_BASE}")
print(f"📂 ARQUIVO DE ENTRADA:    {arquivo_selecionado if arquivo_selecionado else '❌ NENHUM ARQUIVO ENCONTRADO'}")
print(f"📂 MODELO UTILIZADO:      {MODELO_PATH}")
print("-" * 80)

if not arquivo_selecionado:
    print(f"\n[ERRO] Arquivo de entrada não encontrado na pasta Data.")
    print("1. Peça aos gestores o arquivo CSV exportado do HubSpot.")
    print("2. Salve-o na pasta 'Data' com o nome 'base_teste.csv' ou 'base_teste.xlsx'.")
    print("3. Rode este script novamente.")
    exit()

# ==============================================================================
# 1. CARREGAMENTO E PREPARAÇÃO (REPLICANDO A LÓGICA DO TREINO)
# ==============================================================================
print("\n[1/4] Carregando e preparando dados do arquivo de teste...")

try:
    if arquivo_selecionado.endswith('.csv') or arquivo_selecionado.endswith('.csv.csv'):
        try:
            df = pd.read_csv(arquivo_selecionado, encoding='utf-8')
        except UnicodeDecodeError:
            df = pd.read_csv(arquivo_selecionado, encoding='latin1')
    else:
        df = pd.read_excel(arquivo_selecionado)
except Exception as e:
    print(f"[ERRO CRÍTICO] Falha ao ler o arquivo: {e}")
    exit()

df.columns = df.columns.str.strip()
print(f"  → Registros carregados: {len(df):,}")

# --- ENGENHARIA DE FEATURES (IDÊNTICA AO SCRIPT 11) ---

# Datas
df['Data de criação'] = pd.to_datetime(df['Data de criação'], errors='coerce')
df['Data de fechamento'] = pd.to_datetime(df['Data de fechamento'], errors='coerce')

# SANITY CHECK: Data de fechamento (Crucial para o teste histórico)
if df['Data de fechamento'].notna().sum() == 0:
    print("\n⚠️ [AVISO] A coluna 'Data de fechamento' parece vazia ou ausente.")
    print("   Para testar leads ANTIGOS, essa coluna é obrigatória, senão eles parecerão 'zumbis' (velhos e parados).")

# Viagem no Tempo (Essencial para não penalizar leads antigos de sucesso)
data_referencia = df['Data de fechamento'].fillna(datetime.now())
df['Dias_Desde_Criacao'] = (data_referencia - df['Data de criação']).dt.days
df['Dias_Desde_Criacao'] = df['Dias_Desde_Criacao'].clip(lower=0)

# Features Temporais
df['Ano'] = df['Data de criação'].dt.year
df['Mes'] = df['Data de criação'].dt.month
df['DiaSemana'] = df['Data de criação'].dt.dayofweek
df['Trimestre'] = df['Data de criação'].dt.quarter

# Tratamento de Nulos e Tipos
df['Fonte original do tráfego'] = df['Fonte original do tráfego'].fillna('(Sem fonte)')
df['Detalhamento da fonte original do tráfego 1'] = df['Detalhamento da fonte original do tráfego 1'].fillna('(Sem detalhe)')
df['Número de atividades de vendas'] = pd.to_numeric(df['Número de atividades de vendas'], errors='coerce').fillna(0)
df['Unidade Desejada'] = df['Unidade Desejada'].fillna('(Sem unidade)')

for col in ['Ano', 'Mes', 'DiaSemana', 'Trimestre', 'Dias_Desde_Criacao']:
    df[col] = df[col].fillna(0)

# Feature: Densidade
df['Densidade_Atividade'] = df['Número de atividades de vendas'] / (df['Dias_Desde_Criacao'] + 1)

# Feature: Faixa de Atividade
def classificar_atividade_comercial(n):
    if n < 3: return '0_Outlier_Sem_Registro'
    elif n <= 6: return '1_Inicial_3_a_6'
    elif n <= 12: return '2_Engajamento_7_a_12'
    elif n <= 20: return '3_Persistencia_13_a_20'
    elif n <= 35: return '4_Alta_Persistencia_21_a_35'
    elif n <= 60: return '5_Recuperacao_Longa_36_a_60'
    else: return '6_Saturacao_Extrema_60+'

# --- BLINDAGEM E GABARITO ---
gabarito_etapa = None

# 1. Se a coluna existe no arquivo, removemos da memória (Blindagem em tempo de execução)
if 'Etapa do negócio' in df.columns:
    print("\n🔒 [BLINDAGEM] Removendo coluna 'Etapa do negócio' da memória...")
    print("   Isso garante que o modelo NÃO tem acesso à resposta final (Matrícula/Perdido).")
    gabarito_etapa = df['Etapa do negócio'].copy()
    df = df.drop(columns=['Etapa do negócio'])

# 2. Se não existe, tentamos carregar do arquivo externo (opcional)
else:
    caminho_gabarito = os.path.join(CAMINHO_BASE, 'Data', 'gabarito_separado.csv')
    if os.path.exists(caminho_gabarito):
        try:
            df_gab = pd.read_csv(caminho_gabarito)
            # Lógica simplificada de gabarito se necessário
            pass 
        except: pass

df['Faixa_Atividade'] = df['Número de atividades de vendas'].apply(classificar_atividade_comercial)

# Feature: Fonte Categorizada
def categorizar_fonte(fonte):
    fonte = str(fonte).lower()
    if 'off-line' in fonte or 'offline' in fonte: return 'Offline'
    elif 'social pago' in fonte: return 'Meta_Ads'
    elif 'pesquisa paga' in fonte: return 'Google_Ads'
    elif 'pesquisa orgânica' in fonte: return 'Google_Organico'
    elif 'social orgânico' in fonte: return 'Social_Organico'
    elif 'tráfego direto' in fonte: return 'Direto'
    else: return 'Outros'

df['Fonte_Cat'] = df['Fonte original do tráfego'].apply(categorizar_fonte)

# Feature: Segmento ML
def definir_segmento_ml(fonte_cat):
    if fonte_cat == 'Offline': return 'Offline'
    elif fonte_cat in ['Google_Ads', 'Google_Organico']: return 'Digital_Busca'
    elif fonte_cat in ['Meta_Ads', 'Social_Organico']: return 'Digital_Social'
    else: return 'Digital_Direto'

df['Segmento_ML'] = df['Fonte_Cat'].apply(definir_segmento_ml)

# Feature: Detalhe Offline
def categorizar_detalhe_offline(row):
    if row['Fonte_Cat'] != 'Offline': return 'N/A'
    detalhe = str(row['Detalhamento da fonte original do tráfego 1']).upper()
    if 'CRM_UI' in detalhe: return 'CRM_UI'
    elif 'CONTACTS' in detalhe: return 'CONTACTS'
    elif 'IMPORT' in detalhe: return 'IMPORT'
    else: return 'Outros_Offline'

df['Detalhe_Offline'] = df.apply(categorizar_detalhe_offline, axis=1)

# ==============================================================================
# 2. APLICAÇÃO DO MODELO
# ==============================================================================
print("\n[2/4] Carregando cérebro (modelo) e pontuando...")

if not os.path.exists(MODELO_PATH):
    print("[ERRO] Modelo não encontrado. Rode o script 11 primeiro.")
    exit()

with open(MODELO_PATH, 'rb') as f:
    modelo_data = pickle.load(f)

encoders = modelo_data.get('encoders', {})

# Aplicar Encoders
FEATURES_CAT_MODEL = ['Fonte_Cat', 'Detalhe_Offline', 'Unidade Desejada', 'Faixa_Atividade']
for col in FEATURES_CAT_MODEL:
    if col in encoders:
        le = encoders[col]
        df[f'{col}_enc'] = df[col].astype(str).apply(
            lambda x: le.transform([x])[0] if x in le.classes_ else -1
        )
    else:
        # Fallback seguro
        le = LabelEncoder()
        le.fit(df[col].astype(str))
        df[f'{col}_enc'] = le.transform(df[col].astype(str))

# Scoring Segmentado
df['Probabilidade_Conversao'] = 0.0

segmentos_disponiveis = [k.replace('model_', '') for k in modelo_data.keys() if k.startswith('model_')]

for seg in segmentos_disponiveis:
    model_seg = modelo_data[f'model_{seg}']
    features_seg = modelo_data[f'features_{seg}']
    
    mask = df['Segmento_ML'] == seg
    if mask.sum() > 0:
        # Verificar se todas as features existem
        missing_cols = [c for c in features_seg if c not in df.columns]
        if not missing_cols:
            X_seg = df.loc[mask, features_seg].values
            df.loc[mask, 'Probabilidade_Conversao'] = model_seg.predict_proba(X_seg)[:, 1]
        else:
            print(f"  ⚠️  Aviso: Faltam colunas para o segmento {seg}: {missing_cols}")

# Calcular Nota (Régua Factual Calibrada - Igual ao Script 11)
# Faixas: 1(0-5%), 2(5-15%), 3(15-30%), 4(30-50%), 5(>50%)
bins = [0, 0.05, 0.15, 0.30, 0.50, 1.0]
df['Nota_Modelo'] = pd.cut(
    df['Probabilidade_Conversao'],
    bins=bins,
    labels=[1, 2, 3, 4, 5],
    include_lowest=True
).astype(int)

# --- EXPLICABILIDADE (POR QUE DEU ESSA NOTA?) ---
def gerar_explicacao(row):
    nota = row['Nota_Modelo']
    atv = row['Número de atividades de vendas']
    fonte = str(row['Fonte_Cat'])
    
    if nota >= 4:
        if atv >= 8:
            return f"Alta Intensidade Comercial ({int(atv)} atividades)"
        elif fonte in ['Google_Ads', 'Meta_Ads', 'Google_Organico', 'Direto']:
            return f"Fonte Premium ({fonte})"
        else:
            return "Padrão Comportamental de Matrícula"
    elif nota <= 2:
        if atv < 3:
            return "Baixo Engajamento (Poucas atividades)"
        else:
            return "Fonte/Perfil com Baixa Conversão Histórica"
    else:
        return "Perfil Intermediário (Zona Cinzenta)"

df['Motivo_da_Nota'] = df.apply(gerar_explicacao, axis=1)

# --- RESTAURAR GABARITO (APENAS PARA O RELATÓRIO FINAL) ---
if gabarito_etapa is not None:
    print("\n🔓 [RELATÓRIO] Restaurando a coluna 'Etapa do negócio' APENAS para visualização no Excel.")
    df['Etapa do negócio'] = gabarito_etapa

# ==============================================================================
# 3. EXPORTAÇÃO DO RESULTADO
# ==============================================================================
print("\n[3/4] Gerando relatório para os gestores...")

wb = Workbook()
ws = wb.active
ws.title = 'Resultado_Teste_Cego'

# Estilização
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)

# Selecionar colunas para o relatório
cols_export = [
    'Record ID', 'Nome do negócio', 'Etapa do negócio', 
    'Fonte original do tráfego', 'Número de atividades de vendas', 
    'Data de criação', 'Data de fechamento',
    'Segmento_ML', 'Probabilidade_Conversao', 'Nota_Modelo', 'Motivo_da_Nota'
]

# Filtrar apenas as que existem
cols_final = [c for c in cols_export if c in df.columns]
df_export = df[cols_final].sort_values('Probabilidade_Conversao', ascending=False)

# Escrever no Excel
for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), start=1):
    for c_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        
        # Formatar probabilidade
        if 'Probabilidade' in cols_final and c_idx == cols_final.index('Probabilidade_Conversao') + 1 and r_idx > 1:
            cell.number_format = '0.0%'

# Ajustar largura
for col_idx in range(1, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 20

nome_arquivo = f'Resultado_Teste_Gestores_{DATA_HOJE}.xlsx'
caminho_final = os.path.join(PASTA_OUTPUT, nome_arquivo)
wb.save(caminho_final)

print("\n" + "="*80)
print("RESULTADO PRONTO!")
print("="*80)
print(f"Arquivo gerado: {caminho_final}")
print("\nINSTRUÇÕES PARA OS GESTORES:")
print("1. Abra o arquivo Excel gerado.")
print("2. Filtre a coluna 'Nota_Modelo'.")
print("3. Verifique:")
print("   - Os leads Nota 5 e 4 são os que MATRICULARAM?")
print("   - Os leads Nota 1 e 2 são os que PERDERAM?")
print("\nOBS: Leads de balcão (Offline < 3 atividades) podem ter nota baixa corretamente")
print("     pois o modelo avalia o esforço de venda registrado.")
print("="*80)