"""
==============================================================================
VALIDAÇÃO INVERSA: MATRÍCULAS ATUAIS -> PREVISÃO PASSADA
==============================================================================
Script: 17.Validacao_Qualificacao_Pausa.py
Objetivo: Identificar leads que HOJE são Matrícula e verificar se o Hélio
          havia previsto eles corretamente nas listas antigas (Nota 4 ou 5).
==============================================================================
"""

import pandas as pd
import os
import sys
import glob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# ==============================================================================
# CONFIGURAÇÃO DE ARQUIVOS
# ==============================================================================
CAMINHO_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASTA_DATA = os.path.join(CAMINHO_BASE, 'Data')
PASTA_OUTPUT = os.path.join(CAMINHO_BASE, 'Outputs', 'Relatorios_Validacao')
os.makedirs(PASTA_OUTPUT, exist_ok=True)

# Nomes dos arquivos definidos na solicitação
NOME_ARQUIVO_ATUAL  = 'hubspot_leads_atual.csv'        # Base nova (Realidade)
DATA_PREVISAO_STR   = '2025-12-16'                     # Data das listas antigas para validação

CAMINHO_ATUAL = os.path.join(PASTA_DATA, NOME_ARQUIVO_ATUAL)

print("="*80)
print("VALIDAÇÃO INVERSA: QUEM MATRICULOU HOJE ESTAVA PREVISTO?")
print("="*80)

# ==============================================================================
# 1. CARREGAMENTO DAS BASES
# ==============================================================================
print(f"\n[1/4] Carregando bases de dados...")

# Função auxiliar de leitura
def ler_csv_seguro(caminho):
    if not os.path.exists(caminho):
        return None
    try:
        return pd.read_csv(caminho, sep=',', encoding='utf-8')
    except:
        try:
            return pd.read_csv(caminho, sep=';', encoding='utf-8-sig') # Tenta formato scored
        except:
            return pd.read_csv(caminho, encoding='latin1')

# Lista para consolidar histórico (CSV + Excels Antigos)
dfs_historico = []

# 2. Carregar Relatórios Excel Antigos (Busca os 4 arquivos específicos)
PASTA_RELATORIOS_ANTIGOS = os.path.join(PASTA_DATA, 'relatorios_antigos')
excel_files = []

# Tenta buscar na pasta 'relatorios_antigos', se não achar, busca na raiz 'Data'
if os.path.exists(PASTA_RELATORIOS_ANTIGOS):
    # Tenta buscar arquivos com a data específica primeiro
    excel_files = glob.glob(os.path.join(PASTA_RELATORIOS_ANTIGOS, f'Lead_Scoring_*{DATA_PREVISAO_STR}*.xlsx'))
    if not excel_files:
        excel_files = glob.glob(os.path.join(PASTA_RELATORIOS_ANTIGOS, 'Lead_Scoring_*.xlsx'))

if not excel_files:
    # Fallback: Busca na pasta Data se não achou na subpasta
    excel_files = glob.glob(os.path.join(PASTA_DATA, f'Lead_Scoring_*{DATA_PREVISAO_STR}*.xlsx'))
    if not excel_files:
        excel_files = glob.glob(os.path.join(PASTA_DATA, 'Lead_Scoring_*.xlsx'))

# Filtrar arquivos temporários do Excel (~$) para evitar erros
excel_files = [f for f in excel_files if not os.path.basename(f).startswith('~$')]

if excel_files:
    print(f"   -> Encontrados {len(excel_files)} relatórios de histórico (Listas de {DATA_PREVISAO_STR}):")
    for f in excel_files:
        print(f"      📄 Lendo: {os.path.basename(f)}")
        try:
            xls = pd.ExcelFile(f)
            for sheet in xls.sheet_names:
                # Ignora abas de resumo, foca nas listas de leads
                if any(x in sheet for x in ['Resumo', 'Fatores', 'Perfil', 'Projecao', 'Financeira', 'Caveats', 'Matriz']): continue
                
                # Tenta ler com header na linha 6 (padrão dos relatórios do Hélio)
                df_temp = pd.read_excel(xls, sheet_name=sheet, header=5) 
                df_temp.columns = df_temp.columns.astype(str).str.strip() # Limpeza imediata
                
                # Se não achar Record ID, tenta header na linha 0
                if 'Record ID' not in df_temp.columns: 
                    df_temp = pd.read_excel(xls, sheet_name=sheet, header=0)
                    df_temp.columns = df_temp.columns.astype(str).str.strip()
                
                # Fallback: Mapeamento posicional (A=ID, O=Nota) se nomes não baterem
                if 'Record ID' not in df_temp.columns and len(df_temp.columns) >= 15:
                    print(f"         ⚠️  Colunas por nome não encontradas. Tentando posição: Col A=ID, Col O=Nota.")
                    cols = df_temp.columns.tolist()
                    # Renomeia usando o nome atual da coluna nesses índices (0=A, 14=O)
                    df_temp.rename(columns={
                        cols[0]: 'Record ID',
                        cols[14]: 'Nota_1a5'
                    }, inplace=True)
                    if len(cols) > 2:
                        df_temp.rename(columns={cols[2]: 'Etapa do negócio'}, inplace=True)

                if 'Record ID' in df_temp.columns:
                    # Garante que temos a nota
                    if 'Nota_1a5' in df_temp.columns:
                        # Verifica se tem Etapa, se não tiver, cria vazia para não quebrar
                        if 'Etapa do negócio' not in df_temp.columns:
                            df_temp['Etapa do negócio'] = 'Desconhecida'
                        dfs_historico.append(df_temp[['Record ID', 'Nota_1a5', 'Etapa do negócio']])
                    else:
                        print(f"         ⚠️  Aba '{sheet}': Coluna 'Nota_1a5' não encontrada.")
                else:
                    print(f"         ⚠️  Aba '{sheet}': Coluna 'Record ID' não encontrada.")
        except Exception as e:
            print(f"      ⚠️ Erro ao ler {os.path.basename(f)}: {e}")

if not dfs_historico:
    print(f"❌ [ERRO] Arquivos encontrados, mas NENHUM dado válido (Record ID + Nota) foi extraído.")
    print("   Verifique se as colunas 'Record ID' e 'Nota_1a5' existem nas abas dos relatórios.")
    sys.exit()

df_treino = pd.concat(dfs_historico, ignore_index=True)

# Carregar Base Atual (Realidade)
df_atual = ler_csv_seguro(CAMINHO_ATUAL)
if df_atual is None:
    # Tenta fallback para o nome padrão se o específico não existir
    caminho_fallback = os.path.join(PASTA_DATA, 'hubspot_leads.csv')
    print(f"⚠️  Arquivo '{NOME_ARQUIVO_ATUAL}' não encontrado. Tentando 'hubspot_leads.csv'...")
    df_atual = ler_csv_seguro(caminho_fallback)
    if df_atual is None:
        print("❌ [ERRO] Nenhuma base atual encontrada.")
        sys.exit()

# Padronizar colunas
df_atual.columns = df_atual.columns.str.strip()

# Garantir leitura da Data de Criação para validação visual
if 'Data de criação' in df_atual.columns:
    df_atual['Data de criação'] = pd.to_datetime(df_atual['Data de criação'], dayfirst=True, errors='coerce')
else:
    print("⚠️  Coluna 'Data de criação' não encontrada na base atual.")
    df_atual['Data de criação'] = pd.NaT

# Padronizar IDs para cruzamento
df_treino['Record ID'] = df_treino['Record ID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)

# Garantir que a nota seja numérica para ordenação (corrige erro de tipos mistos str/int)
df_treino['Nota_1a5'] = pd.to_numeric(df_treino['Nota_1a5'], errors='coerce').fillna(0)

# Manter a maior nota se houver duplicatas (cenário otimista: se o Hélio viu potencial em algum momento)
df_treino = df_treino.sort_values('Nota_1a5', ascending=False).drop_duplicates(subset='Record ID', keep='first')
df_atual['Record ID'] = df_atual['Record ID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)

print(f"   -> Base Treino (Previsão): {len(df_treino):,} leads")
print(f"   -> Base Atual (Realidade): {len(df_atual):,} leads")

# ==============================================================================
# 2. FILTRAGEM DO CENÁRIO (MATRÍCULAS NA BASE ATUAL)
# ==============================================================================
print(f"\n[2/4] Identificando matrículas realizadas na base atual...")

df_atual['Etapa_Normalizada'] = df_atual['Etapa do negócio'].astype(str).str.upper()
filtro_matricula = df_atual['Etapa_Normalizada'].str.contains('MATRÍCULA|MATRICULA|GANHO|WON', na=False)

df_matriculas_hoje = df_atual[filtro_matricula].copy()

print(f"   -> Total de Matrículas na base atual: {len(df_matriculas_hoje):,}")
if len(df_matriculas_hoje) == 0:
    print("⚠️  Nenhuma matrícula encontrada na base atual.")
    sys.exit()

# ==============================================================================
# 3. CRUZAMENTO COM O PASSADO (QUEM O HÉLIO PREVIU?)
# ==============================================================================
print(f"\n[3/4] Cruzando com previsões antigas...")

cols_hist = ['Record ID', 'Nota_1a5', 'Etapa do negócio'] # Etapa do negócio aqui é a ANTIGA (do dia 05/01)
df_hist_cut = df_treino[[c for c in cols_hist if c in df_treino.columns]].rename(columns={'Etapa do negócio': 'Etapa_Passada', 'Nota_1a5': 'Nota_Prevista'})

# Join (Inner Join: Queremos apenas as matrículas que EXISTIAM nas listas antigas)
df_resultado = pd.merge(df_matriculas_hoje, df_hist_cut, on='Record ID', how='inner')

# ==============================================================================
# 3.1 ANÁLISE TEMPORAL (POR QUE TEM DATA ANTIGA?)
# ==============================================================================
# Converter data de fechamento para datetime
df_resultado['Data de fechamento'] = pd.to_datetime(df_resultado['Data de fechamento'], dayfirst=True, errors='coerce')

# Verificar range de criação dos leads encontrados
if 'Data de criação' in df_resultado.columns:
    max_criacao = df_resultado['Data de criação'].max()
    print(f"   ℹ️  Data de Criação mais recente encontrada no cruzamento: {max_criacao}")

DATA_LISTA = pd.Timestamp(DATA_PREVISAO_STR)

def analisar_retroativo(dt):
    if pd.isna(dt): return "Data Desconhecida"
    if dt < DATA_LISTA: return "⚠️ Fechamento Retroativo (Backdated)"
    return "✅ Conversão Real (Pós-Lista)"

df_resultado['Status_Temporal'] = df_resultado['Data de fechamento'].apply(analisar_retroativo)

# Separar Conversão Real vs Retroativa
df_reais = df_resultado[df_resultado['Status_Temporal'].str.contains('Real', na=False)]
df_retro = df_resultado[df_resultado['Status_Temporal'].str.contains('Retroativo', na=False)]

# Análise de Acerto (Geral - Apenas para constar)
total_encontrados = len(df_resultado)

# Análise de Acerto (Apenas Reais - Métrica Limpa)
total_reais = len(df_reais)
quentes_reais = len(df_reais[df_reais['Nota_Prevista'].isin([4, 5])])
taxa_acerto_reais = quentes_reais / total_reais if total_reais > 0 else 0

print("\n" + "="*60)
print("RESULTADO DA VALIDAÇÃO")
print("="*60)
print(f" Matrículas HOJE (Base Atual):          {len(df_matriculas_hoje):,}")
print(f"🔍 Encontrados nas Listas {DATA_PREVISAO_STR}: {total_encontrados:,}")
print("-" * 60)
print(f"⚠️  Fechamento Retroativo (Backdated):     {len(df_retro):,} (Ignorados na métrica real)")
print(f"✅  Conversão Real (Pós-Lista):            {total_reais:,}")
print("-" * 60)
print(f"🎯 Desses REAIS, previstos como QUENTES:  {quentes_reais:,}")
print(f"📉 Desses REAIS, previstos como FRIOS:    {total_reais - quentes_reais:,}")
print(f"🏆 Taxa de Acerto (Recall) REAL:          {taxa_acerto_reais:.1%}")
print("   (Considerando apenas matrículas que ocorreram DEPOIS da geração da lista)")
print("="*60)

# Distribuição detalhada
print("\nDistribuição das Notas Previstas para quem Matriculou:")
print(df_resultado['Nota_Prevista'].value_counts().sort_index(ascending=False))

# ==============================================================================
# 4. GERAR RELATÓRIO EXCEL
# ==============================================================================
print(f"\n[4/4] Gerando relatório de evidência...")

wb = Workbook()
ws = wb.active
ws.title = "Matriculas_vs_Previsao"

# Estilos
HEADER_FILL = PatternFill(start_color='CE0E2D', end_color='CE0E2D', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
SUCCESS_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
WARNING_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

# Selecionar colunas para exportação
cols_export = [
    'Record ID', 'Nome do negócio', 'Unidade Desejada', 'Data de criação', 'Data de fechamento',
    'Status_Temporal', 'Etapa do negócio', 'Nota_Prevista', 'Etapa_Passada'
]

# Filtrar colunas existentes
cols_final = [c for c in cols_export if c in df_resultado.columns]
df_export = df_resultado[cols_final].sort_values('Nota_Prevista', ascending=False)

# Escrever no Excel
ws.append(cols_final) # Header

for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=False), start=2):
    for c_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.border = THIN_BORDER
        
        # Pintar nota
        col_nota_idx = cols_final.index('Nota_Prevista') + 1 if 'Nota_Prevista' in cols_final else -1
        if c_idx == col_nota_idx:
            nota = value
            if isinstance(nota, (int, float)) and nota >= 4:
                cell.fill = SUCCESS_FILL
            elif isinstance(nota, (int, float)) and nota <= 2:
                cell.fill = WARNING_FILL
        
        # Formatar Data
        if 'Data de fechamento' in cols_final and c_idx == cols_final.index('Data de fechamento') + 1:
            cell.number_format = 'DD/MM/YYYY'
        
        if 'Data de criação' in cols_final and c_idx == cols_final.index('Data de criação') + 1:
            cell.number_format = 'DD/MM/YYYY'

# Formatar Header
for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center')

# Ajustar largura
for col_idx in range(1, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 20

nome_saida = f"Validacao_Inversa_Matriculas_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
caminho_saida = os.path.join(PASTA_OUTPUT, nome_saida)

wb.save(caminho_saida)
print(f"\n✅ Relatório salvo em: {caminho_saida}")
print("   (Abra o Excel para ver a lista nominal dos leads que converteram)")