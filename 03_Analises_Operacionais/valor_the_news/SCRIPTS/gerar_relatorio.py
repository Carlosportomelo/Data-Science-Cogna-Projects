import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Criar workbook
wb = Workbook()

# Estilos
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
title_font = Font(bold=True, size=14, color="1F4E79")
subtitle_font = Font(bold=True, size=11, color="2E75B6")
insight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_data_cells(ws, start_row, end_row, cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

def add_title(ws, row, title, emoji=""):
    cell = ws.cell(row=row, column=1, value=f"{emoji} {title}" if emoji else title)
    cell.font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

def add_insight(ws, row, text, cols=5):
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = insight_fill
    cell.alignment = Alignment(wrap_text=True, vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 40

# ========== ABA 1: RESUMO EXECUTIVO ==========
ws1 = wb.active
ws1.title = "Resumo Executivo"

add_title(ws1, 1, "RELATORIO EXECUTIVO: IMPACTO RED BALLOON", "")
ws1.cell(row=2, column=1, value="Acoes nos veiculos Valor Economico e The News (19 e 20 de Janeiro de 2026)")
ws1.cell(row=2, column=1).font = Font(italic=True, size=10, color="666666")
ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

ws1.cell(row=4, column=1, value="Data de Geracao:").font = Font(bold=True)
ws1.cell(row=4, column=2, value=datetime.now().strftime("%d/%m/%Y %H:%M"))
ws1.cell(row=5, column=1, value="Status:").font = Font(bold=True)
ws1.cell(row=5, column=2, value="Dados validados e prontos para exportacao")

# Principais KPIs
ws1.cell(row=7, column=1, value="PRINCIPAIS KPIs DA ACAO").font = subtitle_font
headers = ["KPI", "Valor", "Variacao"]
for col, h in enumerate(headers, 1):
    ws1.cell(row=8, column=col, value=h)
style_header_row(ws1, 8, 3)

kpis = [
    ("Sessoes nos 2 dias de acao", "5.874", "-"),
    ("Visualizacoes de pagina", "15.374", "-"),
    ("Crescimento Sessoes (Seg+Ter) vs 2025", "+64,59%", "vs 20-21/01/2025"),
    ("Crescimento Views (Seg+Ter) vs 2025", "+43,11%", "vs 20-21/01/2025"),
    ("Crescimento vs Semana Anterior", "+26,04%", "vs 12-18/01"),
    ("Media Views/Sessao", "2,62", "-"),
]
for i, (kpi, valor, var) in enumerate(kpis, 9):
    ws1.cell(row=i, column=1, value=kpi)
    ws1.cell(row=i, column=2, value=valor)
    ws1.cell(row=i, column=3, value=var)
style_data_cells(ws1, 9, 14, 3)

# KPIs de Leads
ws1.cell(row=16, column=1, value="KPIs DE LEADS (CRM HubSpot)").font = subtitle_font
for col, h in enumerate(headers, 1):
    ws1.cell(row=17, column=col, value=h)
style_header_row(ws1, 17, 3)

kpis_leads = [
    ("Leads na Semana da Acao", "1.335", "-"),
    ("Crescimento Leads vs 2025", "+34,17%", "vs 19-25/01/2025"),
    ("Leads Extras Gerados", "+340", "-"),
    ("Crescimento Social Pago", "+105,13%", "312 -> 640"),
]
for i, (kpi, valor, var) in enumerate(kpis_leads, 18):
    ws1.cell(row=i, column=1, value=kpi)
    ws1.cell(row=i, column=2, value=valor)
    ws1.cell(row=i, column=3, value=var)
style_data_cells(ws1, 18, 21, 3)

# Destacar crescimento de leads
ws1.cell(row=19, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ws1.cell(row=19, column=2).font = Font(bold=True, color="006100")
ws1.cell(row=21, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ws1.cell(row=21, column=2).font = Font(bold=True, color="006100")

ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 18

# ========== ABA 2: IMPACTO DIRETO ==========
ws2 = wb.create_sheet("1. Impacto Direto")

add_title(ws2, 1, "IMPACTO DIRETO (O Pico da Acao)")
ws2.cell(row=2, column=1, value="Dados granulares dos dias de publicacao, focando no volume bruto e profundidade de navegacao.")
ws2.cell(row=2, column=1).font = Font(italic=True, size=10)
ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

headers = ["Metrica", "19/01 (Seg - Acao)", "20/01 (Ter - Acao)", "Total Acumulado (2 dias)"]
for col, h in enumerate(headers, 1):
    ws2.cell(row=4, column=col, value=h)
style_header_row(ws2, 4, 4)

data = [
    ("Sessoes (Acessos)", "3.415", "2.459", "5.874"),
    ("Visualizacoes de Pagina", "9.008", "6.366", "15.374"),
    ("Views por Sessao", "2,64", "2,59", "2,62 (Media)"),
]
for i, row_data in enumerate(data, 5):
    for j, val in enumerate(row_data, 1):
        ws2.cell(row=i, column=j, value=val)
style_data_cells(ws2, 5, 7, 4)

add_insight(ws2, 9, "ANALISE DE QUALIDADE: A terca-feira (20/01) manteve taxa de visualizacao por sessao estavel em relacao a segunda, indicando que o trafego residual da acao manteve o mesmo perfil de interesse.", 4)

# Comparativo com 2025
ws2.cell(row=11, column=1, value="COMPARATIVO COM ANO ANTERIOR (Seg+Ter)").font = subtitle_font
headers_comp = ["Metrica", "2025 (20-21/01)", "2026 (19-20/01)", "Crescimento"]
for col, h in enumerate(headers_comp, 1):
    ws2.cell(row=12, column=col, value=h)
style_header_row(ws2, 12, 4)

data_comp = [
    ("Sessoes (2 dias)", "3.569", "5.874", "+64,59%"),
    ("Visualizacoes (2 dias)", "10.744", "15.374", "+43,11%"),
]
for i, row_data in enumerate(data_comp, 13):
    for j, val in enumerate(row_data, 1):
        ws2.cell(row=i, column=j, value=val)
style_data_cells(ws2, 13, 14, 4)

# Destacar crescimento
ws2.cell(row=13, column=4).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ws2.cell(row=13, column=4).font = Font(bold=True, color="006100")
ws2.cell(row=14, column=4).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ws2.cell(row=14, column=4).font = Font(bold=True, color="006100")

add_insight(ws2, 16, "INSIGHT: Comparando os dois dias da acao (Seg+Ter), houve um crescimento de +64,59% em sessoes e +43,11% em visualizacoes em relacao aos dias equivalentes de 2025.", 4)

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 22
ws2.column_dimensions['D'].width = 25

# ========== ABA 3: COMPARACOES DE CRESCIMENTO ==========
ws3 = wb.create_sheet("2. Benchmarks")

add_title(ws3, 1, "COMPARACOES DE CRESCIMENTO (Benchmarks)")
ws3.cell(row=2, column=1, value="Validacao estatistica do impacto do dia 19/01 contra todas as referencias temporais e historicas.")
ws3.cell(row=2, column=1).font = Font(italic=True, size=10)
ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

headers = ["Referencia", "Data Comparada", "Sessoes Base", "Sessoes Acao", "Crescimento (%)"]
for col, h in enumerate(headers, 1):
    ws3.cell(row=4, column=col, value=h)
style_header_row(ws3, 4, 5)

data = [
    ("Segunda Anterior", "12/01/2026", "2.358", "3.415", "+44,83%"),
    ("Segunda Posterior", "26/01/2026", "2.961", "3.415", "+15,33%"),
    ("Segunda Equiv. 2025", "20/01/2025", "1.803", "3.415", "+89,41%"),
]
for i, row_data in enumerate(data, 5):
    for j, val in enumerate(row_data, 1):
        ws3.cell(row=i, column=j, value=val)
style_data_cells(ws3, 5, 7, 5)

# Highlight do maior crescimento
ws3.cell(row=7, column=5).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ws3.cell(row=7, column=5).font = Font(bold=True, color="006100")

add_insight(ws3, 9, "INSIGHT DE BENCHMARK: O salto de 89,41% em relacao ao ano anterior e o indicador mais forte de sucesso da campanha, superando largamente a tendencia de crescimento organico da unidade.", 5)

for col in ['A', 'B', 'C', 'D', 'E']:
    ws3.column_dimensions[col].width = 20

# ========== ABA 4: PERFORMANCE SEMANAL ==========
ws4 = wb.create_sheet("3. Performance Semanal")

add_title(ws4, 1, "PERFORMANCE SEMANAL AGREGADA")
ws4.cell(row=2, column=1, value="Impacto na semana completa (19/01 a 25/01) para medir a sustentacao do trafego.")
ws4.cell(row=2, column=1).font = Font(italic=True, size=10)
ws4.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

headers = ["Periodo", "Sessoes Totais", "Visualizacoes Totais", "Var. % Sessoes"]
for col, h in enumerate(headers, 1):
    ws4.cell(row=4, column=col, value=h)
style_header_row(ws4, 4, 4)

data = [
    ("Semana da Acao (19-25/01/2026)", "16.421", "41.136", "-"),
    ("Semana Anterior (12-18/01/2026)", "13.028", "33.180", "+26,04%"),
    ("Semana Equiv. 2025 (20-26/01)", "11.729", "34.874", "+40,00%"),
]
for i, row_data in enumerate(data, 5):
    for j, val in enumerate(row_data, 1):
        ws4.cell(row=i, column=j, value=val)
style_data_cells(ws4, 5, 7, 4)

add_insight(ws4, 9, "NOTA DE SUSTENTACAO: A semana da acao nao so gerou um pico, como elevou o patamar medio diario de 1.861 para 2.345 sessoes.", 4)

# Tabela de media diaria
ws4.cell(row=11, column=1, value="MEDIA DIARIA").font = subtitle_font
headers2 = ["Periodo", "Media Sessoes/Dia", "Elevacao"]
for col, h in enumerate(headers2, 1):
    ws4.cell(row=12, column=col, value=h)
style_header_row(ws4, 12, 3)

data2 = [
    ("Semana Anterior", "1.861", "-"),
    ("Semana da Acao", "2.345", "+26,0%"),
]
for i, row_data in enumerate(data2, 13):
    for j, val in enumerate(row_data, 1):
        ws4.cell(row=i, column=j, value=val)
style_data_cells(ws4, 13, 14, 3)

ws4.column_dimensions['A'].width = 32
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 20
ws4.column_dimensions['D'].width = 18

# ========== ABA 5: COMPARATIVO DE LEADS ==========
ws_leads = wb.create_sheet("4. Comparativo Leads")

add_title(ws_leads, 1, "COMPARATIVO DE LEADS: SEMANA DA ACAO vs 2025")
ws_leads.cell(row=2, column=1, value="Analise do impacto na geracao de leads comparando a semana da acao (19-25/01/2026) com a semana equivalente de 2025.")
ws_leads.cell(row=2, column=1).font = Font(italic=True, size=10)
ws_leads.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

# Tabela principal de comparativo
ws_leads.cell(row=4, column=1, value="VOLUME DE LEADS POR FONTE").font = subtitle_font
headers = ["Metrica", "2025 (19-25/01)", "2026 (19-25/01)", "Variacao", "Leads Extras"]
for col, h in enumerate(headers, 1):
    ws_leads.cell(row=5, column=col, value=h)
style_header_row(ws_leads, 5, 5)

data_leads = [
    ("Total de Leads", "995", "1.335", "+34,17%", "+340"),
    ("Leads/Dia (media)", "142", "191", "+34,5%", "+49"),
    ("Social Pago", "312", "640", "+105,13%", "+328"),
    ("Pesquisa Paga", "78", "59", "-24,36%", "-19"),
    ("Pesquisa Organica", "37", "18", "-51,35%", "-19"),
    ("Trafego Direto", "73", "80", "+9,59%", "+7"),
    ("Fontes Off-line", "465", "522", "+12,26%", "+57"),
]
for i, row_data in enumerate(data_leads, 6):
    for j, val in enumerate(row_data, 1):
        ws_leads.cell(row=i, column=j, value=val)
style_data_cells(ws_leads, 6, 12, 5)

# Destacar crescimento de Social Pago (linha 8, coluna 4)
ws_leads.cell(row=8, column=4).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ws_leads.cell(row=8, column=4).font = Font(bold=True, color="006100")
ws_leads.cell(row=8, column=5).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ws_leads.cell(row=8, column=5).font = Font(bold=True, color="006100")

# Destacar queda de Pesquisa Organica em vermelho (linha 10)
ws_leads.cell(row=10, column=4).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ws_leads.cell(row=10, column=4).font = Font(bold=True, color="9C0006")
ws_leads.cell(row=10, column=5).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ws_leads.cell(row=10, column=5).font = Font(bold=True, color="9C0006")

add_insight(ws_leads, 14, "INSIGHT PRINCIPAL: O Social Pago mais que dobrou (+105%), indicando que a acao nos veiculos Valor e The News potencializou significativamente as campanhas de midia paga. A queda em Pesquisa Organica (-51%) pode indicar migracao de canal.", 5)

# Leads por dia
ws_leads.cell(row=16, column=1, value="LEADS POR DIA DA SEMANA").font = subtitle_font
headers2 = ["Dia", "2025", "2026", "Variacao"]
for col, h in enumerate(headers2, 1):
    ws_leads.cell(row=17, column=col, value=h)
style_header_row(ws_leads, 17, 4)

# Dados alinhados por dia da semana (Dom a Sab)
leads_dia = [
    ("Domingo (19/01)", "0", "265", "+265"),
    ("Segunda (20/01)", "193", "235", "+21,8%"),
    ("Terca (21/01)", "152", "207", "+36,2%"),
    ("Quarta (22/01)", "227", "250", "+10,1%"),
    ("Quinta (23/01)", "215", "181", "-15,8%"),
    ("Sexta (24/01)", "157", "93", "-40,8%"),
    ("Sabado (25/01)", "51", "104", "+103,9%"),
]
for i, row_data in enumerate(leads_dia, 18):
    for j, val in enumerate(row_data, 1):
        ws_leads.cell(row=i, column=j, value=val)
style_data_cells(ws_leads, 18, 24, 4)

# Destacar domingo
ws_leads.cell(row=18, column=3).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ws_leads.cell(row=18, column=3).font = Font(bold=True, color="006100")

add_insight(ws_leads, 26, "NOTA: O domingo 19/01/2026 (dia da publicacao no Valor) gerou 265 leads - um volume atipico para fim de semana, evidenciando o impacto imediato da acao.", 4)

# Etapa do funil 2026
ws_leads.cell(row=28, column=1, value="DISTRIBUICAO POR ETAPA DO FUNIL (2026)").font = subtitle_font
headers3 = ["Etapa", "Quantidade", "% do Total"]
for col, h in enumerate(headers3, 1):
    ws_leads.cell(row=29, column=col, value=h)
style_header_row(ws_leads, 29, 3)

etapas = [
    ("Em Qualificacao", "756", "56,6%"),
    ("Matricula Concluida", "235", "17,6%"),
    ("Novo Negocio", "143", "10,7%"),
    ("Negocio Perdido", "63", "4,7%"),
    ("Visita Agendada", "54", "4,0%"),
    ("Em Pausa", "40", "3,0%"),
    ("Visita Realizada", "37", "2,8%"),
    ("Lista de Espera", "7", "0,5%"),
]
for i, row_data in enumerate(etapas, 30):
    for j, val in enumerate(row_data, 1):
        ws_leads.cell(row=i, column=j, value=val)
style_data_cells(ws_leads, 30, 37, 3)

ws_leads.column_dimensions['A'].width = 28
ws_leads.column_dimensions['B'].width = 18
ws_leads.column_dimensions['C'].width = 18
ws_leads.column_dimensions['D'].width = 15
ws_leads.column_dimensions['E'].width = 15

# ========== ABA 6: DATA SCIENCE ==========
ws5 = wb.create_sheet("5. Data Science")

add_title(ws5, 1, "ESTRATEGIA PARA MODELAGEM (Data Science)")
ws5.cell(row=2, column=1, value="Diretrizes para aplicacao de modelos de IA com base nos dados consolidados.")
ws5.cell(row=2, column=1).font = Font(italic=True, size=10)
ws5.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

# Modelo A
ws5.cell(row=4, column=1, value="A. MODELO DE REGRESSAO (Causalidade)").font = subtitle_font
ws5.merge_cells(start_row=4, start_column=1, end_row=4, end_column=4)

headers = ["Item", "Descricao"]
for col, h in enumerate(headers, 1):
    ws5.cell(row=5, column=col, value=h)
style_header_row(ws5, 5, 2)

data = [
    ("Objetivo", "Isolar o efeito 'Spike' da acao da sazonalidade natural de janeiro"),
    ("Variavel Dependente", "Sessoes diarias"),
    ("Variaveis Independentes", "Data da publicacao (dummy), Investimento em midia, Temperatura/Clima, Historico 2025"),
    ("Foco", "Explicar os 1.612 acessos extras identificados na segunda-feira"),
]
for i, row_data in enumerate(data, 6):
    for j, val in enumerate(row_data, 1):
        ws5.cell(row=i, column=j, value=val)
style_data_cells(ws5, 6, 9, 2)

# Modelo B
ws5.cell(row=11, column=1, value="B. MODELO DE CLASSIFICACAO (Qualidade de Leads)").font = subtitle_font
ws5.merge_cells(start_row=11, start_column=1, end_row=11, end_column=4)

for col, h in enumerate(headers, 1):
    ws5.cell(row=12, column=col, value=h)
style_header_row(ws5, 12, 2)

data2 = [
    ("Ponto de Partida", "Queda na profundidade de navegacao (2,64 vs 3,07 em 2025) indica trafego de 'topo de funil'"),
    ("Target", "Conversao em lead (Phidelis/CRM)"),
    ("Features Sugeridas", "Duracao media da sessao, Visualizacoes por sessao, Origem do trafego"),
]
for i, row_data in enumerate(data2, 13):
    for j, val in enumerate(row_data, 1):
        ws5.cell(row=i, column=j, value=val)
style_data_cells(ws5, 13, 15, 2)

ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 70

# ========== ABA 7: FONTES E METADADOS ==========
ws6 = wb.create_sheet("6. Fontes e Metadados")

add_title(ws6, 1, "FONTES E METADADOS")

ws6.cell(row=3, column=1, value="FONTES DE DADOS").font = subtitle_font
headers = ["Fonte", "Arquivos", "Descricao"]
for col, h in enumerate(headers, 1):
    ws6.cell(row=4, column=col, value=h)
style_header_row(ws6, 4, 3)

data = [
    ("Google Analytics 4 (GA4)", "download(18-21).csv", "Metricas de sessoes e visualizacoes"),
    ("HubSpot CRM", "hubspot_leads_25.csv", "Leads e negocios Jan/2025"),
    ("HubSpot CRM", "hubspot-core-report-*.csv", "Leads e negocios Jan/2026"),
    ("Excel Consolidado", "Analise de Impacto...xlsx", "Dados consolidados manualmente"),
]
for i, row_data in enumerate(data, 5):
    for j, val in enumerate(row_data, 1):
        ws6.cell(row=i, column=j, value=val)
style_data_cells(ws6, 5, 8, 3)

# Limitacoes
ws6.cell(row=10, column=1, value="LIMITACOES IDENTIFICADAS").font = subtitle_font
ws6.cell(row=11, column=1, value="- A metrica 'Novos Usuarios' nao esta presente nas exportacoes do GA4")
ws6.cell(row=12, column=1, value="- Recomenda-se nova extracao para refinar a taxa de aquisicao liquida")
ws6.cell(row=13, column=1, value="- Dados de conversao CRM precisam de cruzamento com periodo exato")

# Status
ws6.cell(row=15, column=1, value="STATUS FINAL").font = subtitle_font
status_cell = ws6.cell(row=16, column=1, value="DADOS VALIDADOS E PRONTOS PARA EXPORTACAO")
status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
status_cell.font = Font(bold=True, color="006100")

ws6.column_dimensions['A'].width = 30
ws6.column_dimensions['B'].width = 35
ws6.column_dimensions['C'].width = 40

# Salvar arquivo
output_path = r"c:\Users\a483650\Projetos\Analise_performance_valor_the_news\OUTPUTS\Relatorio_Executivo_RedBalloon_Jan2026_v3.xlsx"
wb.save(output_path)
print(f"Relatorio salvo em: {output_path}")
